#!/usr/bin/env python3
"""Prepare k6 credential/token pools from a local XLSX file.

The workbook is never committed. Expected columns:
role, loginName, password, tenantCode, accessToken, enabled

role: STUDENT or TEACHER. Either accessToken or loginName+password is required.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

HEADERS = ("role", "loginName", "password", "tenantCode", "accessToken", "enabled")


def _text(value: Any) -> str:
    return str(value or "").strip()


def create_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "capacity-accounts"
    ws.append(HEADERS)
    ws.append(("STUDENT", "", "", "", "", True))
    ws.append(("TEACHER", "", "", "", "", True))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F3"
    widths = {"A": 14, "B": 24, "C": 20, "D": 22, "E": 48, "F": 12}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    wb.save(path)
    print(f"template_created={path}")


def read_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = tuple(_text(value) for value in next(rows))
    except StopIteration as exc:
        raise SystemExit("XLSX is empty") from exc
    if header[: len(HEADERS)] != HEADERS:
        raise SystemExit(f"Invalid headers: expected {HEADERS}, got {header}")

    result: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for row_no, row in enumerate(rows, start=2):
        values = dict(zip(HEADERS, row, strict=False))
        enabled = _text(values.get("enabled")).lower()
        if enabled in {"false", "0", "no", "否"}:
            continue
        role = _text(values.get("role")).upper()
        if not role and not any(_text(value) for value in row):
            continue
        if role not in {"STUDENT", "TEACHER"}:
            raise SystemExit(f"row {row_no}: role must be STUDENT or TEACHER")
        login_name = _text(values.get("loginName"))
        password = _text(values.get("password"))
        tenant_code = _text(values.get("tenantCode"))
        token = _text(values.get("accessToken"))
        if not token and not (login_name and password):
            raise SystemExit(f"row {row_no}: accessToken or loginName+password is required")
        identity = (role, token or f"{tenant_code}:{login_name}")
        if identity in seen:
            raise SystemExit(f"row {row_no}: duplicate account/token")
        seen.add(identity)
        result.append({
            "role": role,
            "loginName": login_name,
            "password": password,
            "tenantCode": tenant_code,
            "accessToken": token,
        })
    return result


def write_secret_files(rows: list[dict[str, str]], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    counts: dict[str, int] = {}
    for role in ("STUDENT", "TEACHER"):
        role_rows = [row for row in rows if row["role"] == role]
        tokens = [row["accessToken"] for row in role_rows if row["accessToken"]]
        credentials = [
            {
                "loginName": row["loginName"],
                "password": row["password"],
                **({"tenantCode": row["tenantCode"]} if row["tenantCode"] else {}),
            }
            for row in role_rows
            if not row["accessToken"]
        ]
        prefix = role.lower()
        (out_dir / f"{prefix}-tokens.json").write_text(
            json.dumps(tokens, ensure_ascii=False), encoding="utf-8"
        )
        (out_dir / f"{prefix}-credentials.json").write_text(
            json.dumps(credentials, ensure_ascii=False), encoding="utf-8"
        )
        counts[f"{prefix}_tokens"] = len(tokens)
        counts[f"{prefix}_credentials"] = len(credentials)
    (out_dir / ".gitignore").write_text("*\n!.gitignore\n", encoding="utf-8")
    print("prepared=" + json.dumps(counts, ensure_ascii=False, sort_keys=True))
    print("security=output files contain secrets; do not commit or upload as artifacts")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build k6 secret pools from XLSX")
    parser.add_argument("--input", type=Path, help="local capacity accounts XLSX")
    parser.add_argument("--out", type=Path, default=Path("performance/secrets"))
    parser.add_argument("--template", type=Path, help="create an empty XLSX template and exit")
    args = parser.parse_args()
    if args.template:
        create_template(args.template)
        return
    if not args.input:
        parser.error("--input is required unless --template is used")
    if not args.input.exists():
        raise SystemExit(f"Input not found: {args.input}")
    write_secret_files(read_rows(args.input), args.out)


if __name__ == "__main__":
    main()
