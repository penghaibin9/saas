"""师生账号唯一导入入口的 .xlsx 模板、解析、预检批次与错误回执。

本服务刻意只接受标准 xlsx，不接受 csv/xls/xlsm。预检后的结构化数据写入数据库共享批次，
确认时按租户与操作者再次校验，并使用短租约防止多实例重复确认。
"""
from __future__ import annotations

import hashlib
import io
import secrets
import zipfile
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.exceptions import AppException, not_found
from app.db.session import db_enabled, get_sessionmaker
from app.services.saas_role_templates import role_catalog
from app.services.xlsx_util import build_ledger_xlsx, pack_xlsx_result

MAX_FILE_BYTES = 20 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 2000
MAX_ROWS = 2000
BATCH_TTL_SECONDS = 24 * 60 * 60
CONFIRMED_BATCH_TTL_SECONDS = 7 * 24 * 60 * 60
CLAIM_STALE_SECONDS = 5 * 60

HEADERS = (
    "账号类型", "工号/学号", "姓名",
    "所属学院（学生）", "所属专业（学生）", "班级名称（学生）", "年级（学生）", "性别（学生）",
    "所属部门（教师）", "岗位名称（教师）", "预设角色编码（教师）",
    "数据范围类型（教师）", "数据范围引用（教师）",
)
REQUIRED_HEADERS = {"账号类型", "工号/学号", "姓名"}

# ── 学生 / 教师拆分模板（学生导入与账号开通、教师导入两个独立入口）──────────
# 拆分理由：混合模板靠「账号类型」列区分，学生行要跳过教师列、教师行要跳过学生列，
# 学校填表时极易串列；且两类导入的权限、结果统计、后续流程完全不同。
# 两套模板共用本文件的归档校验、行数上限、公式注入防护、批次与回执能力，不另造框架。
STUDENT_HEADERS = ("学号", "姓名", "所属学院", "所属专业", "班级名称", "年级", "性别", "身份证号")
STUDENT_REQUIRED_HEADERS = {"学号", "姓名", "班级名称"}

TEACHER_HEADERS = ("工号", "姓名", "所属部门", "岗位名称", "预设角色编码",
                   "数据范围类型", "数据范围引用")
TEACHER_REQUIRED_HEADERS = {"工号", "姓名", "预设角色编码"}
RELATION_HEADERS = ("关系类型", "主体工号", "对象编号/学号", "业务批次编号", "备注")
RELATION_REQUIRED_HEADERS = {"关系类型", "主体工号", "对象编号/学号"}
RELATION_TYPES = {
    "COUNSELOR_CLASS": "辅导员—班级",
    "HEAD_TEACHER_CLASS": "班主任—班级",
    "TEACHER_TEACHING_TASK": "任课教师—教学班",
    "GRADUATION_MENTOR_STUDENT": "毕设导师—学生",
    "INTERNSHIP_ADVISOR_STUDENT": "实习指导教师—学生",
    "DORM_MANAGER_BUILDING": "宿管—楼栋",
}
_RELATION_TYPE_ALIASES = {
    **{code: code for code in RELATION_TYPES},
    **{name: code for code, name in RELATION_TYPES.items()},
}
_TYPE_ALIASES = {
    "STUDENT": "STUDENT", "学生": "STUDENT",
    "TEACHER": "TEACHER", "教师": "TEACHER", "老师": "TEACHER",
}


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value) -> str:
    return _cell_text(value).rstrip(" *").strip()


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ARCHIVE_ENTRIES:
                raise AppException("FILE_TOO_LARGE", "Excel 内部文件数量异常，请使用系统标准模板")
            if sum(info.file_size for info in infos) > MAX_UNCOMPRESSED_BYTES:
                raise AppException("FILE_TOO_LARGE", "Excel 解压后超过 100MB 上限，请拆分后重试")
            names = {info.filename.lower() for info in infos}
            if any(name.endswith("vbaproject.bin") for name in names):
                raise AppException("FILE_TYPE_NOT_ALLOWED", "师生账号导入禁止包含宏代码")
    except AppException:
        raise
    except (zipfile.BadZipFile, OSError) as exc:
        raise AppException("FILE_TYPE_NOT_ALLOWED", "文件不是有效的标准 .xlsx，请重新下载模板") from exc


def _user_key(user: dict) -> str:
    return str(user.get("userId") or user.get("sub") or user.get("loginName") or "")


def _require_database() -> None:
    if not db_enabled():
        raise AppException(
            "SERVER_ERROR",
            "共享导入批次需要启用数据库；当前环境不能安全保存预检结果",
        )


def _tenant_number(value: object) -> int:
    try:
        return int(str(value or "0"))
    except (TypeError, ValueError) as exc:
        raise AppException("VALIDATION_ERROR", "当前学校标识无效，无法创建导入批次") from exc


def _batch_entry(row) -> dict:
    return {
        "batchNo": row.batch_no,
        "tenantId": str(row.tenant_id),
        "userKey": row.operator_key,
        "payload": row.payload_json or {},
        "rawRows": row.raw_rows_json or [],
        "errors": row.errors_json or [],
        "report": row.report_json or {},
        "relationships": row.relationships_json or [],
        "relationErrors": row.relation_errors_json or [],
        "preErrors": row.pre_errors_json or [],
        "fileName": row.file_name,
        "fileSha256": row.file_sha256,
        "status": row.status,
        "identityConfirmed": row.status == "IDENTITY_CONFIRMED",
        "publicResult": row.public_result_json or {},
        "expiresAt": row.expires_at.isoformat(timespec="seconds") if row.expires_at else None,
    }


def _owned_row(db, user: dict, tenant_id: object, batch_no: str, *, lock: bool = False):
    from sqlalchemy import select
    from app.models import IdentityImportBatch

    stmt = select(IdentityImportBatch).where(
        IdentityImportBatch.tenant_id == _tenant_number(tenant_id),
        IdentityImportBatch.batch_no == str(batch_no or ""),
        IdentityImportBatch.operator_key == _user_key(user),
        IdentityImportBatch.is_deleted.is_(False),
    )
    if lock:
        stmt = stmt.with_for_update()
    row = db.scalar(stmt)
    if not row or row.expires_at <= datetime.utcnow() or row.status == "EXPIRED":
        if row and row.status != "EXPIRED":
            row.status = "EXPIRED"
            row.claim_token = None
            row.claim_started_at = None
            db.commit()
        raise not_found("导入批次不存在或已过期，请重新上传预检")
    return row


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append([f"{h} *" if h in REQUIRED_HEADERS else h for h in HEADERS])
    fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    widths = [16, 20, 16, 24, 24, 24, 18, 16, 24, 24, 30, 26, 28]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width
    ws.freeze_panes = "A2"
    ws.column_dimensions["B"].number_format = "@"
    type_validation = DataValidation(type="list", formula1='"STUDENT,TEACHER"', allow_blank=False)
    ws.add_data_validation(type_validation)
    type_validation.add(f"A2:A{MAX_ROWS + 1}")

    notes = wb.create_sheet("填写说明")
    instructions = [
        "只允许上传本模板生成的 .xlsx 文件，不支持 CSV、旧版 .xls 或启用宏的 .xlsm。",
        "账号类型只能填写 STUDENT（学生）或 TEACHER（教师）。",
        "学生账号固定绑定 STUDENT 角色，不填写预设角色编码。",
        "学生尽量填写学院、专业、班级和年级；实施中心会先生成组织树候选，存在歧义时必须人工确认。",
        "教师可填写所属部门和岗位名称；岗位只用于推荐预设角色，最终角色仍需学校确认。",
        "教师必须填写预设角色编码；多个角色用中文/英文逗号、分号或竖线分隔。",
        "工号/学号请设置为文本，避免前导零丢失；同一学校内必须唯一。",
        "辅导员必须填写 CLASS 或 ADVISOR 数据范围及对应引用。",
        "预检通过后才能整批确认；任何错误都会阻止整批写入。",
    ]
    notes.append(["填写说明"])
    notes["A1"].font = Font(bold=True, size=12)
    for index, item in enumerate(instructions, 2):
        notes.cell(row=index, column=1, value=item)
    notes.column_dimensions["A"].width = 100

    samples = wb.create_sheet("填写示例（不要导入）")
    samples.append(list(HEADERS))
    samples.append(["STUDENT", "20260001", "张同学", "信息工程学院", "软件技术", "软件2601", "2026", "男", "", "", "", "", ""])
    samples.append(["TEACHER", "T2026001", "李老师", "", "", "", "", "", "教务处", "任课教师", "ACADEMIC_TEACHER", "", ""])
    for cell in samples[1]:
        cell.font = Font(bold=True)
        cell.fill = fill

    roles = wb.create_sheet("教师预设角色")
    roles.append(["角色编码", "角色名称", "默认数据范围", "分类"])
    for item in role_catalog(teacher_only=True)["items"]:
        roles.append([item["roleCode"], item["roleName"], item["defaultScope"], item["category"]])
    for cell in roles[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for col, width in zip(("A", "B", "C", "D"), (30, 24, 26, 22)):
        roles.column_dimensions[col].width = width
    roles.freeze_panes = "A2"

    relations = wb.create_sheet("业务关系")
    relations.append([f"{h} *" if h in RELATION_REQUIRED_HEADERS else h for h in RELATION_HEADERS])
    for cell in relations[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for index, width in enumerate((34, 22, 28, 24, 40), 1):
        relations.column_dimensions[relations.cell(row=1, column=index).column_letter].width = width
    relations.freeze_panes = "A2"
    relation_validation = DataValidation(
        type="list", formula1='"' + ",".join(RELATION_TYPES) + '"', allow_blank=False)
    relations.add_data_validation(relation_validation)
    relation_validation.add(f"A2:A{MAX_ROWS + 1}")

    relation_notes = wb.create_sheet("业务关系说明")
    relation_notes.append(["关系类型", "写入的真实业务主表", "对象编号/学号填写方式", "业务批次编号"])
    relation_rows = (
        ("COUNSELOR_CLASS", "班级主表 counselor_id", "班级编码；无编码时可填唯一班级名称", "留空"),
        ("HEAD_TEACHER_CLASS", "班级主表 head_teacher_id", "班级编码；无编码时可填唯一班级名称", "留空"),
        ("TEACHER_TEACHING_TASK", "教学任务主表 teacher_id/teacher_key", "教学班编码", "留空"),
        ("GRADUATION_MENTOR_STUDENT", "毕设学生及导师分配历史表", "学号", "同一学生多批次时必填毕设批次编号"),
        ("INTERNSHIP_ADVISOR_STUDENT", "实习记录主表 advisor_user_id", "学号", "同一学生多批次时必填实习批次编号"),
        ("DORM_MANAGER_BUILDING", "宿舍楼栋主表 manager_teacher_key", "楼栋编码；无编码时可填唯一楼栋名称", "留空"),
    )
    for row in relation_rows:
        relation_notes.append(row)
    for cell in relation_notes[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for col, width in zip(("A", "B", "C", "D"), (34, 38, 44, 38)):
        relation_notes.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _style_header(ws, headers, required, fill, widths) -> None:
    ws.append([f"{h} *" if h in required else h for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width
    ws.freeze_panes = "A2"


def _append_notes(wb, title: str, instructions: list) -> None:
    notes = wb.create_sheet(title)
    notes.append([title])
    notes["A1"].font = Font(bold=True, size=12)
    for index, item in enumerate(instructions, 2):
        notes.cell(row=index, column=1, value=item)
    notes.column_dimensions["A"].width = 100


def build_student_template() -> bytes:
    """学生导入模板：只含学生字段，不出现任何教师列（角色/部门/岗位/数据范围）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    fill = PatternFill("solid", fgColor="DCE6F1")
    _style_header(ws, STUDENT_HEADERS, STUDENT_REQUIRED_HEADERS, fill,
                  [20, 16, 24, 24, 24, 12, 10, 24])
    ws.column_dimensions["A"].number_format = "@"  # 学号按文本，避免前导零丢失
    ws.column_dimensions["H"].number_format = "@"
    gender = DataValidation(type="list", formula1='"男,女"', allow_blank=True)
    ws.add_data_validation(gender)
    gender.add(f"G2:G{MAX_ROWS + 1}")

    _append_notes(wb, "填写说明", [
        "只允许上传本模板生成的 .xlsx 文件，不支持 CSV、旧版 .xls 或启用宏的 .xlsm。",
        "本模板只导入学生。教师请使用「教师导入」的专用模板，两者不可混用。",
        "学生角色由系统固定为 STUDENT，无需也无法在此指定角色。",
        "班级名称必填：学生必须归属完整的学院、专业、班级，系统会按班级自动补全专业与学院。",
        "若校内存在同名班级，请同时填写所属学院与所属专业以便唯一定位。",
        "学号请设置为文本格式，学校内必须唯一；学号一经建立在本校内永久唯一，作废后不可另建新档。",
        "已存在的学生会自动复用原档案并补齐空缺信息，不会重复建档；院系班变更请走「学籍异动」。",
        "身份证号选填；填写后用于识别「同一人两个学号」等异常，不会明文展示。",
        "导入将同时创建登录账号并生成一次性初始密码，请在结果页及时下载凭据回执。",
        "预检通过后才能整批确认；任何错误都会阻止整批写入。",
    ])

    samples = wb.create_sheet("填写示例（不要导入）")
    samples.append(list(STUDENT_HEADERS))
    samples.append(["20260001", "张同学", "信息工程学院", "软件技术", "软件2601", "2026", "男", ""])
    for cell in samples[1]:
        cell.font = Font(bold=True)
        cell.fill = fill

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


def build_teacher_template() -> bytes:
    """教师导入模板：只含教职工字段，不出现学生学院/专业/班级/年级/学籍状态。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    fill = PatternFill("solid", fgColor="DCE6F1")
    _style_header(ws, TEACHER_HEADERS, TEACHER_REQUIRED_HEADERS, fill,
                  [20, 16, 24, 20, 30, 22, 28])
    ws.column_dimensions["A"].number_format = "@"
    scope = DataValidation(type="list", formula1='"SCHOOL,COLLEGE,CLASS,ADVISOR"', allow_blank=True)
    ws.add_data_validation(scope)
    scope.add(f"F2:F{MAX_ROWS + 1}")

    _append_notes(wb, "填写说明", [
        "只允许上传本模板生成的 .xlsx 文件，不支持 CSV、旧版 .xls 或启用宏的 .xlsm。",
        "本模板只导入教师。学生请使用「学生导入与账号开通」的专用模板，两者不可混用。",
        "预设角色编码必填；多个角色用中文/英文逗号、分号或竖线分隔，可选值见「教师预设角色」页。",
        "辅导员必须填写 CLASS 或 ADVISOR 数据范围类型及对应的数据范围引用（班级名称）。",
        "工号请设置为文本格式，学校内必须唯一；工号被学生或其它账号占用时会整批阻断。",
        "导入将创建登录账号并生成一次性初始密码，请在结果页及时下载凭据回执。",
        "预检通过后才能整批确认；任何错误都会阻止整批写入。",
    ])

    samples = wb.create_sheet("填写示例（不要导入）")
    samples.append(list(TEACHER_HEADERS))
    samples.append(["T2026001", "李老师", "教务处", "任课教师", "ACADEMIC_TEACHER", "", ""])
    samples.append(["T2026002", "王老师", "学生工作处", "辅导员", "COUNSELOR", "CLASS", "软件2601"])
    for cell in samples[1]:
        cell.font = Font(bold=True)
        cell.fill = fill

    roles = wb.create_sheet("教师预设角色")
    roles.append(["角色编码", "角色名称", "默认数据范围", "分类"])
    for item in role_catalog(teacher_only=True)["items"]:
        roles.append([item["roleCode"], item["roleName"], item["defaultScope"], item["category"]])
    for cell in roles[1]:
        cell.font = Font(bold=True)
        cell.fill = fill

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


def _open_single_sheet(content: bytes, filename: str, headers: tuple, required: set,
                       what: str) -> tuple:
    """共用的 xlsx 打开与表头校验（归档校验、大小与公式防护与混合模板一致）。"""
    if not str(filename or "").lower().endswith(".xlsx"):
        raise AppException("FILE_TYPE_NOT_ALLOWED", f"{what}只支持标准 .xlsx 文件")
    if not content:
        raise AppException("VALIDATION_ERROR", "上传文件为空")
    if len(content) > MAX_FILE_BYTES:
        raise AppException("FILE_TOO_LARGE", "导入文件超过 20MB 上限，请拆分后重试")
    _validate_xlsx_archive(content)
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_links=False)
        ws = wb["导入模板"] if "导入模板" in wb.sheetnames else wb.worksheets[0]
        iterator = ws.iter_rows(values_only=True)
        raw_headers = next(iterator)
    except (StopIteration, KeyError):
        raise AppException("VALIDATION_ERROR", "Excel 没有可导入的工作表或表头")
    except Exception as exc:  # noqa: BLE001
        raise AppException("FILE_TYPE_NOT_ALLOWED", "文件不是有效的标准 .xlsx，请重新下载模板") from exc

    parsed_headers = [_normalize_header(v) for v in raw_headers]
    dup = sorted({h for h in parsed_headers if h and parsed_headers.count(h) > 1})
    if dup:
        wb.close()
        raise AppException("VALIDATION_ERROR", f"Excel 表头重复：{','.join(dup)}")
    missing = [h for h in required if h not in parsed_headers]
    unknown = [h for h in parsed_headers if h and h not in headers]
    if missing or unknown:
        wb.close()
        parts = []
        if missing:
            parts.append(f"缺少表头：{','.join(missing)}")
        if unknown:
            # 串用模板是最常见的误操作，直接点破
            parts.append(f"不支持的表头：{','.join(unknown)}（请确认没有把另一类模板传到这里）")
        raise AppException("VALIDATION_ERROR", "；".join(parts) + "。请使用系统下载的最新版模板")
    return wb, iterator, parsed_headers


def _row_cells(values, headers, header_index, row_no, errors, entity):
    """取一行并做公式注入防护；返回 (cells, 是否空行)。"""
    cells = {}
    for name in headers:
        index = header_index.get(name)
        value = values[index] if index is not None and index < len(values) else ""
        if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
            errors.append({"row": row_no, "entity": entity, "field": name,
                           "error": "单元格禁止公式或可执行前缀，请改为纯文本"})
            value = ""
        cells[name] = _cell_text(value)
    return cells, not any(cells.values())


def parse_student_xlsx(content: bytes, filename: str) -> dict:
    """解析学生专用模板。输出结构与 parse_xlsx 一致，下游批次/预检/确认完全复用。"""
    wb, iterator, headers = _open_single_sheet(
        content, filename, STUDENT_HEADERS, STUDENT_REQUIRED_HEADERS, "学生导入")
    header_index = {n: headers.index(n) for n in STUDENT_HEADERS if n in headers}
    students, raw_rows, errors = [], [], []
    total = 0
    for row_no, values in enumerate(iterator, 2):
        cells, empty = _row_cells(values, STUDENT_HEADERS, header_index, row_no, errors, "student")
        if empty:
            continue
        total += 1
        if total > MAX_ROWS:
            wb.close()
            raise AppException("VALIDATION_ERROR", f"单次最多导入 {MAX_ROWS} 行，请拆分文件")
        no, name = cells["学号"], cells["姓名"]
        raw_rows.append({"row": row_no, "accountType": "STUDENT", "accountNo": no, "name": name})
        if not no:
            errors.append({"row": row_no, "entity": "student", "field": "学号", "error": "学号必填"})
        if not name:
            errors.append({"row": row_no, "entity": "student", "field": "姓名", "error": "姓名必填"})
        if not cells["班级名称"]:
            errors.append({"row": row_no, "entity": "student", "field": "班级名称",
                           "error": "班级必填：学生必须归属完整的学院、专业、班级"})
        students.append({
            "_rowNo": row_no, "studentNo": no, "name": name,
            "collegeName": cells["所属学院"], "majorName": cells["所属专业"],
            "className": cells["班级名称"], "grade": cells["年级"],
            "gender": cells["性别"], "idCard": cells["身份证号"],
        })
    wb.close()
    if total == 0:
        raise AppException("VALIDATION_ERROR", "Excel 没有数据行，请填写后再上传")
    return {"students": students, "teachers": [], "rawRows": raw_rows,
            "relationships": [], "relationErrors": [], "errors": errors, "totalRows": total,
            "importKind": "STUDENT",
            "fileName": filename, "fileSha256": hashlib.sha256(content).hexdigest()}


def parse_teacher_xlsx(content: bytes, filename: str) -> dict:
    """解析教师专用模板。角色、数据范围、辅导员班级范围等既有能力全部保留。"""
    wb, iterator, headers = _open_single_sheet(
        content, filename, TEACHER_HEADERS, TEACHER_REQUIRED_HEADERS, "教师导入")
    header_index = {n: headers.index(n) for n in TEACHER_HEADERS if n in headers}
    teachers, raw_rows, errors = [], [], []
    total = 0
    for row_no, values in enumerate(iterator, 2):
        cells, empty = _row_cells(values, TEACHER_HEADERS, header_index, row_no, errors, "teacher")
        if empty:
            continue
        total += 1
        if total > MAX_ROWS:
            wb.close()
            raise AppException("VALIDATION_ERROR", f"单次最多导入 {MAX_ROWS} 行，请拆分文件")
        no, name = cells["工号"], cells["姓名"]
        raw_rows.append({"row": row_no, "accountType": "TEACHER", "accountNo": no, "name": name})
        if not no:
            errors.append({"row": row_no, "entity": "teacher", "field": "工号", "error": "工号必填"})
        if not name:
            errors.append({"row": row_no, "entity": "teacher", "field": "姓名", "error": "姓名必填"})
        if not cells["预设角色编码"]:
            errors.append({"row": row_no, "entity": "teacher", "field": "预设角色编码",
                           "error": "教师必须指定预设角色编码"})
        teachers.append({
            "_rowNo": row_no, "loginName": no, "name": name,
            "departmentName": cells["所属部门"], "positionName": cells["岗位名称"],
            "roleCodes": cells["预设角色编码"],
            "scopeType": cells["数据范围类型"], "scopeRef": cells["数据范围引用"],
        })
    wb.close()
    if total == 0:
        raise AppException("VALIDATION_ERROR", "Excel 没有数据行，请填写后再上传")
    return {"students": [], "teachers": teachers, "rawRows": raw_rows,
            "relationships": [], "relationErrors": [], "errors": errors, "totalRows": total,
            "importKind": "TEACHER",
            "fileName": filename, "fileSha256": hashlib.sha256(content).hexdigest()}


def parse_xlsx(content: bytes, filename: str) -> dict:
    if not str(filename or "").lower().endswith(".xlsx"):
        raise AppException("FILE_TYPE_NOT_ALLOWED", "师生账号导入只支持标准 .xlsx 文件")
    if not content:
        raise AppException("VALIDATION_ERROR", "上传文件为空")
    if len(content) > MAX_FILE_BYTES:
        raise AppException("FILE_TOO_LARGE", "导入文件超过 20MB 上限，请拆分后重试")
    _validate_xlsx_archive(content)
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_links=False)
        ws = wb["导入模板"] if "导入模板" in wb.sheetnames else wb.worksheets[0]
        iterator = ws.iter_rows(values_only=True)
        raw_headers = next(iterator)
    except (StopIteration, KeyError):
        raise AppException("VALIDATION_ERROR", "Excel 没有可导入的工作表或表头")
    except Exception as exc:  # noqa: BLE001
        raise AppException("FILE_TYPE_NOT_ALLOWED", "文件不是有效的标准 .xlsx，请重新下载模板") from exc

    headers = [_normalize_header(value) for value in raw_headers]
    duplicate_headers = sorted({h for h in headers if h and headers.count(h) > 1})
    if duplicate_headers:
        wb.close()
        raise AppException("VALIDATION_ERROR", f"Excel 表头重复：{','.join(duplicate_headers)}")
    # 新增的学院/专业/部门/岗位均为可选列，旧版标准模板继续兼容；三项身份主键仍必需。
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    unknown = [h for h in headers if h and h not in HEADERS]
    if missing or unknown:
        wb.close()
        parts = []
        if missing:
            parts.append(f"缺少表头：{','.join(missing)}")
        if unknown:
            parts.append(f"不支持的表头：{','.join(unknown)}")
        raise AppException("VALIDATION_ERROR", "；".join(parts) + "。请使用系统下载的最新版模板")

    header_index = {name: headers.index(name) for name in HEADERS if name in headers}
    students, teachers, raw_rows, errors = [], [], [], []
    total = 0
    for row_no, values in enumerate(iterator, 2):
        cells = {}
        formula_fields = []
        for name in HEADERS:
            index = header_index.get(name)
            value = values[index] if index is not None and index < len(values) else ""
            if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
                formula_fields.append(name)
                value = ""
            cells[name] = _cell_text(value)
        if not any(cells.values()):
            continue
        total += 1
        if total > MAX_ROWS:
            wb.close()
            raise AppException("VALIDATION_ERROR", f"单次最多导入 {MAX_ROWS} 行，请拆分文件")
        account_type = _TYPE_ALIASES.get(cells["账号类型"].upper())
        account_no = cells["工号/学号"]
        name = cells["姓名"]
        raw_rows.append({"row": row_no, "accountType": cells["账号类型"],
                         "accountNo": account_no, "name": name})
        for field in formula_fields:
            errors.append({"row": row_no, "entity": "file", "field": field,
                           "error": "单元格禁止公式或可执行前缀，请改为纯文本"})
        if not account_type:
            errors.append({"row": row_no, "entity": "file", "field": "账号类型",
                           "error": "账号类型只能填写 STUDENT 或 TEACHER"})
            continue
        # 旧版 9 列模板兼容：把历史位置映射到当前 13 列模型。
        row_values = [cells[name] for name in HEADERS]
        if account_type == "STUDENT":
            if (not row_values[7] and row_values[6] in {"男", "女", "未知"}
                    and row_values[5].isdigit() and row_values[4]):
                row_values[5], row_values[6], row_values[7] = row_values[4], row_values[5], row_values[6]
                row_values[4] = ""
            if row_values[3] and row_values[3].isupper() and "_" in row_values[3]:
                row_values[10] = row_values[3]
                row_values[3] = ""
        elif row_values[7].upper() in {"CLASS", "COLLEGE", "SCHOOL"} and row_values[3]:
            row_values[10], row_values[11], row_values[12] = row_values[3], row_values[7], row_values[8]
            row_values[3], row_values[7], row_values[8] = "", "", ""
        for index, header in enumerate(HEADERS):
            cells[header] = row_values[index]
        if account_type == "STUDENT":
            role = cells["预设角色编码（教师）"]
            if role and role.upper() != "STUDENT":
                errors.append({"row": row_no, "entity": "student", "field": "预设角色编码（教师）",
                               "error": "学生角色由系统固定为 STUDENT，请留空"})
            students.append({
                "_rowNo": row_no, "studentNo": account_no, "name": name,
                "collegeName": cells["所属学院（学生）"], "majorName": cells["所属专业（学生）"],
                "className": cells["班级名称（学生）"], "grade": cells["年级（学生）"],
                "gender": cells["性别（学生）"],
            })
        else:
            teachers.append({
                "_rowNo": row_no, "loginName": account_no, "name": name,
                "departmentName": cells["所属部门（教师）"], "positionName": cells["岗位名称（教师）"],
                "roleCodes": cells["预设角色编码（教师）"],
                "scopeType": cells["数据范围类型（教师）"],
                "scopeRef": cells["数据范围引用（教师）"],
            })
    relationships, relation_errors = [], []
    if "业务关系" in wb.sheetnames:
        relation_ws = wb["业务关系"]
        relation_iterator = relation_ws.iter_rows(values_only=True)
        try:
            relation_raw_headers = next(relation_iterator)
        except StopIteration:
            relation_raw_headers = ()
        relation_headers = [_normalize_header(value) for value in relation_raw_headers]
        relation_missing = [h for h in RELATION_REQUIRED_HEADERS if h not in relation_headers]
        relation_unknown = [h for h in relation_headers if h and h not in RELATION_HEADERS]
        if relation_missing or relation_unknown:
            relation_errors.append({"row": 1, "field": "表头", "error":
                "业务关系表头不正确：" + (f"缺少 {','.join(relation_missing)}；" if relation_missing else "")
                + (f"不支持 {','.join(relation_unknown)}" if relation_unknown else "")})
        else:
            relation_index = {name: relation_headers.index(name) for name in RELATION_HEADERS}
            for row_no, values in enumerate(relation_iterator, 2):
                cells, formula_fields = {}, []
                for name in RELATION_HEADERS:
                    index = relation_index[name]
                    value = values[index] if index < len(values) else ""
                    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
                        formula_fields.append(name); value = ""
                    cells[name] = _cell_text(value)
                if not any(cells.values()):
                    continue
                relation_type = _RELATION_TYPE_ALIASES.get(cells["关系类型"].upper()) \
                    or _RELATION_TYPE_ALIASES.get(cells["关系类型"])
                if not relation_type:
                    relation_errors.append({"row": row_no, "field": "关系类型",
                                            "error": f"不支持的关系类型：{cells['关系类型']}"})
                for field in formula_fields:
                    relation_errors.append({"row": row_no, "field": field,
                                            "error": "单元格禁止公式或可执行前缀，请改为纯文本"})
                if not cells["主体工号"]:
                    relation_errors.append({"row": row_no, "field": "主体工号", "error": "主体工号必填"})
                if not cells["对象编号/学号"]:
                    relation_errors.append({"row": row_no, "field": "对象编号/学号", "error": "对象编号/学号必填"})
                relationships.append({
                    "_rowNo": row_no, "relationType": relation_type or cells["关系类型"],
                    "subjectRef": cells["主体工号"], "objectRef": cells["对象编号/学号"],
                    "contextRef": cells["业务批次编号"], "remark": cells["备注"],
                })
    wb.close()
    if total == 0:
        raise AppException("VALIDATION_ERROR", "Excel 没有数据行，请填写后再上传")
    return {
        "students": students, "teachers": teachers, "rawRows": raw_rows,
        "relationships": relationships, "relationErrors": relation_errors,
        "errors": errors, "totalRows": total,
        "fileName": filename, "fileSha256": hashlib.sha256(content).hexdigest(),
    }


def _relation_suggestion_count(parsed: dict) -> int:
    keys = {(str(x.get("relationType") or ""), str(x.get("subjectRef") or ""),
             str(x.get("objectRef") or "")) for x in parsed.get("relationships") or []}
    for teacher in parsed.get("teachers") or []:
        roles = str(teacher.get("roleCodes") or "").upper()
        scope_type = str(teacher.get("scopeType") or "").upper()
        relation_type = ("COUNSELOR_CLASS" if "COUNSELOR" in roles and scope_type == "CLASS" else
                         "DORM_MANAGER_BUILDING" if "DORM_MANAGER" in roles and scope_type == "DORM_BUILDING" else "")
        if relation_type and teacher.get("scopeRef"):
            keys.add((relation_type, str(teacher.get("loginName") or ""), str(teacher.get("scopeRef") or "")))
    return len(keys)


def create_batch(user: dict, parsed: dict, report: dict) -> dict:
    _require_database()
    from app.models import IdentityImportBatch

    batch_no = f"IDIMP{datetime.now():%Y%m%d%H%M%S}{secrets.token_hex(3).upper()}"
    errors = list(report.get("errors") or [])
    invalid_rows = {int(item.get("row") or 0) for item in errors if int(item.get("row") or 0) > 0}
    total = int(parsed["totalRows"])
    invalid_count = len(invalid_rows) if invalid_rows else (1 if errors else 0)
    payload = {
            "tenantId": report.get("tenantId"), "students": parsed["students"],
            "teachers": parsed["teachers"], "atomic": True,
    }
    session = get_sessionmaker()()
    try:
        session.add(IdentityImportBatch(
            tenant_id=_tenant_number(report.get("tenantId")),
            batch_no=batch_no,
            operator_key=_user_key(user),
            file_name=parsed["fileName"],
            file_sha256=parsed["fileSha256"],
            status="VALIDATED",
            payload_json=payload,
            raw_rows_json=parsed["rawRows"],
            errors_json=errors,
            pre_errors_json=list(parsed.get("errors") or []),
            report_json=report,
            relationships_json=parsed.get("relationships") or [],
            relation_errors_json=parsed.get("relationErrors") or [],
            expires_at=datetime.utcnow() + timedelta(seconds=BATCH_TTL_SECONDS),
        ))
        session.commit()
    finally:
        session.close()
    return {
        "batchNo": batch_no, "fileName": parsed["fileName"],
        "total": total, "valid": max(total - invalid_count, 0),
        "invalid": invalid_count, "errors": [
            {"row": item.get("row", 0), "entity": item.get("entity", ""),
             "field": item.get("field", ""), "message": item.get("error") or item.get("message") or "校验失败"}
            for item in errors
        ],
        "roleTemplateVersion": report.get("roleTemplateVersion"),
        "entities": report.get("entities") or {},
        "relations": {"total": len(parsed.get("relationships") or []),
                      "suggested": _relation_suggestion_count(parsed),
                      "invalid": len({int(x.get('row') or 0) for x in parsed.get('relationErrors') or []}),
                      "errors": parsed.get("relationErrors") or []},
        "expiresInMinutes": BATCH_TTL_SECONDS // 60,
    }


def get_batch(user: dict, tenant_id: object, batch_no: str, *, require_valid: bool = False) -> dict:
    _require_database()
    session = get_sessionmaker()()
    try:
        row = _owned_row(session, user, tenant_id, batch_no)
        entry = _batch_entry(row)
        if require_valid and entry["errors"]:
            raise AppException("VALIDATION_ERROR", "该批次存在错误，禁止确认导入")
        return entry
    finally:
        session.close()


def claim_batch(user: dict, tenant_id: object, batch_no: str) -> tuple[dict, str | None, bool]:
    """Atomically acquire a confirmation lease shared by all backend instances."""
    _require_database()
    session = get_sessionmaker()()
    try:
        row = _owned_row(session, user, tenant_id, batch_no, lock=True)
        if row.errors_json:
            raise AppException("VALIDATION_ERROR", "该批次存在错误，禁止确认导入")
        if row.status == "IDENTITY_CONFIRMED":
            return _batch_entry(row), None, True
        now = datetime.utcnow()
        if row.status == "CONFIRMING" and row.claim_started_at \
                and row.claim_started_at > now - timedelta(seconds=CLAIM_STALE_SECONDS):
            raise AppException("DATA_CONFLICT", "该导入批次正在另一服务实例确认，请稍后查看结果")
        if row.status not in ("VALIDATED", "CONFIRMING"):
            raise AppException("DATA_CONFLICT", f"当前批次状态 {row.status} 不允许确认")
        token = secrets.token_hex(24)
        row.status = "CONFIRMING"
        row.claim_token = token
        row.claim_started_at = now
        row.last_error = None
        row.version = int(row.version or 0) + 1
        session.commit()
        return _batch_entry(row), token, False
    finally:
        session.close()


def mark_confirmed(user: dict, tenant_id: object, batch_no: str, claim_token: str,
                   public_result: dict) -> None:
    _require_database()
    session = get_sessionmaker()()
    try:
        row = _owned_row(session, user, tenant_id, batch_no, lock=True)
        if row.status == "IDENTITY_CONFIRMED":
            return
        if row.status != "CONFIRMING" or row.claim_token != claim_token:
            raise AppException("DATA_CONFLICT", "导入确认租约已失效，请刷新后重试")
        now = datetime.utcnow()
        row.status = "IDENTITY_CONFIRMED"
        row.confirmed_at = now
        row.public_result_json = public_result
        row.claim_token = None
        row.claim_started_at = None
        row.last_error = None
        row.expires_at = now + timedelta(seconds=CONFIRMED_BATCH_TTL_SECONDS)
        row.version = int(row.version or 0) + 1
        session.commit()
    finally:
        session.close()


def release_claim(user: dict, tenant_id: object, batch_no: str, claim_token: str,
                  error: str = "") -> None:
    """Release only the caller's lease so another instance can safely retry."""
    _require_database()
    session = get_sessionmaker()()
    try:
        row = _owned_row(session, user, tenant_id, batch_no, lock=True)
        if row.status == "CONFIRMING" and row.claim_token == claim_token:
            row.status = "VALIDATED"
            row.claim_token = None
            row.claim_started_at = None
            row.last_error = str(error or "")[:2000] or None
            row.version = int(row.version or 0) + 1
            session.commit()
    finally:
        session.close()


def refresh_batch_report(user: dict, tenant_id: object, batch_no: str, report: dict) -> dict:
    """组织/角色候选安装后持久化重新校验结果，任一实例都能继续确认。"""
    _require_database()
    session = get_sessionmaker()()
    try:
        row = _owned_row(session, user, tenant_id, batch_no, lock=True)
        if row.status != "VALIDATED":
            raise AppException("DATA_CONFLICT", "该批次已进入确认阶段，不能再刷新预检结果")
        errors = list(report.get("errors") or [])
        row.report_json = report
        row.errors_json = errors
        row.expires_at = datetime.utcnow() + timedelta(seconds=BATCH_TTL_SECONDS)
        row.version = int(row.version or 0) + 1
        total = len(row.raw_rows_json or [])
        session.commit()
        invalid_rows = {int(item.get("row") or 0) for item in errors if int(item.get("row") or 0) > 0}
        return {
        "batchNo": row.batch_no, "fileName": row.file_name, "total": total,
        "valid": max(total - len(invalid_rows), 0), "invalid": len(invalid_rows),
        "errors": [{"row": item.get("row", 0), "entity": item.get("entity", ""),
                    "field": item.get("field", ""),
                    "message": item.get("error") or item.get("message") or "校验失败"}
                   for item in errors],
        "roleTemplateVersion": report.get("roleTemplateVersion"),
        "entities": report.get("entities") or {}, "expiresInMinutes": BATCH_TTL_SECONDS // 60,
        }
    finally:
        session.close()


def build_error_workbook(entry: dict) -> bytes:
    raw = {int(row["row"]): row for row in entry.get("rawRows") or []}
    rows = []
    for item in entry.get("errors") or []:
        row_no = int(item.get("row") or 0)
        source = raw.get(row_no, {})
        rows.append([
            row_no or "全局", source.get("accountType", ""), source.get("accountNo", ""),
            source.get("name", ""), item.get("entity", ""), item.get("field", ""),
            item.get("error") or item.get("message") or "校验失败",
        ])
    return build_ledger_xlsx(
        "师生账号导入错误", ["Excel行号", "账号类型", "工号/学号", "姓名", "对象", "错误字段", "错误原因"],
        rows, watermark="本文件仅含预检失败行；修正原导入模板后重新上传，不会自动创建账号。")


def build_credential_receipt(entry: dict, report: dict) -> dict | None:
    """把仅本次返回的初始密码封装为 xlsx 回执；调用方不得记录返回内容。"""
    student_names = {str(row.get("studentNo")): str(row.get("name") or "")
                     for row in entry["payload"].get("students") or []}
    teacher_names = {str(row.get("loginName")): str(row.get("name") or "")
                     for row in entry["payload"].get("teachers") or []}
    credential_rows = [
        ["STUDENT", item["studentNo"], student_names.get(str(item["studentNo"]), ""),
         item["initialPassword"], "首次登录必须修改密码"]
        for item in report.get("studentCredentials") or []
    ] + [
        ["TEACHER", item["loginName"], teacher_names.get(str(item["loginName"]), ""),
         item["initialPassword"], "首次登录必须修改密码"]
        for item in report.get("teacherCredentials") or []
    ]
    if not credential_rows:
        return None
    return pack_xlsx_result(
        build_ledger_xlsx(
            "初始账号凭据", ["账号类型", "工号/学号", "姓名", "初始密码", "安全要求"],
            credential_rows,
            watermark="初始密码仅本次生成和显示；请安全转交，禁止通过公开群聊传播。"),
        f"师生账号初始凭据_{entry['batchNo']}.xlsx", len(credential_rows))
