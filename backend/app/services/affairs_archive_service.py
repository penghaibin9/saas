"""学工归档正式服务：批次、真实学生档案文件、公共版本与 Manifest。

状态：DRAFT→COLLECTING→COLLEGE_REVIEW→SA_CONFIRM→ARCHIVED。
本文件直接实现范围、节点权限、档案生成和归档清单，不再由 guard monkey-patch。
"""
from __future__ import annotations

import hashlib
import io
import json
import uuid
from datetime import datetime, timedelta

from openpyxl import Workbook
from sqlalchemy import and_, func, or_, select

from app.core.exceptions import AppException, not_found
from app.core.optimistic_lock import atomic_claim_version
from app.models.file import FileAsset, FileBinding, FileObject, FileVersion
from app.services import file_service
from app.services.db_service import _iso, _tid, session

BATCH_FLOW = ["DRAFT", "COLLECTING", "COLLEGE_REVIEW", "SA_CONFIRM", "ARCHIVED"]
_PACKAGE_LEASE_SECONDS = 10 * 60
_PACKAGE_MAX_ATTEMPTS = 5


def archive_system_user() -> dict:
    """后台归档 worker 的显式系统身份；仅在已设置租户上下文后使用。"""
    return {
        "userId": "0",
        "loginName": "system-affairs-archive",
        "realName": "系统归档任务",
        "userType": "TEACHER",
        "currentRoleCode": "SCHOOL_ADMIN",
    }


def _op():
    from app.core.context import get_current_user_ctx

    user = get_current_user_ctx() or {}
    return (
        user.get("realName") or "系统",
        user.get("currentRoleCode") or "",
        str(user.get("userId") or ""),
    )


def _audit(db, biz_id, action, detail=""):
    from app.models import AffairsAuditTrail

    name, role, uid = _op()
    db.add(AffairsAuditTrail(
        tenant_id=_tid(), biz_type="ARCHIVE",
        biz_id=int(biz_id) if biz_id else None, action=action,
        operator=name or uid, role_name=role, detail=detail,
        occurred_at=datetime.utcnow(),
    ))


def _batch_row(batch) -> dict:
    return {
        "batchId": str(batch.id), "batchName": batch.batch_name,
        "yearCode": batch.year_code or "", "status": batch.status,
        "confirmBy": batch.confirm_by or "", "confirmAt": _iso(batch.confirm_at),
        "version": int(batch.version or 0),
    }


def _excel_text(value):
    if value is None:
        return ""
    text = value if isinstance(value, str) else json.dumps(value, ensure_ascii=False, default=str)
    return "'" + text if text[:1] in ("=", "+", "-", "@") else text


def _package_bytes(student_id: int, user) -> bytes:
    from app.services import affairs_profile_service as profile

    data = profile.get_profile(student_id, user)
    timeline: list[dict] = []
    page = 1
    while True:
        rows, total = profile.get_timeline(student_id, user, page=page, page_size=200)
        timeline.extend(rows)
        if not rows or len(timeline) >= int(total or 0):
            break
        page += 1
    workbook = Workbook()
    summary = workbook.active
    summary.title = "学工档案摘要"
    summary.append(["字段", "内容"])
    for key, value in data.items():
        summary.append([_excel_text(key), _excel_text(value)])
    events = workbook.create_sheet("成长时间线")
    events.append(["时间", "模块", "事件", "说明"])
    for item in timeline:
        events.append([
            _excel_text(item.get("occurredAt")), _excel_text(item.get("module")),
            _excel_text(item.get("title")), _excel_text(item.get("detail")),
        ])
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _manifest_bytes(batch, packages, students) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "归档清单"
    sheet.append([f"学工归档 · {batch.batch_name} · 生成时间 {datetime.now():%Y-%m-%d %H:%M}"])
    sheet.append([
        "学生ID", "学号", "姓名", "档案状态", "档案文件ID",
        "Manifest ID", "Manifest Revision", "Manifest SHA-256", "缺项",
    ])
    for package in packages:
        student = students.get(int(package.student_id))
        sheet.append([
            str(package.student_id), _excel_text(student.student_no if student else ""),
            _excel_text(student.real_name if student else ""), package.status,
            str(package.package_file_id or ""), str(package.manifest_id or ""),
            str(package.manifest_revision or ""), package.manifest_sha256 or "",
            _excel_text(package.missing_items_json or "[]"),
        ])
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _package_asset_code(package_id: int) -> str:
    return f"AFFAIRS_ARCHIVE_PACKAGE:{_tid()}:{int(package_id)}"


def _register_package_version(
    package_id: int,
    file_id: int,
    user,
    *,
    lease_token: str | None = None,
    expected_version: int | None = None,
) -> bool:
    from app.models import ArchivePackage, StudentProfile
    from app.modules.student_affairs.services import affairs_material_center_service as center

    with session() as db:
        package = db.scalars(select(ArchivePackage).where(
            ArchivePackage.tenant_id == _tid(), ArchivePackage.id == int(package_id),
            ArchivePackage.is_deleted.is_(False),
        ).with_for_update()).first()
        if not package:
            raise not_found("档案包不存在")
        if lease_token is not None:
            if (
                package.status != "GENERATING"
                or package.generation_lease_token != lease_token
                or (expected_version is not None and int(package.version or 0) != int(expected_version))
            ):
                return False
        file_obj = db.scalars(select(FileObject).where(
            FileObject.tenant_id == _tid(), FileObject.id == int(file_id),
            FileObject.is_deleted.is_(False),
        ).with_for_update()).first()
        if not file_obj:
            raise not_found("档案文件不存在")
        center._require_file_ready(file_obj)
        student = db.get(StudentProfile, int(package.student_id))
        if not student or student.is_deleted or student.tenant_id != _tid():
            raise not_found("学生主档不存在")
        asset = db.scalars(select(FileAsset).where(
            FileAsset.tenant_id == _tid(), FileAsset.asset_code == _package_asset_code(package.id),
            FileAsset.is_deleted.is_(False),
        ).with_for_update()).first()
        if asset is None:
            asset = FileAsset(
                tenant_id=_tid(), asset_code=_package_asset_code(package.id),
                title=f"{student.real_name}学工档案包", category_code="AFFAIRS_ARCHIVE_PACKAGE",
                owner_type="ARCHIVE_PACKAGE", owner_id=str(package.id),
                lifecycle_status="ACTIVE", version_count=0, sensitivity_level="SENSITIVE",
            )
            db.add(asset)
            db.flush()
        old_versions = db.scalars(select(FileVersion).where(
            FileVersion.tenant_id == _tid(), FileVersion.asset_id == int(asset.id),
            FileVersion.is_current.is_(True), FileVersion.is_deleted.is_(False),
        ).with_for_update()).all()
        for old in old_versions:
            old.is_current = False
            old.status = "INVALIDATED"
            old.invalidated_at = datetime.utcnow()
            old.invalidated_by = center._actor_name(user)
            old.invalid_reason = "重新生成学工档案包"
        version_no = int(db.scalar(select(func.max(FileVersion.version_no)).where(
            FileVersion.tenant_id == _tid(), FileVersion.asset_id == int(asset.id),
        )) or 0) + 1
        version = FileVersion(
            tenant_id=_tid(), asset_id=int(asset.id), file_object_id=int(file_obj.id),
            version_no=version_no, source_channel="SYSTEM_GENERATED",
            uploader_user_id=str(center._actor_id(user) or "") or None,
            uploader_name_snapshot=center._actor_name(user),
            submit_comment="学工档案快照", status="APPROVED", is_current=True,
            submitted_at=datetime.utcnow(), created_by=center._actor_id(user) or None,
        )
        db.add(version)
        db.flush()
        old_bindings = db.scalars(select(FileBinding).where(
            FileBinding.tenant_id == _tid(), FileBinding.asset_id == int(asset.id),
            FileBinding.is_current.is_(True), FileBinding.is_deleted.is_(False),
        ).with_for_update()).all()
        for old in old_bindings:
            old.is_current = False
            old.status = "SUPERSEDED"
            old.invalidated_at = datetime.utcnow()
        binding = FileBinding(
            tenant_id=_tid(), file_id=int(file_obj.id), biz_type="AFFAIRS_ARCHIVE",
            biz_id=str(student.id), relation_type="ARCHIVE_PACKAGE",
            subject_type="STUDENT", subject_id=str(student.id),
            batch_id=str(package.batch_id), version_no=version_no,
            is_current=True, status="ACTIVE",
            scope_json={"studentId": str(student.id), "archivePackageId": str(package.id)},
            asset_id=int(asset.id), version_id=int(version.id), module_code="student-affairs",
            student_id=int(student.id), college_id=getattr(student, "college_id", None),
            class_id=getattr(student, "class_id", None),
            data_scope_snapshot_json={
                "studentId": str(student.id), "collegeId": str(getattr(student, "college_id", None) or ""),
                "classId": str(getattr(student, "class_id", None) or ""),
            }, created_by=center._actor_id(user) or None,
        )
        db.add(binding)
        db.flush()
        asset.current_version_id = int(version.id)
        asset.version_count = version_no
        package.package_file_id = int(file_obj.id)
        package.package_asset_id = int(asset.id)
        package.package_version_id = int(version.id)
        package.missing_items_json = "[]"
        package.status = "SUBMITTED"
        package.generation_error = None
        package.generation_lease_token = None
        package.generation_lease_until = None
        package.version = int(package.version or 0) + 1
        file_obj.biz_type = "AFFAIRS_ARCHIVE"
        file_obj.biz_id = str(student.id)
        file_obj.visibility = "BIZ_SCOPED"
        file_obj.security_level = "SENSITIVE"
        db.commit()
        return True


def _claim_pending_packages(*, limit: int = 2, batch_id: int | None = None) -> list[dict]:
    from app.models import ArchivePackage

    now = datetime.utcnow()
    with session() as db:
        conditions = [
            ArchivePackage.tenant_id == _tid(),
            ArchivePackage.is_deleted.is_(False),
            ArchivePackage.package_version_id.is_(None),
            ArchivePackage.generation_attempts < _PACKAGE_MAX_ATTEMPTS,
            or_(
                ArchivePackage.status == "PENDING_GEN",
                and_(
                    ArchivePackage.status == "GENERATING",
                    ArchivePackage.generation_lease_until.is_not(None),
                    ArchivePackage.generation_lease_until <= now,
                ),
            ),
        ]
        if batch_id is not None:
            conditions.append(ArchivePackage.batch_id == int(batch_id))
        rows = db.scalars(
            select(ArchivePackage)
            .where(*conditions)
            .order_by(ArchivePackage.id)
            .with_for_update(skip_locked=True)
            .limit(max(1, min(int(limit or 2), 20)))
        ).all()
        claimed: list[dict] = []
        for package in rows:
            token = uuid.uuid4().hex
            package.status = "GENERATING"
            package.generation_attempts = int(package.generation_attempts or 0) + 1
            package.generation_error = None
            package.generation_lease_token = token
            package.generation_lease_until = now + timedelta(seconds=_PACKAGE_LEASE_SECONDS)
            package.version = int(package.version or 0) + 1
            claimed.append({
                "packageId": int(package.id),
                "batchId": int(package.batch_id),
                "studentId": int(package.student_id),
                "leaseToken": token,
                "version": int(package.version or 0),
            })
        db.commit()
        return claimed


def _fail_package_generation(
    package_id: int, lease_token: str, expected_version: int, error: Exception | str,
) -> bool:
    from app.models import ArchivePackage

    with session() as db:
        package = db.scalars(select(ArchivePackage).where(
            ArchivePackage.tenant_id == _tid(), ArchivePackage.id == int(package_id),
            ArchivePackage.is_deleted.is_(False),
        ).with_for_update()).first()
        if not package:
            return False
        if (
            package.status != "GENERATING"
            or package.generation_lease_token != lease_token
            or int(package.version or 0) != int(expected_version)
        ):
            return False
        message = str(error or "档案生成失败")[:1000]
        package.generation_error = message
        package.generation_lease_token = None
        package.generation_lease_until = None
        if int(package.generation_attempts or 0) >= _PACKAGE_MAX_ATTEMPTS:
            package.status = "PENDING_SUPPLEMENT"
            package.missing_items_json = json.dumps(
                ["档案生成连续失败，请管理员核查后重试"], ensure_ascii=False,
            )
        else:
            package.status = "PENDING_GEN"
        package.version = int(package.version or 0) + 1
        db.commit()
        return True


def run_pending_packages(*, limit: int = 2, batch_id: int | None = None, user=None) -> dict:
    """领取并异步生成档案包；租约防重复，过期任务可由下一 worker 回收。"""
    from app.models import StudentProfile

    actor = user or archive_system_user()
    claimed = _claim_pending_packages(limit=limit, batch_id=batch_id)
    succeeded = 0
    failed = 0
    stale = 0
    for item in claimed:
        package_id = int(item["packageId"])
        token = str(item["leaseToken"])
        version = int(item["version"])
        try:
            with session() as db:
                student = db.scalars(select(StudentProfile).where(
                    StudentProfile.tenant_id == _tid(),
                    StudentProfile.id == int(item["studentId"]),
                    StudentProfile.is_deleted.is_(False),
                )).first()
                if not student:
                    raise not_found("学生主档不存在")
                student_id = int(student.id)
                student_no = student.student_no or str(student.id)
            payload = _package_bytes(student_id, actor)
            meta = file_service.store_bytes(
                payload, f"学工档案_{student_no}.xlsx", biz_type="AFFAIRS_ARCHIVE",
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                biz_id=str(student_id), user=actor, visibility="BIZ_SCOPED", security_level="SENSITIVE",
            )
            if meta.get("status") not in ("AVAILABLE", "STORED", "available", None):
                raise AppException("DATA_CONFLICT", f"学生{student_no}档案文件安全校验未通过")
            if _register_package_version(
                package_id, int(meta["fileId"]), actor,
                lease_token=token, expected_version=version,
            ):
                succeeded += 1
            else:
                stale += 1
        except Exception as exc:  # noqa: BLE001
            if _fail_package_generation(package_id, token, version, exc):
                failed += 1
            else:
                stale += 1
    return {
        "claimed": len(claimed),
        "succeeded": succeeded,
        "failed": failed,
        "stale": stale,
    }


def list_batches(user, status=None, page=1, page_size=50):
    from app.models import ArchiveBatch, ArchivePackage

    page = max(1, int(page or 1))
    page_size = max(1, min(int(page_size or 50), 200))
    with session() as db:
        conds = [ArchiveBatch.tenant_id == _tid(), ArchiveBatch.is_deleted.is_(False)]
        if status:
            conds.append(ArchiveBatch.status == status)
        total = int(db.scalar(select(func.count()).select_from(ArchiveBatch).where(*conds)) or 0)
        package_counts = (
            select(
                ArchivePackage.batch_id.label("batch_id"),
                func.count(ArchivePackage.id).label("package_count"),
            )
            .where(
                ArchivePackage.tenant_id == _tid(),
                ArchivePackage.is_deleted.is_(False),
            )
            .group_by(ArchivePackage.batch_id)
            .subquery()
        )
        rows = db.execute(
            select(ArchiveBatch, func.coalesce(package_counts.c.package_count, 0))
            .outerjoin(package_counts, package_counts.c.batch_id == ArchiveBatch.id)
            .where(*conds)
            .order_by(ArchiveBatch.id.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        ).all()
        result = []
        for batch, package_count in rows:
            item = _batch_row(batch)
            item["packageCount"] = int(package_count or 0)
            result.append(item)
        return result, total


def create_batch(body, user) -> dict:
    from app.models import ArchiveBatch

    with session() as db:
        batch = ArchiveBatch(
            tenant_id=_tid(), batch_name=body.batchName,
            year_code=getattr(body, "yearCode", None),
            scope_json=json.dumps(getattr(body, "scope", {}) or {}, ensure_ascii=False),
            status="DRAFT",
        )
        db.add(batch)
        db.flush()
        _audit(db, batch.id, "BATCH_CREATE")
        db.commit()
        db.refresh(batch)
        return _batch_row(batch)


def collect(batch_id, user, student_ids, expected_version=None) -> dict:
    from app.core.affairs_security import build_affairs_context, no_data_scope
    from app.models import ArchiveBatch, ArchivePackage, StudentProfile

    raw = list(student_ids or [])
    if not raw:
        raise AppException("VALIDATION_ERROR", "至少圈定一名学生")
    try:
        student_ids = list(dict.fromkeys(int(value) for value in raw))
    except (TypeError, ValueError) as exc:
        raise AppException("VALIDATION_ERROR", "学生ID必须为有效数字") from exc
    with session() as db:
        batch = db.scalars(select(ArchiveBatch).where(
            ArchiveBatch.tenant_id == _tid(), ArchiveBatch.id == int(batch_id),
            ArchiveBatch.is_deleted.is_(False),
        ).with_for_update()).first()
        if not batch:
            raise not_found("归档批次不存在")
        if batch.status not in ("DRAFT", "COLLECTING"):
            raise AppException("APPROVAL_VERSION_CONFLICT", "该批次不可再收集")
        atomic_claim_version(db, batch, expected_version)
        context = build_affairs_context(user, db)
        students = db.scalars(select(StudentProfile).where(
            StudentProfile.tenant_id == _tid(),
            StudentProfile.id.in_(student_ids),
            StudentProfile.is_deleted.is_(False),
        )).all()
        student_map = {int(student.id): student for student in students}
        missing = [student_id for student_id in student_ids if student_id not in student_map]
        if missing:
            raise not_found("学生不存在")
        allowed_classes = context.allowed_class_ids(db)
        allowed_students = context.psychology_student_ids | context.student_ids
        for student_id in student_ids:
            student = student_map[student_id]
            if allowed_classes is None:
                continue
            if context.scope_type == "SELF":
                if context.self_student_id and int(context.self_student_id) == student_id:
                    continue
                raise no_data_scope("学生只能访问本人数据")
            if context.scope_type == "STUDENT":
                if student_id in allowed_students:
                    continue
                raise no_data_scope("该学生不在您的授权范围内")
            if student.class_id not in allowed_classes:
                raise no_data_scope("该学生不在您的数据范围内")
        existing_ids = set(db.scalars(select(ArchivePackage.student_id).where(
            ArchivePackage.tenant_id == _tid(),
            ArchivePackage.batch_id == batch.id,
            ArchivePackage.student_id.in_(student_ids),
            ArchivePackage.is_deleted.is_(False),
        )).all())
        made = 0
        for student_id in student_ids:
            if student_id in existing_ids:
                continue
            db.add(ArchivePackage(
                tenant_id=_tid(), batch_id=batch.id, student_id=student_id,
                missing_items_json="[]", status="PENDING_GEN",
            ))
            made += 1
        batch.status = "COLLECTING"
        batch.version = int(batch.version or 0) + 1
        _audit(db, batch.id, "COLLECT", f"{made}生;异步生成")
        db.commit()
        pending = int(db.scalar(select(func.count()).select_from(ArchivePackage).where(
            ArchivePackage.tenant_id == _tid(), ArchivePackage.batch_id == int(batch_id),
            ArchivePackage.status != "SUBMITTED", ArchivePackage.is_deleted.is_(False),
        )) or 0)
        return {
            "batchId": str(batch_id), "packagesCreated": made,
            "packagesQueued": made, "packagesGenerated": 0, "packagesPending": pending,
            "status": "COLLECTING", "version": int(batch.version or 0),
        }


def advance(batch_id, user, action="APPROVE", expected_version=None) -> dict:
    from app.core.affairs_security import build_affairs_context
    from app.models import ArchiveBatch, ArchivePackage, ExportTask, StudentProfile
    from app.modules.student_affairs.services import affairs_material_center_service as center
    from app.services.message_identity import resolve_message_user_id

    if str(action or "").upper() != "APPROVE":
        raise AppException("VALIDATION_ERROR", "归档流转仅支持APPROVE，退回须使用专用退回流程")
    with session() as db:
        batch = db.scalars(select(ArchiveBatch).where(
            ArchiveBatch.tenant_id == _tid(), ArchiveBatch.id == int(batch_id),
            ArchiveBatch.is_deleted.is_(False),
        ).with_for_update()).first()
        if not batch:
            raise not_found("归档批次不存在")
        if batch.status not in ("COLLECTING", "COLLEGE_REVIEW", "SA_CONFIRM"):
            raise AppException("APPROVAL_VERSION_CONFLICT", "该批次当前状态不可流转")
        context = build_affairs_context(user, db)
        if batch.status == "COLLEGE_REVIEW" and context.scope_type not in ("COLLEGE", "TENANT_ALL"):
            raise AppException("NO_PERMISSION", "仅学院学工或全域管理员可完成学院审核")
        if batch.status == "SA_CONFIRM" and context.scope_type != "TENANT_ALL":
            raise AppException("NO_PERMISSION", "仅学校/学工处全域管理员可确认归档")
        if batch.status != "SA_CONFIRM":
            atomic_claim_version(db, batch, expected_version)
            next_status = BATCH_FLOW[BATCH_FLOW.index(batch.status) + 1]
            batch.status = next_status
            batch.version = int(batch.version or 0) + 1
            _audit(db, batch.id, "ADVANCE", f"->{next_status}")
            db.commit()
            db.refresh(batch)
            return _batch_row(batch)

    with session() as db:
        batch = db.scalars(select(ArchiveBatch).where(
            ArchiveBatch.tenant_id == _tid(), ArchiveBatch.id == int(batch_id),
            ArchiveBatch.is_deleted.is_(False),
        ).with_for_update()).first()
        if not batch or batch.status != "SA_CONFIRM":
            raise AppException("APPROVAL_VERSION_CONFLICT", "归档批次状态已变化，请刷新")
        atomic_claim_version(db, batch, expected_version)
        packages = db.scalars(select(ArchivePackage).where(
            ArchivePackage.tenant_id == _tid(), ArchivePackage.batch_id == batch.id,
            ArchivePackage.is_deleted.is_(False),
        ).order_by(ArchivePackage.id).with_for_update()).all()
        if not packages:
            raise AppException("DATA_CONFLICT", "归档批次没有学生档案包")
        not_ready = [package for package in packages if package.status != "SUBMITTED" or not package.package_version_id]
        if not_ready:
            raise AppException("DATA_CONFLICT", f"仍有{len(not_ready)}份档案包未生成完成，不能归档")
        for package in packages:
            center.freeze_archive_manifest(db, package, user)
            package.status = "ARCHIVED"
            package.version = int(package.version or 0) + 1
        student_ids = {int(package.student_id) for package in packages}
        students = {
            int(student.id): student
            for student in db.scalars(select(StudentProfile).where(
                StudentProfile.tenant_id == _tid(), StudentProfile.id.in_(student_ids),
                StudentProfile.is_deleted.is_(False),
            )).all()
        }
        payload = _manifest_bytes(batch, packages, students)
        digest = hashlib.sha256(payload).hexdigest()
        batch.status = "ARCHIVED"
        batch.confirm_by = center._actor_name(user)
        batch.confirm_at = datetime.utcnow()
        batch.version = int(batch.version or 0) + 1
        db.commit()

    meta = file_service.store_bytes(
        payload, f"学工归档清单_{batch_id}.xlsx", biz_type="AFFAIRS_ARCHIVE_MANIFEST",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        biz_id=str(batch_id), user=user, visibility="BIZ_SCOPED", security_level="SENSITIVE",
    )
    with session() as db:
        batch = db.get(ArchiveBatch, int(batch_id))
        packages = db.scalars(select(ArchivePackage).where(
            ArchivePackage.tenant_id == _tid(), ArchivePackage.batch_id == int(batch_id),
            ArchivePackage.is_deleted.is_(False),
        )).all()
        task = ExportTask(
            tenant_id=_tid(), export_mode="ARCHIVE_MANIFEST", module_code="student-affairs",
            row_count=len(packages), purpose=f"学工归档清单：{batch.batch_name}",
            file_hash=digest, status="SUCCESS", remark=f"file-object:{meta['fileId']}",
            created_by=resolve_message_user_id(user) or None,
        )
        db.add(task)
        db.flush()
        for package in packages:
            package.export_task_id = int(task.id)
        _audit(
            db, batch.id, "ARCHIVED",
            f"export_task={task.id};file={meta['fileId']};sha256={digest};manifests={len(packages)}",
        )
        db.commit()
        db.refresh(batch)
        result = _batch_row(batch)
        result.update({
            "exportTaskId": str(task.id), "fileId": str(meta["fileId"]),
            "fileName": meta.get("fileName") or f"学工归档清单_{batch_id}.xlsx",
            "rowCount": len(packages),
        })
        return result


def get_batch(batch_id, user) -> dict:
    from app.models import ArchiveBatch, ArchivePackage, StudentProfile
    from app.services.affairs_dashboard_service import _allowed_class_ids

    with session() as db:
        batch = db.get(ArchiveBatch, int(batch_id))
        if not batch or batch.is_deleted or batch.tenant_id != _tid():
            raise not_found("归档批次不存在")
        allowed, _ = _allowed_class_ids(db, user)
        packages = db.scalars(select(ArchivePackage).where(
            ArchivePackage.tenant_id == _tid(), ArchivePackage.batch_id == batch.id,
            ArchivePackage.is_deleted.is_(False),
        ).order_by(ArchivePackage.id)).all()
        student_ids = {int(package.student_id) for package in packages if package.student_id}
        students = {
            int(student.id): student
            for student in db.scalars(select(StudentProfile).where(
                StudentProfile.tenant_id == _tid(),
                StudentProfile.id.in_(student_ids) if student_ids else StudentProfile.id == -1,
                StudentProfile.is_deleted.is_(False),
            )).all()
        }
        result = _batch_row(batch)
        visible = []
        for package in packages:
            student = students.get(int(package.student_id)) if package.student_id else None
            if allowed is not None and (not student or student.class_id not in allowed):
                continue
            visible.append({
                "packageId": str(package.id), "studentId": str(package.student_id),
                "studentNo": student.student_no if student else "",
                "studentName": student.real_name if student else "",
                "status": package.status, "exportTaskId": str(package.export_task_id or ""),
                "packageFileId": str(package.package_file_id or ""),
                "packageVersionId": str(package.package_version_id or ""),
                "manifestId": str(package.manifest_id or ""),
                "manifestRevision": int(package.manifest_revision or 0),
                "manifestSha256": package.manifest_sha256 or "",
                "generationAttempts": int(package.generation_attempts or 0),
                "generationError": package.generation_error or "",
                "missingItems": json.loads(package.missing_items_json or "[]"),
            })
        result["packages"] = visible
        return result
