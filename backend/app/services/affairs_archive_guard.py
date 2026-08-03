"""学工归档真实化安全门：范围强校验、真实档案快照、节点权限与真实导出任务。"""
from __future__ import annotations

import hashlib
import io
import json
import uuid
from datetime import datetime

from openpyxl import Workbook
from sqlalchemy import select

from app.core.exceptions import AppException, not_found
from app.services.db_service import _tid, session

_INSTALLED = False


def _excel_text(value):
    if value is None:
        return ""
    text = value if isinstance(value, str) else json.dumps(value, ensure_ascii=False, default=str)
    return "'" + text if text[:1] in ("=", "+", "-", "@") else text


def _package_bytes(student_id: int, user) -> bytes:
    from app.services import affairs_profile_service as profile

    data = profile.get_profile(student_id, user)
    timeline: list[dict] = []
    page = 1
    while True:
        rows, total = profile.get_timeline(student_id, user, page=page, page_size=200)
        timeline.extend(rows)
        if not rows or len(timeline) >= int(total or 0):
            break
        page += 1
    wb = Workbook()
    ws = wb.active
    ws.title = "学工档案摘要"
    ws.append(["字段", "内容"])
    for key, value in data.items():
        ws.append([_excel_text(key), _excel_text(value)])
    ts = wb.create_sheet("成长时间线")
    ts.append(["时间", "模块", "事件", "说明"])
    for item in timeline:
        ts.append([
            _excel_text(item.get("occurredAt")),
            _excel_text(item.get("module")),
            _excel_text(item.get("title")),
            _excel_text(item.get("detail")),
        ])
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _generate_pending_packages(batch_id: int, user) -> int:
    from app.models import ArchivePackage, StudentProfile
    from app.services import file_service

    generated = 0
    with session() as db:
        package_ids = list(db.scalars(select(ArchivePackage.id).where(
            ArchivePackage.tenant_id == _tid(),
            ArchivePackage.batch_id == int(batch_id),
            ArchivePackage.package_file_id.is_(None),
            ArchivePackage.is_deleted.is_(False),
        )).all())
    for package_id in package_ids:
        with session() as db:
            package = db.get(ArchivePackage, int(package_id))
            if (
                not package
                or package.is_deleted
                or package.tenant_id != _tid()
                or package.package_file_id
            ):
                continue
            student = db.get(StudentProfile, int(package.student_id))
            if not student or student.is_deleted or student.tenant_id != _tid():
                package.status = "PENDING_SUPPLEMENT"
                package.missing_items_json = json.dumps(["学生主档不存在"], ensure_ascii=False)
                db.commit()
                continue
            student_id = int(student.id)
            student_no = student.student_no or str(student.id)
        payload = _package_bytes(student_id, user)
        meta = file_service.store_bytes(
            payload,
            f"学工档案_{student_no}.xlsx",
            biz_type="AFFAIRS_ARCHIVE",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            biz_id=str(student_id),
            user=user,
            visibility="BIZ_SCOPED",
            security_level="SENSITIVE",
        )
        if meta.get("status") not in ("AVAILABLE", "available", None):
            raise AppException("DATA_CONFLICT", f"学生{student_no}档案文件安全校验未通过")
        with session() as db:
            package = db.get(ArchivePackage, int(package_id))
            if (
                package
                and not package.is_deleted
                and package.tenant_id == _tid()
                and not package.package_file_id
            ):
                package.package_file_id = int(meta["fileId"])
                package.missing_items_json = "[]"
                package.status = "SUBMITTED"
                package.version = int(package.version or 0) + 1
                db.commit()
                generated += 1
    return generated


def _manifest_bytes(batch, packages, students) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "归档清单"
    ws.append([f"学工归档 · {batch.batch_name} · 生成时间 {datetime.now():%Y-%m-%d %H:%M}"])
    ws.append(["学生ID", "学号", "姓名", "档案状态", "档案文件ID", "缺项"])
    for package in packages:
        student = students.get(int(package.student_id))
        ws.append([
            str(package.student_id),
            _excel_text(student.student_no if student else ""),
            _excel_text(student.real_name if student else ""),
            package.status,
            str(package.package_file_id or ""),
            _excel_text(package.missing_items_json or "[]"),
        ])
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def install() -> None:
    global _INSTALLED
    if _INSTALLED:
        return
    from app.models import ArchiveBatch, ArchivePackage, ExportTask, StudentProfile
    from app.services import affairs_archive_service as archive

    old_collect = archive.collect
    old_advance = archive.advance
    old_get_batch = archive.get_batch

    def collect(batch_id, user, student_ids, expected_version=None):
        raw = list(student_ids or [])
        if not raw:
            raise AppException("VALIDATION_ERROR", "至少圈定一名学生")
        try:
            ids = list(dict.fromkeys(int(value) for value in raw))
        except (TypeError, ValueError) as exc:
            raise AppException("VALIDATION_ERROR", "学生ID必须为有效数字") from exc
        from app.core.affairs_security import build_affairs_context

        with session() as db:
            context = build_affairs_context(user, db)
            for student_id in ids:
                context.require_student(db, student_id)
        result = old_collect(batch_id, user, ids, expected_version)
        result["packagesGenerated"] = _generate_pending_packages(int(batch_id), user)
        with session() as db:
            pending = db.scalars(select(ArchivePackage.id).where(
                ArchivePackage.tenant_id == _tid(),
                ArchivePackage.batch_id == int(batch_id),
                ArchivePackage.status != "SUBMITTED",
                ArchivePackage.is_deleted.is_(False),
            )).all()
        result["packagesPending"] = len(pending)
        return result

    def advance(batch_id, user, action="APPROVE", expected_version=None):
        if str(action or "").upper() != "APPROVE":
            raise AppException("VALIDATION_ERROR", "归档流转仅支持APPROVE，退回须使用专用退回流程")
        from app.core.affairs_security import build_affairs_context

        with session() as db:
            batch = db.get(ArchiveBatch, int(batch_id))
            if not batch or batch.is_deleted or batch.tenant_id != _tid():
                raise not_found("归档批次不存在")
            context = build_affairs_context(user, db)
            if batch.status == "COLLEGE_REVIEW" and context.scope_type not in ("COLLEGE", "TENANT_ALL"):
                raise AppException("NO_PERMISSION", "仅学院学工或全域管理员可完成学院审核")
            if batch.status == "SA_CONFIRM" and context.scope_type != "TENANT_ALL":
                raise AppException("NO_PERMISSION", "仅学校/学工处全域管理员可确认归档")
            if batch.status != "SA_CONFIRM":
                return old_advance(batch_id, user, "APPROVE", expected_version)

        _generate_pending_packages(int(batch_id), user)
        with session() as db:
            batch = db.scalars(select(ArchiveBatch).where(
                ArchiveBatch.tenant_id == _tid(),
                ArchiveBatch.id == int(batch_id),
                ArchiveBatch.is_deleted.is_(False),
            ).with_for_update()).first()
            if not batch or batch.status != "SA_CONFIRM":
                raise AppException("APPROVAL_VERSION_CONFLICT", "归档批次状态已变化，请刷新")
            archive.atomic_claim_version(db, batch, expected_version)
            packages = db.scalars(select(ArchivePackage).where(
                ArchivePackage.tenant_id == _tid(),
                ArchivePackage.batch_id == batch.id,
                ArchivePackage.is_deleted.is_(False),
            )).all()
            if not packages:
                raise AppException("DATA_CONFLICT", "归档批次没有学生档案包")
            not_ready = [p for p in packages if p.status != "SUBMITTED" or not p.package_file_id]
            if not_ready:
                raise AppException("DATA_CONFLICT", f"仍有{len(not_ready)}份档案包未生成完成，不能归档")
            student_ids = {int(p.student_id) for p in packages}
            students = {
                int(s.id): s
                for s in db.scalars(select(StudentProfile).where(
                    StudentProfile.tenant_id == _tid(),
                    StudentProfile.id.in_(student_ids),
                    StudentProfile.is_deleted.is_(False),
                )).all()
            }
            payload = _manifest_bytes(batch, packages, students)
            from app.services.import_export_service import upload_dir

            key = f"exports/{datetime.now():%Y%m%d}/affairs_archive_{batch.id}_{uuid.uuid4().hex[:8]}.xlsx"
            target = upload_dir() / key
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(payload)
            digest = hashlib.sha256(payload).hexdigest()
            from app.services.message_identity import resolve_message_user_id

            task = ExportTask(
                tenant_id=_tid(),
                export_mode="ARCHIVE_MANIFEST",
                module_code="student-affairs",
                row_count=len(packages),
                purpose=f"学工归档清单：{batch.batch_name}",
                file_hash=digest,
                status="SUCCESS",
                remark=key,
                created_by=resolve_message_user_id(user) or None,
            )
            db.add(task)
            db.flush()
            for package in packages:
                package.export_task_id = task.id
                package.status = "ARCHIVED"
                package.version = int(package.version or 0) + 1
            operator, _role, _uid = archive._op()
            batch.status = "ARCHIVED"
            batch.confirm_by = operator
            batch.confirm_at = datetime.utcnow()
            batch.version = int(batch.version or 0) + 1
            archive._audit(db, batch.id, "ARCHIVED", f"export_task={task.id};sha256={digest}")
            db.commit()
            db.refresh(batch)
            data = archive._batch_row(batch)
            data.update({
                "exportTaskId": str(task.id),
                "fileName": target.name,
                "rowCount": len(packages),
            })
            return data

    def get_batch(batch_id, user):
        data = old_get_batch(batch_id, user)
        ids = [
            int(row["packageId"])
            for row in data.get("packages", [])
            if str(row.get("packageId") or "").isdigit()
        ]
        if not ids:
            return data
        with session() as db:
            rows = {
                int(row.id): row
                for row in db.scalars(select(ArchivePackage).where(
                    ArchivePackage.tenant_id == _tid(),
                    ArchivePackage.id.in_(ids),
                    ArchivePackage.is_deleted.is_(False),
                )).all()
            }
        for item in data["packages"]:
            row = rows.get(int(item["packageId"]))
            if row:
                item["packageFileId"] = str(row.package_file_id or "")
                item["missingItems"] = json.loads(row.missing_items_json or "[]")
        return data

    archive.collect = collect
    archive.advance = advance
    archive.get_batch = get_batch
    _INSTALLED = True
