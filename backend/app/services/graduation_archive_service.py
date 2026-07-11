"""毕业设计中心 · 毕设归档服务（对齐老系统 filingModel 0待生成/1待提交/2已提交/3已备案/4驳回语义）。

生成清单（按实际业务完成情况自动核验必备材料）→ 提交 → 核验归档 → 驳回补交；支持批量归档。
隔离说明：不引用实习/迎新域文件；Excel 台账导出用 openpyxl 直连。
"""
from __future__ import annotations

from datetime import datetime

from sqlalchemy import func, select

from app.core.context import get_current_user_ctx
from app.core.exceptions import AppException, not_found
from app.models import (GraduationArchiveRecord, GraduationAuditTrail, GraduationDefenseScore,
                        GraduationFinal, GraduationGrade, GraduationMidterm, GraduationProposal,
                        GraduationStudent, GraduationTaskBook)
from app.services.db_service import _iso, _tid, session

STATUS_LABEL = {"NOT_GENERATED": "待生成", "PENDING_SUBMIT": "待提交", "SUBMITTED": "已提交",
                "FILED": "已备案", "REJECTED": "已驳回"}
STATUS_TONE = {"NOT_GENERATED": "default", "PENDING_SUBMIT": "warning", "SUBMITTED": "warning",
               "FILED": "success", "REJECTED": "danger"}
CHECKLIST_ITEMS = [
    ("taskbook", "任务书（已确认）"), ("proposal", "开题报告（已通过）"), ("midterm", "中期检查（已通过）"),
    ("final", "成果定稿（已通过）"), ("review", "教师评阅（已完成）"), ("defenseScore", "答辩评分（已确认）"),
    ("grade", "成绩（已发布）"),
]


def _op() -> tuple[str, str]:
    u = get_current_user_ctx() or {}
    return u.get("realName") or "系统", u.get("roleName") or u.get("currentRoleCode") or ""


def _audit(db, bid, action, detail="", before="", after=""):
    n, r = _op()
    db.add(GraduationAuditTrail(tenant_id=_tid(), biz_type="ARCHIVE", biz_id=str(bid), action=action,
                                operator=n, role_name=r, detail=detail, before_val=before, after_val=after,
                                occurred_at=datetime.utcnow()))


def _stu(db, sid) -> GraduationStudent:
    s = db.get(GraduationStudent, int(sid))
    if not s or s.is_deleted or s.tenant_id != _tid():
        raise not_found("毕设学生不存在或不在当前数据范围内")
    return s


def _get_or_create(db, stu: GraduationStudent) -> GraduationArchiveRecord:
    a = db.scalars(select(GraduationArchiveRecord).where(
        GraduationArchiveRecord.tenant_id == _tid(), GraduationArchiveRecord.gd_student_id == stu.id,
        GraduationArchiveRecord.is_deleted.is_(False))).first()
    if not a:
        a = GraduationArchiveRecord(tenant_id=_tid(), gd_student_id=stu.id, status="NOT_GENERATED")
        db.add(a)
        db.flush()
    return a


def _check_completeness(db, stu: GraduationStudent) -> tuple[list[dict], list[str]]:
    tb = db.scalars(select(GraduationTaskBook).where(
        GraduationTaskBook.tenant_id == _tid(), GraduationTaskBook.gd_student_id == stu.id,
        GraduationTaskBook.status == "CONFIRMED")).first()
    prop = db.scalars(select(GraduationProposal).where(
        GraduationProposal.tenant_id == _tid(), GraduationProposal.gd_student_id == stu.id,
        GraduationProposal.status == "APPROVED", GraduationProposal.is_deleted.is_(False))).first()
    mt = db.scalars(select(GraduationMidterm).where(
        GraduationMidterm.tenant_id == _tid(), GraduationMidterm.gd_student_id == stu.id,
        GraduationMidterm.status.in_(("CHECKED_PASS", "RECTIFIED_PASS")))).first()
    fin = db.scalars(select(GraduationFinal).where(
        GraduationFinal.tenant_id == _tid(), GraduationFinal.gd_student_id == stu.id,
        GraduationFinal.status == "APPROVED", GraduationFinal.is_deleted.is_(False))).first()
    from app.models import GraduationReview
    rv = db.scalars(select(GraduationReview).where(
        GraduationReview.tenant_id == _tid(), GraduationReview.gd_student_id == stu.id,
        GraduationReview.status == "COMPLETED", GraduationReview.is_deleted.is_(False))).first()
    ds = db.scalars(select(GraduationDefenseScore).where(
        GraduationDefenseScore.tenant_id == _tid(), GraduationDefenseScore.gd_student_id == stu.id,
        GraduationDefenseScore.status == "CONFIRMED", GraduationDefenseScore.is_deleted.is_(False))).first()
    gr = db.scalars(select(GraduationGrade).where(
        GraduationGrade.tenant_id == _tid(), GraduationGrade.gd_student_id == stu.id,
        GraduationGrade.status == "PUBLISHED")).first()
    present = {"taskbook": bool(tb), "proposal": bool(prop), "midterm": bool(mt), "final": bool(fin),
              "review": bool(rv), "defenseScore": bool(ds), "grade": bool(gr)}
    checklist = [{"item": key, "label": label, "present": present[key]} for key, label in CHECKLIST_ITEMS]
    missing = [label for key, label in CHECKLIST_ITEMS if not present[key]]
    return checklist, missing


def _row(a: GraduationArchiveRecord, stu=None) -> dict:
    return {"id": str(a.id), "gdStudentId": str(a.gd_student_id),
            "studentName": stu.name if stu else "", "studentNo": stu.student_no if stu else "",
            "status": a.status, "statusLabel": STATUS_LABEL.get(a.status, a.status),
            "statusTone": STATUS_TONE.get(a.status, "default"),
            "checklist": a.checklist_json or [], "missingItems": a.missing_items or [],
            "generatedAt": _iso(a.generated_at), "submittedAt": _iso(a.submitted_at),
            "verifiedBy": a.verified_by or "", "filedAt": _iso(a.filed_at),
            "archiveBatchNo": a.archive_batch_no or "", "rejectReason": a.reject_reason or "",
            "updatedAt": _iso(a.updated_at)}


def list_archives(page: int, page_size: int, keyword=None, status=None) -> tuple[list[dict], int]:
    with session() as db:
        q = select(GraduationArchiveRecord).where(GraduationArchiveRecord.tenant_id == _tid(),
                                                   GraduationArchiveRecord.is_deleted.is_(False))
        if status:
            q = q.where(GraduationArchiveRecord.status == status)
        rows = db.scalars(q.order_by(GraduationArchiveRecord.id.desc())).all()
        sids = {a.gd_student_id for a in rows}
        smap = {s.id: s for s in db.scalars(select(GraduationStudent).where(
            GraduationStudent.id.in_(sids))).all()} if sids else {}
        items = []
        for a in rows:
            stu = smap.get(a.gd_student_id)
            if keyword and (not stu or keyword.strip() not in (stu.name or "")):
                continue
            items.append(_row(a, stu))
        total = len(items)
        start = (max(1, page) - 1) * page_size
        return items[start:start + page_size], total


def get_archive(gd_student_id) -> dict:
    with session() as db:
        stu = _stu(db, gd_student_id)
        a = _get_or_create(db, stu)
        db.commit()
        return _row(a, stu)


def generate_archive(gd_student_id) -> dict:
    with session() as db:
        stu = _stu(db, gd_student_id)
        a = _get_or_create(db, stu)
        if a.status == "FILED":
            raise AppException("DATA_CONFLICT", "已备案归档不可重新生成")
        checklist, missing = _check_completeness(db, stu)
        a.checklist_json = checklist
        a.missing_items = missing
        a.status = "PENDING_SUBMIT"
        a.generated_at = datetime.utcnow()
        _audit(db, a.id, "生成归档清单", detail=f"缺失 {len(missing)} 项")
        db.commit()
        return _row(a, stu)


def submit_archive(gd_student_id) -> dict:
    with session() as db:
        stu = _stu(db, gd_student_id)
        a = _get_or_create(db, stu)
        if a.status != "PENDING_SUBMIT":
            raise AppException("DATA_CONFLICT", "仅「待提交」记录可提交")
        if a.missing_items:
            raise AppException("DATA_CONFLICT", f"仍缺 {len(a.missing_items)} 项材料，不能提交")
        a.status = "SUBMITTED"
        a.submitted_at = datetime.utcnow()
        _audit(db, a.id, "提交归档")
        db.commit()
        return _row(a, stu)


def verify_and_file(gd_student_id, archive_batch_no: str = None) -> dict:
    with session() as db:
        stu = _stu(db, gd_student_id)
        a = _get_or_create(db, stu)
        if a.status != "SUBMITTED":
            raise AppException("DATA_CONFLICT", "仅「已提交」记录可核验归档")
        n, _ = _op()
        a.status = "FILED"
        a.verified_by = n
        a.filed_at = datetime.utcnow()
        a.archive_batch_no = archive_batch_no or f"GDARCH-{datetime.now():%Y%m%d}"
        _audit(db, a.id, "核验归档", detail=a.archive_batch_no)
        db.commit()
        return _row(a, stu)


def reject_archive(gd_student_id, reason: str) -> dict:
    if not reason or len(reason.strip()) < 5:
        raise AppException("VALIDATION_ERROR", "退回原因必填且不少于 5 字")
    with session() as db:
        stu = _stu(db, gd_student_id)
        a = _get_or_create(db, stu)
        if a.status != "SUBMITTED":
            raise AppException("DATA_CONFLICT", "仅「已提交」记录可驳回")
        a.status = "REJECTED"
        a.reject_reason = reason.strip()
        _audit(db, a.id, "驳回归档", reason.strip())
        db.commit()
        return _row(a, stu)


def batch_file(archive_batch_no: str = None) -> dict:
    """批量归档一键操作：对全部「已提交」的归档记录一键核验备案。"""
    with session() as db:
        subs = db.scalars(select(GraduationArchiveRecord).where(
            GraduationArchiveRecord.tenant_id == _tid(),
            GraduationArchiveRecord.status == "SUBMITTED",
            GraduationArchiveRecord.is_deleted.is_(False))).all()
        n, _ = _op()
        batch_no = archive_batch_no or f"GDARCH-{datetime.now():%Y%m%d}"
        for a in subs:
            a.status = "FILED"
            a.verified_by = n
            a.filed_at = datetime.utcnow()
            a.archive_batch_no = batch_no
            _audit(db, a.id, "批量核验归档", detail=batch_no)
        db.commit()
        return {"filed": len(subs), "archiveBatchNo": batch_no}


def batch_generate_submit() -> dict:
    """批量归档一键操作：对材料齐全的在册学生一键生成+提交归档（缺材料的跳过）。"""
    with session() as db:
        stus = db.scalars(select(GraduationStudent).where(
            GraduationStudent.tenant_id == _tid(), GraduationStudent.is_deleted.is_(False),
            GraduationStudent.record_status == "ACTIVE")).all()
        submitted, skipped = 0, 0
        for stu in stus:
            a = _get_or_create(db, stu)
            if a.status in ("FILED", "SUBMITTED"):
                continue
            checklist, missing = _check_completeness(db, stu)
            a.checklist_json = checklist
            a.missing_items = missing
            a.generated_at = datetime.utcnow()
            if missing:
                a.status = "PENDING_SUBMIT"
                skipped += 1
            else:
                a.status = "SUBMITTED"
                a.submitted_at = datetime.utcnow()
                submitted += 1
        _audit(db, "batch", "批量生成并提交归档", detail=f"提交{submitted}/跳过{skipped}")
        db.commit()
        return {"submitted": submitted, "skipped": skipped}


def archive_stats() -> dict:
    with session() as db:
        base = [GraduationArchiveRecord.tenant_id == _tid(), GraduationArchiveRecord.is_deleted.is_(False)]
        total = int(db.scalar(select(func.count()).select_from(GraduationArchiveRecord).where(*base)) or 0)
        by_status = [{"status": s, "label": STATUS_LABEL[s],
                      "count": int(db.scalar(select(func.count()).select_from(GraduationArchiveRecord).where(
                          *base, GraduationArchiveRecord.status == s)) or 0)} for s in STATUS_LABEL]
        filed = int(db.scalar(select(func.count()).select_from(GraduationArchiveRecord).where(
            *base, GraduationArchiveRecord.status == "FILED")) or 0)
        student_total = int(db.scalar(select(func.count()).select_from(GraduationStudent).where(
            GraduationStudent.tenant_id == _tid(), GraduationStudent.is_deleted.is_(False),
            GraduationStudent.record_status == "ACTIVE")) or 0)
        rate = round(filed / student_total * 100, 1) if student_total else 0
        return {"total": total, "byStatus": by_status, "filedCount": filed, "studentTotal": student_total,
                "archiveRate": rate}


def export_archives_xlsx(status=None) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    items, _ = list_archives(1, 100000, status=status)
    headers = ["学生", "学号", "状态", "缺失材料数", "提交时间", "备案时间", "归档批次号"]
    operator, _role = _op()
    title = f"毕设归档台账　导出时间：{datetime.now():%Y-%m-%d %H:%M}　导出人：{operator}"
    wb = Workbook()
    ws = wb.active
    ws.title = "毕设归档台账"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="555555", size=10)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="DCE6F1")
    for c in ws[2]:
        c.font = Font(bold=True); c.fill = fill
    for it in items:
        ws.append([it["studentName"], it["studentNo"], it["statusLabel"], len(it["missingItems"]),
                  (it["submittedAt"] or "")[:19], (it["filedAt"] or "")[:19], it["archiveBatchNo"]])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 18
    ws.freeze_panes = "A3"
    import base64
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return {"filename": f"毕设归档台账_{datetime.now():%Y%m%d_%H%M}.xlsx",
            "contentBase64": base64.b64encode(buf.getvalue()).decode("ascii"), "rowCount": len(items),
            "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
