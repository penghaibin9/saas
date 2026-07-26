"""
学生导出（真实 xlsx + 水印 + t_export_task）。
口径：tenant_id 隔离；导出敏感字段恒脱敏；用途必填写审计。
"""
from __future__ import annotations

import uuid
from datetime import datetime
from pathlib import Path

from app.core.context import current_tenant_id, get_current_user_ctx
from app.core.exceptions import AppException, not_found
from app.db.session import db_enabled, get_sessionmaker
from app.services.file_service import upload_dir

EXPORT_PAGE_SIZE = 2000


def _tid() -> int:
    try:
        tenant_id = int(current_tenant_id() or 0)
    except (TypeError, ValueError):
        tenant_id = 0
    if not tenant_id:
        raise AppException("TENANT_CONTEXT_REQUIRED", "缺少租户上下文，拒绝导入导出写入")
    return tenant_id


def _excel_safe(value):
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def parse_upload_rows(content: bytes, ext: str) -> list[dict]:
    """解析上传文件为行字典列表。正式学校批量业务只接受 xlsx。"""
    if ext != "xlsx":
        raise AppException("FILE_TYPE_NOT_ALLOWED", "学校批量导入仅支持标准 .xlsx 模板")
    import io
    from openpyxl import load_workbook
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as exc:  # noqa: BLE001
        raise AppException("FILE_TYPE_NOT_ALLOWED", "文件不是有效的 xlsx，请用标准模板导出后再上传") from exc
    if not rows:
        return []
    headers = [str(header or "").strip() for header in rows[0]]
    if not any(headers):
        raise AppException("VALIDATION_ERROR", "xlsx 首行字段名不能为空")
    if len(headers) != len(set(headers)):
        raise AppException("VALIDATION_ERROR", "xlsx 字段名不能重复")
    result = []
    for row_index, row in enumerate(rows[1:], start=2):
        values = []
        for column_index, cell in enumerate(row, start=1):
            if isinstance(cell, str) and cell.lstrip().startswith(("=", "+", "@")):
                raise AppException(
                    "VALIDATION_ERROR",
                    f"第{row_index}行第{column_index}列包含公式或危险表达式，禁止导入",
                )
            values.append(cell if cell is not None else "")
        if any(value not in (None, "") for value in values):
            result.append(dict(zip(headers, values)))
    return result


# 旧学生导入的 dry_run / confirm 已随 /import/students/* 一并删除。


def _mask_phone(value: str | None) -> str:
    value = value or ""
    return value[:3] + "****" + value[-4:] if len(value) >= 7 else ("***" if value else "")


def create_students_export(purpose: str, user: dict | None = None) -> dict:
    """真实导出 xlsx（数据范围 + 脱敏 + 公式转义 + 水印），写 t_export_task。"""
    _ensure_feature("studentExport", "学生导出")
    need = _rule("export", "exportNeedPurpose")
    min_len = int(_rule("export", "exportPurposeMinLength") or 0)
    if need and (not purpose or len(purpose.strip()) < min_len):
        raise AppException("VALIDATION_ERROR", f"导出用途必填且不少于 {min_len} 字（平台规则中心配置）")
    user = user or get_current_user_ctx() or {}
    from openpyxl import Workbook
    from app.core.affairs_security import student_directory_scope
    from app.services import db_service

    class_ids, student_ids = student_directory_scope(user)
    items, total = (
        db_service.list_students(1, EXPORT_PAGE_SIZE, class_ids=class_ids, student_ids=student_ids)
        if db_enabled() else ([], 0)
    )
    max_rows = int(_rule("export", "exportMaxRows") or 10000)
    if total > max_rows:
        raise AppException(
            "VALIDATION_ERROR",
            f"导出数据量 {total} 行超过单次上限 {max_rows} 行，请按学院、班级或年级分批导出",
        )
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title="学生主档")
    watermark = (
        f"高校学生全生命周期管理平台 · 导出人：{user.get('realName', '-')} · "
        f"时间：{datetime.now():%Y-%m-%d %H:%M} · 用途：{purpose.strip()} · 敏感字段已脱敏"
    )
    sheet.append([_excel_safe(watermark)])
    sheet.append(["学号", "姓名", "性别", "年级", "班级", "阶段", "学籍状态", "手机号(脱敏)", "风险"])
    page = 1
    while True:
        for row in items:
            sheet.append([
                _excel_safe(row["studentNo"]), _excel_safe(row["realName"]), _excel_safe(row["gender"]),
                _excel_safe(row["grade"]), _excel_safe(row["className"]), _excel_safe(row["currentStage"]),
                _excel_safe(row["studentStatus"]), _excel_safe(row["phoneMasked"]), _excel_safe(row["riskLevel"]),
            ])
        if page * EXPORT_PAGE_SIZE >= total:
            break
        page += 1
        items, _ = db_service.list_students(
            page, EXPORT_PAGE_SIZE, class_ids=class_ids, student_ids=student_ids, count_total=False,
        )
    key = f"exports/{datetime.now():%Y%m%d}/students_{uuid.uuid4().hex[:8]}.xlsx"
    target = upload_dir() / key
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    task = {
        "taskId": "", "status": "SUCCESS", "rowCount": total, "fileKey": key,
        "fileName": target.name, "purpose": purpose.strip(),
        "securityNotice": "导出含学生信息：已脱敏 + 首行水印 + 审计留痕；不得随意外发",
    }
    if db_enabled():
        import hashlib
        from app.models import ExportTask
        from app.services.message_identity import resolve_message_user_id
        db = get_sessionmaker()()
        try:
            record = ExportTask(
                tenant_id=_tid(), export_mode="LIST", module_code="student", row_count=total,
                purpose=purpose.strip(), file_hash=hashlib.sha256(target.read_bytes()).hexdigest(),
                status="SUCCESS", remark=key, created_by=resolve_message_user_id(user) or None,
            )
            db.add(record)
            db.commit()
            db.refresh(record)
            task["taskId"] = str(record.id)
        except Exception:
            db.rollback()
            target.unlink(missing_ok=True)
            raise
        finally:
            db.close()
    else:
        target.unlink(missing_ok=True)
        raise AppException("SERVER_ERROR", "学生导出必须启用MySQL数据库")
    return task


def export_file_path(task_id: str, user: dict | None = None) -> Path:
    from app.core.import_export_auth import assert_export_download
    user = user or get_current_user_ctx() or {}
    tenant_id = _tid()
    if db_enabled() and task_id.isdigit():
        from sqlalchemy import select
        from app.models import ExportTask
        db = get_sessionmaker()()
        try:
            record = db.scalars(select(ExportTask).where(
                ExportTask.id == int(task_id),
                ExportTask.tenant_id == tenant_id,
                ExportTask.is_deleted.is_(False),
            )).first()
            if not record:
                raise not_found("导出任务不存在或文件已清理")
            assert_export_download(
                user, task_tenant_id=record.tenant_id, task_created_by=record.created_by,
                module_code=record.module_code, current_tenant_id=tenant_id,
            )
            if not record.remark:
                raise not_found("导出任务不存在或文件已清理")
            root = upload_dir().resolve()
            candidate = (root / record.remark).resolve()
            try:
                candidate.relative_to(root)
            except ValueError:
                raise not_found("导出任务不存在或文件已清理")
            if candidate.exists() and candidate.is_file() and candidate.suffix.lower() == ".xlsx":
                return candidate
            raise not_found("导出任务不存在或文件已清理")
        finally:
            db.close()
    raise not_found("导出任务不存在或文件已清理")


def _rule(group: str, key: str):
    from app.services.platform_service import safe_rule
    return safe_rule(_tid(), group, key)


def _ensure_feature(key: str, label: str) -> None:
    from app.core.config import settings
    from app.services.platform_service import feature_enabled
    if not feature_enabled(_tid(), key):
        raise AppException(
            "MODULE_NOT_AUTHORIZED",
            f"当前学校套餐未开通「{label}」功能，请联系{settings.support_contact_display}",
        )
