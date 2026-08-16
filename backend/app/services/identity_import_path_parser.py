"""学生/教师身份导入 XLSX 的路径型解析器。

只在 FileObject 已通过安全扫描后调用。ZIP 结构、宏、表头、公式注入与行数规则与既有
identity_import_file_service 保持一致；openpyxl 直接读取本地路径，SHA-256 分块计算，
不把整个 XLSX 拼入内存。
"""
from __future__ import annotations

import hashlib
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from app.core.exceptions import AppException
from app.services.identity_import_file_service import (
    HEADERS,
    MAX_ARCHIVE_ENTRIES,
    MAX_FILE_BYTES,
    MAX_ROWS,
    MAX_UNCOMPRESSED_BYTES,
    RELATION_HEADERS,
    RELATION_REQUIRED_HEADERS,
    REQUIRED_HEADERS,
    STUDENT_HEADERS,
    STUDENT_REQUIRED_HEADERS,
    TEACHER_HEADERS,
    TEACHER_REQUIRED_HEADERS,
    _RELATION_TYPE_ALIASES,
    _TYPE_ALIASES,
    _cell_text,
    _normalize_header,
    _row_cells,
)

CHUNK_SIZE = 1024 * 1024


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while True:
            chunk = stream.read(CHUNK_SIZE)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _validate_archive(path: Path) -> None:
    if not path.is_file() or path.stat().st_size <= 0:
        raise AppException("VALIDATION_ERROR", "上传文件为空")
    if path.stat().st_size > MAX_FILE_BYTES:
        raise AppException("FILE_TOO_LARGE", "导入文件超过 20MB 上限，请拆分后重试")
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ARCHIVE_ENTRIES:
                raise AppException("FILE_TOO_LARGE", "Excel 内部文件数量异常，请使用系统标准模板")
            if sum(max(0, int(info.file_size or 0)) for info in infos) > MAX_UNCOMPRESSED_BYTES:
                raise AppException("FILE_TOO_LARGE", "Excel 解压后超过 100MB 上限，请拆分后重试")
            names = {info.filename.lower() for info in infos}
            if any(name.endswith("vbaproject.bin") for name in names):
                raise AppException("FILE_TYPE_NOT_ALLOWED", "师生账号导入禁止包含宏代码")
            for info in infos:
                normalized = info.filename.replace("\\", "/")
                if normalized.startswith("/") or ".." in normalized.split("/"):
                    raise AppException("FILE_TYPE_NOT_ALLOWED", "Excel 内部路径不安全")
    except AppException:
        raise
    except (zipfile.BadZipFile, OSError) as exc:
        raise AppException("FILE_TYPE_NOT_ALLOWED", "文件不是有效的标准 .xlsx，请重新下载模板") from exc


def _open(path: Path, filename: str, *, headers: tuple, required: set, what: str):
    if not str(filename or "").lower().endswith(".xlsx"):
        raise AppException("FILE_TYPE_NOT_ALLOWED", f"{what}只支持标准 .xlsx 文件")
    _validate_archive(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        sheet = workbook["导入模板"] if "导入模板" in workbook.sheetnames else workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        raw_headers = next(iterator)
    except (StopIteration, KeyError):
        raise AppException("VALIDATION_ERROR", "Excel 没有可导入的工作表或表头")
    except Exception as exc:  # noqa: BLE001
        raise AppException("FILE_TYPE_NOT_ALLOWED", "文件不是有效的标准 .xlsx，请重新下载模板") from exc
    parsed_headers = [_normalize_header(value) for value in raw_headers]
    duplicates = sorted({item for item in parsed_headers if item and parsed_headers.count(item) > 1})
    missing = [item for item in required if item not in parsed_headers]
    unknown = [item for item in parsed_headers if item and item not in headers]
    if duplicates or missing or unknown:
        workbook.close()
        parts = []
        if duplicates:
            parts.append(f"Excel 表头重复：{','.join(duplicates)}")
        if missing:
            parts.append(f"缺少表头：{','.join(missing)}")
        if unknown:
            parts.append(f"不支持的表头：{','.join(unknown)}（请确认没有把另一类模板传到这里）")
        raise AppException("VALIDATION_ERROR", "；".join(parts) + "。请使用系统下载的最新版模板")
    return workbook, iterator, parsed_headers


def parse_identity_xlsx_path(path: str | Path, filename: str, kind: str) -> dict:
    file_path = Path(path)
    kind_up = str(kind or "").upper()
    if kind_up == "STUDENT":
        headers, required, what = STUDENT_HEADERS, STUDENT_REQUIRED_HEADERS, "学生导入"
    elif kind_up == "TEACHER":
        headers, required, what = TEACHER_HEADERS, TEACHER_REQUIRED_HEADERS, "教师导入"
    else:
        raise AppException("VALIDATION_ERROR", "身份导入类型仅支持 STUDENT 或 TEACHER")

    workbook, iterator, parsed_headers = _open(
        file_path, filename, headers=headers, required=required, what=what
    )
    header_index = {name: parsed_headers.index(name) for name in headers if name in parsed_headers}
    students: list[dict] = []
    teachers: list[dict] = []
    raw_rows: list[dict] = []
    errors: list[dict] = []
    total = 0
    try:
        for row_no, values in enumerate(iterator, 2):
            entity = "student" if kind_up == "STUDENT" else "teacher"
            cells, empty = _row_cells(values, headers, header_index, row_no, errors, entity)
            if empty:
                continue
            total += 1
            if total > MAX_ROWS:
                raise AppException("VALIDATION_ERROR", f"单次最多导入 {MAX_ROWS} 行，请拆分文件")
            if kind_up == "STUDENT":
                account_no, name = cells["学号"], cells["姓名"]
                raw_rows.append({"row": row_no, "accountType": "STUDENT", "accountNo": account_no, "name": name})
                if not account_no:
                    errors.append({"row": row_no, "entity": entity, "field": "学号", "error": "学号必填"})
                if not name:
                    errors.append({"row": row_no, "entity": entity, "field": "姓名", "error": "姓名必填"})
                if not cells["班级名称"]:
                    errors.append({
                        "row": row_no,
                        "entity": entity,
                        "field": "班级名称",
                        "error": "班级必填：学生必须归属完整的学院、专业、班级",
                    })
                students.append({
                    "_rowNo": row_no,
                    "studentNo": account_no,
                    "name": name,
                    "collegeName": cells["所属学院"],
                    "majorName": cells["所属专业"],
                    "className": cells["班级名称"],
                    "grade": cells["年级"],
                    "gender": cells["性别"],
                    "idCard": cells["身份证号"],
                })
            else:
                account_no, name = cells["工号"], cells["姓名"]
                raw_rows.append({"row": row_no, "accountType": "TEACHER", "accountNo": account_no, "name": name})
                if not account_no:
                    errors.append({"row": row_no, "entity": entity, "field": "工号", "error": "工号必填"})
                if not name:
                    errors.append({"row": row_no, "entity": entity, "field": "姓名", "error": "姓名必填"})
                if not cells["预设角色编码"]:
                    errors.append({
                        "row": row_no,
                        "entity": entity,
                        "field": "预设角色编码",
                        "error": "教师必须指定预设角色编码",
                    })
                teachers.append({
                    "_rowNo": row_no,
                    "loginName": account_no,
                    "name": name,
                    "departmentName": cells["所属部门"],
                    "positionName": cells["岗位名称"],
                    "roleCodes": cells["预设角色编码"],
                    "scopeType": cells["数据范围类型"],
                    "scopeRef": cells["数据范围引用"],
                })
    finally:
        workbook.close()
    if total == 0:
        raise AppException("VALIDATION_ERROR", "Excel 没有数据行，请填写后再上传")
    return {
        "students": students,
        "teachers": teachers,
        "rawRows": raw_rows,
        "relationships": [],
        "relationErrors": [],
        "errors": errors,
        "totalRows": total,
        "importKind": kind_up,
        "fileName": filename,
        "fileSha256": _sha256(file_path),
    }


def parse_mixed_identity_xlsx_path(path: str | Path, filename: str) -> dict:
    """解析实施中心历史 mixed 师生模板，但保持 FileObject 扫描后的路径型安全边界。"""
    file_path = Path(path)
    workbook, iterator, headers = _open(
        file_path,
        filename,
        headers=HEADERS,
        required=REQUIRED_HEADERS,
        what="师生账号导入",
    )
    header_index = {name: headers.index(name) for name in HEADERS if name in headers}
    students: list[dict] = []
    teachers: list[dict] = []
    raw_rows: list[dict] = []
    errors: list[dict] = []
    relationships: list[dict] = []
    relation_errors: list[dict] = []
    total = 0
    try:
        for row_no, values in enumerate(iterator, 2):
            cells: dict[str, str] = {}
            formula_fields: list[str] = []
            for name in HEADERS:
                index = header_index.get(name)
                value = values[index] if index is not None and index < len(values) else ""
                if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
                    formula_fields.append(name)
                    value = ""
                cells[name] = _cell_text(value)
            if not any(cells.values()):
                continue
            total += 1
            if total > MAX_ROWS:
                raise AppException("VALIDATION_ERROR", f"单次最多导入 {MAX_ROWS} 行，请拆分文件")

            account_type = _TYPE_ALIASES.get(cells["账号类型"].upper())
            account_no = cells["工号/学号"]
            name = cells["姓名"]
            raw_rows.append({
                "row": row_no,
                "accountType": cells["账号类型"],
                "accountNo": account_no,
                "name": name,
            })
            for field in formula_fields:
                errors.append({
                    "row": row_no,
                    "entity": "file",
                    "field": field,
                    "error": "单元格禁止公式或可执行前缀，请改为纯文本",
                })
            if not account_type:
                errors.append({
                    "row": row_no,
                    "entity": "file",
                    "field": "账号类型",
                    "error": "账号类型只能填写 STUDENT 或 TEACHER",
                })
                continue

            row_values = [cells[item] for item in HEADERS]
            if account_type == "STUDENT":
                if (
                    not row_values[7]
                    and row_values[6] in {"男", "女", "未知"}
                    and row_values[5].isdigit()
                    and row_values[4]
                ):
                    row_values[5], row_values[6], row_values[7] = (
                        row_values[4], row_values[5], row_values[6]
                    )
                    row_values[4] = ""
                if row_values[3] and row_values[3].isupper() and "_" in row_values[3]:
                    row_values[10] = row_values[3]
                    row_values[3] = ""
            elif row_values[7].upper() in {"CLASS", "COLLEGE", "SCHOOL"} and row_values[3]:
                row_values[10], row_values[11], row_values[12] = (
                    row_values[3], row_values[7], row_values[8]
                )
                row_values[3], row_values[7], row_values[8] = "", "", ""
            for index, header in enumerate(HEADERS):
                cells[header] = row_values[index]

            if account_type == "STUDENT":
                role = cells["预设角色编码（教师）"]
                if role and role.upper() != "STUDENT":
                    errors.append({
                        "row": row_no,
                        "entity": "student",
                        "field": "预设角色编码（教师）",
                        "error": "学生角色由系统固定为 STUDENT，请留空",
                    })
                students.append({
                    "_rowNo": row_no,
                    "studentNo": account_no,
                    "name": name,
                    "collegeName": cells["所属学院（学生）"],
                    "majorName": cells["所属专业（学生）"],
                    "className": cells["班级名称（学生）"],
                    "grade": cells["年级（学生）"],
                    "gender": cells["性别（学生）"],
                })
            else:
                teachers.append({
                    "_rowNo": row_no,
                    "loginName": account_no,
                    "name": name,
                    "departmentName": cells["所属部门（教师）"],
                    "positionName": cells["岗位名称（教师）"],
                    "roleCodes": cells["预设角色编码（教师）"],
                    "scopeType": cells["数据范围类型（教师）"],
                    "scopeRef": cells["数据范围引用（教师）"],
                })

        if "业务关系" in workbook.sheetnames:
            relation_sheet = workbook["业务关系"]
            relation_iterator = relation_sheet.iter_rows(values_only=True)
            try:
                relation_raw_headers = next(relation_iterator)
            except StopIteration:
                relation_raw_headers = ()
            relation_headers = [_normalize_header(value) for value in relation_raw_headers]
            relation_missing = [
                item for item in RELATION_REQUIRED_HEADERS if item not in relation_headers
            ]
            relation_unknown = [
                item for item in relation_headers if item and item not in RELATION_HEADERS
            ]
            if relation_missing or relation_unknown:
                relation_errors.append({
                    "row": 1,
                    "field": "表头",
                    "error": "业务关系表头不正确："
                    + (f"缺少 {','.join(relation_missing)}；" if relation_missing else "")
                    + (f"不支持 {','.join(relation_unknown)}" if relation_unknown else ""),
                })
            else:
                relation_index = {
                    name: relation_headers.index(name) for name in RELATION_HEADERS
                }
                for row_no, values in enumerate(relation_iterator, 2):
                    cells: dict[str, str] = {}
                    formula_fields: list[str] = []
                    for name in RELATION_HEADERS:
                        index = relation_index[name]
                        value = values[index] if index < len(values) else ""
                        if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
                            formula_fields.append(name)
                            value = ""
                        cells[name] = _cell_text(value)
                    if not any(cells.values()):
                        continue
                    relation_type = (
                        _RELATION_TYPE_ALIASES.get(cells["关系类型"].upper())
                        or _RELATION_TYPE_ALIASES.get(cells["关系类型"])
                    )
                    if not relation_type:
                        relation_errors.append({
                            "row": row_no,
                            "field": "关系类型",
                            "error": f"不支持的关系类型：{cells['关系类型']}",
                        })
                    for field in formula_fields:
                        relation_errors.append({
                            "row": row_no,
                            "field": field,
                            "error": "单元格禁止公式或可执行前缀，请改为纯文本",
                        })
                    if not cells["主体工号"]:
                        relation_errors.append({
                            "row": row_no,
                            "field": "主体工号",
                            "error": "主体工号必填",
                        })
                    if not cells["对象编号/学号"]:
                        relation_errors.append({
                            "row": row_no,
                            "field": "对象编号/学号",
                            "error": "对象编号/学号必填",
                        })
                    relationships.append({
                        "_rowNo": row_no,
                        "relationType": relation_type or cells["关系类型"],
                        "subjectRef": cells["主体工号"],
                        "objectRef": cells["对象编号/学号"],
                        "contextRef": cells["业务批次编号"],
                        "remark": cells["备注"],
                    })
    finally:
        workbook.close()

    if total == 0:
        raise AppException("VALIDATION_ERROR", "Excel 没有数据行，请填写后再上传")
    return {
        "students": students,
        "teachers": teachers,
        "rawRows": raw_rows,
        "relationships": relationships,
        "relationErrors": relation_errors,
        "errors": errors,
        "totalRows": total,
        "fileName": filename,
        "fileSha256": _sha256(file_path),
    }
