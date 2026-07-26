"""Excel(.xlsx) 导入/模板/台账导出通用工具（openpyxl）。

- read_xlsx(file_bytes, header_map)：解析首个 sheet，首行为表头，按「表头文字→字段key」映射为 list[dict]。
- build_template_xlsx(headers, sample)：生成带表头(+示例行)的 .xlsx 字节，供「下载模板」。
- build_ledger_xlsx / build_error_rows_xlsx / pack_xlsx_result：台账导出与错误行 Excel 下载。
不做业务校验（脏数据拦截仍由各域 import_dry_run 负责），只做「文件→行数据」的解析。
"""
from __future__ import annotations

import base64
import io
import zipfile
from datetime import datetime

from openpyxl import Workbook, load_workbook
from app.core.exceptions import AppException

MAX_ROWS = 5000  # 单次导入行数上限，防超大文件
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ZIP_ENTRIES = 2000


async def read_safe_upload(upload) -> bytes:
    """Bound upload memory and reject non-xlsx content before workbook parsing."""
    filename = (getattr(upload, "filename", "") or "").lower()
    if not filename.endswith(".xlsx"):
        raise AppException("VALIDATION_ERROR", "仅支持 .xlsx 文件")
    chunks: list[bytes] = []
    size = 0
    while True:
        chunk = await upload.read(1024 * 1024)
        if not chunk:
            break
        size += len(chunk)
        if size > MAX_UPLOAD_BYTES:
            raise AppException("VALIDATION_ERROR", "Excel 文件不得超过 10MB")
        chunks.append(chunk)
    content = b"".join(chunks)
    validate_xlsx_package(content)
    return content


def validate_xlsx_package(file_bytes: bytes) -> None:
    """Reject zip bombs, path traversal, macros, OLE and external workbook links."""
    if not file_bytes or len(file_bytes) > MAX_UPLOAD_BYTES:
        raise AppException("VALIDATION_ERROR", "Excel 文件为空或超过 10MB")
    stream = io.BytesIO(file_bytes)
    if not zipfile.is_zipfile(stream):
        raise AppException("VALIDATION_ERROR", "文件内容不是有效的 XLSX")
    try:
        with zipfile.ZipFile(stream) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_ENTRIES:
                raise AppException("VALIDATION_ERROR", "XLSX 内部文件数量异常")
            total = 0
            for member in members:
                name = member.filename.replace("\\", "/").lower()
                if name.startswith("/") or ".." in name.split("/"):
                    raise AppException("VALIDATION_ERROR", "XLSX 包含非法路径")
                total += member.file_size
                if total > MAX_UNCOMPRESSED_BYTES:
                    raise AppException("VALIDATION_ERROR", "XLSX 解压后体积异常")
                if member.file_size > 1024 * 1024 and member.compress_size > 0:
                    if member.file_size / member.compress_size > 100:
                        raise AppException("VALIDATION_ERROR", "XLSX 压缩比异常")
                if (
                    name.endswith(".bin")
                    or "vbaproject" in name
                    or "/embeddings/" in name
                    or "/oleobjects/" in name
                    or "/externallinks/" in name
                ):
                    raise AppException("VALIDATION_ERROR", "XLSX 不允许宏、嵌入对象或外部链接")
    except zipfile.BadZipFile:
        raise AppException("VALIDATION_ERROR", "文件内容不是有效的 XLSX") from None


def read_xlsx(file_bytes: bytes, header_map: dict[str, str]) -> list[dict]:
    """.xlsx 字节 → list[dict]。header_map 例：{"企业名称": "name", "统一社会信用代码": "creditCode"}。"""
    validate_xlsx_package(file_bytes)
    wb = load_workbook(
        io.BytesIO(file_bytes), read_only=True, data_only=True, keep_links=False,
    )
    if len(wb.worksheets) > 10:
        wb.close()
        raise AppException("VALIDATION_ERROR", "XLSX 工作表不得超过 10 个")
    # 优先取「导入模板」页（模板含「填写说明」第二页；避免误读说明页）
    ws = wb["导入模板"] if "导入模板" in wb.sheetnames else wb.worksheets[0]
    if ws.max_column > 100:
        wb.close()
        raise AppException("VALIDATION_ERROR", "XLSX 单表列数不得超过 100")
    if ws.max_row > MAX_ROWS + 1:
        wb.close()
        raise AppException("VALIDATION_ERROR", f"单次导入不得超过 {MAX_ROWS} 行")
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    idx_key: dict[int, str] = {}
    for i, h in enumerate(header):
        name = str(h).strip() if h is not None else ""
        name = name.rstrip(" *").strip()  # 去掉表头必填标记「 *」，与 header_map 对齐
        if name in header_map:
            idx_key[i] = header_map[name]
    out: list[dict] = []
    for row in it:
        if row is None:
            continue
        d, empty = {}, True
        for i, key in idx_key.items():
            v = row[i] if i < len(row) else None
            s = "" if v is None else str(v).strip()
            d[key] = s
            if s:
                empty = False
        if not empty:
            out.append(d)
        if len(out) > MAX_ROWS:
            wb.close()
            raise AppException("VALIDATION_ERROR", f"单次导入不得超过 {MAX_ROWS} 行")
    wb.close()
    return out


def build_template_xlsx(headers: list[str], sample: list | None = None,
                        samples: list[list] | None = None,
                        notes: list[str] | None = None,
                        required: list[str] | None = None) -> bytes:
    """生成预制导入模板 .xlsx。
    - headers：表头（第一行）；required：必填表头集合（会在表头后加「*」）。
    - sample/samples：示例行（sample 单行；samples 多行）。
    - notes：填写说明（生成第二个 sheet「填写说明」，一行一条）。
    """
    from openpyxl.styles import Font, PatternFill
    req = set(required or [])
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append([(h + " *") if h in req else h for h in headers])
    head_fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = head_fill
    for row in ([sample] if sample else []) + (samples or []):
        ws.append(row)
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i)].width = 24
    ws.freeze_panes = "A2"
    if notes:
        ns = wb.create_sheet("填写说明")
        ns.append(["填写说明（请阅读后删除本页或忽略，仅导入「导入模板」页）"])
        ns["A1"].font = Font(bold=True, size=12)
        for i, n in enumerate(notes, start=2):
            ns.cell(row=i, column=1, value=n)
        ns.column_dimensions["A"].width = 80
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_ledger_xlsx(sheet_title: str, headers: list[str], data_rows: list[list],
                      watermark: str | None = None) -> bytes:
    """生成台账 .xlsx（可选首行水印 + 表头 + 数据行）。"""
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "台账")[:28]
    row_idx = 1
    if watermark:
        ws.append([watermark])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        ws["A1"].font = Font(bold=True, color="666666", size=10)
        row_idx = 2
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)
        cell.fill = head_fill
    for row in data_rows:
        ws.append(row)
    for i in range(1, len(headers) + 1):
        col = chr(64 + i) if i <= 26 else "A"
        ws.column_dimensions[col].width = 18
    if watermark:
        ws.freeze_panes = "A3"
    else:
        ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_error_rows_xlsx(template_headers: list[str], rows: list[dict], errors: list[dict],
                          row_values: callable) -> bytes:
    """错误行 Excel：原表头 + 错误字段 + 错误原因（仅含出错行）。"""
    err_by_row: dict[int, list[dict]] = {}
    for e in errors or []:
        err_by_row.setdefault(int(e.get("rowNo") or 0), []).append(e)
    headers = template_headers + ["错误字段", "错误原因"]
    data: list[list] = []
    for i, r in enumerate(rows or []):
        row_no = i + 1
        if row_no not in err_by_row:
            continue
        msgs = err_by_row[row_no]
        fields = "；".join(str(x.get("field") or "") for x in msgs)
        reason = "；".join(str(x.get("message") or "") for x in msgs)
        data.append(list(row_values(r)) + [fields, reason])
    return build_ledger_xlsx("导入错误行", headers, data,
                             watermark="以下为未通过预校验的行，请修正后重新导入（仅导入「导入模板」页）")


def pack_xlsx_result(content: bytes, filename: str, row_count: int,
                     tenant_label: str = "") -> dict:
    """API 响应：xlsx 二进制转 base64，文件名补时间戳。"""
    name = filename if filename.endswith(".xlsx") else f"{filename}.xlsx"
    if not any(c.isdigit() for c in name[-20:]):
        stem = name[:-5] if name.endswith(".xlsx") else name
        suffix = f"_{tenant_label}" if tenant_label else ""
        name = f"{stem}{suffix}_{datetime.now():%Y%m%d_%H%M}.xlsx"
    return {
        "filename": name,
        "contentBase64": base64.b64encode(content).decode("ascii"),
        "rowCount": row_count,
        "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
