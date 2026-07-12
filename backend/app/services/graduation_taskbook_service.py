"""毕业设计中心 · 任务书服务。

导师下达→学生确认→变更需重确认（对齐老系统 t_gd_task_book 语义与商业系统"任务书"惯例）。
一生一条当前记录；变更时旧内容快照进 history_json，版本号自增。
确认后若学生阶段仍处于 TASKBOOK_CONFIRM，则推进到 GUIDING（指导中）。

隔离说明：不引用实习/迎新域文件；Excel 台账导出用 openpyxl 直连。
"""
from __future__ import annotations

from datetime import datetime

from sqlalchemy import func, select

from app.core.context import get_current_user_ctx
from app.core.exceptions import AppException, not_found
from app.models import GraduationAuditTrail, GraduationStudent, GraduationTaskBook
from app.services.db_service import _iso, _tid, session
from app.services.graduation_scope_service import accessible_student_ids, assert_student_access

STATUS_LABEL = {"PENDING_CONFIRM": "待学生确认", "CONFIRMED": "已确认", "CHANGE_PENDING": "变更待确认"}
STATUS_TONE = {"PENDING_CONFIRM": "warning", "CONFIRMED": "success", "CHANGE_PENDING": "warning"}


def _op() -> tuple[str, str]:
    u = get_current_user_ctx() or {}
    return u.get("realName") or "系统", u.get("roleName") or u.get("currentRoleCode") or ""


def _audit(db, bid, action, detail="", before="", after=""):
    n, r = _op()
    db.add(GraduationAuditTrail(tenant_id=_tid(), biz_type="TASKBOOK", biz_id=str(bid), action=action,
                                operator=n, role_name=r, detail=detail, before_val=before,
                                after_val=after, occurred_at=datetime.utcnow()))


def _stu(db, sid) -> GraduationStudent:
    s = db.get(GraduationStudent, int(sid))
    if not s or s.is_deleted or s.tenant_id != _tid():
        raise not_found("毕设学生不存在或不在当前数据范围内")
    return assert_student_access(db, s, "taskbook")


def _row(t: GraduationTaskBook, stu=None) -> dict:
    return {
        "id": str(t.id), "gdStudentId": str(t.gd_student_id),
        "studentName": stu.name if stu else "", "studentNo": stu.student_no if stu else "",
        "advisorName": stu.advisor_name if stu else "",
        "objective": t.objective or "", "content": t.content or "",
        "progressPlan": t.progress_plan or "", "outcomeRequirement": t.outcome_requirement or "",
        "taskbookVersion": t.taskbook_version, "status": t.status,
        "statusLabel": STATUS_LABEL.get(t.status, t.status), "statusTone": STATUS_TONE.get(t.status, "default"),
        "issuedBy": t.issued_by or "", "issuedAt": _iso(t.issued_at), "confirmedAt": _iso(t.confirmed_at),
        "changeReason": t.change_reason or "", "history": t.history_json or [],
        "updatedAt": _iso(t.updated_at),
    }


def list_taskbooks(page: int, page_size: int, keyword=None, status=None) -> tuple[list[dict], int]:
    with session() as db:
        scope_ids = accessible_student_ids(db, _tid())
        q = select(GraduationTaskBook).where(GraduationTaskBook.tenant_id == _tid(),
                                              GraduationTaskBook.is_deleted.is_(False),
                                              GraduationTaskBook.gd_student_id.in_(scope_ids or [-1]))
        if status:
            q = q.where(GraduationTaskBook.status == status)
        rows = db.scalars(q.order_by(GraduationTaskBook.id.desc())).all()
        items = []
        for t in rows:
            stu = db.get(GraduationStudent, t.gd_student_id)
            if keyword and (not stu or keyword.strip() not in (stu.name or "")):
                continue
            items.append(_row(t, stu))
        total = len(items)
        start = (max(1, page) - 1) * page_size
        return items[start:start + page_size], total


def get_taskbook(gd_student_id) -> dict:
    with session() as db:
        stu = _stu(db, gd_student_id)
        t = db.scalars(select(GraduationTaskBook).where(
            GraduationTaskBook.tenant_id == _tid(), GraduationTaskBook.gd_student_id == stu.id,
            GraduationTaskBook.is_deleted.is_(False))).first()
        if not t:
            return {"exists": False, "gdStudentId": str(stu.id), "studentName": stu.name,
                    "advisorName": stu.advisor_name or ""}
        return {"exists": True, **_row(t, stu)}


def issue_taskbook(gd_student_id, body: dict) -> dict:
    with session() as db:
        stu = _stu(db, gd_student_id)
        if not stu.mentor_id and not stu.advisor_name:
            raise AppException("DATA_CONFLICT", "该生尚未分配导师，无法下达任务书")
        dup = db.scalars(select(GraduationTaskBook).where(
            GraduationTaskBook.tenant_id == _tid(), GraduationTaskBook.gd_student_id == stu.id,
            GraduationTaskBook.is_deleted.is_(False))).first()
        if dup:
            raise AppException("DATA_CONFLICT", "该生已有任务书，如需调整请使用「变更」")
        n, _ = _op()
        t = GraduationTaskBook(
            tenant_id=_tid(), gd_student_id=stu.id, mentor_id=stu.mentor_id,
            objective=body.get("objective"), content=body.get("content"),
            progress_plan=body.get("progressPlan"), outcome_requirement=body.get("outcomeRequirement"),
            taskbook_version=1, status="PENDING_CONFIRM", issued_by=n, issued_at=datetime.utcnow(),
            history_json=[])
        db.add(t)
        db.flush()
        _audit(db, t.id, "下达任务书", detail=f"{stu.name} v1")
        db.commit()
        return _row(t, stu)


def confirm_taskbook(gd_student_id) -> dict:
    with session() as db:
        stu = _stu(db, gd_student_id)
        t = db.scalars(select(GraduationTaskBook).where(
            GraduationTaskBook.tenant_id == _tid(), GraduationTaskBook.gd_student_id == stu.id,
            GraduationTaskBook.is_deleted.is_(False))).first()
        if not t:
            raise not_found("尚未下达任务书")
        if t.status not in ("PENDING_CONFIRM", "CHANGE_PENDING"):
            raise AppException("DATA_CONFLICT", "当前任务书无需确认")
        before = t.status
        t.status = "CONFIRMED"
        t.confirmed_at = datetime.utcnow()
        if stu.stage == "TASKBOOK_CONFIRM":
            stu.stage = "GUIDING"
        _audit(db, t.id, "学生确认任务书", before=before, after="CONFIRMED")
        db.commit()
        return _row(t, stu)


def change_taskbook(gd_student_id, body: dict) -> dict:
    reason = (body.get("reason") or "").strip()
    if len(reason) < 5:
        raise AppException("VALIDATION_ERROR", "变更原因必填且不少于 5 字")
    with session() as db:
        stu = _stu(db, gd_student_id)
        t = db.scalars(select(GraduationTaskBook).where(
            GraduationTaskBook.tenant_id == _tid(), GraduationTaskBook.gd_student_id == stu.id,
            GraduationTaskBook.is_deleted.is_(False))).first()
        if not t:
            raise not_found("尚未下达任务书")
        if t.status != "CONFIRMED":
            raise AppException("DATA_CONFLICT", "仅「已确认」任务书可发起变更")
        history = t.history_json or []
        history.append({"version": t.taskbook_version, "objective": t.objective, "content": t.content,
                        "progressPlan": t.progress_plan, "outcomeRequirement": t.outcome_requirement,
                        "confirmedAt": _iso(t.confirmed_at)})
        t.history_json = history
        for src, col in [("objective", "objective"), ("content", "content"),
                         ("progressPlan", "progress_plan"), ("outcomeRequirement", "outcome_requirement")]:
            if body.get(src) is not None:
                setattr(t, col, body[src])
        t.taskbook_version += 1
        t.status = "CHANGE_PENDING"
        t.change_reason = reason
        t.confirmed_at = None
        _audit(db, t.id, "变更任务书（待重新确认）", reason)
        db.commit()
        return _row(t, stu)


def taskbook_stats() -> dict:
    with session() as db:
        scope_ids = accessible_student_ids(db, _tid())
        base = [GraduationTaskBook.tenant_id == _tid(), GraduationTaskBook.is_deleted.is_(False),
                GraduationTaskBook.gd_student_id.in_(scope_ids or [-1])]
        total = int(db.scalar(select(func.count()).select_from(GraduationTaskBook).where(*base)) or 0)
        by_status = [{"status": s, "label": STATUS_LABEL[s],
                      "count": int(db.scalar(select(func.count()).select_from(GraduationTaskBook).where(
                          *base, GraduationTaskBook.status == s)) or 0)} for s in STATUS_LABEL]
        no_taskbook = int(db.scalar(select(func.count()).select_from(GraduationStudent).where(
            GraduationStudent.tenant_id == _tid(), GraduationStudent.is_deleted.is_(False),
            GraduationStudent.record_status == "ACTIVE", GraduationStudent.mentor_id.isnot(None),
            GraduationStudent.stage != "TOPIC_SELECTING",
            GraduationStudent.id.in_(scope_ids or [-1]))) or 0) - total
        return {"total": total, "byStatus": by_status, "noTaskbookYet": max(0, no_taskbook)}


def export_taskbooks_xlsx(status=None) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    items, _ = list_taskbooks(1, 100000, status=status)
    headers = ["学生", "学号", "导师", "版本", "状态", "任务目标", "成果要求", "确认时间"]
    operator, _role = _op()
    title = f"任务书台账　导出时间：{datetime.now():%Y-%m-%d %H:%M}　导出人：{operator}"
    wb = Workbook()
    ws = wb.active
    ws.title = "任务书台账"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="555555", size=10)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="DCE6F1")
    for c in ws[2]:
        c.font = Font(bold=True); c.fill = fill
    for it in items:
        ws.append([it["studentName"], it["studentNo"], it["advisorName"], it["taskbookVersion"],
                  it["statusLabel"], it["objective"], it["outcomeRequirement"], (it["confirmedAt"] or "")[:19]])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 20
    ws.freeze_panes = "A3"
    import base64
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return {"filename": f"任务书台账_{datetime.now():%Y%m%d_%H%M}.xlsx",
            "contentBase64": base64.b64encode(buf.getvalue()).decode("ascii"), "rowCount": len(items),
            "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
