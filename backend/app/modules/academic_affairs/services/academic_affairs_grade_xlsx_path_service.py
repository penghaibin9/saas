"""C-W4 Path-backed grade XLSX parser.

Upload bytes are streamed to a temporary file and validated by the shared
``xlsx_util.validate_xlsx_path`` security policy merged from PR #145. This module
contains only grade-business parsing limits; it deliberately does not duplicate
ZIP traversal/bomb/macro/external-link checks.
"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from app.core.exceptions import AppException
from app.services import xlsx_util

MAX_GRADE_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_GRADE_ROWS = 5000
MAX_GRADE_COLUMNS = 100


async def read_grade_upload_path(upload, header_map: dict[str, str]) -> list[dict]:
    filename = str(getattr(upload, "filename", "") or "").lower()
    if not filename.endswith(".xlsx"):
        raise AppException("VALIDATION_ERROR", "仅支持 .xlsx 文件")

    temp_path: Path | None = None
    size = 0
    try:
        with tempfile.NamedTemporaryFile(prefix="aa-grade-", suffix=".xlsx", delete=False) as handle:
            temp_path = Path(handle.name)
            while True:
                chunk = await upload.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > MAX_GRADE_UPLOAD_BYTES:
                    raise AppException("FILE_TOO_LARGE", "成绩导入文件超过 20MB，请拆分后重试")
                handle.write(chunk)
            handle.flush()
            os.fsync(handle.fileno())

        xlsx_util.validate_xlsx_path(
            temp_path,
            max_bytes=MAX_GRADE_UPLOAD_BYTES,
            too_large_code="FILE_TOO_LARGE",
            too_large_message="成绩导入文件超过 20MB，请拆分后重试",
        )
        workbook = load_workbook(temp_path, read_only=True, data_only=True, keep_links=False)
        try:
            if len(workbook.worksheets) > 10:
                raise AppException("VALIDATION_ERROR", "XLSX 工作表不得超过 10 个")
            sheet = workbook["导入模板"] if "导入模板" in workbook.sheetnames else workbook.worksheets[0]
            if sheet.max_column > MAX_GRADE_COLUMNS:
                raise AppException("VALIDATION_ERROR", "XLSX 单表列数不得超过 100")
            if sheet.max_row > MAX_GRADE_ROWS + 1:
                raise AppException("VALIDATION_ERROR", f"单次导入不得超过 {MAX_GRADE_ROWS} 行")

            iterator = sheet.iter_rows(values_only=True)
            try:
                header = next(iterator)
            except StopIteration:
                return []
            index_map: dict[int, str] = {}
            for index, value in enumerate(header):
                title = str(value).strip() if value is not None else ""
                title = title.rstrip(" *").strip()
                if title in header_map:
                    index_map[index] = header_map[title]
            if not index_map:
                raise AppException("VALIDATION_ERROR", "Excel 表头与成绩导入模板不匹配")

            rows: list[dict] = []
            for values in iterator:
                item: dict[str, str] = {}
                empty = True
                for index, key in index_map.items():
                    value = values[index] if index < len(values) else None
                    normalized = "" if value is None else str(value).strip()
                    item[key] = normalized
                    if normalized:
                        empty = False
                if not empty:
                    rows.append(item)
                if len(rows) > MAX_GRADE_ROWS:
                    raise AppException("VALIDATION_ERROR", f"单次导入不得超过 {MAX_GRADE_ROWS} 行")
            return rows
        finally:
            workbook.close()
    finally:
        if temp_path is not None:
            try:
                temp_path.unlink(missing_ok=True)
            except OSError:
                pass
