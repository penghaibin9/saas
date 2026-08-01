"""教务文件与数据交换中心（阶段 7 收口）。

核心约束：
- 上传请求只登记原始 FileObject 和 SCANNING ImportJob，绝不解析隔离文件；
- 任务详情在文件 CLEAN/AVAILABLE 后推进 PARSING → VALIDATED/VALIDATION_FAILED；
- 预检只持久化摘要与脱敏错误，原始 rows 不进入任务 JSON；
- 确认只接受 jobId + expectedVersion，并重新读取同一 FileObject、复算 rowDigest、重新预检；
- 学籍、成绩、排课写入继续委托原领域事务，公共层负责租约、文件安全和任务生命周期；
- 导出生成 FileObject + ExportJob，支持过期、撤销和一次性下载票据。
"""
from __future__ import annotations

import hashlib
import json
import uuid
import zipfile
from datetime import timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.exceptions import AppException, not_found
from app.db.session import get_sessionmaker
from app.services import data_exchange_job_service as jobs

ACADEMIC_MODULE_CODE = "ACADEMIC_AFFAIRS"
ACADEMIC_ROSTER_IMPORT = "ACADEMIC_ROSTER"
ACADEMIC_GRADE_IMPORT = "ACADEMIC_GRADE"
ACADEMIC_SCHEDULE_IMPORT = "ACADEMIC_SCHEDULE"
ACADEMIC_ROSTER_EXPORT = "ACADEMIC_ROSTER"
MAX_IMPORT_BYTES = 20 * 1024 * 1024


def _json_safe(value: Any) -> Any:
    return json.loads(json.dumps(value, ensure_ascii=False, default=str))


def _row_digest(rows: list[dict]) -> str:
    return hashlib.sha256(
        json.dumps(_json_safe(rows), ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()


def _error_items(preview: dict) -> list[dict]:
    raw = preview.get("errors") or preview.get("errorList") or []
    if isinstance(raw, dict):
        raw = [{"row": key, "message": value} for key, value in raw.items()]
    return [item if isinstance(item, dict) else {"message": str(item)} for item in raw]


def _preview_counts(rows: list[dict], preview: dict) -> tuple[int, int, int]:
    total = int(preview.get("totalRows") or preview.get("total") or len(rows))
    invalid = int(preview.get("invalidRows") or preview.get("invalid") or preview.get("errorRows") or 0)
    if not invalid:
        invalid = len({str(item.get("row") or item.get("rowNo") or "") for item in _error_items(preview)})
    valid_raw = preview.get("validRows") if preview.get("validRows") is not None else preview.get("valid")
    valid = int(valid_raw if valid_raw is not None else max(0, total - invalid))
    return total, valid, invalid


def _redacted_rows(rows: list[dict], limit: int = 200) -> list[dict]:
    sensitive = {"idCard", "id_card", "phone", "mobile", "bankAccount", "bank_account"}
    return [
        {key: value for key, value in dict(row).items() if key not in sensitive}
        for row in rows[:limit]
    ]


def create_academic_import_job(
    *,
    filename: str,
    source_file_id: int,
    import_type: str,
    context: dict[str, Any] | None,
    user: dict,
) -> dict:
    """只登记权威文件与 SCANNING Job；解析只能由 refresh_import_job 在安全门后执行。"""
    from app.models.data_exchange import ImportJob
    from app.models.file import FileObject

    import_type = str(import_type or "").upper()
    if import_type not in {ACADEMIC_ROSTER_IMPORT, ACADEMIC_GRADE_IMPORT, ACADEMIC_SCHEDULE_IMPORT}:
        raise AppException("VALIDATION_ERROR", "不支持的教务导入类型")
    db = get_sessionmaker()()
    try:
        file_row = db.scalars(select(FileObject).where(
            FileObject.id == int(source_file_id),
            FileObject.tenant_id == jobs._tenant_id(),
            FileObject.is_deleted.is_(False),
        )).first()
        if not file_row:
            raise not_found("导入原始文件不存在")
        if int(file_row.size_bytes or 0) <= 0:
            raise AppException("VALIDATION_ERROR", "导入文件不能为空")
        if int(file_row.size_bytes or 0) > MAX_IMPORT_BYTES:
            raise AppException("FILE_TOO_LARGE", "教务导入文件超过 20MB，请拆分后重试")
        actor_id = jobs._actor_id(user)
        row = ImportJob(
            tenant_id=jobs._tenant_id(),
            module_code=ACADEMIC_MODULE_CODE,
            import_type=import_type,
            source_file_id=int(source_file_id),
            adapter_type=jobs.IMPORT_ADAPTER_EXCEL,
            adapter_ref=f"AA-{import_type}-{uuid.uuid4().hex}",
            template_version="v1",
            status="SCANNING",
            total_rows=0,
            valid_rows=0,
            invalid_rows=0,
            operator_id=actor_id,
            operator_name=jobs._actor_name(user),
            expires_at=jobs._now() + timedelta(hours=jobs.IMPORT_JOB_TTL_HOURS),
            source_snapshot_json={
                "authority": "SOURCE_FILE_OBJECT",
                "fileName": str(filename or file_row.file_name or "academic_import.xlsx"),
                "fileObjectId": str(source_file_id),
                "fileSha256": str(file_row.sha256 or ""),
                "spec": f"academicAffairs.{import_type.lower()}.v1",
                "context": _json_safe(context),
                "preview": {"totalRows": 0, "validRows": 0, "invalidRows": 0},
            },
            created_by=actor_id,
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        job_id = str(row.id)
    finally:
        db.close()
    return refresh_import_job(job_id, user=user)


def create_roster_import_job(
    *,
    filename: str,
    source_file_id: int,
    user: dict,
    content: bytes | None = None,
) -> dict:
    """兼容旧服务调用签名；content 故意不参与解析或权威判断。"""
    _ = content
    return create_academic_import_job(
        filename=filename,
        source_file_id=source_file_id,
        import_type=ACADEMIC_ROSTER_IMPORT,
        context={},
        user=user,
    )


def _source_file_path(import_row, user: dict) -> Path:
    from app.models.file import FileObject
    from app.services.file_scan_service import assert_file_ready_for_business
    from app.services.storage import get_backend

    if not import_row.source_file_id:
        raise AppException("DATA_CONFLICT", "导入任务缺少原始文件，请重新上传")
    assert_file_ready_for_business(str(import_row.source_file_id), user=user)
    db = get_sessionmaker()()
    try:
        file_row = db.scalars(select(FileObject).where(
            FileObject.id == int(import_row.source_file_id),
            FileObject.tenant_id == jobs._tenant_id(),
            FileObject.is_deleted.is_(False),
        )).first()
        if not file_row:
            raise not_found("导入原始文件不存在")
        if int(file_row.size_bytes or 0) > MAX_IMPORT_BYTES:
            raise AppException("FILE_TOO_LARGE", "教务导入文件超过 20MB，请拆分后重试")
        source_path = get_backend().fetch_local(file_row.file_key)
        if not source_path or not source_path.exists():
            raise not_found("导入原始文件不存在或已清理")
        source_path = Path(source_path)
        if source_path.stat().st_size <= 0:
            raise AppException("VALIDATION_ERROR", "导入文件不能为空")
        if source_path.stat().st_size > MAX_IMPORT_BYTES:
            raise AppException("FILE_TOO_LARGE", "教务导入文件超过 20MB，请拆分后重试")
        return source_path
    finally:
        db.close()


def _read_xlsx_path(source_path: Path, header_map: dict[str, str]) -> list[dict]:
    if not zipfile.is_zipfile(source_path):
        raise AppException("VALIDATION_ERROR", "文件内容不是有效的 XLSX")
    try:
        with zipfile.ZipFile(source_path) as archive:
            members = archive.infolist()
            if len(members) > 2000:
                raise AppException("VALIDATION_ERROR", "XLSX 内部文件数量异常")
            uncompressed = 0
            for member in members:
                name = member.filename.replace("\\", "/").lower()
                if name.startswith("/") or ".." in name.split("/"):
                    raise AppException("VALIDATION_ERROR", "XLSX 包含非法路径")
                uncompressed += member.file_size
                if uncompressed > 50 * 1024 * 1024:
                    raise AppException("VALIDATION_ERROR", "XLSX 解压后体积异常")
                if member.file_size > 1024 * 1024 and member.compress_size > 0:
                    if member.file_size / member.compress_size > 100:
                        raise AppException("VALIDATION_ERROR", "XLSX 压缩比异常")
                if (
                    name.endswith(".bin")
                    or "vbaproject" in name
                    or "/embeddings/" in name
                    or "/oleobjects/" in name
                    or "/externallinks/" in name
                ):
                    raise AppException("VALIDATION_ERROR", "XLSX 不允许宏、嵌入对象或外部链接")
    except zipfile.BadZipFile:
        raise AppException("VALIDATION_ERROR", "文件内容不是有效的 XLSX") from None

    workbook = load_workbook(source_path, read_only=True, data_only=True, keep_links=False)
    try:
        if len(workbook.worksheets) > 10:
            raise AppException("VALIDATION_ERROR", "XLSX 工作表不得超过 10 个")
        sheet = workbook["导入模板"] if "导入模板" in workbook.sheetnames else workbook.worksheets[0]
        if sheet.max_column > 100:
            raise AppException("VALIDATION_ERROR", "XLSX 单表列数不得超过 100")
        if sheet.max_row > 5001:
            raise AppException("VALIDATION_ERROR", "单次导入不得超过 5000 行")
        iterator = sheet.iter_rows(values_only=True)
        try:
            header = next(iterator)
        except StopIteration:
            return []
        index_map: dict[int, str] = {}
        for index, value in enumerate(header):
            title = str(value).strip() if value is not None else ""
            title = title.rstrip(" *").strip()
            if title in header_map:
                index_map[index] = header_map[title]
        rows: list[dict] = []
        for values in iterator:
            item: dict[str, str] = {}
            empty = True
            for index, key in index_map.items():
                value = values[index] if index < len(values) else None
                normalized = "" if value is None else str(value).strip()
                item[key] = normalized
                if normalized:
                    empty = False
            if not empty:
                rows.append(item)
            if len(rows) > 5000:
                raise AppException("VALIDATION_ERROR", "单次导入不得超过 5000 行")
        return rows
    finally:
        workbook.close()


def _parse_and_validate(import_row, source_path: Path, user: dict) -> tuple[list[dict], dict]:
    snapshot = dict(import_row.source_snapshot_json or {})
    context = dict(snapshot.get("context") or {})
    if import_row.import_type == ACADEMIC_ROSTER_IMPORT:
        from app.modules.academic_affairs.services import academic_affairs_service as roster
        rows = _read_xlsx_path(source_path, roster.build_roster_import_spec().header_map)
        return rows, roster.roster_import_dry_run(rows)
    if import_row.import_type == ACADEMIC_GRADE_IMPORT:
        from app.modules.academic_affairs.services import academic_affairs_grade_service as grade
        from app.services import xlsx_util
        task_id = int(context.get("taskId") or 0)
        if not task_id:
            raise AppException("DATA_CONFLICT", "成绩导入任务缺少成绩任务编号")
        rows = _read_xlsx_path(source_path, grade.IMPORT_HEADER_MAP)
        return rows, grade.grade_import_dry_run(task_id, user, rows)
    if import_row.import_type == ACADEMIC_SCHEDULE_IMPORT:
        from app.modules.academic_affairs.services import academic_affairs_schedule_service as schedule
        from app.services import xlsx_util
        batch_id = int(context.get("batchId") or 0)
        if not batch_id:
            raise AppException("DATA_CONFLICT", "排课导入任务缺少排课批次编号")
        rows = _read_xlsx_path(source_path, schedule.IMPORT_HEADER_MAP)
        if len(rows) > schedule.IMPORT_MAX_ROWS:
            raise AppException("VALIDATION_ERROR", f"单批导入行数不得超过 {schedule.IMPORT_MAX_ROWS} 行")
        rows = schedule.sanitize_import_rows(rows)
        return rows, {
            "totalRows": len(rows),
            "validRows": len(rows),
            "invalidRows": 0,
            "errors": [],
        }
    raise AppException("DATA_CONFLICT", "未知教务导入类型")


def _detail(row, errors: list[dict] | None = None, rows: list[dict] | None = None) -> dict:
    result = jobs._import_row(row)
    snapshot = dict(row.source_snapshot_json or {})
    result["preview"] = {
        **dict(snapshot.get("preview") or {}),
        "errors": errors or [],
        "rows": _redacted_rows(rows or []),
    }
    return result


def _stored_errors(db, job_id: int) -> list[dict]:
    from app.models.data_exchange import ImportRowError
    rows = db.scalars(select(ImportRowError).where(
        ImportRowError.tenant_id == jobs._tenant_id(),
        ImportRowError.import_job_id == int(job_id),
        ImportRowError.is_deleted.is_(False),
    ).order_by(ImportRowError.row_no, ImportRowError.id)).all()
    return [
        {
            "row": item.row_no,
            "field": item.field_code,
            "code": item.error_code,
            "message": item.error_message,
        }
        for item in rows
    ]


def refresh_import_job(job_id: str, *, user: dict) -> dict:
    """在详情轮询中推进安全扫描后的解析；并发实例通过状态锁只允许一个解析者。"""
    from app.models.data_exchange import ImportRowError

    db = get_sessionmaker()()
    try:
        row = jobs._owned_import(db, job_id, user)
        if row.status not in {"SCANNING", "PARSING"}:
            return _detail(row, _stored_errors(db, int(row.id)))
        if row.expires_at and row.expires_at <= jobs._now():
            row.status = "EXPIRED"
            row.version = int(row.version or 0) + 1
            db.commit()
            return _detail(row)
    finally:
        db.close()

    try:
        db = get_sessionmaker()()
        try:
            current = jobs._owned_import(db, job_id, user)
            source_path = _source_file_path(current, user)
        finally:
            db.close()
    except AppException as exc:
        if exc.code == "FILE_NOT_READY":
            db = get_sessionmaker()()
            try:
                return _detail(jobs._owned_import(db, job_id, user))
            finally:
                db.close()
        db = get_sessionmaker()()
        try:
            row = jobs._owned_import(db, job_id, user, lock=True)
            row.status = "VALIDATION_FAILED"
            row.invalid_rows = max(1, int(row.invalid_rows or 0))
            row.error_message = str(exc)[:4000]
            row.version = int(row.version or 0) + 1
            db.commit()
            return _detail(row, [{"code": exc.code, "message": str(exc)}])
        finally:
            db.close()

    db = get_sessionmaker()()
    try:
        row = jobs._owned_import(db, job_id, user, lock=True)
        if row.status not in {"SCANNING", "PARSING"}:
            return _detail(row, _stored_errors(db, int(row.id)))
        row.status = "PARSING"
        row.version = int(row.version or 0) + 1
        db.commit()
    finally:
        db.close()

    try:
        db = get_sessionmaker()()
        try:
            row = jobs._owned_import(db, job_id, user)
            rows, preview = _parse_and_validate(row, source_path, user)
        finally:
            db.close()
        errors = _error_items(preview)
        total, valid, invalid = _preview_counts(rows, preview)
        digest = _row_digest(rows)
    except Exception as exc:
        db = get_sessionmaker()()
        try:
            row = jobs._owned_import(db, job_id, user, lock=True)
            row.status = "VALIDATION_FAILED"
            row.invalid_rows = max(1, int(row.invalid_rows or 0))
            row.error_message = str(exc)[:4000]
            row.version = int(row.version or 0) + 1
            db.commit()
            return _detail(row, [{"code": getattr(exc, "code", "PARSE_ERROR"), "message": str(exc)}])
        finally:
            db.close()

    db = get_sessionmaker()()
    try:
        row = jobs._owned_import(db, job_id, user, lock=True)
        db.execute(delete(ImportRowError).where(
            ImportRowError.tenant_id == jobs._tenant_id(),
            ImportRowError.import_job_id == int(row.id),
        ))
        actor_id = jobs._actor_id(user)
        for item in errors:
            raw = dict(item.get("raw") or item.get("rowData") or {})
            for field in ("idCard", "id_card", "phone", "mobile", "bankAccount", "bank_account"):
                raw.pop(field, None)
            db.add(ImportRowError(
                tenant_id=jobs._tenant_id(),
                import_job_id=row.id,
                sheet_name=str(item.get("sheetName") or "导入模板")[:100],
                row_no=int(item.get("row") or item.get("rowNo") or 0) or None,
                field_code=str(item.get("field") or item.get("fieldCode") or "")[:100] or None,
                error_code=str(item.get("code") or item.get("errorCode") or "VALIDATION_ERROR")[:80],
                error_message=str(item.get("message") or item.get("reason") or "预检失败")[:1000],
                raw_snapshot_json=_json_safe(raw) if raw else None,
                created_by=actor_id,
            ))
        snapshot = dict(row.source_snapshot_json or {})
        snapshot["rowDigest"] = digest
        snapshot["preview"] = {"totalRows": total, "validRows": valid, "invalidRows": invalid}
        row.source_snapshot_json = snapshot
        row.total_rows = total
        row.valid_rows = valid
        row.invalid_rows = invalid
        row.status = "VALIDATED" if invalid == 0 else "VALIDATION_FAILED"
        row.error_message = None if invalid == 0 else "预检存在错误，请修正原始文件后重新上传"
        row.version = int(row.version or 0) + 1
        db.commit()
        db.refresh(row)
        return _detail(row, errors, rows)
    finally:
        db.close()


def confirm_academic_import(job_id: str, *, lease: str, user: dict) -> dict:
    """从同一 FileObject 重读、复算摘要并委托原领域事务确认。"""
    db = get_sessionmaker()()
    try:
        row = jobs._owned_import(db, job_id, user)
        if row.adapter_type != jobs.IMPORT_ADAPTER_EXCEL:
            raise AppException("DATA_CONFLICT", "该任务不是教务 Excel 导入")
        if row.lease_token != lease:
            raise AppException("DATA_CONFLICT", "导入任务确认租约已失效")
        snapshot = dict(row.source_snapshot_json or {})
        context = dict(snapshot.get("context") or {})
        import_type = row.import_type
        source_path = _source_file_path(row, user)
        rows, preview = _parse_and_validate(row, source_path, user)
    finally:
        db.close()

    digest = _row_digest(rows)
    if snapshot.get("rowDigest") and snapshot["rowDigest"] != digest:
        raise AppException("DATA_CONFLICT", "导入文件解析结果已变化，请重新上传预检")
    _total, _valid, invalid = _preview_counts(rows, preview)
    if invalid:
        raise AppException("VALIDATION_ERROR", "确认前重新预检发现错误，请重新上传修正后的文件")

    if import_type == ACADEMIC_ROSTER_IMPORT:
        from app.modules.academic_affairs.services import academic_affairs_service as roster
        result = roster.roster_import_confirm(rows)
    elif import_type == ACADEMIC_GRADE_IMPORT:
        from app.modules.academic_affairs.services import academic_affairs_grade_service as grade
        result = grade.grade_import_confirm(int(context.get("taskId") or 0), user, rows)
    elif import_type == ACADEMIC_SCHEDULE_IMPORT:
        from app.modules.academic_affairs.services import academic_affairs_schedule_service as schedule
        result = schedule.import_items(int(context.get("batchId") or 0), user, rows)
    else:
        raise AppException("DATA_CONFLICT", "未知教务导入类型")
    if not isinstance(result, dict):
        result = {"result": result}
    result = dict(result)
    result.setdefault("confirmedRows", len(rows))
    return result


def confirm_roster_import(job_id: str, *, lease: str, user: dict) -> dict:
    return confirm_academic_import(job_id, lease=lease, user=user)


def create_roster_export_job(
    *,
    user: dict,
    purpose: str,
    keyword: str | None = None,
    status: str | None = None,
) -> dict:
    from app.models.data_exchange import ExportJob
    from app.modules.academic_affairs.services import academic_affairs_service as roster

    purpose = str(purpose or "").strip()
    if len(purpose) < 5:
        raise AppException("VALIDATION_ERROR", "导出用途必填且不少于 5 个字")
    content = roster.export_roster_xlsx(user, purpose, keyword, status)
    _rows, total = roster.roster(user, keyword, status, page=1, page_size=1)
    adapter_ref = f"AA-ROSTER-EXPORT-{uuid.uuid4().hex}"
    file_id = jobs._write_generated_file(
        content,
        "roster_ledger.xlsx",
        biz_id=adapter_ref,
        user=user,
        security_level="SENSITIVE",
    )
    actor_id = jobs._actor_id(user)
    db = get_sessionmaker()()
    try:
        row = ExportJob(
            tenant_id=jobs._tenant_id(),
            module_code=ACADEMIC_MODULE_CODE,
            export_type=ACADEMIC_ROSTER_EXPORT,
            purpose=purpose,
            adapter_type="ACADEMIC_EXPORT",
            adapter_ref=adapter_ref,
            filter_snapshot_json={"keyword": keyword or "", "status": status or ""},
            data_scope_snapshot_json={
                "actorUserId": str(user.get("userId") or ""),
                "roleCode": str(user.get("currentRoleCode") or ""),
            },
            status="SUCCEEDED",
            progress=100,
            row_count=int(total or 0),
            file_object_id=file_id,
            expires_at=jobs._now() + timedelta(hours=jobs.RECEIPT_TTL_HOURS),
            operator_id=actor_id,
            created_by=actor_id,
            finished_at=jobs._now(),
            result_json={"fileObjectId": str(file_id)},
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        return jobs._export_row(row)
    finally:
        db.close()


def list_academic_jobs(*, user: dict, job_type: str = "", status: str = "", page: int = 1, page_size: int = 20) -> dict:
    from app.models.data_exchange import ExportJob, ImportJob

    page = max(1, int(page or 1))
    page_size = min(100, max(1, int(page_size or 20)))
    actor_id = jobs._actor_id(user)
    wanted = str(job_type or "").upper()
    wanted_status = str(status or "").upper()
    items: list[dict] = []
    db = get_sessionmaker()()
    try:
        if wanted in {"", "IMPORT"}:
            stmt = select(ImportJob).where(
                ImportJob.tenant_id == jobs._tenant_id(),
                ImportJob.module_code == ACADEMIC_MODULE_CODE,
                ImportJob.is_deleted.is_(False),
            )
            if actor_id:
                stmt = stmt.where(ImportJob.operator_id == actor_id)
            if wanted_status:
                stmt = stmt.where(ImportJob.status == wanted_status)
            items.extend(jobs._import_row(row) for row in db.scalars(stmt.order_by(ImportJob.id.desc())).all())
        if wanted in {"", "EXPORT"}:
            stmt = select(ExportJob).where(
                ExportJob.tenant_id == jobs._tenant_id(),
                ExportJob.module_code == ACADEMIC_MODULE_CODE,
                ExportJob.is_deleted.is_(False),
            )
            if actor_id:
                stmt = stmt.where(ExportJob.operator_id == actor_id)
            if wanted_status:
                stmt = stmt.where(ExportJob.status == wanted_status)
            items.extend(jobs._export_row(row) for row in db.scalars(stmt.order_by(ExportJob.id.desc())).all())
        items.sort(key=lambda item: (item.get("createdAt") or "", int(item["id"])), reverse=True)
        total = len(items)
        start = (page - 1) * page_size
        return {"list": items[start:start + page_size], "total": total, "page": page, "pageSize": page_size}
    finally:
        db.close()
