"""Local six-sheet XLSX adapter for ordinary Program imports.

XLSX package security remains owned by ``app.services.xlsx_util``.  This module
only adds the Program-specific multi-sheet contract that the shared first-sheet
reader cannot express: exact sheet/header sets, one bounded total row budget and
conversion into the already-frozen grouped/normalized Program source shape.

It owns no FileObject/ImportJob lifecycle, database session, domain preflight or
public dispatcher.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.core.exceptions import AppException
from app.services import xlsx_util

from .academic_affairs_school_setup_program_file_exchange_spec import (
    PROGRAM_FILLING_NOTES,
    PROGRAM_GROUP_BY_SHEET,
    PROGRAM_HEADER_MAP_BY_GROUP,
    PROGRAM_SHEET_ORDER,
    _required_headers,
)
from .academic_affairs_school_setup_program_import_adapter import (
    normalize_program_import_rows,
)

_NOTES_SHEET = "填写说明"
_ALLOWED_SHEETS = frozenset((*PROGRAM_SHEET_ORDER, _NOTES_SHEET))

_SAMPLE_BY_SHEET = {
    "培养方案": {
        "培养方案系列键": "SOFTWARE-2026",
        "版本": 1,
        "方案名称": "软件技术2026级人才培养方案",
        "专业ID": 1001,
        "适用年级": "2026",
        "毕业总学分": 3,
        "学制年限(断言)": 3,
    },
    "方案课程": {
        "培养方案系列键": "SOFTWARE-2026",
        "版本": 1,
        "课程代码": "CS101",
        "课程版本": 1,
        "开课学期": 1,
        "课程模块": "MAJOR_CORE",
        "编班方式": "ADMIN_FIXED",
        "学分快照(断言)": "",
    },
    "学分要求": {
        "培养方案系列键": "SOFTWARE-2026",
        "版本": 1,
        "课程模块": "MAJOR_CORE",
        "目标学分": 3,
    },
    "实践环节": {
        "培养方案系列键": "SOFTWARE-2026",
        "版本": 1,
        "实践环节名称": "综合实训",
        "实践环节类型": "COURSE_DESIGN",
        "开设学期": 1,
        "周数": 1,
        "学分": 1,
        "组织方式": "CENTRALIZED",
        "考核方式": "CHECK",
        "地点/承担单位": "校内实训中心",
        "排序": 1,
    },
    "毕业要求": {
        "培养方案系列键": "SOFTWARE-2026",
        "版本": 1,
        "毕业要求类别": "ABILITY",
        "毕业要求内容": "完成专业综合项目并通过考核",
        "排序": 1,
    },
    "适用范围": {
        "培养方案系列键": "SOFTWARE-2026",
        "版本": 1,
        "专业ID": 1001,
        "适用年级": "2026",
        "绑定范围": "MAJOR_GRADE",
        "班级ID": "",
    },
}


def build_program_import_template() -> bytes:
    """Generate the canonical six-sheet Program workbook plus filling notes."""
    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="DCE6F1")
    for sheet_name in PROGRAM_SHEET_ORDER:
        group = PROGRAM_GROUP_BY_SHEET[sheet_name]
        header_map = PROGRAM_HEADER_MAP_BY_GROUP[group]
        required = set(_required_headers(group))
        ws = wb.create_sheet(sheet_name)
        ws.append([
            f"{title} *" if title in required else title
            for title in header_map
        ])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = head_fill
        sample = _SAMPLE_BY_SHEET[sheet_name]
        ws.append([sample.get(title, "") for title in header_map])
        ws.freeze_panes = "A2"
        for index in range(1, len(header_map) + 1):
            ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = 24

    notes = wb.create_sheet(_NOTES_SHEET)
    notes.append(["填写说明（Program 普通导入；请勿修改工作表名称和模板表头）"])
    notes["A1"].font = Font(bold=True, size=12)
    for index, note in enumerate(PROGRAM_FILLING_NOTES, start=2):
        notes.cell(row=index, column=1, value=note)
    notes.column_dimensions["A"].width = 100

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _header_title(value: object) -> str:
    return str(value or "").strip().rstrip(" *").strip()


def _read_sheet_rows(ws, *, group: str) -> list[dict]:
    header_map = PROGRAM_HEADER_MAP_BY_GROUP[group]
    required_headers = set(_required_headers(group))
    if ws.max_column > 100:
        raise AppException("VALIDATION_ERROR", f"工作表“{ws.title}”列数不得超过 100")
    if ws.max_row > xlsx_util.MAX_ROWS + 1:
        raise AppException(
            "VALIDATION_ERROR",
            f"工作表“{ws.title}”单表不得超过 {xlsx_util.MAX_ROWS} 行数据",
        )

    iterator = ws.iter_rows(values_only=True)
    try:
        raw_header = next(iterator)
    except StopIteration:
        raw_header = ()
    headers = tuple(_header_title(value) for value in raw_header if _header_title(value))
    if len(headers) != len(set(headers)):
        raise AppException("VALIDATION_ERROR", f"工作表“{ws.title}”存在重复表头")
    missing = [title for title in required_headers if title not in headers]
    unknown = [title for title in headers if title not in header_map]
    if missing or unknown:
        parts = []
        if missing:
            parts.append("缺少必填表头：" + "、".join(sorted(missing)))
        if unknown:
            parts.append("存在未知表头：" + "、".join(sorted(unknown)))
        raise AppException(
            "VALIDATION_ERROR",
            f"工作表“{ws.title}”表头与模板不一致；" + "；".join(parts),
        )

    # Optional columns may be omitted, but supplied columns must retain canonical
    # names. This keeps source compatibility without silently accepting typos.
    header_index = {
        index: header_map[_header_title(value)]
        for index, value in enumerate(raw_header)
        if _header_title(value) in header_map
    }
    out: list[dict] = []
    for raw_row in iterator:
        row: dict[str, str] = {}
        non_empty = False
        for index, key in header_index.items():
            value = raw_row[index] if index < len(raw_row) else None
            text = "" if value is None else str(value).strip()
            row[key] = text
            non_empty = non_empty or bool(text)
        if non_empty:
            out.append(row)
    return out


def read_program_workbook(
    file_bytes: bytes,
    *,
    max_bytes: int = xlsx_util.MAX_UPLOAD_BYTES,
) -> dict[str, list[dict]]:
    """Validate one Program workbook and return rows keyed by logical group."""
    xlsx_util.validate_xlsx_package(file_bytes, max_bytes=max_bytes)
    wb = load_workbook(
        BytesIO(file_bytes),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        names = tuple(wb.sheetnames)
        missing = [name for name in PROGRAM_SHEET_ORDER if name not in names]
        unexpected = [name for name in names if name not in _ALLOWED_SHEETS]
        if missing or unexpected:
            parts = []
            if missing:
                parts.append("缺少工作表：" + "、".join(missing))
            if unexpected:
                parts.append("存在未知工作表：" + "、".join(unexpected))
            raise AppException(
                "VALIDATION_ERROR",
                "培养方案工作簿结构与模板不一致；" + "；".join(parts),
            )
        if tuple(name for name in names if name in PROGRAM_SHEET_ORDER) != PROGRAM_SHEET_ORDER:
            raise AppException(
                "VALIDATION_ERROR",
                "培养方案六张业务工作表顺序必须与服务端模板一致",
            )

        grouped: dict[str, list[dict]] = {}
        total_rows = 0
        for sheet_name in PROGRAM_SHEET_ORDER:
            group = PROGRAM_GROUP_BY_SHEET[sheet_name]
            rows = _read_sheet_rows(wb[sheet_name], group=group)
            grouped[group] = rows
            total_rows += len(rows)
            if total_rows > xlsx_util.MAX_ROWS:
                raise AppException(
                    "VALIDATION_ERROR",
                    f"培养方案工作簿六表合计不得超过 {xlsx_util.MAX_ROWS} 行数据",
                )
        return grouped
    finally:
        wb.close()


def parse_and_normalize_program_workbook(
    file_bytes: bytes,
    *,
    max_bytes: int = xlsx_util.MAX_UPLOAD_BYTES,
) -> tuple[dict[str, list[dict]], list[dict]]:
    """Return both grouped source rows and the frozen canonical normalized rows."""
    grouped = read_program_workbook(file_bytes, max_bytes=max_bytes)
    normalized = normalize_program_import_rows(grouped)
    return grouped, normalized
