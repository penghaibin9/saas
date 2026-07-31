"""教务同步导出到公共 ExportJob 的兼容适配服务（阶段 7 收口）。

旧页面仍按原 URL 接收文件；本服务在返回文件之前强制创建 FileObject + ExportJob，
再通过公共短时一次性票据读取结果。文件字节响应统一委托公共文件权威合同。
"""
from __future__ import annotations

import io
import json
import uuid
import zipfile
from datetime import timedelta
from typing import Any

from fastapi.responses import FileResponse
from openpyxl import load_workbook

from app.api.v1.file_contract import validated_local_file_response
from app.db.session import get_sessionmaker
from app.services import data_exchange_job_service as jobs

ACADEMIC_MODULE_CODE = "ACADEMIC_AFFAIRS"
ADAPTER_TYPE = "ACADEMIC_COMPAT_EXPORT"
_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _safe_snapshot(value: dict[str, Any] | None) -> dict[str, Any]:
    return json.loads(json.dumps(value or {}, ensure_ascii=False, default=str))


def _row_count(content: bytes, filename: str) -> int:
    lower = str(filename or "").lower()
    try:
        if lower.endswith(".xlsx"):
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            try:
                sheet = workbook.active
                return max(0, int(sheet.max_row or 0) - 1)
            finally:
                workbook.close()
        if lower.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(content), "r") as archive:
                return len([item for item in archive.infolist() if not item.is_dir()])
    except Exception:
        # 行数只是任务展示信息，生成文件本身仍由领域服务和后续下载校验负责。
        return 0
    return 0


def task_backed_file_response(
    *,
    content: bytes,
    filename: str,
    export_type: str,
    purpose: str,
    user: dict,
    parameters: dict[str, Any] | None = None,
    media_type: str = _XLSX_MEDIA,
) -> FileResponse:
    """把领域导出字节纳入公共任务生命周期，并保持旧页面下载响应不变。"""
    from app.models.data_exchange import ExportJob

    purpose = str(purpose or "").strip()
    if len(purpose) < 5:
        from app.core.exceptions import AppException
        raise AppException("VALIDATION_ERROR", "导出用途必填且不少于 5 个字")
    safe_filename = str(filename or "academic_export.xlsx").replace("/", "_").replace("\\", "_")
    adapter_ref = f"AA-COMPAT-{uuid.uuid4().hex}"
    file_id = jobs._write_generated_file(
        bytes(content),
        safe_filename,
        biz_id=adapter_ref,
        user=user,
        security_level="SENSITIVE",
    )
    actor_id = jobs._actor_id(user)
    now = jobs._now()
    db = get_sessionmaker()()
    try:
        row = ExportJob(
            tenant_id=jobs._tenant_id(),
            module_code=ACADEMIC_MODULE_CODE,
            export_type=str(export_type or "ACADEMIC_EXPORT").upper()[:80],
            purpose=purpose[:500],
            adapter_type=ADAPTER_TYPE,
            adapter_ref=adapter_ref,
            filter_snapshot_json=_safe_snapshot(parameters),
            data_scope_snapshot_json={
                "actorUserId": str(user.get("userId") or ""),
                "roleCode": str(user.get("currentRoleCode") or ""),
                "dataScope": str(user.get("dataScope") or ""),
                "collegeId": str(user.get("collegeId") or ""),
            },
            status="SUCCEEDED",
            progress=100,
            row_count=_row_count(content, safe_filename),
            file_object_id=file_id,
            expires_at=now + timedelta(hours=jobs.RECEIPT_TTL_HOURS),
            operator_id=actor_id,
            created_by=actor_id,
            finished_at=now,
            result_json={
                "fileObjectId": str(file_id),
                "compatibilityDownload": True,
                "filename": safe_filename,
            },
        )
        db.add(row)
        db.commit()
        db.refresh(row)
        job = jobs._export_row(row)
    finally:
        db.close()

    ticket = jobs.create_download_ticket(
        job["id"],
        expected_version=int(job["version"]),
        user=user,
    )
    path, resolved_name = jobs.consume_download_ticket(job["id"], ticket["ticket"], user=user)
    return validated_local_file_response(
        path,
        filename=resolved_name or safe_filename,
        media_type=media_type,
        audit_action="ACADEMIC_EXPORT_COMPAT_DOWNLOAD",
        audit_target=f"academic-export-job:{job['id']}",
        audit_detail={"jobId": str(job["id"]), "exportType": str(export_type or "")},
        headers={
            "X-Export-Job-Id": str(job["id"]),
            "X-File-Center-Contract": "EXPORT_JOB",
            "Deprecation": "true",
            "Link": '</api/v1/academic-affairs/file-exchange/jobs>; rel="successor-version"',
        },
    )
