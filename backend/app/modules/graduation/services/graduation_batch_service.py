"""毕业设计中心 · 毕设批次服务（DB_ENABLED=true 走本模块）。

一届/一轮毕业设计的组织容器：阶段时间轴 + 规则配置（查重/评阅/答辩/成绩）+ 状态机
DRAFT→RUNNING→CLOSED→ARCHIVED（草稿可作废 VOIDED）。租户隔离 + 软删 + 审计 t_gd_audit_trail。

隔离说明（与 Cursor 的 Excel 导入导出工作互不干扰）：
本文件不引用实习域/迎新域文件，也不依赖 app.services.xlsx_util（Excel 导出用 openpyxl 直连）。
"""
from __future__ import annotations

import base64
import io
from datetime import datetime

from sqlalchemy import func, or_, select

from app.core.context import get_current_user_ctx
from app.core.exceptions import AppException, not_found
from app.models import GraduationAuditTrail, GraduationBatch
from app.services.db_service import _iso, _tid, session

STATUS_LABEL = {"DRAFT": "草稿", "RUNNING": "进行中", "CLOSED": "已结束",
                "ARCHIVED": "已归档", "VOIDED": "已作废"}
STATUS_TONE = {"DRAFT": "default", "RUNNING": "success", "CLOSED": "warning",
               "ARCHIVED": "default", "VOIDED": "danger"}

DEFAULT_STAGES = [
    {"code": "TOPIC", "name": "选题", "startDate": None, "endDate": None},
    {"code": "PROPOSAL", "name": "开题", "startDate": None, "endDate": None},
    {"code": "MIDTERM", "name": "中期检查", "startDate": None, "endDate": None},
    {"code": "SUBMISSION", "name": "成果提交", "startDate": None, "endDate": None},
    {"code": "PLAGIARISM", "name": "查重", "startDate": None, "endDate": None},
    {"code": "REVIEW", "name": "评阅", "startDate": None, "endDate": None},
    {"code": "DEFENSE", "name": "答辩", "startDate": None, "endDate": None},
    {"code": "GRADE", "name": "成绩", "startDate": None, "endDate": None},
]
DEFAULT_RULES = {
    "plagiarism": {"thresholdPercent": 30, "mustPassToDefense": True},
    "review": {"mode": "DOUBLE_BLIND", "minReviewers": 2},
    "defense": {"groupSize": 5, "passScore": 60},
    "score": {"advisorWeight": 0.4, "reviewerWeight": 0.3, "defenseWeight": 0.3},
}


def _op() -> tuple[str, str]:
    u = get_current_user_ctx() or {}
    return u.get("realName") or "系统", u.get("roleName") or ""


def _audit(db, bid, action, detail="", before="", after=""):
    n, r = _op()
    db.add(GraduationAuditTrail(tenant_id=_tid(), biz_type="BATCH", biz_id=str(bid), action=action,
                                operator=n, role_name=r, detail=detail, before_val=before,
                                after_val=after, occurred_at=datetime.utcnow()))


def _parse_dt(v):
    if not v:
        return None
    s = str(v).strip().replace("Z", "").replace("/", "-")
    if not s:
        return None
    try:
        return datetime.fromisoformat(s[:19])
    except ValueError:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            return None


def _get(db, bid) -> GraduationBatch:
    b = db.get(GraduationBatch, int(bid))
    if not b or b.is_deleted or b.tenant_id != _tid():
        raise not_found("毕设批次不存在或不在当前数据范围内")
    return b


def _row(b: GraduationBatch) -> dict:
    return {
        "id": str(b.id), "batchName": b.batch_name, "batchNo": b.batch_no,
        "academicYear": b.academic_year or "", "gradeYear": b.grade_year or "",
        "collegeScope": b.college_scope or "", "plannedCount": b.planned_count,
        "startDate": _iso(b.start_date), "endDate": _iso(b.end_date),
        "status": b.status, "statusLabel": STATUS_LABEL.get(b.status, b.status),
        "statusTone": STATUS_TONE.get(b.status, "default"),
        "archiveStatus": b.archive_status, "remark": b.remark or "",
        "updatedAt": _iso(b.updated_at),
    }


# ═══════════ 列表 / 详情 ═══════════

def list_batches(page: int, page_size: int, keyword=None, status=None) -> tuple[list[dict], int]:
    with session() as db:
        q = select(GraduationBatch).where(GraduationBatch.tenant_id == _tid(),
                                          GraduationBatch.is_deleted.is_(False))
        if keyword:
            like = f"%{keyword.strip()}%"
            q = q.where(or_(GraduationBatch.batch_name.like(like),
                            GraduationBatch.batch_no.like(like),
                            GraduationBatch.grade_year.like(like)))
        if status:
            q = q.where(GraduationBatch.status == status)
        total = int(db.scalar(select(func.count()).select_from(q.subquery())) or 0)
        rows = db.scalars(q.order_by(GraduationBatch.id.desc())
                          .offset((max(1, page) - 1) * page_size).limit(page_size)).all()
        return [_row(b) for b in rows], total


def get_batch(bid) -> dict:
    with session() as db:
        b = _get(db, bid)
        trail = db.scalars(select(GraduationAuditTrail).where(
            GraduationAuditTrail.tenant_id == _tid(), GraduationAuditTrail.biz_type == "BATCH",
            GraduationAuditTrail.biz_id == str(b.id)).order_by(
            GraduationAuditTrail.id.desc()).limit(30)).all()
        return {
            **_row(b),
            "stages": b.stage_config or DEFAULT_STAGES,
            "rules": b.rules_config or DEFAULT_RULES,
            "previousStatus": b.previous_status or "", "lastTransitionAt": _iso(b.last_transition_at),
            "lastTransitionBy": b.last_transition_by or "", "transitionReason": b.transition_reason or "",
            "archivedAt": _iso(b.archived_at), "archivedBy": b.archived_by or "",
            "voidReason": b.void_reason or "",
            "auditTrail": [{"action": a.action, "operator": a.operator or "", "roleName": a.role_name or "",
                            "detail": a.detail or "", "occurredAt": _iso(a.occurred_at)} for a in trail],
        }


# ═══════════ 增 / 改 ═══════════

def create_batch(body: dict) -> dict:
    with session() as db:
        name = (body.get("batchName") or "").strip()
        no = (body.get("batchNo") or "").strip()
        if not name or not no:
            raise AppException("VALIDATION_ERROR", "批次名称与批次编号必填")
        dup = db.scalars(select(GraduationBatch).where(
            GraduationBatch.tenant_id == _tid(), GraduationBatch.batch_no == no,
            GraduationBatch.is_deleted.is_(False))).first()
        if dup:
            raise AppException("DATA_CONFLICT", f"批次编号已存在：{no}")
        b = GraduationBatch(
            tenant_id=_tid(), batch_name=name, batch_no=no,
            academic_year=body.get("academicYear"), grade_year=body.get("gradeYear"),
            college_scope=body.get("collegeScope"),
            start_date=_parse_dt(body.get("startDate")), end_date=_parse_dt(body.get("endDate")),
            planned_count=int(body.get("plannedCount") or 0),
            stage_config=_validate_stages(body.get("stages") or DEFAULT_STAGES),
            rules_config=_validate_and_merge_rules(None, body.get("rules") or DEFAULT_RULES),
            remark=body.get("remark"), status="DRAFT", archive_status="NOT_ARCHIVED")
        db.add(b)
        db.flush()
        _audit(db, b.id, "CREATE", detail=f"{name}/{no}")
        db.commit()
        return _row(b)


def update_batch(bid, body: dict) -> dict:
    with session() as db:
        b = _get(db, bid)
        if b.status in ("CLOSED", "ARCHIVED", "VOIDED"):
            raise AppException("DATA_CONFLICT", f"「{STATUS_LABEL.get(b.status)}」批次不可编辑")
        for src, col in [("batchName", "batch_name"), ("academicYear", "academic_year"),
                         ("gradeYear", "grade_year"), ("collegeScope", "college_scope"),
                         ("remark", "remark")]:
            if src in body and body[src] is not None:
                setattr(b, col, body[src])
        if "plannedCount" in body and body["plannedCount"] is not None:
            b.planned_count = int(body["plannedCount"])
        if "startDate" in body:
            b.start_date = _parse_dt(body["startDate"])
        if "endDate" in body:
            b.end_date = _parse_dt(body["endDate"])
        if body.get("stages"):
            b.stage_config = _validate_stages(body["stages"])
        if body.get("rules"):
            b.rules_config = _validate_and_merge_rules(b.rules_config, body["rules"])
        b.version += 1
        _audit(db, b.id, "UPDATE")
        db.commit()
        return _row(b)


def _validate_stages(stages: list) -> list:
    """Validate stage date windows: each stage end>=start, and stages do not reverse chronological order."""
    if not stages:
        return DEFAULT_STAGES
    cleaned = []
    prev_end = None
    for idx, raw in enumerate(stages):
        if not isinstance(raw, dict):
            raise AppException("VALIDATION_ERROR", f"第 {idx + 1} 个阶段格式无效")
        code = str(raw.get("code") or "").strip()
        name = str(raw.get("name") or code or f"阶段{idx + 1}").strip()
        if not code:
            raise AppException("VALIDATION_ERROR", f"第 {idx + 1} 个阶段缺少 code")
        start = _parse_dt(raw.get("startDate"))
        end = _parse_dt(raw.get("endDate"))
        if start and end and end < start:
            raise AppException("VALIDATION_ERROR", f"阶段「{name}」结束日期不能早于开始日期")
        if prev_end and start and start < prev_end:
            raise AppException(
                "VALIDATION_ERROR",
                f"阶段顺序错误：阶段「{name}」开始日期早于上一阶段结束日期",
            )
        if end:
            prev_end = end
        elif start:
            prev_end = start
        item = dict(raw)
        item["code"], item["name"] = code, name
        cleaned.append(item)
    return cleaned


def _validate_and_merge_rules(existing: dict | None, patch: dict | None) -> dict:
    """Deep-merge rules and enforce score weights sum to 100% (±0.5pp) when score block present."""
    base = {**(DEFAULT_RULES or {}), **(existing or {})}
    patch = patch or {}
    merged = dict(base)
    for key, value in patch.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = {**merged[key], **value}
        else:
            merged[key] = value
    score = merged.get("score") or {}
    weights = [
        float(score.get("advisorWeight") or 0),
        float(score.get("reviewerWeight") or 0),
        float(score.get("defenseWeight") or 0),
    ]
    total = round(sum(weights) * 100, 4)
    if abs(total - 100.0) > 0.5:
        raise AppException(
            "VALIDATION_ERROR",
            f"成绩权重之和必须等于 100%（当前 {total:.1f}%："
            f"指导教师 {weights[0] * 100:.1f}% + 评阅 {weights[1] * 100:.1f}% + 答辩 {weights[2] * 100:.1f}%）",
        )
    plag = merged.get("plagiarism") or {}
    if "thresholdPercent" in plag:
        thr = float(plag.get("thresholdPercent") or 0)
        if thr < 0 or thr > 100:
            raise AppException("VALIDATION_ERROR", "查重阈值须在 0–100 之间")
    return merged


def set_stages(bid, stages: list) -> dict:
    with session() as db:
        b = _get(db, bid)
        if b.status in ("ARCHIVED", "VOIDED"):
            raise AppException("DATA_CONFLICT", "已归档/已作废批次不可改阶段配置")
        b.stage_config = _validate_stages(stages or DEFAULT_STAGES)
        _audit(db, b.id, "SET_STAGES")
        db.commit()
        return get_batch(b.id)


def set_rules(bid, rules: dict) -> dict:
    with session() as db:
        b = _get(db, bid)
        if b.status in ("ARCHIVED", "VOIDED"):
            raise AppException("DATA_CONFLICT", "已归档/已作废批次不可改规则配置")
        b.rules_config = _validate_and_merge_rules(b.rules_config, rules)
        _audit(db, b.id, "SET_RULES")
        db.commit()
        return get_batch(b.id)


# ═══════════ 状态机 ═══════════

def _transition(db, b: GraduationBatch, to_status: str, reason: str = ""):
    b.previous_status = b.status
    b.status = to_status
    b.last_transition_at = datetime.utcnow()
    b.last_transition_by, _ = _op()
    b.transition_reason = reason


def activate_batch(bid) -> dict:
    with session() as db:
        b = _get(db, bid)
        if b.status != "DRAFT":
            raise AppException("DATA_CONFLICT", "仅「草稿」批次可启用")
        _transition(db, b, "RUNNING")
        _audit(db, b.id, "ACTIVATE")
        db.commit()
        return _row(b)


def close_batch(bid) -> dict:
    with session() as db:
        b = _get(db, bid)
        if b.status != "RUNNING":
            raise AppException("DATA_CONFLICT", "仅「进行中」批次可结束")
        _transition(db, b, "CLOSED")
        _audit(db, b.id, "CLOSE")
        db.commit()
        return _row(b)


def archive_batch(bid) -> dict:
    with session() as db:
        b = _get(db, bid)
        if b.status != "CLOSED":
            raise AppException("DATA_CONFLICT", "仅「已结束」批次可归档")
        _transition(db, b, "ARCHIVED")
        b.archive_status = "ARCHIVED"
        b.archived_at = datetime.utcnow()
        b.archived_by, _ = _op()
        _audit(db, b.id, "ARCHIVE")
        db.commit()
        return _row(b)


def void_batch(bid, reason: str) -> dict:
    if not reason or len(reason.strip()) < 5:
        raise AppException("VALIDATION_ERROR", "作废原因必填且不少于 5 个字")
    with session() as db:
        b = _get(db, bid)
        if b.status != "DRAFT":
            raise AppException("DATA_CONFLICT", "仅「草稿」批次可作废")
        _transition(db, b, "VOIDED", reason=reason.strip())
        b.void_reason = reason.strip()
        _audit(db, b.id, "VOID", detail=reason.strip())
        db.commit()
        return _row(b)


# ═══════════ 统计 ═══════════

def batch_stats() -> dict:
    with session() as db:
        base = [GraduationBatch.tenant_id == _tid(), GraduationBatch.is_deleted.is_(False)]
        total = int(db.scalar(select(func.count()).select_from(GraduationBatch).where(*base)) or 0)
        by_status = [{"status": s, "label": STATUS_LABEL[s],
                      "count": int(db.scalar(select(func.count()).select_from(GraduationBatch).where(
                          *base, GraduationBatch.status == s)) or 0)} for s in STATUS_LABEL]
        planned = int(db.scalar(select(func.coalesce(func.sum(GraduationBatch.planned_count), 0)).where(
            *base, GraduationBatch.status == "RUNNING")) or 0)
        return {"total": total, "byStatus": by_status, "runningPlanned": planned}


# ═══════════ Excel 台账导出（openpyxl 直连，不依赖 xlsx_util）═══════════

def export_batches_xlsx(keyword=None, status=None) -> dict:
    """毕设批次台账导出（.xlsx，按当前筛选，含抬头/来源行；调用方写审计）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    items, _ = list_batches(1, 100000, keyword=keyword, status=status)
    headers = ["批次名称", "批次编号", "学年", "届", "适用范围", "计划人数",
               "开始", "结束", "状态", "备注"]
    operator, _role = _op()
    cond = "、".join(filter(None, [
        f"状态={STATUS_LABEL.get(status, status)}" if status else "",
        f"关键词={keyword}" if keyword else ""])) or "全部"
    title = (f"毕设批次台账　导出时间：{datetime.now():%Y-%m-%d %H:%M}　"
             f"筛选：{cond}　导出人：{operator}")
    wb = Workbook()
    ws = wb.active
    ws.title = "毕设批次台账"
    # 第 1 行抬头/来源（合并跨列）
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="555555", size=10)
    # 第 2 行表头
    ws.append(headers)
    fill = PatternFill("solid", fgColor="DCE6F1")
    for c in ws[2]:
        c.font = Font(bold=True)
        c.fill = fill
    for it in items:
        ws.append([it["batchName"], it["batchNo"], it["academicYear"], it["gradeYear"],
                   it["collegeScope"], it["plannedCount"], (it["startDate"] or "")[:10],
                   (it["endDate"] or "")[:10], it["statusLabel"], it["remark"]])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 18
    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    return {
        "filename": f"毕设批次台账_{datetime.now():%Y%m%d_%H%M}.xlsx",
        "contentBase64": base64.b64encode(buf.getvalue()).decode("ascii"),
        "rowCount": len(items),
        "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
