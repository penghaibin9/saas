"""毕业设计中心服务层。

V9.2 在这里保留模块内兼容绑定：历史调用方仍引用原 service 模块，
但开题/成果/学生/成绩/归档列表统一切到 SQL 读模型。绑定只替换只读入口，
不触碰提交、批阅、状态机等 canonical 写链。
"""
from __future__ import annotations


def _install_proposal_read_model() -> None:
    from app.modules.graduation.services import graduation_proposal_read_service as proposal_read
    from app.modules.graduation.services import graduation_service as service

    def _list_proposals_sql(page, ps, keyword=None, status=None, batch_id=None):
        with service.session() as db:
            return proposal_read.list_proposals(
                db,
                service._tid(),
                page,
                ps,
                keyword=keyword,
                status=status,
                batch_id=batch_id,
            )

    def _proposal_stats_sql(batch_id=None):
        with service.session() as db:
            return proposal_read.proposal_stats(db, service._tid(), batch_id=batch_id)

    def _export_proposals_sql(status=None, keyword=None, batch_id=None):
        service.enforce_permission(
            service.get_current_user_ctx() or {},
            "graduationDesign.proposal.export",
        )
        if not batch_id:
            raise service.AppException("VALIDATION_ERROR", "导出前必须选择毕业设计批次")

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import base64
        import io

        headers = ["学生", "班级", "课题", "指导教师", "版本", "是否重交", "提交时间", "状态"]
        operator, _role = service._op()
        title = f"开题材料台账　导出时间：{service.datetime.now():%Y-%m-%d %H:%M}　导出人：{operator}"
        wb = Workbook()
        ws = wb.active
        ws.title = "开题材料台账"
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"].font = Font(bold=True, color="555555", size=10)
        ws.append(headers)
        fill = PatternFill("solid", fgColor="DCE6F1")
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = fill

        written = 0
        with service.session() as db:
            for item in proposal_read.iter_proposals(
                db,
                service._tid(),
                keyword=keyword,
                status=status,
                batch_id=batch_id,
                chunk_size=200,
            ):
                ws.append([
                    item["studentName"],
                    item.get("className", ""),
                    item.get("topicTitle", ""),
                    item.get("advisorName", ""),
                    item.get("version", ""),
                    "是" if item.get("isResubmit") else "否",
                    (item.get("submitAt") or "")[:19],
                    item.get("statusLabel", ""),
                ])
                written += 1

        for index in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + index)].width = 18
        ws.freeze_panes = "A3"
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        digest = service.hashlib.sha256(content).hexdigest()

        with service.session() as db:
            trail = service._audit(
                db,
                "PROPOSAL",
                "export",
                "导出开题材料台账",
                f"共 {written} 行，状态={status or '全部'}，批次={batch_id}，sha256={digest}",
            )
            trail.batch_id = int(batch_id)
            db.commit()

        return {
            "filename": f"开题材料台账_{service.datetime.now():%Y%m%d_%H%M}.xlsx",
            "contentBase64": base64.b64encode(content).decode("ascii"),
            "rowCount": written,
            "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    _list_proposals_sql.__name__ = "list_proposals"
    _proposal_stats_sql.__name__ = "proposal_stats"
    _export_proposals_sql.__name__ = "export_proposals_xlsx"

    service.list_proposals = _list_proposals_sql
    service.proposal_stats = _proposal_stats_sql
    service.export_proposals_xlsx = service.sanitize_xlsx_export(_export_proposals_sql)


def _install_final_read_model() -> None:
    from app.modules.graduation.services import graduation_final_read_service as final_read
    from app.modules.graduation.services import graduation_service as service

    def _list_finals_sql(page, ps, keyword=None, status=None, batch_id=None):
        with service.session() as db:
            return final_read.list_finals(
                db,
                service._tid(),
                page,
                ps,
                keyword=keyword,
                status=status,
                batch_id=batch_id,
            )

    def _final_stats_sql(batch_id=None):
        with service.session() as db:
            return final_read.final_stats(db, service._tid(), batch_id=batch_id)

    def _export_finals_sql(status=None, keyword=None, batch_id=None):
        service.enforce_permission(
            service.get_current_user_ctx() or {},
            "graduationDesign.final.export",
        )
        if not batch_id:
            raise service.AppException("VALIDATION_ERROR", "导出前必须选择毕业设计批次")

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import base64
        import io

        headers = ["学生", "班级", "课题", "指导教师", "成果类型", "版本", "查重率", "查重状态", "提交时间", "状态"]
        operator, _role = service._op()
        title = f"成果提交台账　导出时间：{service.datetime.now():%Y-%m-%d %H:%M}　导出人：{operator}"
        wb = Workbook()
        ws = wb.active
        ws.title = "成果提交台账"
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"].font = Font(bold=True, color="555555", size=10)
        ws.append(headers)
        fill = PatternFill("solid", fgColor="DCE6F1")
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = fill

        written = 0
        with service.session() as db:
            for item in final_read.iter_finals(
                db,
                service._tid(),
                keyword=keyword,
                status=status,
                batch_id=batch_id,
                chunk_size=200,
            ):
                ws.append([
                    item["studentName"],
                    item.get("className", ""),
                    item.get("topicTitle", ""),
                    item.get("advisorName", ""),
                    item.get("type", ""),
                    item.get("version", ""),
                    item.get("plagiarismRate", ""),
                    item.get("plagiarismStatus", ""),
                    (item.get("submitAt") or "")[:19],
                    item.get("statusLabel", ""),
                ])
                written += 1

        for index in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + index)].width = 16
        ws.freeze_panes = "A3"
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        digest = service.hashlib.sha256(content).hexdigest()

        with service.session() as db:
            trail = service._audit(
                db,
                "FINAL",
                "export",
                "导出成果提交台账",
                f"共 {written} 行，状态={status or '全部'}，批次={batch_id}，sha256={digest}",
            )
            trail.batch_id = int(batch_id)
            db.commit()

        return {
            "filename": f"成果提交台账_{service.datetime.now():%Y%m%d_%H%M}.xlsx",
            "contentBase64": base64.b64encode(content).decode("ascii"),
            "rowCount": written,
            "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    _list_finals_sql.__name__ = "list_finals"
    _final_stats_sql.__name__ = "final_stats"
    _export_finals_sql.__name__ = "export_finals_xlsx"

    service.list_finals = _list_finals_sql
    service.final_stats = _final_stats_sql
    service.export_finals_xlsx = service.sanitize_xlsx_export(_export_finals_sql)


def _install_student_read_model() -> None:
    from app.modules.graduation.services import graduation_student_read_service as student_read
    from app.modules.graduation.services import graduation_student_service as service

    service.list_students = student_read.list_students


def _install_grade_read_model() -> None:
    from app.modules.graduation.services import graduation_grade_read_service as grade_read
    from app.modules.graduation.services import graduation_grade_service as service

    service.list_grades = grade_read.list_grades


def _install_archive_read_model() -> None:
    from app.modules.graduation.services import graduation_archive_read_service as archive_read
    from app.modules.graduation.services import graduation_archive_service as service

    service.list_archives = archive_read.list_archives


_install_proposal_read_model()
_install_final_read_model()
_install_student_read_model()
_install_grade_read_model()
_install_archive_read_model()
del _install_proposal_read_model
del _install_final_read_model
del _install_student_read_model
del _install_grade_read_model
del _install_archive_read_model
