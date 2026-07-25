"""毕业设计域真实数据服务。租户过滤 + 脱敏 + 审计留痕 + 开题批阅/答辩发布闭环。"""
from __future__ import annotations

from datetime import datetime, timezone

from sqlalchemy import func, select

from app.core.context import get_current_user_ctx
from app.core.exceptions import AppException, no_permission, not_found
from app.models import (GraduationAuditTrail, GraduationBatch, GraduationDefenseGroup, GraduationFinal,
                        GraduationProposal, GraduationStudent, GraduationTopic)
from app.modules.graduation.services import graduation_student_service as gd_stu_svc
from app.modules.graduation.services.graduation_scope_service import (
    accessible_student_ids, assert_student_access, can_access_student, has_full_scope,
)
from app.services.db_service import _iso, _mask_phone, _tid, session

# 学生阶段中文名。COMPLETED 由 graduation_grade_service.publish() 真实写入（该文件 L276），
# 此前本表漏登记，导致：① 学生列表 stageLabel 回落显示英文 "COMPLETED"；② 看板流程条按 L_STAGE
# 的键迭代，已完成的学生整段不显示。补齐后两处同时修复。顺序即流程条展示顺序，勿随意调整。
L_STAGE = {"TOPIC_SELECTING": "选题中", "TASKBOOK_CONFIRM": "任务书确认", "GUIDING": "指导中",
           "MIDTERM": "中期检查", "FINAL_CHECK": "成果检查", "DEFENSE": "答辩中",
           "COMPLETED": "已完成", "ARCHIVED": "已归档"}

# 批次阶段（学校日程，graduation_batch_service.DEFAULT_STAGES 的 8 个 code）
# → 学生阶段（个人进度，L_STAGE 的键）。两套词汇不是一一对应：
# 成果提交/查重/评阅三个批次阶段，学生都停在「成果检查」。
_BATCH_STAGE_TO_STUDENT_STAGE = {
    "TOPIC": "TOPIC_SELECTING", "PROPOSAL": "TASKBOOK_CONFIRM", "MIDTERM": "MIDTERM",
    "SUBMISSION": "FINAL_CHECK", "PLAGIARISM": "FINAL_CHECK", "REVIEW": "FINAL_CHECK",
    "DEFENSE": "DEFENSE", "GRADE": "COMPLETED",
}


def _active_student_stage(batch):
    """按批次阶段时间轴推「今天该走哪一步」，供看板流程条高亮当前节点。

    返回 L_STAGE 的键；批次不存在、未配阶段日期、或今天不落在任何阶段区间内时返回 None
    —— 此时流程条不高亮任何节点。新建批次的 stage_config 日期默认全为 None，
    不能因为「总得亮一个」就假高亮一个阶段。
    """
    if not batch:
        return None
    from datetime import date, datetime

    def _d(v):
        if not v:
            return None
        if isinstance(v, date) and not isinstance(v, datetime):
            return v
        try:
            return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
        except ValueError:
            return None

    today = date.today()
    for st in (batch.stage_config or []):
        start, end = _d(st.get("startDate")), _d(st.get("endDate"))
        if start and end and start <= today <= end:
            return _BATCH_STAGE_TO_STUDENT_STAGE.get(st.get("code"))
    return None
L_RISK = {"NONE": "无风险", "LOW": "低风险", "MEDIUM": "中风险", "HIGH": "高风险"}
L_MAT = {"PENDING_REVIEW": "待审阅", "APPROVED": "已通过", "REJECTED": "已驳回", "NOT_SUBMITTED": "未提交"}
L_TOPIC = {"CONFIRMED": "已确认", "PENDING_CONFIRM": "待确认", "DISABLED": "已停用"}


def _op():
    u = get_current_user_ctx() or {}
    return u.get("realName") or "系统", u.get("currentRoleCode") or ""


def _audit(db, bt, bid, action, detail="", before="", after=""):
    n, r = _op()
    db.add(GraduationAuditTrail(tenant_id=_tid(), biz_type=bt, biz_id=str(bid), action=action,
                                operator=n, role_name=r, detail=detail, before_val=before,
                                after_val=after, occurred_at=datetime.now(timezone.utc)))


def _page(items, page, ps):
    total = len(items)
    start = (max(1, page) - 1) * ps
    return items[start:start + ps], total


def _att_id(x) -> str:
    """兼容附件存储形态：既支持裸 file_id，也支持 {fileId|id} 对象。"""
    if isinstance(x, dict):
        return str(x.get("fileId") or x.get("id") or "")
    return str(x or "")


def _resolve_attachments(ids, *, student_channel: bool = False) -> list:
    """把附件 file_id 列表解析为可下载展示项（文件中心 t_file_object，租户内可见）。
    student_channel=True → /mobile/graduation/materials（学生本人）；
    默认 → /graduation/materials（管理端业务关系鉴权）。"""
    out = []
    if not ids:
        return out
    from app.services import file_service
    base = ("/api/v1/mobile/graduation/materials" if student_channel
            else "/api/v1/graduation/materials")
    for raw in ids:
        fid = _att_id(raw)
        if not fid:
            continue
        try:
            v = file_service.attachment_view(fid)
        except Exception:  # noqa: BLE001
            v = None
        if not v:
            continue
        out.append({**v, "downloadUrl": f"{base}/{v['fileId']}/download"})
    return out


def _validate_final_attachments(attachments, *, require_nonempty: bool = True) -> list[str]:
    """Reject forged cross-tenant IDs and unsafe thesis material types before binding.
    require_nonempty=True（成果）必须至少 1 个附件；开题可传空列表。"""
    from app.services import file_service
    normalized: list[str] = []
    for raw in attachments or []:
        fid = _att_id(raw)
        if not fid or fid in normalized:
            continue
        meta = file_service.get_file_meta(fid)
        if not meta:
            raise AppException("VALIDATION_ERROR", "Thesis attachment is missing or outside the current tenant")
        if (meta.get("ext") or "").lower() not in {"pdf", "doc", "docx", "zip"}:
            raise AppException("FILE_TYPE_NOT_ALLOWED", "Only PDF, Word, or ZIP thesis files are allowed")
        normalized.append(fid)
    if require_nonempty and not normalized:
        raise AppException("VALIDATION_ERROR", "请先上传论文/成果附件再提交")
    if len(normalized) > 10:
        raise AppException("VALIDATION_ERROR", "A thesis submission may contain at most 10 attachments")
    return normalized


def _mark_material_files(db, attachment_ids: list[str]) -> None:
    if not attachment_ids:
        return
    from app.models import FileObject
    file_rows = db.scalars(select(FileObject).where(
        FileObject.tenant_id == _tid(),
        FileObject.id.in_([int(fid) for fid in attachment_ids]),
        FileObject.is_deleted.is_(False),
    ).with_for_update()).all()
    if len(file_rows) != len(attachment_ids):
        raise AppException("VALIDATION_ERROR", "One or more graduation attachments are invalid")
    for file_row in file_rows:
        file_row.biz_type = "GRADUATION_MATERIAL"


def _stu_of(db, sid):
    return db.get(GraduationStudent, sid)


def resolve_material_download(file_id: str):
    """Resolve only when the file is bound to an accessible graduation proposal/final."""
    from app.services import file_service
    with session() as db:
        candidates = []
        candidates.extend(db.scalars(select(GraduationProposal).where(
            GraduationProposal.tenant_id == _tid(), GraduationProposal.is_deleted.is_(False),
        )).all())
        candidates.extend(db.scalars(select(GraduationFinal).where(
            GraduationFinal.tenant_id == _tid(), GraduationFinal.is_deleted.is_(False),
        )).all())
        for material in candidates:
            bound = {_att_id(raw) for raw in (material.attachments_json or [])}
            if file_id not in bound:
                continue
            student = db.get(GraduationStudent, material.gd_student_id)
            assert_student_access(db, student, "graduation.material.download")
            return file_service.resolve_download(file_id, allow_graduation_material=True)
    return None


def _stu_row(s: GraduationStudent) -> dict:
    return {"id": str(s.id), "studentId": str(s.student_id or s.id), "name": s.name,
            "studentNo": s.student_no or "", "className": s.class_name or "", "classId": s.class_id or "",
            "topicTitle": s.topic_title or "（未确认选题）", "topicSource": s.topic_source or "",
            "advisorName": s.advisor_name or "", "stage": s.stage,
            "stageLabel": L_STAGE.get(s.stage, s.stage), "materialSummary": s.material_summary or "",
            "plagiarismRate": s.plagiarism_rate or "—",
            "plagiarismTone": "danger" if (s.plagiarism_rate and _rate_over(s.plagiarism_rate)) else "success",
            "riskLevel": s.risk_level, "riskLabel": L_RISK.get(s.risk_level, s.risk_level),
            "phone": _mask_phone(s.phone_encrypted), "recordStatus": s.record_status,
            "updateTime": _iso(s.updated_at)}


def _rate_over(r, threshold=30):
    try:
        return float(str(r).replace("%", "")) > threshold
    except ValueError:
        return False


# ═══ 学生 ═══

def list_students(page, ps, keyword=None, class_id=None, stage=None, risk_level=None):
    items, total = gd_stu_svc.list_students(page, ps, keyword=keyword, class_id=class_id,
                                            stage=stage, risk_level=risk_level)
    return items, total


def get_student_detail(sid) -> dict:
    detail = gd_stu_svc.get_student(sid)
    student = {k: detail[k] for k in detail if k not in (
        "taskbook", "batch", "topic", "proposals", "midterm", "finals", "plagiarisms",
        "defense", "stateFlow", "auditTrail")}
    return {
        "student": student,
        "proposals": detail.get("proposals", []),
        "midterm": detail.get("midterm", {}),
        "finals": detail.get("finals", []),
        "defense": detail.get("defense", {}),
        "auditTrail": detail.get("auditTrail", []),
    }


# ═══ 选题 ═══

def list_topics(page, ps, keyword=None, status=None):
    from app.modules.graduation.services import graduation_topic_service as topic_svc
    review_status = None
    op_status = status
    if status == "CONFIRMED":
        review_status = "APPROVED"
        op_status = "CONFIRMED"
    elif status == "PENDING_CONFIRM":
        review_status = None
    items, total = topic_svc.list_topics(page, ps, keyword=keyword, review_status=review_status,
                                         status=op_status, archive_view="active")
    # 兼容旧字段 students
    for it in items:
        if "students" not in it or not it["students"]:
            it["students"] = []
    return items, total


# ═══ 开题 ═══

def _prop_row(p: GraduationProposal, stu=None) -> dict:
    return {"id": str(p.id), "projectId": str(p.gd_student_id),
            "studentName": stu.name if stu else "", "className": stu.class_name if stu else "",
            "topicTitle": stu.topic_title if stu else "", "advisorName": stu.advisor_name if stu else "",
            "version": p.version or "—", "isResubmit": p.is_resubmit, "submitAt": _iso(p.submit_at) or "",
            "attachments": len(p.attachments_json or []), "status": p.status,
            "statusLabel": L_MAT.get(p.status, p.status)}



def _match_batch(stu, batch_id) -> bool:
    """学生是否属于指定批次；batch_id 为空时不限制。"""
    if not batch_id:
        return True
    if not stu or not getattr(stu, "batch_id", None):
        return False
    return str(stu.batch_id) == str(batch_id)

def list_proposals(page, ps, keyword=None, status=None, batch_id=None):
    """开题材料列表。status=NOT_SUBMITTED 时派生"已过选题但尚未提交开题报告"的学生行（用于催交）。"""
    with session() as db:
        if status == "NOT_SUBMITTED":
            return _page(_not_submitted_proposals(db, keyword, batch_id=batch_id), page, ps)
        q = select(GraduationProposal).where(GraduationProposal.tenant_id == _tid(),
                                             GraduationProposal.is_deleted.is_(False))
        if status:
            q = q.where(GraduationProposal.status == status)
        rows = db.scalars(q.order_by(GraduationProposal.id.desc())).all()
        items = []
        for p in rows:
            stu = _stu_of(db, p.gd_student_id)
            if not stu or not can_access_student(db, stu):
                continue
            if not _match_batch(stu, batch_id):
                continue
            if keyword and (not stu or keyword.strip() not in (stu.name or "")):
                continue
            items.append(_prop_row(p, stu))
        # 全部页签时也把"未提交"学生并入，便于一屏监管
        if not status:
            items += _not_submitted_proposals(db, keyword, batch_id=batch_id)
        return _page(items, page, ps)


def _not_submitted_proposals(db, keyword=None, batch_id=None) -> list:
    """派生未提交开题报告的学生：已确认选题（topic_id 存在或阶段已过选题中）且无任何开题记录。"""
    have = {r for (r,) in db.execute(select(GraduationProposal.gd_student_id).where(
        GraduationProposal.tenant_id == _tid(), GraduationProposal.is_deleted.is_(False))).all()}
    stus = db.scalars(select(GraduationStudent).where(
        GraduationStudent.tenant_id == _tid(), GraduationStudent.is_deleted.is_(False),
        GraduationStudent.record_status == "ACTIVE").order_by(GraduationStudent.id)).all()
    rows = []
    for s in stus:
        if not can_access_student(db, s):
            continue
        if not _match_batch(s, batch_id):
            continue
        if s.id in have:
            continue
        confirmed_topic = bool(s.topic_id) or s.stage not in ("TOPIC_SELECTING", None, "")
        if not confirmed_topic:
            continue
        if keyword and keyword.strip() not in (s.name or ""):
            continue
        rows.append({"id": f"S{s.id}", "projectId": str(s.id), "gdStudentId": str(s.id),
                     "studentName": s.name, "className": s.class_name or "",
                     "topicTitle": s.topic_title or "（未确认选题）", "advisorName": s.advisor_name or "",
                     "version": "—", "isResubmit": False, "submitAt": "", "attachments": 0,
                     "status": "NOT_SUBMITTED", "statusLabel": L_MAT["NOT_SUBMITTED"]})
    return rows


def get_proposal_detail(pid) -> dict:
    with session() as db:
        p = db.get(GraduationProposal, int(pid))
        if not p or p.is_deleted or p.tenant_id != _tid():
            raise not_found("开题材料不存在")
        stu = _stu_of(db, p.gd_student_id)
        assert_student_access(db, stu, "proposal.detail")
        logs = db.scalars(select(GraduationAuditTrail).where(GraduationAuditTrail.tenant_id == _tid(),
                          GraduationAuditTrail.biz_type == "PROPOSAL",
                          GraduationAuditTrail.biz_id == str(p.id)).order_by(GraduationAuditTrail.id)).all()
        # 历史版本：同一学生的全部开题报告（含被驳回的旧版），按版本先后
        vers = db.scalars(select(GraduationProposal).where(
            GraduationProposal.tenant_id == _tid(), GraduationProposal.gd_student_id == p.gd_student_id,
            GraduationProposal.is_deleted.is_(False)).order_by(GraduationProposal.id)).all()
        version_tone = {"APPROVED": "success", "REJECTED": "danger", "PENDING_REVIEW": "processing"}
        row = _prop_row(p, stu)
        row.update({"content": {"background": p.background or "", "plan": p.plan or "",
                                "outcome": p.outcome or ""},
                    "attachmentsList": _resolve_attachments(p.attachments_json or []),
                    "reviewComment": p.review_comment or "",
                    "defenseResult": p.defense_result or "", "defenseComment": p.defense_comment or "",
                    "defenseAt": _iso(p.defense_at) or "",
                    "versions": [{"title": (v.version or "v?") + " · " + L_MAT.get(v.status, v.status)
                                  + (" · 重交" if v.is_resubmit else ""),
                                  "desc": (v.review_comment or "") if v.status == "REJECTED" else "",
                                  "time": _iso(v.submit_at) or _iso(v.created_at) or "",
                                  "tone": version_tone.get(v.status, "info")} for v in vers],
                    "trail": [{"who": x.operator or "系统", "time": _iso(x.occurred_at),
                               "action": x.action, "affected": x.detail or ""} for x in logs]})
        return row


def review_proposal(pid, action, comment=None) -> dict:
    if action not in ("APPROVE", "REJECT"):
        raise AppException("VALIDATION_ERROR", "action 必须是 APPROVE/REJECT")
    if action == "REJECT" and (not comment or len(comment.strip()) < 5):
        raise AppException("VALIDATION_ERROR", "驳回原因必填且不少于 5 字")
    with session() as db:
        p = db.get(GraduationProposal, int(pid))
        if not p or p.is_deleted or p.tenant_id != _tid():
            raise not_found("开题材料不存在")
        stu = _stu_of(db, p.gd_student_id)
        assert_student_access(db, stu, "proposal.review")
        if p.status in ("APPROVED", "REJECTED"):
            raise AppException("DATA_CONFLICT", "该开题已批阅，请刷新")
        before = p.status
        n, _ = _op()
        target = "APPROVED" if action == "APPROVE" else "REJECTED"
        p.status = target
        p.reviewer = n
        p.review_comment = (comment or "").strip()
        p.review_time = datetime.now(timezone.utc)
        p.version = p.version or "v1"
        _audit(db, "PROPOSAL", p.id, "批阅开题-" + ("通过" if action == "APPROVE" else "驳回"),
               (comment or "").strip(), before, target)
        from app.modules.graduation.services import graduation_todo_helper as gd_todo
        gd_todo.todo_done(db, biz_id=p.id, todo_type=gd_todo.TODO_PROPOSAL)
        # 开题通过仅在任务书已确认时推进到指导中，禁止跳过任务书确认
        stu = _stu_of(db, p.gd_student_id)
        if stu and action == "APPROVE" and stu.stage in ("TOPIC_SELECTING", "TASKBOOK_CONFIRM"):
            from app.models import GraduationTaskBook
            tb = db.scalars(select(GraduationTaskBook).where(
                GraduationTaskBook.tenant_id == _tid(), GraduationTaskBook.gd_student_id == stu.id,
                GraduationTaskBook.is_deleted.is_(False), GraduationTaskBook.status == "CONFIRMED",
            ).limit(1)).first()
            if tb:
                stu.stage = "GUIDING"
            elif stu.stage == "TOPIC_SELECTING":
                stu.stage = "TASKBOOK_CONFIRM"
        db.commit()
        return {"id": str(p.id), "status": target, "statusLabel": L_MAT.get(target, target)}


def proposal_stats(batch_id=None) -> dict:
    """开题统计：按状态分布 + 未提交数（与列表/页签同一批次与数据范围）。"""
    with session() as db:
        scope_ids = accessible_student_ids(db, _tid(), batch_id=batch_id)
        base = [GraduationProposal.tenant_id == _tid(), GraduationProposal.is_deleted.is_(False),
                GraduationProposal.gd_student_id.in_(scope_ids or [-1])]
        total = int(db.scalar(select(func.count()).select_from(GraduationProposal).where(*base)) or 0)
        by_status = [{"status": s, "label": L_MAT.get(s, s),
                     "count": int(db.scalar(select(func.count()).select_from(GraduationProposal).where(
                         *base, GraduationProposal.status == s)) or 0)}
                     for s in ("PENDING_REVIEW", "APPROVED", "REJECTED")]
        not_submitted = len(_not_submitted_proposals(db, batch_id=batch_id))
        return {"total": total, "byStatus": by_status, "notSubmitted": not_submitted,
                "batchId": str(batch_id) if batch_id else None}


def final_stats(batch_id=None) -> dict:
    """成果统计：按状态分布 + 查重超标数（与列表/页签同一批次与数据范围）。"""
    with session() as db:
        scope_ids = accessible_student_ids(db, _tid(), batch_id=batch_id)
        base = [GraduationFinal.tenant_id == _tid(), GraduationFinal.is_deleted.is_(False),
                GraduationFinal.gd_student_id.in_(scope_ids or [-1])]
        total = int(db.scalar(select(func.count()).select_from(GraduationFinal).where(*base)) or 0)
        by_status = [{"status": s, "label": L_MAT.get(s, s),
                     "count": int(db.scalar(select(func.count()).select_from(GraduationFinal).where(
                         *base, GraduationFinal.status == s)) or 0)}
                     for s in ("PENDING_REVIEW", "APPROVED", "REJECTED")]
        overs = [f for f in db.scalars(select(GraduationFinal).where(*base)).all() if _rate_over(f.plagiarism_rate)]
        return {"total": total, "byStatus": by_status, "plagiarismOver": len(overs),
                "batchId": str(batch_id) if batch_id else None}


def submit_proposal(gd_student_id, background, plan, outcome, attachments=None) -> dict:
    """学生提交/重交开题报告。已有待审/已通过时不可重复提交；被驳回后可重交（版本自增 + is_resubmit）。"""
    attachment_ids = _validate_final_attachments(attachments, require_nonempty=False)
    if not (background and background.strip()):
        raise AppException("VALIDATION_ERROR", "选题背景不能为空")
    if not (plan and plan.strip()):
        raise AppException("VALIDATION_ERROR", "研究方案与进度不能为空")
    with session() as db:
        stu = _stu_of(db, int(gd_student_id))
        if not stu or stu.is_deleted or stu.tenant_id != _tid():
            raise not_found("毕设学生档案不存在")
        if not stu.topic_id:
            raise AppException("DATA_CONFLICT", "请先完成选题确认后再提交开题报告")
        elig = getattr(stu, "eligibility_status", None) or "PENDING"
        if elig == "UNQUALIFIED":
            raise AppException("DATA_CONFLICT", "资格不合格，不能提交开题报告")
        from app.models import GraduationTaskBook
        tb = db.scalars(select(GraduationTaskBook).where(
            GraduationTaskBook.tenant_id == _tid(), GraduationTaskBook.gd_student_id == stu.id,
            GraduationTaskBook.is_deleted.is_(False), GraduationTaskBook.status == "CONFIRMED",
        ).limit(1)).first()
        if not tb:
            raise AppException("DATA_CONFLICT", "请先确认任务书后再提交开题报告")
        _mark_material_files(db, attachment_ids)
        existing = db.scalars(select(GraduationProposal).where(
            GraduationProposal.tenant_id == _tid(), GraduationProposal.gd_student_id == stu.id,
            GraduationProposal.is_deleted.is_(False)).order_by(GraduationProposal.id.desc())).all()
        latest = existing[0] if existing else None
        if latest and latest.status == "PENDING_REVIEW":
            raise AppException("DATA_CONFLICT", "已有待审阅的开题报告，请等待指导教师批阅")
        if latest and latest.status == "APPROVED":
            raise AppException("DATA_CONFLICT", "开题报告已通过，无需重复提交")
        is_resubmit = latest is not None  # 存在被驳回的旧版即为重交
        version = f"v{len(existing) + 1}"
        p = GraduationProposal(
            tenant_id=_tid(), gd_student_id=stu.id, version=version, is_resubmit=is_resubmit,
            submit_at=datetime.now(timezone.utc), background=background.strip(), plan=plan.strip(),
            outcome=(outcome or "").strip(), attachments_json=attachment_ids, status="PENDING_REVIEW")
        db.add(p)
        db.flush()
        _audit(db, "PROPOSAL", p.id, "提交开题报告-" + ("重交" if is_resubmit else "首次"),
               f"{stu.name} {version}", "", "PENDING_REVIEW")
        from app.modules.graduation.services import graduation_todo_helper as gd_todo
        gd_todo.push_proposal_todo(db, p, stu)
        db.commit()
        return {"id": str(p.id), "version": version, "isResubmit": is_resubmit, "status": "PENDING_REVIEW"}


def hold_proposal_defense(pid, result, comment=None) -> dict:
    """开题答辩（现场环节）：录入通过/不通过 + 评语。仅书面开题已通过（APPROVED）才可进行开题答辩。"""
    if result not in ("PASS", "FAIL"):
        raise AppException("VALIDATION_ERROR", "开题答辩结果必须是 PASS/FAIL")
    if result == "FAIL" and (not comment or len(comment.strip()) < 5):
        raise AppException("VALIDATION_ERROR", "开题答辩不通过时评语必填且不少于 5 字")
    with session() as db:
        p = db.get(GraduationProposal, int(pid))
        if not p or p.is_deleted or p.tenant_id != _tid():
            raise not_found("开题材料不存在")
        stu = _stu_of(db, p.gd_student_id)
        assert_student_access(db, stu, "proposal.defense")
        if p.status != "APPROVED":
            raise AppException("DATA_CONFLICT", "仅书面开题审核通过后方可进行开题答辩")
        p.defense_result = result
        p.defense_comment = (comment or "").strip() or None
        p.defense_at = datetime.now(timezone.utc)
        _audit(db, "PROPOSAL", p.id, "开题答辩-" + ("通过" if result == "PASS" else "不通过"),
               (comment or "").strip())
        db.commit()
        return {"id": str(p.id), "defenseResult": result}


def remind_proposal(gd_student_id, channel="站内消息") -> dict:
    """开题催交（GD-R04 联动）：对已过选题但未提交开题的学生留痕催办。真实写审计，不 mock 冒充。"""
    with session() as db:
        stu = _stu_of(db, int(gd_student_id))
        if not stu or stu.is_deleted or stu.tenant_id != _tid():
            raise not_found("毕设学生档案不存在")
        done = db.scalar(select(func.count()).select_from(GraduationProposal).where(
            GraduationProposal.tenant_id == _tid(), GraduationProposal.gd_student_id == stu.id,
            GraduationProposal.status.in_(("PENDING_REVIEW", "APPROVED")),
            GraduationProposal.is_deleted.is_(False))) or 0
        if done:
            raise AppException("DATA_CONFLICT", "该生已提交开题报告，无需催交")
        _audit(db, "PROPOSAL", f"remind-{stu.id}", "开题催交",
               f"催办 {stu.name} 提交开题报告（{channel}）")
        db.commit()
        return {"gdStudentId": str(stu.id), "studentName": stu.name, "reminded": True}


def export_proposals_xlsx(status=None, keyword=None, batch_id=None) -> dict:
    """开题材料台账 Excel 导出（含导出人/时间抬头，写导出审计）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    items, total = list_proposals(1, 100000, keyword=keyword, status=status, batch_id=batch_id)
    headers = ["学生", "班级", "课题", "指导教师", "版本", "是否重交", "提交时间", "状态"]
    operator, _role = _op()
    title = f"开题材料台账　导出时间：{datetime.now():%Y-%m-%d %H:%M}　导出人：{operator}"
    wb = Workbook()
    ws = wb.active
    ws.title = "开题材料台账"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="555555", size=10)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="DCE6F1")
    for c in ws[2]:
        c.font = Font(bold=True); c.fill = fill
    for it in items:
        ws.append([it["studentName"], it.get("className", ""), it.get("topicTitle", ""),
                  it.get("advisorName", ""), it.get("version", ""),
                  "是" if it.get("isResubmit") else "否", (it.get("submitAt") or "")[:19],
                  it.get("statusLabel", "")])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 18
    ws.freeze_panes = "A3"
    import base64
    import io
    with session() as db:
        _audit(db, "PROPOSAL", "export", "导出开题材料台账", f"共 {total} 行，状态={status or '全部'}，批次={batch_id or '全部'}")
        db.commit()
    buf = io.BytesIO()
    wb.save(buf)
    return {"filename": f"开题材料台账_{datetime.now():%Y%m%d_%H%M}.xlsx",
            "contentBase64": base64.b64encode(buf.getvalue()).decode("ascii"), "rowCount": total,
            "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}


# ═══ 成果 ═══

def _final_row(f: GraduationFinal, stu=None) -> dict:
    return {"id": str(f.id), "projectId": str(f.gd_student_id),
            "studentName": stu.name if stu else "", "className": stu.class_name if stu else "",
            "topicTitle": stu.topic_title if stu else "", "advisorName": stu.advisor_name if stu else "",
            "type": f.final_type, "version": f.version or "", "submitAt": _iso(f.submit_at) or "",
            "plagiarismRate": f.plagiarism_rate or "—", "plagiarismStatus": f.plagiarism_status or "未检测",
            "plagiarismTone": "danger" if _rate_over(f.plagiarism_rate) else "success",
            "status": f.status, "statusLabel": L_MAT.get(f.status, f.status)}


FINAL_TYPES = ("初稿", "定稿")


def list_finals(page, ps, keyword=None, status=None, batch_id=None):
    """成果提交列表。status=NOT_SUBMITTED 时派生"已进入指导/中期/成果阶段但未提交论文"的学生行。"""
    with session() as db:
        if status == "NOT_SUBMITTED":
            return _page(_not_submitted_finals(db, keyword, batch_id=batch_id), page, ps)
        q = select(GraduationFinal).where(GraduationFinal.tenant_id == _tid(),
                                          GraduationFinal.is_deleted.is_(False))
        if status:
            q = q.where(GraduationFinal.status == status)
        rows = db.scalars(q.order_by(GraduationFinal.id.desc())).all()
        items = []
        for f in rows:
            stu = _stu_of(db, f.gd_student_id)
            if not stu or not can_access_student(db, stu):
                continue
            if not _match_batch(stu, batch_id):
                continue
            if keyword and (not stu or keyword.strip() not in (stu.name or "")):
                continue
            items.append(_final_row(f, stu))
        if not status:
            items += _not_submitted_finals(db, keyword, batch_id=batch_id)
        return _page(items, page, ps)


def _not_submitted_finals(db, keyword=None, batch_id=None) -> list:
    """派生未提交论文的学生：已进入指导/中期/成果检查/答辩阶段且无任何成果记录。"""
    have = {r for (r,) in db.execute(select(GraduationFinal.gd_student_id).where(
        GraduationFinal.tenant_id == _tid(), GraduationFinal.is_deleted.is_(False))).all()}
    stus = db.scalars(select(GraduationStudent).where(
        GraduationStudent.tenant_id == _tid(), GraduationStudent.is_deleted.is_(False),
        GraduationStudent.record_status == "ACTIVE").order_by(GraduationStudent.id)).all()
    rows = []
    for s in stus:
        if not can_access_student(db, s):
            continue
        if not _match_batch(s, batch_id):
            continue
        if s.id in have or s.stage not in ("GUIDING", "MIDTERM", "FINAL_CHECK", "DEFENSE"):
            continue
        if keyword and keyword.strip() not in (s.name or ""):
            continue
        rows.append({"id": f"S{s.id}", "projectId": str(s.id), "gdStudentId": str(s.id),
                     "studentName": s.name, "className": s.class_name or "",
                     "topicTitle": s.topic_title or "（未确认选题）", "advisorName": s.advisor_name or "",
                     "type": "—", "version": "—", "submitAt": "", "plagiarismRate": "—",
                     "plagiarismStatus": "未提交", "plagiarismTone": "warning",
                     "status": "NOT_SUBMITTED", "statusLabel": L_MAT["NOT_SUBMITTED"]})
    return rows


def midterm_allows_final_submit(mid) -> bool:
    """中期检查已通过（含整改复核通过）才允许提交成果。仅认状态，不单靠 conclusion。"""
    if mid is None:
        return False
    status = str(getattr(mid, "status", "") or "")
    return status in ("CHECKED_PASS", "RECTIFIED_PASS")


def submit_final(gd_student_id, final_type, attachments=None) -> dict:
    """学生提交/重交论文成果。初稿→定稿有序（定稿须初稿已通过）；有待审时不可重复提交。
    attachments：论文/材料附件 file_id 列表（文件中心 t_file_object.id），存 attachments_json。"""
    attachment_ids = _validate_final_attachments(attachments)
    if final_type not in FINAL_TYPES:
        raise AppException("VALIDATION_ERROR", "成果类型必须是 初稿/定稿")
    with session() as db:
        stu = _stu_of(db, int(gd_student_id))
        if not stu or stu.is_deleted or stu.tenant_id != _tid():
            raise not_found("毕设学生档案不存在")
        if stu.stage not in ("FINAL_CHECK", "DEFENSE"):
            raise AppException("DATA_CONFLICT", "当前阶段不可提交成果（须进入成果检查阶段）")
        if not stu.topic_id:
            raise AppException("DATA_CONFLICT", "请先完成选题确认后再提交成果")
        elig = getattr(stu, "eligibility_status", None) or "PENDING"
        if elig == "UNQUALIFIED":
            raise AppException("DATA_CONFLICT", "资格不合格，不能提交成果")
        # 中期必须已通过（与归档完整性口径一致）；缺记录/待检查/整改中/不通过均拦截
        from app.models import GraduationMidterm
        mid = db.scalars(select(GraduationMidterm).where(
            GraduationMidterm.tenant_id == _tid(),
            GraduationMidterm.gd_student_id == stu.id,
            GraduationMidterm.is_deleted.is_(False),
        ).order_by(GraduationMidterm.id.desc())).first()
        if not midterm_allows_final_submit(mid):
            raise AppException(
                "DATA_CONFLICT",
                "中期检查未通过或尚未完成，不能提交成果",
            )
        _mark_material_files(db, attachment_ids)
        existing = db.scalars(select(GraduationFinal).where(
            GraduationFinal.tenant_id == _tid(), GraduationFinal.gd_student_id == stu.id,
            GraduationFinal.is_deleted.is_(False)).order_by(GraduationFinal.id.desc())).all()
        if any(f.status == "PENDING_REVIEW" for f in existing):
            raise AppException("DATA_CONFLICT", "已有待审阅的成果，请等待指导教师批阅")
        if final_type == "定稿":
            has_draft_approved = any(f.final_type == "初稿" and f.status == "APPROVED" for f in existing)
            if not has_draft_approved:
                raise AppException("DATA_CONFLICT", "请先提交初稿并通过后再提交定稿")
            if any(f.final_type == "定稿" and f.status == "APPROVED" for f in existing):
                raise AppException("DATA_CONFLICT", "定稿已通过，无需重复提交")
        same_type = [f for f in existing if f.final_type == final_type]
        version = f"v{len(same_type) + 1}"
        f = GraduationFinal(tenant_id=_tid(), gd_student_id=stu.id, final_type=final_type,
                            version=version, submit_at=datetime.now(timezone.utc),
                            plagiarism_rate=None,
                            plagiarism_status="未检测",
                            attachments_json=attachment_ids,
                            status="PENDING_REVIEW")
        db.add(f)
        db.flush()
        _audit(db, "FINAL", f.id, f"提交成果-{final_type}", f"{stu.name} {final_type} {version}",
               "", "PENDING_REVIEW")
        from app.modules.graduation.services import graduation_todo_helper as gd_todo
        gd_todo.push_final_todo(db, f, stu)
        db.commit()
        return {"id": str(f.id), "finalType": final_type, "version": version, "status": "PENDING_REVIEW"}


def review_final(fid, action, comment=None) -> dict:
    if action not in ("APPROVE", "REJECT"):
        raise AppException("VALIDATION_ERROR", "action 必须是 APPROVE/REJECT")
    if action == "REJECT" and (not comment or len(comment.strip()) < 5):
        raise AppException("VALIDATION_ERROR", "驳回原因必填且不少于 5 字")
    with session() as db:
        f = db.scalars(select(GraduationFinal).where(
            GraduationFinal.id == int(fid),
            GraduationFinal.tenant_id == _tid(),
            GraduationFinal.is_deleted.is_(False),
        ).with_for_update()).first()
        if not f or f.is_deleted or f.tenant_id != _tid():
            raise not_found("成果不存在")
        stu = _stu_of(db, f.gd_student_id)
        assert_student_access(db, stu, "final.review")
        if f.status in ("APPROVED", "REJECTED"):
            raise AppException("DATA_CONFLICT", "该成果已批阅，请刷新")
        # GD-R09: approval relies on the server-side check record, never a client-supplied rate.
        if action == "APPROVE":
            from app.models import GraduationPlagiarismCheck
            check = db.scalars(select(GraduationPlagiarismCheck).where(
                GraduationPlagiarismCheck.tenant_id == _tid(),
                GraduationPlagiarismCheck.gd_final_id == f.id,
                GraduationPlagiarismCheck.is_deleted.is_(False),
            ).order_by(GraduationPlagiarismCheck.id.desc()).with_for_update()).first()
            if f.final_type == "定稿" and (not check or check.status != "DONE"):
                raise AppException("DATA_CONFLICT", "查重尚未完成，不能通过成果审核")
            if check and check.status == "DONE" and check.over_threshold and check.dispute_status != "APPROVED":
                raise AppException(
                    "DATA_CONFLICT",
                    f"查重率 {check.rate} 超标（GD-R09），须退回修改或完成复查特例审批",
                )
        before = f.status
        n, _ = _op()
        target = "APPROVED" if action == "APPROVE" else "REJECTED"
        f.status = target
        f.reviewer = n
        f.review_comment = (comment or "").strip()
        f.review_time = datetime.now(timezone.utc)
        _audit(db, "FINAL", f.id, "批阅成果-" + ("通过" if action == "APPROVE" else "退回修改"),
               (comment or "").strip(), before, target)
        from app.modules.graduation.services import graduation_todo_helper as gd_todo
        gd_todo.todo_done(db, biz_id=f.id, todo_type=gd_todo.TODO_FINAL)
        db.commit()
        return {"id": str(f.id), "status": target, "statusLabel": L_MAT.get(target, target)}


def get_final_detail(fid) -> dict:
    """成果批阅详情：本条 + 同生历史版本 + 退回意见 + 真实附件（文件中心解析）。供教师移动端批阅前查看。"""
    with session() as db:
        f = db.get(GraduationFinal, int(fid))
        if not f or f.is_deleted or f.tenant_id != _tid():
            raise not_found("成果不存在")
        stu = _stu_of(db, f.gd_student_id)
        assert_student_access(db, stu, "final.detail")
        vers = db.scalars(select(GraduationFinal).where(
            GraduationFinal.tenant_id == _tid(), GraduationFinal.gd_student_id == f.gd_student_id,
            GraduationFinal.is_deleted.is_(False)).order_by(GraduationFinal.id)).all()
        version_tone = {"APPROVED": "success", "REJECTED": "danger", "PENDING_REVIEW": "processing"}
        row = _final_row(f, stu)
        row.update({"studentNo": stu.student_no if stu else "",
                    "reviewComment": f.review_comment or "",
                    "attachmentsList": _resolve_attachments(f.attachments_json or []),
                    "versions": [{"title": (v.final_type or "") + " " + (v.version or "") + " · "
                                  + L_MAT.get(v.status, v.status),
                                  "desc": (v.review_comment or "") if v.status == "REJECTED" else "",
                                  "time": _iso(v.submit_at) or "", "tone": version_tone.get(v.status, "info")}
                                 for v in vers]})
        return row


def remind_final(gd_student_id, channel="站内消息") -> dict:
    """成果催交：对已进入指导/中期/成果阶段但未提交论文的学生留痕催办。"""
    with session() as db:
        stu = _stu_of(db, int(gd_student_id))
        if not stu or stu.is_deleted or stu.tenant_id != _tid():
            raise not_found("毕设学生档案不存在")
        done = db.scalar(select(func.count()).select_from(GraduationFinal).where(
            GraduationFinal.tenant_id == _tid(), GraduationFinal.gd_student_id == stu.id,
            GraduationFinal.is_deleted.is_(False))) or 0
        if done:
            raise AppException("DATA_CONFLICT", "该生已提交成果，无需催交")
        _audit(db, "FINAL", f"remind-{stu.id}", "成果催交", f"催办 {stu.name} 提交论文成果（{channel}）")
        db.commit()
        return {"gdStudentId": str(stu.id), "studentName": stu.name, "reminded": True}


def export_finals_xlsx(status=None, keyword=None, batch_id=None) -> dict:
    """成果提交台账 Excel 导出（含导出人/时间抬头，写导出审计）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    items, total = list_finals(1, 100000, keyword=keyword, status=status, batch_id=batch_id)
    headers = ["学生", "班级", "课题", "指导教师", "成果类型", "版本", "查重率", "查重状态", "提交时间", "状态"]
    operator, _role = _op()
    title = f"成果提交台账　导出时间：{datetime.now():%Y-%m-%d %H:%M}　导出人：{operator}"
    wb = Workbook()
    ws = wb.active
    ws.title = "成果提交台账"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="555555", size=10)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="DCE6F1")
    for c in ws[2]:
        c.font = Font(bold=True); c.fill = fill
    for it in items:
        ws.append([it["studentName"], it.get("className", ""), it.get("topicTitle", ""),
                  it.get("advisorName", ""), it.get("type", ""), it.get("version", ""),
                  it.get("plagiarismRate", ""), it.get("plagiarismStatus", ""),
                  (it.get("submitAt") or "")[:19], it.get("statusLabel", "")])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 16
    ws.freeze_panes = "A3"
    import base64
    import io
    with session() as db:
        _audit(db, "FINAL", "export", "导出成果提交台账", f"共 {total} 行，状态={status or '全部'}，批次={batch_id or '全部'}")
        db.commit()
    buf = io.BytesIO()
    wb.save(buf)
    return {"filename": f"成果提交台账_{datetime.now():%Y%m%d_%H%M}.xlsx",
            "contentBase64": base64.b64encode(buf.getvalue()).decode("ascii"), "rowCount": total,
            "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}


# ═══ 答辩 ═══

def _def_row(g: GraduationDefenseGroup) -> dict:
    from app.modules.graduation.services import graduation_identity as gid
    members = [gid.normalize_member(m) for m in (g.members_json or [])]
    return {"id": str(g.id), "groupName": g.group_name,
            "batchId": str(g.batch_id) if g.batch_id else None,
            "date": g.defense_date or "待定",
            "location": g.location or "待定", "chair": g.chair or "待指定",
            "chairMentorId": str(g.chair_mentor_id) if getattr(g, "chair_mentor_id", None) else None,
            "members": members or (g.members_json or []),
            "secretary": g.secretary or "待指定",
            "secretaryMentorId": str(g.secretary_mentor_id) if getattr(g, "secretary_mentor_id", None) else None,
            "studentCount": g.student_count, "conflict": g.conflict or "",
            "published": g.published,
            "publishedLabel": "已发布（学生端 P17 可见）" if g.published else "待调整后发布"}


def _can_access_defense_group(db, group: GraduationDefenseGroup) -> bool:
    if has_full_scope():
        return True
    user = get_current_user_ctx() or {}
    role = (user.get("currentRoleCode") or user.get("userType") or "").strip().upper()
    real_name = (user.get("realName") or "").strip()
    from app.modules.graduation.services import graduation_identity as gid
    from app.modules.graduation.services.graduation_scope_service import _name_is_ambiguous
    me = gid.current_user_mentor(db, group.tenant_id)
    if role == "GD_DEFENSE_SECRETARY":
        if not gid.user_is_secretary(group, mentor=me, real_name=real_name):
            return False
        if getattr(group, "secretary_mentor_id", None):
            return True
        return not _name_is_ambiguous(db, group.tenant_id, real_name)
    if role == "GD_DEFENSE_EXPERT":
        if not gid.user_on_judge_panel(group, mentor=me, real_name=real_name):
            return False
        matched_by_id = bool(
            me and any(
                s.get("mentorId") is not None and int(me.id) == int(s["mentorId"])
                for s in gid.judge_panel_seats(group)
            )
        )
        if matched_by_id:
            return True
        return not _name_is_ambiguous(db, group.tenant_id, real_name)
    return any(can_access_student(db, student) for student in _assigned_students(db, group.id))


def list_defense_groups(page, ps, keyword=None, batch_id=None):
    """按答辩组自身 batch_id 过滤；空组不再靠学生推断跨批出现。"""
    with session() as db:
        q = select(GraduationDefenseGroup).where(
            GraduationDefenseGroup.tenant_id == _tid(),
            GraduationDefenseGroup.is_deleted.is_(False))
        if batch_id is not None and batch_id != "":
            q = q.where(GraduationDefenseGroup.batch_id == int(batch_id))
        rows = db.scalars(q.order_by(GraduationDefenseGroup.id)).all()
        items = []
        for g in rows:
            if not _can_access_defense_group(db, g):
                continue
            items.append(_def_row(g))
        if keyword:
            kw = keyword.strip()
            items = [i for i in items if kw in i["groupName"]]
        return _page(items, page, ps)


MAX_DEFENSE_STUDENTS = 30


def _assigned_students(db, gid) -> list:
    return db.scalars(select(GraduationStudent).where(
        GraduationStudent.tenant_id == _tid(), GraduationStudent.defense_group_id == gid,
        GraduationStudent.is_deleted.is_(False)).order_by(GraduationStudent.id)).all()


def _recompute_defense(db, g):
    """按已分配学生重算人数 + 评委回避冲突（评委/组长不得是本组学生的指导教师）。"""
    from app.modules.graduation.services import graduation_identity as gid
    panel_ids = gid.panel_mentor_ids(g)
    panel_names = gid.panel_names(g)
    students = _assigned_students(db, g.id)
    clashes = []
    for s in students:
        hit = gid.assert_member_not_advisor(db, s, panel_ids, panel_names)
        if hit and hit not in clashes:
            clashes.append(hit)
    g.student_count = len(students)
    g.conflict = ("评委与指导教师冲突：" + "、".join(clashes) + "，须回避") if clashes else ""


def _require_defense_batch(db, batch_id) -> GraduationBatch:
    if batch_id is None or batch_id == "":
        raise AppException("VALIDATION_ERROR", "新建答辩组必须指定毕设批次 batchId")
    b = db.get(GraduationBatch, int(batch_id))
    if not b or b.is_deleted or b.tenant_id != _tid():
        raise not_found("毕设批次不存在")
    if b.status in ("ARCHIVED", "VOIDED"):
        raise AppException("DATA_CONFLICT", "已归档/已作废批次不可新建答辩组")
    return b


def _apply_defense_people(db, g, *, chair=None, secretary=None, members=None,
                          chair_mentor_id=None, secretary_mentor_id=None, member_mentor_ids=None,
                          preserve_existing: bool = False):
    """写主席/秘书/评委。update 时 preserve_existing=True：未传 ID 且未传姓名则保留原快照。"""
    from app.modules.graduation.services import graduation_identity as gid
    cid, cname = gid.merge_person_fields(
        existing_mentor_id=getattr(g, "chair_mentor_id", None),
        existing_name=getattr(g, "chair", None),
        mentor_id=chair_mentor_id, name=chair, preserve_existing=preserve_existing, db=db,
    )
    sid, sname = gid.merge_person_fields(
        existing_mentor_id=getattr(g, "secretary_mentor_id", None),
        existing_name=getattr(g, "secretary", None),
        mentor_id=secretary_mentor_id, name=secretary, preserve_existing=preserve_existing, db=db,
    )
    g.chair_mentor_id = cid
    g.chair = cname
    g.secretary_mentor_id = sid
    g.secretary = sname
    g.members_json = gid.merge_members_fields(
        db,
        existing_members=getattr(g, "members_json", None),
        member_ids=member_mentor_ids,
        legacy_members=members,
        preserve_existing=preserve_existing,
    )


def create_defense_group(group_name, defense_date=None, location=None, chair=None,
                         members=None, secretary=None, batch_id=None,
                         chair_mentor_id=None, secretary_mentor_id=None,
                         member_mentor_ids=None) -> dict:
    if not (group_name and group_name.strip()):
        raise AppException("VALIDATION_ERROR", "答辩组名称不能为空")
    with session() as db:
        batch = _require_defense_batch(db, batch_id)
        name = group_name.strip()
        dup = db.scalar(select(func.count()).select_from(GraduationDefenseGroup).where(
            GraduationDefenseGroup.tenant_id == _tid(),
            GraduationDefenseGroup.batch_id == batch.id,
            GraduationDefenseGroup.group_name == name,
            GraduationDefenseGroup.is_deleted.is_(False))) or 0
        if dup:
            raise AppException("DATA_CONFLICT", "当前批次已存在同名答辩组")
        g = GraduationDefenseGroup(
            tenant_id=_tid(), batch_id=batch.id, group_name=name,
            defense_date=(defense_date or "").strip() or None,
            location=(location or "").strip() or None,
            student_count=0, conflict="", published=False)
        _apply_defense_people(
            db, g, chair=chair, secretary=secretary, members=members,
            chair_mentor_id=chair_mentor_id, secretary_mentor_id=secretary_mentor_id,
            member_mentor_ids=member_mentor_ids)
        db.add(g)
        db.flush()
        _recompute_defense(db, g)
        _audit(db, "DEFENSE", g.id, "新建答辩组", f"{g.group_name} batch={batch.id}")
        db.commit()
        return get_defense_group_detail(g.id)


def update_defense_group(gid, group_name=None, defense_date=None, location=None, chair=None,
                         members=None, secretary=None,
                         chair_mentor_id=None, secretary_mentor_id=None,
                         member_mentor_ids=None) -> dict:
    """编辑不可改 batch_id（禁止跨批迁移）。"""
    with session() as db:
        g = db.get(GraduationDefenseGroup, int(gid))
        if not g or g.is_deleted or g.tenant_id != _tid():
            raise not_found("答辩组不存在")
        if group_name and group_name.strip():
            new_name = group_name.strip()
            if new_name != g.group_name and g.batch_id:
                dup = db.scalar(select(func.count()).select_from(GraduationDefenseGroup).where(
                    GraduationDefenseGroup.tenant_id == _tid(),
                    GraduationDefenseGroup.batch_id == g.batch_id,
                    GraduationDefenseGroup.group_name == new_name,
                    GraduationDefenseGroup.is_deleted.is_(False),
                    GraduationDefenseGroup.id != g.id)) or 0
                if dup:
                    raise AppException("DATA_CONFLICT", "当前批次已存在同名答辩组")
            g.group_name = new_name
        g.defense_date = (defense_date or "").strip() or None
        g.location = (location or "").strip() or None
        _apply_defense_people(
            db, g, chair=chair, secretary=secretary, members=members,
            chair_mentor_id=chair_mentor_id, secretary_mentor_id=secretary_mentor_id,
            member_mentor_ids=member_mentor_ids, preserve_existing=True)
        was_published = g.published
        g.published = False  # 编辑后须重新发布，学生端重新通知
        _recompute_defense(db, g)
        _audit(db, "DEFENSE", g.id, "编辑答辩组" + ("（撤回已发布，需重新发布）" if was_published else ""),
               g.group_name)
        db.commit()
        return get_defense_group_detail(g.id)


def get_defense_group_detail(gid) -> dict:
    with session() as db:
        g = db.get(GraduationDefenseGroup, int(gid))
        if not g or g.is_deleted or g.tenant_id != _tid():
            raise not_found("答辩组不存在")
        if not _can_access_defense_group(db, g):
            raise no_permission("Defense group is outside the current graduation-design scope")
        from app.modules.graduation.services import graduation_identity as gid_id
        panel_ids = gid_id.panel_mentor_ids(g)
        panel_names = gid_id.panel_names(g)
        row = _def_row(g)
        row["students"] = []
        for s in _assigned_students(db, g.id):
            if not can_access_student(db, s):
                continue
            conflict = bool(gid_id.assert_member_not_advisor(db, s, panel_ids, panel_names))
            row["students"].append({
                "id": str(s.id), "name": s.name, "className": s.class_name or "",
                "topicTitle": s.topic_title or "", "advisorName": s.advisor_name or "",
                "mentorId": str(s.mentor_id) if s.mentor_id else None,
                "conflict": conflict,
            })
        return row


def list_defense_eligible_students(gid=None, keyword=None) -> list:
    """可分配到答辩组的学生：须已进入成果检查及以后阶段，且与答辩组同批次。"""
    with session() as db:
        if not has_full_scope():
            raise no_permission("Only graduation managers can list defense assignment candidates")
        group_batch = None
        gid_int = int(gid) if gid else None
        if gid_int:
            g = db.get(GraduationDefenseGroup, gid_int)
            if not g or g.is_deleted or g.tenant_id != _tid():
                raise not_found("答辩组不存在")
            group_batch = g.batch_id
        stus = db.scalars(select(GraduationStudent).where(
            GraduationStudent.tenant_id == _tid(), GraduationStudent.is_deleted.is_(False),
            GraduationStudent.record_status == "ACTIVE").order_by(GraduationStudent.id)).all()
        out = []
        for s in stus:
            if s.stage not in ("FINAL_CHECK", "DEFENSE", "COMPLETED"):
                continue
            if group_batch is not None and s.batch_id != group_batch:
                continue
            if not (s.topic_id or s.stage not in ("TOPIC_SELECTING", None, "")):
                continue
            if keyword and keyword.strip() not in (s.name or ""):
                continue
            out.append({"id": str(s.id), "name": s.name, "className": s.class_name or "",
                        "topicTitle": s.topic_title or "", "advisorName": s.advisor_name or "",
                        "currentGroup": s.defense_group or "",
                        "assignedHere": s.defense_group_id == gid_int if gid_int else False,
                        "assignedElsewhere": bool(s.defense_group_id) and s.defense_group_id != gid_int})
        return out


def assign_defense_students(gid, student_ids) -> dict:
    with session() as db:
        g = db.get(GraduationDefenseGroup, int(gid))
        if not g or g.is_deleted or g.tenant_id != _tid():
            raise not_found("答辩组不存在")
        if not g.batch_id:
            raise AppException("DATA_CONFLICT", "答辩组未绑定批次，请先迁移/重建后再分配学生")
        current = len(_assigned_students(db, g.id))
        add_ids = [int(x) for x in (student_ids or [])]
        for sid in add_ids:
            s = db.get(GraduationStudent, sid)
            if not s or s.is_deleted or s.tenant_id != _tid():
                raise not_found(f"学生 {sid} 不存在")
            if s.defense_group_id == g.id:
                continue
            if s.batch_id != g.batch_id:
                raise AppException(
                    "DATA_CONFLICT",
                    f"学生 {s.name or sid} 与答辩组不在同一毕设批次，不能跨批分配",
                )
            if s.stage not in ("FINAL_CHECK", "DEFENSE", "COMPLETED"):
                raise AppException(
                    "DATA_CONFLICT",
                    f"学生 {s.name or sid} 须进入成果检查阶段后方可分配答辩组",
                )
            assert_student_access(db, s, "defense.assign")
            current += 1
            if current > MAX_DEFENSE_STUDENTS:
                raise AppException("DATA_CONFLICT", f"单个答辩组学生数不得超过 {MAX_DEFENSE_STUDENTS} 人")
            s.defense_group_id = g.id
            s.defense_group = g.group_name
            if s.stage == "FINAL_CHECK":
                final_ok = db.scalars(select(GraduationFinal).where(
                    GraduationFinal.tenant_id == _tid(), GraduationFinal.gd_student_id == s.id,
                    GraduationFinal.is_deleted.is_(False), GraduationFinal.final_type == "定稿",
                    GraduationFinal.status == "APPROVED",
                ).limit(1)).first()
                if final_ok:
                    s.stage = "DEFENSE"
        db.flush()
        _recompute_defense(db, g)
        g.published = False
        _audit(db, "DEFENSE", g.id, "分配学生进答辩组", f"{g.group_name} +{len(add_ids)} 人")
        db.commit()
        return get_defense_group_detail(g.id)


def unassign_defense_students(gid, student_ids) -> dict:
    with session() as db:
        g = db.get(GraduationDefenseGroup, int(gid))
        if not g or g.is_deleted or g.tenant_id != _tid():
            raise not_found("答辩组不存在")
        for sid in [int(x) for x in (student_ids or [])]:
            s = db.get(GraduationStudent, sid)
            if s and s.defense_group_id == g.id:
                s.defense_group_id = None
                s.defense_group = None
        db.flush()
        _recompute_defense(db, g)
        g.published = False
        _audit(db, "DEFENSE", g.id, "移出答辩组学生", g.group_name)
        db.commit()
        return get_defense_group_detail(g.id)


def publish_defense(gid) -> dict:
    with session() as db:
        g = db.get(GraduationDefenseGroup, int(gid))
        if not g or g.is_deleted or g.tenant_id != _tid():
            raise not_found("答辩组不存在")
        _recompute_defense(db, g)  # 发布前按最新分配重算冲突/人数
        if g.conflict:
            raise AppException("VALIDATION_ERROR", "存在评委与导师冲突，调整评委或学生后方可发布")
        if (g.chair or "待指定") in ("待指定", "") or (g.location or "待定") in ("待定", ""):
            raise AppException("VALIDATION_ERROR", "评委或地点未安排完整，暂不能发布")
        if g.student_count <= 0:
            raise AppException("VALIDATION_ERROR", "尚未分配答辩学生，暂不能发布")
        g.published = True
        g.version += 1
        _audit(db, "DEFENSE", g.id, "发布答辩安排", f"{g.group_name}（{g.student_count} 人）")
        db.commit()
        return {"id": str(g.id), "published": True}


def notify_defense_group(gid, user=None) -> dict:
    """向已发布答辩组学生投递站内信（receiver_id=StudentProfile.id）。"""
    from app.services.message_event_outbox_service import emit_message_event, process_pending_outbox

    if not gid:
        raise AppException("VALIDATION_ERROR", "defenseGroupId 必填")
    with session() as db:
        g = db.get(GraduationDefenseGroup, int(gid))
        if not g or g.is_deleted or g.tenant_id != _tid():
            raise not_found("答辩组不存在")
        if not _can_access_defense_group(db, g):
            raise no_permission("Defense group is outside the current graduation-design scope")
        if not g.published:
            return {"notified": 0, "skipped": 0, "groupName": g.group_name or "",
                    "message": "该答辩组未发布，暂不能通知"}
        students = [s for s in _assigned_students(db, g.id) if can_access_student(db, s)]
        when = (g.defense_date or "").strip() or "待通知"
        where = (g.location or "").strip() or "待定"
        title = f"答辩安排通知：{g.group_name or '答辩组'}"
        notified = 0
        skipped = 0
        for s in students:
            rid = int(s.student_id or 0)
            if rid <= 0:
                skipped += 1
                continue
            content = (
                f"同学你好，你的毕业设计答辩组「{g.group_name or ''}」已发布。"
                f"时间：{when}；地点：{where}；主席：{g.chair or '待指定'}。"
                f"请按时参加，如有冲突请尽快联系指导教师。"
            )
            emit_message_event(
                db,
                event_code="GRADUATION_DESIGN.DEFENSE_ARRANGED",
                source_module="graduation",
                source_biz_type="defense_group",
                source_biz_id=int(g.id),
                recipient_refs=[{"studentId": rid}],
                content=content,
                title=title,
                dedup_key=f"GRADUATION_DESIGN.DEFENSE_ARRANGED:{g.id}:student:{rid}",
            )
            notified += 1
        _audit(db, "DEFENSE", g.id, "发送答辩通知",
               f"{g.group_name} notified={notified} skipped={skipped}")
        db.commit()
        try:
            process_pending_outbox(limit=50, worker_id="graduation-inline")
        except Exception:  # noqa: BLE001
            pass
        if notified <= 0:
            msg = "暂无可投递学生（缺学籍关联或组内无学生）"
        else:
            msg = f"已向 {notified} 名学生发送答辩通知"
            if skipped:
                msg += f"（{skipped} 人缺学籍关联已跳过）"
        return {"notified": notified, "skipped": skipped, "groupName": g.group_name or "",
                "message": msg}


def export_defense_xlsx(batch_id=None) -> dict:
    """答辩安排台账 Excel 导出（一组一行；与列表同一 batch_id 口径）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    items, _ = list_defense_groups(1, 100000, batch_id=batch_id)
    headers = ["答辩组", "时间", "地点", "组长", "评委", "秘书", "学生数", "冲突", "发布状态"]
    operator, _role = _op()
    title = f"答辩安排台账　导出时间：{datetime.now():%Y-%m-%d %H:%M}　导出人：{operator}"
    wb = Workbook()
    ws = wb.active
    ws.title = "答辩安排台账"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="555555", size=10)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="DCE6F1")
    for c in ws[2]:
        c.font = Font(bold=True); c.fill = fill
    for it in items:
        ws.append([it["groupName"], it["date"], it["location"], it["chair"],
                  "、".join(it["members"]), it["secretary"], it["studentCount"],
                  it["conflict"] or "无", "已发布" if it["published"] else "未发布"])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 16
    ws.freeze_panes = "A3"
    import base64
    import io
    with session() as db:
        _audit(db, "DEFENSE", "export", "导出答辩安排台账",
               f"共 {len(items)} 组，批次={batch_id or '全部'}")
        db.commit()
    buf = io.BytesIO()
    wb.save(buf)
    return {"filename": f"答辩安排台账_{datetime.now():%Y%m%d_%H%M}.xlsx",
            "contentBase64": base64.b64encode(buf.getvalue()).decode("ascii"), "rowCount": len(items),
            "batchId": str(batch_id) if batch_id else None,
            "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}


def student_defense_view(gd_student_id) -> dict:
    """学生端查看本人答辩安排（仅已发布才展示时间/地点/评委）。"""
    with session() as db:
        s = db.get(GraduationStudent, int(gd_student_id))
        if not s or s.is_deleted or s.tenant_id != _tid():
            return {"hasData": False}
        if not s.defense_group_id:
            return {"hasData": True, "assigned": False, "message": "答辩分组尚未安排"}
        g = db.get(GraduationDefenseGroup, s.defense_group_id)
        if not g or g.is_deleted:
            return {"hasData": True, "assigned": False, "message": "答辩分组尚未安排"}
        if not g.published:
            return {"hasData": True, "assigned": True, "published": False,
                    "groupName": g.group_name, "message": "答辩安排编制中，发布后可见时间地点"}
        return {"hasData": True, "assigned": True, "published": True, "groupName": g.group_name,
                "date": g.defense_date or "待定", "location": g.location or "待定",
                "chair": g.chair or "", "members": g.members_json or [], "secretary": g.secretary or ""}


# ═══ 审计 + 看板 ═══

def list_audit(page, ps, biz_type=None, keyword=None):
    with session() as db:
        q = select(GraduationAuditTrail).where(GraduationAuditTrail.tenant_id == _tid())
        if biz_type:
            q = q.where(GraduationAuditTrail.biz_type == biz_type)
        rows = db.scalars(q.order_by(GraduationAuditTrail.id.desc())).all()
        if keyword:
            kw = keyword.strip()
            rows = [r for r in rows if kw in (r.action or "") or kw in (r.detail or "")]
        items = [{"id": str(x.id), "time": _iso(x.occurred_at), "operator": x.operator or "",
                  "roleName": x.role_name or "", "bizType": x.biz_type, "bizId": x.biz_id or "",
                  "action": x.action, "detail": x.detail or "", "before": x.before_val or "",
                  "after": x.after_val or "", "requestId": x.request_id or "",
                  "requestPath": x.request_path or "", "clientIp": x.client_ip or ""} for x in rows]
        return _page(items, page, ps)


def get_dashboard(batch_id=None) -> dict:
    with session() as db:
        visible_ids = accessible_student_ids(db, _tid(), batch_id=batch_id)
        total = len(visible_ids)
        scope = visible_ids or [-1]
        pend_prop = db.scalar(select(func.count()).select_from(GraduationProposal).where(
            GraduationProposal.tenant_id == _tid(), GraduationProposal.status == "PENDING_REVIEW",
            GraduationProposal.is_deleted.is_(False),
            GraduationProposal.gd_student_id.in_(scope))) or 0
        pend_final = db.scalar(select(func.count()).select_from(GraduationFinal).where(
            GraduationFinal.tenant_id == _tid(), GraduationFinal.status == "PENDING_REVIEW",
            GraduationFinal.is_deleted.is_(False),
            GraduationFinal.gd_student_id.in_(scope))) or 0
        high_risk = db.scalar(select(func.count()).select_from(GraduationStudent).where(
            GraduationStudent.tenant_id == _tid(), GraduationStudent.risk_level == "HIGH",
            GraduationStudent.is_deleted.is_(False), GraduationStudent.id.in_(scope))) or 0
        flow = {}
        for s in db.scalars(select(GraduationStudent).where(
                GraduationStudent.tenant_id == _tid(),
                GraduationStudent.is_deleted.is_(False),
                GraduationStudent.id.in_(scope))).all():
            flow[s.stage] = flow.get(s.stage, 0) + 1
        # 答辩待发布：按答辩组自身 batch_id（与列表/导出一致）
        bid = int(batch_id) if batch_id else None
        defense_q = select(GraduationDefenseGroup).where(
            GraduationDefenseGroup.tenant_id == _tid(), GraduationDefenseGroup.published.is_(False),
            GraduationDefenseGroup.is_deleted.is_(False))
        if bid is not None:
            defense_q = defense_q.where(GraduationDefenseGroup.batch_id == bid)
        defense_groups = db.scalars(defense_q).all()
        pend_defense = 0
        for group in defense_groups:
            if not _can_access_defense_group(db, group):
                continue
            pend_defense += 1
        # 未提交开题：与开题列表同一批次
        not_submitted = len(_not_submitted_proposals(db, batch_id=batch_id))
        # 真实风险预警（未关闭；限定当前批次）
        risk_alerts = []
        try:
            from app.modules.graduation.services import graduation_risk_service as risk_svc
            items, _ = risk_svc.list_risks(1, 50, batch_id=batch_id)
            for r in items:
                if r.get("status") == "CLOSED":
                    continue
                risk_alerts.append({"id": r["id"], "code": r["riskCode"], "title": r["riskName"],
                                    "level": r.get("level") or "MEDIUM",
                                    "detail": f"{r.get('studentName') or '—'}"
                                              + (f" · 指导 {r['advisorName']}" if r.get("advisorName") else "")
                                              + f" · {r.get('statusLabel') or ''}",
                                    "time": r.get("detectedAt") or ""})
                if len(risk_alerts) >= 6:
                    break
        except Exception:  # noqa: BLE001 - 风险模块异常不应影响看板主体
            risk_alerts = []
        # 跨模块统计：与综合统计同一批次
        module_stats = []
        try:
            from app.modules.graduation.services import graduation_stats_service as stats_svc
            ov = stats_svc.overview_stats(batch_id=batch_id)
            m, gu, mt = ov.get("mentor", {}), ov.get("guidance", {}), ov.get("midterm", {})
            rv, gr, ar = ov.get("review", {}), ov.get("grade", {}), ov.get("archive", {})

            def _done(stat, key):
                return next((x["count"] for x in stat.get("byStatus", []) if x["status"] == key), 0)
            module_stats = [
                {"label": "导师已合格", "value": str(m.get("qualifiedCount", 0)),
                 "hint": f"未分配学生 {m.get('unassignedStudents', 0)} · 满员 {m.get('fullCapacityCount', 0)}"},
                {"label": "指导平均次数", "value": str(gu.get("avgCount", 0)),
                 "hint": f"频次不足 {gu.get('insufficientCount', 0)} 人"},
                {"label": "中期检查", "value": str(mt.get("total", 0)),
                 "hint": f"待检 {_done(mt, 'PENDING')}"},
                {"label": "教师评阅", "value": str(rv.get("total", 0)),
                 "hint": f"已完成 {_done(rv, 'COMPLETED')}"},
                {"label": "成绩已发布均分", "value": str(gr.get("publishedAvg") or "—"),
                 "hint": f"优秀 {gr.get('excellentCount', 0)} 人"},
                {"label": "归档率", "value": f"{ar.get('archiveRate', 0)}%",
                 "hint": f"已备案 {ar.get('filedCount', 0)}/{ar.get('studentTotal', 0)}"},
            ]
        except Exception:  # noqa: BLE001 - 统计异常不影响看板主体
            module_stats = []
        # 当前批次信息：优先用请求的 batch_id；否则回落 RUNNING / 最近有效批次
        _BATCH_LABEL = {"DRAFT": "草稿", "RUNNING": "进行中", "CLOSED": "已结束",
                        "ARCHIVED": "已归档", "VOIDED": "已作废"}
        cur_batch = None
        if batch_id:
            cur_batch = db.get(GraduationBatch, int(batch_id))
            if cur_batch and (cur_batch.tenant_id != _tid() or cur_batch.is_deleted):
                cur_batch = None
        if not cur_batch:
            cur_batch = db.scalars(select(GraduationBatch).where(
                GraduationBatch.tenant_id == _tid(), GraduationBatch.is_deleted.is_(False),
                GraduationBatch.status == "RUNNING").order_by(GraduationBatch.id.desc())).first()
        if not cur_batch:
            cur_batch = db.scalars(select(GraduationBatch).where(
                GraduationBatch.tenant_id == _tid(), GraduationBatch.is_deleted.is_(False),
                GraduationBatch.status != "VOIDED").order_by(GraduationBatch.id.desc())).first()
        if cur_batch:
            _bs, _be = cur_batch.start_date, cur_batch.end_date
            batch_range = " ~ ".join(d.strftime("%Y-%m-%d") for d in (_bs, _be) if d) \
                or (cur_batch.academic_year or "")
            batch_name = cur_batch.batch_name
            batch_status = _BATCH_LABEL.get(cur_batch.status, cur_batch.status)
        else:
            batch_name, batch_range, batch_status = "暂无毕设批次", "", "未开始"
        _active_stage = _active_student_stage(cur_batch)
        return {"batchId": str(batch_id) if batch_id else (str(cur_batch.id) if cur_batch else None),
                "batchName": batch_name, "batchRange": batch_range,
                "moduleStats": module_stats,
                "batchStatus": batch_status,
                "stats": [
                    {"label": "毕设学生", "value": str(total), "trend": "", "trendQuality": "neutral"},
                    {"label": "开题待审阅", "value": str(pend_prop),
                     "trend": f"待批 {pend_prop}", "trendQuality": "bad" if pend_prop else "good"},
                    {"label": "成果待审阅", "value": str(pend_final),
                     "trend": f"待批 {pend_final}", "trendQuality": "neutral"},
                    {"label": "答辩待发布", "value": str(pend_defense),
                     "trend": f"未发布 {pend_defense} 组", "trendQuality": "neutral"},
                    {"label": "高风险学生", "value": str(high_risk),
                     "trend": "", "trendQuality": "bad" if high_risk else "good"},
                ],
                # active 按当前批次阶段时间轴真实推算（此前写死 FINAL_CHECK，无论毕设走到哪
                # 都恒亮「成果检查」）。未配阶段日期时 _active_student_stage 返回 None，不高亮。
                "flow": [{"label": L_STAGE[k], "value": flow.get(k, 0), "active": k == _active_stage}
                         for k in L_STAGE],
                "todos": [
                    {"id": "t1", "label": "开题材料待审阅", "count": pend_prop, "tone": "danger",
                     "route": "/admin/graduation/proposals", "hint": "指导教师批阅开题报告"},
                    {"id": "t2", "label": "开题未提交催交", "count": not_submitted, "tone": "warning",
                     "route": "/admin/graduation/proposals", "hint": "已确认选题但未交开题"},
                    {"id": "t3", "label": "成果待审阅", "count": pend_final, "tone": "warning",
                     "route": "/admin/graduation/finals", "hint": "论文初稿/定稿批阅"},
                    {"id": "t4", "label": "答辩组待发布", "count": pend_defense, "tone": "warning",
                     "route": "/admin/graduation/defense", "hint": "分组排期完成后发布"},
                    {"id": "t5", "label": "未处理风险", "count": len(risk_alerts), "tone": "danger",
                     "route": "/admin/graduation/risk-archive", "hint": "受理并处置过程风险"},
                ],
                "riskAlerts": risk_alerts}
