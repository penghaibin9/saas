from __future__ import annotations

import base64
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select

from app.core.context import get_current_user_ctx
from app.core.exceptions import AppException
from app.core.permissions import enforce_permission
from app.models import GraduationArchiveRecord, GraduationStudent
from app.modules.graduation.services.graduation_archive_data_quality import identity_anomaly_reasons
from app.modules.graduation.services.graduation_export_security import sanitize_xlsx_export
from app.modules.graduation.services.graduation_proposal_read_service import student_scope_select
from app.services.db_service import _tid, session


def _row(archive, student):
    from app.modules.graduation.services import graduation_archive_service as service

    item = service._row(archive, student)
    reasons = identity_anomaly_reasons(student)
    if not str(student.name or "").strip():
        item["studentName"] = "历史数据异常（姓名缺失）"
    if not str(student.student_no or "").strip():
        item["studentNo"] = "历史数据异常（学号缺失）"
    item["dataAnomaly"] = bool(reasons)
    item["anomalyReasons"] = reasons
    if reasons:
        # 仅改变读模型展示，不篡改数据库真实状态。前端现有状态按钮条件会自然失效，
        # 同时服务端写守卫提供最终 fail-closed，避免只靠隐藏按钮。
        item["status"] = "DATA_ANOMALY"
        item["statusLabel"] = "历史数据异常 · 只读"
        item["statusTone"] = "danger"
        item["missingItems"] = list(dict.fromkeys([*(item.get("missingItems") or []), *reasons]))
        item["allowedActions"] = []
    return item


def _query_parts(db, tid: int, *, keyword=None, status=None, batch_id=None):
    scope = student_scope_select(db, tid, batch_id=batch_id)
    join_on = GraduationStudent.id == GraduationArchiveRecord.gd_student_id
    filters = [
        GraduationArchiveRecord.tenant_id == tid,
        GraduationArchiveRecord.is_deleted.is_(False),
        GraduationStudent.tenant_id == tid,
        GraduationStudent.is_deleted.is_(False),
        GraduationStudent.record_status == "ACTIVE",
        GraduationStudent.id.in_(scope),
    ]
    if batch_id:
        filters.append(GraduationStudent.batch_id == int(batch_id))
    if status:
        filters.append(GraduationArchiveRecord.status == status)
    value = str(keyword or "").strip()
    if value:
        filters.append(or_(GraduationStudent.name.contains(value), GraduationStudent.student_no.contains(value)))
    return join_on, filters


def _page(db, tid: int, *, offset: int, limit: int, keyword=None, status=None, batch_id=None):
    if int(limit) <= 0:
        return []
    join_on, filters = _query_parts(
        db, tid, keyword=keyword, status=status, batch_id=batch_id,
    )
    rows = db.execute(
        select(GraduationArchiveRecord, GraduationStudent)
        .join(GraduationStudent, join_on)
        .where(*filters)
        .order_by(GraduationArchiveRecord.id.desc())
        .offset(max(0, int(offset)))
        .limit(int(limit))
    ).all()
    return [_row(archive, student) for archive, student in rows]


def list_archives(page: int, page_size: int, keyword=None, status=None, batch_id=None):
    tid = _tid()
    p = max(1, int(page))
    size = min(200, max(1, int(page_size)))
    with session() as db:
        join_on, filters = _query_parts(
            db, tid, keyword=keyword, status=status, batch_id=batch_id,
        )
        total = int(db.scalar(
            select(func.count(func.distinct(GraduationArchiveRecord.id)))
            .select_from(GraduationArchiveRecord)
            .join(GraduationStudent, join_on)
            .where(*filters)
        ) or 0)
        rows = _page(
            db, tid, offset=(p - 1) * size, limit=size,
            keyword=keyword, status=status, batch_id=batch_id,
        )
        return rows, total


def iter_archives(*, keyword=None, status=None, batch_id=None, chunk_size: int = 200):
    """Stream the authoritative archive read model in bounded SQL pages for XLSX export."""
    tid = _tid()
    size = min(200, max(1, int(chunk_size)))
    with session() as db:
        offset = 0
        while True:
            rows = _page(
                db, tid, offset=offset, limit=size,
                keyword=keyword, status=status, batch_id=batch_id,
            )
            if not rows:
                break
            yield from rows
            offset += len(rows)
            if len(rows) < size:
                break


@sanitize_xlsx_export
def export_archives_xlsx(status=None, keyword=None, batch_id=None) -> dict:
    enforce_permission(get_current_user_ctx() or {}, "graduationDesign.archive.export")
    if not batch_id:
        raise AppException("VALIDATION_ERROR", "导出前必须选择毕业设计批次")

    headers = ["学生", "学号", "状态", "缺失材料数", "提交时间", "备案时间", "归档批次号"]
    actor = get_current_user_ctx() or {}
    operator = actor.get("realName") or actor.get("name") or actor.get("loginName") or "系统"
    title = f"毕设归档台账　导出时间：{datetime.now():%Y-%m-%d %H:%M}　导出人：{operator}"

    wb = Workbook()
    ws = wb.active
    ws.title = "毕设归档台账"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="555555", size=10)
    ws.append(headers)
    fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = fill

    written = 0
    for item in iter_archives(
        keyword=keyword,
        status=status,
        batch_id=batch_id,
        chunk_size=200,
    ):
        ws.append([
            item["studentName"], item["studentNo"], item["statusLabel"], len(item["missingItems"]),
            (item["submittedAt"] or "")[:19], (item["filedAt"] or "")[:19], item["archiveBatchNo"],
        ])
        written += 1

    for index in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + index)].width = 18
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    return {
        "filename": f"毕设归档台账_{datetime.now():%Y%m%d_%H%M}.xlsx",
        "contentBase64": base64.b64encode(content).decode("ascii"),
        "rowCount": written,
        "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
