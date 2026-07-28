"""毕业设计域全部 XLSX 导出的二次公式注入净化。

公共 Excel 底座已经在写单元格时防护；本安装器继续覆盖历史上直接使用
openpyxl 的毕业设计导出，避免某个旧台账绕过公共底座。
"""
from __future__ import annotations

import base64
import hashlib
import io
from functools import wraps

from openpyxl import load_workbook

_INSTALLED = False
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_SERVICE_MODULES = (
    "graduation_service",
    "graduation_batch_service",
    "graduation_student_service",
    "graduation_topic_service",
    "graduation_topic_round_service",
    "graduation_taskbook_service",
    "graduation_guidance_service",
    "graduation_midterm_service",
    "graduation_review_service",
    "graduation_defense_score_service",
    "graduation_grade_service",
    "graduation_risk_service",
    "graduation_archive_service",
    "graduation_mentor_service",
    "graduation_more_service",
)


def sanitize_xlsx_result(result):
    if not isinstance(result, dict):
        return result
    media_type = str(result.get("mediaType") or "").lower()
    encoded = result.get("contentBase64")
    if "spreadsheetml" not in media_type or not encoded:
        return result
    try:
        raw = base64.b64decode(encoded, validate=True)
        wb = load_workbook(io.BytesIO(raw), data_only=False, keep_links=False)
    except Exception:
        # 原导出若不是有效 XLSX，应由原调用链报错；这里不把损坏文件伪装成功。
        return result

    changed = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if cell.data_type == "f":
                    cell.value = "'" + str(value or "")
                    changed = True
                elif isinstance(value, str) and value.lstrip().startswith(_FORMULA_PREFIXES):
                    cell.value = "'" + value
                    changed = True
    if not changed:
        wb.close()
        return result

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    content = buf.getvalue()
    packed = dict(result)
    packed["contentBase64"] = base64.b64encode(content).decode("ascii")
    if "sha256" in packed:
        packed["sha256"] = hashlib.sha256(content).hexdigest()
    return packed


def _wrap(fn):
    if getattr(fn, "_gd_xlsx_sanitized", False):
        return fn

    @wraps(fn)
    def wrapped(*args, **kwargs):
        return sanitize_xlsx_result(fn(*args, **kwargs))

    wrapped._gd_xlsx_sanitized = True
    return wrapped


def install_graduation_export_security() -> None:
    global _INSTALLED
    if _INSTALLED:
        return
    import importlib

    for short_name in _SERVICE_MODULES:
        try:
            module = importlib.import_module(f"app.modules.graduation.services.{short_name}")
        except ModuleNotFoundError:
            continue
        for name, value in list(vars(module).items()):
            lowered = name.lower()
            if not callable(value) or not (lowered.startswith("export") or "_export" in lowered):
                continue
            setattr(module, name, _wrap(value))
    _INSTALLED = True
