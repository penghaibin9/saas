"""Read-only verification of the browser-downloaded internship leave XLSX ledger."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

EXPECTED_HEADERS = [
    "学号", "姓名", "校内指导教师", "请假类型", "开始", "结束",
    "天数", "事由", "证明材料", "状态", "审批人", "审批意见",
]


def fail(message: str) -> None:
    raise SystemExit(f"[internship-leave-xlsx-audit] FAIL: {message}")


def text(value) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    if len(sys.argv) != 7:
        fail("usage: <xlsx> <row_count> <first_reason> <resubmit_reason> <reject_reason> <student_no>")

    path = Path(sys.argv[1]).resolve()
    try:
        expected_row_count = int(sys.argv[2])
    except ValueError:
        fail(f"invalid row_count: {sys.argv[2]}")
    first_reason, resubmit_reason, reject_reason, student_no = sys.argv[3:7]

    if path.suffix.lower() != ".xlsx" or not path.is_file():
        fail(f"browser download is not an existing .xlsx file: {path}")
    if path.stat().st_size <= 0:
        fail("browser-downloaded workbook is empty")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:  # noqa: BLE001 - evidence verifier must fail closed on malformed XLSX
        fail(f"openpyxl could not open browser-downloaded workbook: {exc}")

    try:
        if "请假审批台账" not in workbook.sheetnames:
            fail(f"missing worksheet 请假审批台账: {workbook.sheetnames}")
        sheet = workbook["请假审批台账"]

        watermark = text(sheet.cell(row=1, column=1).value)
        for token in ("岗位实习中心·请假审批台账", "导出人：", "导出留痕"):
            if token not in watermark:
                fail(f"watermark missing {token!r}: {watermark!r}")

        headers = [text(sheet.cell(row=2, column=index).value) for index in range(1, len(EXPECTED_HEADERS) + 1)]
        if headers != EXPECTED_HEADERS:
            fail(f"header contract mismatch: {headers}")

        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(min_row=3, max_col=len(EXPECTED_HEADERS), values_only=True):
            if not any(text(value) for value in values):
                continue
            rows.append({header: text(value) for header, value in zip(EXPECTED_HEADERS, values)})

        if len(rows) != expected_row_count:
            fail(f"workbook row count {len(rows)} != export API rowCount {expected_row_count}")

        rejected = next((row for row in rows if row["事由"] == first_reason), None)
        if not rejected:
            fail("browser-created rejected leave is missing from workbook")
        if rejected["学号"] != student_no:
            fail(f"rejected row studentNo {rejected['学号']} != {student_no}")
        if rejected["状态"] != "已驳回":
            fail(f"rejected row status is {rejected['状态']!r}, expected 已驳回")
        if rejected["审批意见"] != reject_reason:
            fail("rejected row approval comment does not match browser-entered rejection reason")

        returned = next((row for row in rows if row["事由"] == resubmit_reason), None)
        if not returned:
            fail("browser-created returned leave is missing from workbook")
        if returned["学号"] != student_no:
            fail(f"returned row studentNo {returned['学号']} != {student_no}")
        if returned["状态"] != "已销假":
            fail(f"returned row status is {returned['状态']!r}, expected 已销假")

        print("[internship-leave-xlsx-audit] XLSX_EVIDENCE_OK")
        print(
            f"sheet=请假审批台账 rows={len(rows)} studentNo={student_no} "
            f"rejected={rejected['状态']} returned={returned['状态']}"
        )
    finally:
        workbook.close()


if __name__ == "__main__":
    main()
