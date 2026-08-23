"""Bootstrap the minimal Academic C grade Browser-First identities through canonical School IAM.

This E2E helper intentionally follows the production identity-import authority:
separate TEACHER/STUDENT workbooks -> FileObject -> explicit process -> canonical confirm.
The retired mixed parser is never used and authentication throttling stays enabled;
CI logins use distinct trusted X-Forwarded-For identities instead of weakening rate limits.
"""
from __future__ import annotations

import io
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.identity_import_file_service import build_student_template, build_teacher_template  # noqa: E402
from scripts.e2e_bootstrap_academic_affairs_accounts import (  # noqa: E402
    ADMIN,
    CLASS_A1,
    COLLEGE_A,
    GRADE,
    MAJOR_A,
    STABLE_PWD,
    TENANT,
    _req,
    ensure_org,
    multipart,
)

OUT_DIR = Path(__file__).resolve().parents[1] / "tmp"
OUT_DIR.mkdir(exist_ok=True)
CRED_PATH = OUT_DIR / "e2e_academic_affairs_credentials.local.json"

TEACHERS = [
    ["e2e_aa_college_a", "E2E教务测试学院教务老师A", COLLEGE_A, "学院教务", "COLLEGE_ADMIN", "COLLEGE", COLLEGE_A],
    ["e2e_aa_teacher_a", "E2E教务测试任课教师A", COLLEGE_A, "任课教师", "ACADEMIC_TEACHER", "COLLEGE", COLLEGE_A],
    ["e2e_aa_teacher_b", "E2E教务测试任课教师B", COLLEGE_A, "任课教师", "ACADEMIC_TEACHER", "COLLEGE", COLLEGE_A],
    ["e2e_aa_grade", "E2E教务测试成绩审核发布", "教务处", "成绩审核发布", "ACADEMIC_ADMIN", "", ""],
]
STUDENTS = [
    ["E2EAA20260001", "E2E教务测试学生A", COLLEGE_A, MAJOR_A, CLASS_A1, GRADE, "男", ""],
    ["E2EAA20260002", "E2E教务测试学生B", COLLEGE_A, MAJOR_A, CLASS_A1, GRADE, "女", ""],
]
REQUIRED_LOGINS = [row[0] for row in TEACHERS] + [row[0] for row in STUDENTS]


def _admin_login() -> str:
    result = _req(
        "POST",
        "/auth/login",
        headers={"X-Forwarded-For": "10.252.0.10"},
        body={"loginName": ADMIN[0], "password": ADMIN[1], "tenantCode": TENANT},
    )
    if result.get("code") != 0:
        raise SystemExit("academic grade admin login failed: " + json.dumps(result, ensure_ascii=False))
    return result["data"]["accessToken"]


def _workbook(template: bytes, rows: list[list[str]]) -> bytes:
    wb = load_workbook(io.BytesIO(template))
    ws = wb["导入模板"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def _canonical_import(token: str, *, kind: str, content: bytes) -> dict:
    filename = f"e2e_academic_grade_{kind}.xlsx"
    raw, boundary = multipart(content, filename)
    created = _req(
        "POST",
        f"/data-exchange/imports/identity/{kind}/validate-file",
        token=token,
        raw=raw,
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Idempotency-Key": f"academic-grade-browser-{kind}-upload-v1",
        },
    )
    if created.get("code") != 0:
        raise SystemExit(f"canonical {kind} upload failed: {json.dumps(created, ensure_ascii=False)}")
    item = dict(created.get("data") or {})
    job_id = str(item.get("id") or item.get("jobId") or "").strip()
    if not job_id:
        raise SystemExit(f"canonical {kind} upload returned no job id")
    status = str(item.get("status") or "").upper()
    if status != "SUCCEEDED" and status != "VALIDATED":
        processed = _req("POST", f"/data-exchange/imports/{job_id}/process", token=token)
        if processed.get("code") != 0:
            raise SystemExit(f"canonical {kind} process failed: {json.dumps(processed, ensure_ascii=False)}")
        item = dict(processed.get("data") or {})
        status = str(item.get("status") or "").upper()
    if status == "SUCCEEDED":
        return item
    if status != "VALIDATED":
        raise SystemExit(
            f"canonical {kind} did not validate: "
            + json.dumps(
                {
                    "jobId": job_id,
                    "status": item.get("status"),
                    "invalidRows": item.get("invalidRows"),
                    "errorMessage": item.get("errorMessage") or item.get("message"),
                },
                ensure_ascii=False,
            )
        )
    confirmed = _req(
        "POST",
        f"/data-exchange/imports/{job_id}/confirm",
        token=token,
        body={"expectedVersion": int(item.get("version") or 0)},
        headers={"Idempotency-Key": f"academic-grade-browser-{kind}-confirm-v1"},
    )
    if confirmed.get("code") != 0:
        raise SystemExit(f"canonical {kind} confirm failed: {json.dumps(confirmed, ensure_ascii=False)}")
    return dict(confirmed.get("data") or {})


def _find_user(token: str, login_name: str) -> dict | None:
    result = _req("GET", f"/system/users?keyword={login_name}&page=1&page_size=20", token=token)
    if result.get("code") != 0:
        return None
    for row in (result.get("data") or {}).get("list") or []:
        if str(row.get("loginName") or row.get("userNo") or "") == login_name:
            return row
    return None


def _normalize_password(token: str, login_name: str, index: int) -> dict:
    user = _find_user(token, login_name)
    if not user:
        return {"loginName": login_name, "ok": False, "message": "user_not_found"}
    user_id = user.get("id") or user.get("userId")
    reset = _req("POST", f"/system/users/{user_id}/reset-password", token=token, body={})
    temp = (reset.get("data") or {}).get("tempPassword") or (reset.get("data") or {}).get("temporaryPassword")
    if reset.get("code") != 0 or not temp:
        return {"loginName": login_name, "ok": False, "message": reset.get("message") or "no_temp_password"}

    first = _req(
        "POST",
        "/auth/login",
        headers={"X-Forwarded-For": f"10.252.1.{20 + index}"},
        body={"loginName": login_name, "password": temp, "tenantCode": TENANT},
    )
    if first.get("code") != 0:
        return {"loginName": login_name, "ok": False, "message": f"temp_login:{first.get('message')}"}
    changed = _req(
        "POST",
        "/auth/change-password",
        token=first["data"]["accessToken"],
        body={"oldPassword": temp, "newPassword": STABLE_PWD, "confirmPassword": STABLE_PWD},
    )
    if changed.get("code") != 0:
        return {"loginName": login_name, "ok": False, "message": f"change:{changed.get('message')}"}
    verified = _req(
        "POST",
        "/auth/login",
        headers={"X-Forwarded-For": f"10.252.2.{20 + index}"},
        body={"loginName": login_name, "password": STABLE_PWD, "tenantCode": TENANT},
    )
    return {
        "loginName": login_name,
        "ok": verified.get("code") == 0,
        "role": (((verified.get("data") or {}).get("currentRole") or {}).get("roleCode")),
        "message": None if verified.get("code") == 0 else verified.get("message"),
    }


def main() -> int:
    token = _admin_login()
    org = ensure_org(token)
    print("org:", json.dumps(org, ensure_ascii=False))
    teacher_receipt = _canonical_import(token, kind="teachers", content=_workbook(build_teacher_template(), TEACHERS))
    student_receipt = _canonical_import(token, kind="students", content=_workbook(build_student_template(), STUDENTS))
    results = [_normalize_password(token, login_name, idx) for idx, login_name in enumerate(REQUIRED_LOGINS)]
    print("account verification:", json.dumps(results, ensure_ascii=False, indent=2))
    ok_logins = {item["loginName"] for item in results if item.get("ok")}
    CRED_PATH.write_text(
        json.dumps(
            {
                "tenantCode": TENANT,
                "stablePassword": STABLE_PWD,
                "passwords": {login_name: STABLE_PWD for login_name in REQUIRED_LOGINS if login_name in ok_logins},
                "loginResults": results,
                "canonicalImport": {
                    "teachers": str(teacher_receipt.get("id") or teacher_receipt.get("jobId") or "confirmed"),
                    "students": str(student_receipt.get("id") or student_receipt.get("jobId") or "confirmed"),
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return 0 if len(ok_logins) == len(REQUIRED_LOGINS) else 1


if __name__ == "__main__":
    raise SystemExit(main())
