"""E2E: bootstrap 岗位实习多角色账号 + 双学院组织（sandbox-school）。

经官方 identity-import 创建，凭据仅写入 backend/tmp/（gitignore）。
前缀统一：E2E岗位实习测试 / e2e_ix_* / E2EIX*
"""
from __future__ import annotations

import io
import json
import sys
import time
import uuid
from pathlib import Path

import urllib.error
import urllib.request
from openpyxl import load_workbook

BASE = "http://127.0.0.1:8011/api/v1"
TENANT = "sandbox-school"
ADMIN = ("admin2", "123456")
STABLE = "E2eTest@2026"
OUT_DIR = Path(__file__).resolve().parents[1] / "tmp"
OUT_DIR.mkdir(exist_ok=True)
CRED_PATH = OUT_DIR / "e2e_internship_credentials.local.json"

COLLEGE_A = "E2E岗位实习测试智能制造学院"
COLLEGE_B = "E2E岗位实习测试信息工程学院"
MAJOR_A = "E2E岗位实习测试工业机器人技术"
MAJOR_B = "E2E岗位实习测试软件技术"
CLASS_A1 = "E2E岗位实习测试机器人2401班"
CLASS_A2 = "E2E岗位实习测试机器人2402班"
CLASS_B1 = "E2E岗位实习测试软件2401班"
GRADE = "2024"

TEACHERS = [
    # login, name, dept, title, roles, scopeType, scopeName
    ("e2e_ix_admin", "E2E岗位实习管理员", "教务处", "岗位实习管理员",
     "COLLEGE_ADMIN", "COLLEGE", COLLEGE_A),
    ("e2e_ix_college_a", "E2E学院实习负责人A", COLLEGE_A, "学院实习负责人",
     "COLLEGE_ADMIN", "COLLEGE", COLLEGE_A),
    ("e2e_ix_college_b", "E2E学院实习负责人B", COLLEGE_B, "学院实习负责人",
     "COLLEGE_ADMIN", "COLLEGE", COLLEGE_B),
    ("e2e_ix_mentor_a", "刘敏", COLLEGE_A, "学校指导教师",
     "INTERN_MENTOR", "", ""),
    ("e2e_ix_mentor_b", "王建国", COLLEGE_A, "学校指导教师",
     "INTERN_MENTOR", "", ""),
    ("e2e_ix_counselor", "E2E辅导员A", COLLEGE_A, "辅导员",
     "COUNSELOR", "CLASS", CLASS_A1),
    ("e2e_ix_score", "E2E评价审核人员", COLLEGE_A, "评价审核",
     "INTERN_MENTOR", "", ""),
    ("e2e_ix_employment", "E2E就业归档管理员", "就业办", "就业归档",
     "EMPLOYMENT_TEACHER", "", ""),
]

STUDENTS = [
    # studentNo, name, college, major, class, gender
    ("E2EIX20260001", "陈晓雨", COLLEGE_A, MAJOR_A, CLASS_A1, "女"),
    ("E2EIX20260002", "李明远", COLLEGE_A, MAJOR_A, CLASS_A1, "男"),
    ("E2EIX20260003", "周安然", COLLEGE_A, MAJOR_A, CLASS_A2, "男"),
]

ALL_LOGINS = ["admin2"] + [t[0] for t in TEACHERS] + [s[0] for s in STUDENTS]


def _req(method: str, path: str, token: str | None = None, body: dict | None = None,
         raw: bytes | None = None, headers: dict | None = None):
    data = raw
    hdrs = dict(headers or {})
    if token:
        hdrs["Authorization"] = f"Bearer {token}"
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        hdrs.setdefault("Content-Type", "application/json")
    req = urllib.request.Request(f"{BASE}{path}", data=data, headers=hdrs, method=method)
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            raw_body = resp.read()
            return json.loads(raw_body.decode("utf-8")) if raw_body else {"code": 0, "data": None}
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        try:
            return json.loads(detail)
        except json.JSONDecodeError:
            return {"code": exc.code, "message": detail}


def login_admin(retries: int = 6) -> str:
    for i in range(retries):
        r = _req("POST", "/auth/login", body={
            "loginName": ADMIN[0], "password": ADMIN[1], "tenantCode": TENANT,
        })
        if r.get("code") == 0:
            return r["data"]["accessToken"]
        if r.get("bizCode") == "RATE_LIMITED" or r.get("code") == 429001:
            print("rate limited, sleep 65")
            time.sleep(65)
            continue
        raise SystemExit(f"admin2 login failed: {r}")
    raise SystemExit("admin2 login failed after retries")


def _find_named(nodes: list, name: str):
    for n in nodes or []:
        if n.get("name") == name:
            return n
    return None


def ensure_org(token: str) -> dict:
    tree = _req("GET", "/system/org-tree", token=token)
    nodes = tree.get("data") or []
    created = []

    def ensure_college(name: str, code: str):
        c = _find_named(nodes, name)
        if c:
            return c
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "COLLEGE", "name": name, "code": code,
        })
        assert r.get("code") == 0, r
        created.append(name)
        # refresh tree
        t2 = _req("GET", "/system/org-tree", token=token)
        return _find_named(t2.get("data") or [], name)

    def ensure_child(parent: dict, typ: str, name: str, code: str):
        kids = parent.get("children") or []
        hit = _find_named(kids, name)
        if hit:
            return hit
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": typ, "name": name, "code": code, "parentId": parent["id"],
        })
        assert r.get("code") == 0, r
        created.append(name)
        det = _req("GET", f"/system/org-nodes/{parent['id']}", token=token)
        # fallback: re-fetch tree
        t2 = _req("GET", "/system/org-tree", token=token)
        col = _find_named(t2.get("data") or [], parent.get("name"))
        if not col:
            return {"id": r["data"]["id"], "name": name, "children": []}
        if typ == "MAJOR":
            return _find_named(col.get("children") or [], name) or {"id": r["data"]["id"], "name": name}
        # CLASS under major
        for m in col.get("children") or []:
            hit = _find_named(m.get("children") or [], name)
            if hit:
                return hit
        return {"id": r["data"]["id"], "name": name}

    ca = ensure_college(COLLEGE_A, "E2E-IX-COL-A")
    cb = ensure_college(COLLEGE_B, "E2E-IX-COL-B")
    # refresh after colleges
    tree = _req("GET", "/system/org-tree", token=token)
    nodes = tree.get("data") or []
    ca = _find_named(nodes, COLLEGE_A) or ca
    cb = _find_named(nodes, COLLEGE_B) or cb

    ma = ensure_child(ca, "MAJOR", MAJOR_A, "E2E-IX-MAJ-A")
    # refresh for major children
    tree = _req("GET", "/system/org-tree", token=token)
    ca = _find_named(tree.get("data") or [], COLLEGE_A)
    ma = _find_named((ca or {}).get("children") or [], MAJOR_A) or ma
    c1 = ensure_child(ma, "CLASS", CLASS_A1, "E2E-IX-CLS-A1")
    tree = _req("GET", "/system/org-tree", token=token)
    ca = _find_named(tree.get("data") or [], COLLEGE_A)
    ma = _find_named((ca or {}).get("children") or [], MAJOR_A) or ma
    c2 = ensure_child(ma, "CLASS", CLASS_A2, "E2E-IX-CLS-A2")

    mb = ensure_child(cb, "MAJOR", MAJOR_B, "E2E-IX-MAJ-B")
    tree = _req("GET", "/system/org-tree", token=token)
    cb = _find_named(tree.get("data") or [], COLLEGE_B)
    mb = _find_named((cb or {}).get("children") or [], MAJOR_B) or mb
    cb1 = ensure_child(mb, "CLASS", CLASS_B1, "E2E-IX-CLS-B1")

    return {
        "collegeA": (ca or {}).get("id"),
        "collegeB": (cb or {}).get("id"),
        "majorA": (ma or {}).get("id"),
        "majorB": (mb or {}).get("id"),
        "classA1": (c1 or {}).get("id"),
        "classA2": (c2 or {}).get("id"),
        "classB1": (cb1 or {}).get("id"),
        "created": created,
    }


def build_xlsx(token: str) -> bytes:
    tpl = urllib.request.urlopen(urllib.request.Request(
        f"{BASE}/system/identity-import/template",
        headers={"Authorization": f"Bearer {token}"},
    )).read()
    wb = load_workbook(io.BytesIO(tpl))
    ws = wb["导入模板"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for login, name, dept, title, roles, scope_type, scope_name in TEACHERS:
        ws.append([
            "TEACHER", login, name, "", "", "", "", "",
            dept, title, roles, scope_type, scope_name,
        ])
    for sno, name, college, major, clazz, gender in STUDENTS:
        ws.append([
            "STUDENT", sno, name, college, major, clazz, GRADE, gender,
            "", "", "", "", "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def multipart(content: bytes, filename: str) -> tuple[bytes, str]:
    boundary = "----Bound" + uuid.uuid4().hex
    parts = [
        f"--{boundary}".encode(),
        f'Content-Disposition: form-data; name="file"; filename="{filename}"'.encode(),
        b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        b"",
        content,
        f"--{boundary}--".encode(),
        b"",
    ]
    return b"\r\n".join(parts), boundary


def import_accounts(token: str) -> dict:
    content = build_xlsx(token)
    body, boundary = multipart(content, "e2e_internship_identity.xlsx")
    validated = _req(
        "POST", "/system/identity-import/validate-file", token=token, raw=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    print("validate summary:", json.dumps({
        k: (validated.get("data") or {}).get(k)
        for k in ("batchNo", "total", "valid", "invalid", "errorCount", "canConfirm", "status")
    }, ensure_ascii=False))
    if validated.get("code") != 0:
        raise SystemExit(validated)
    data = validated.get("data") or {}
    batch_no = data.get("batchNo")
    if not batch_no:
        raise SystemExit(f"no batchNo: {validated}")
    errors = data.get("errors") or data.get("errorRows") or []
    if errors:
        print("precheck errors sample:", json.dumps(errors[:10], ensure_ascii=False, indent=2))
    confirmed = _req("POST", "/system/identity-import/confirm-batch", token=token, body={
        "batchNo": batch_no,
    })
    print("confirm:", confirmed.get("code"), confirmed.get("message"))
    if confirmed.get("code") != 0:
        print("confirm soft-fail (may already exist):", confirmed.get("message"))
        return {"batchNo": batch_no, "confirmed": False, "detail": confirmed}
    receipt = confirmed.get("data") or {}
    CRED_PATH.write_text(json.dumps(receipt, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"batchNo": batch_no, "confirmed": True, "detail": receipt}


def _extract_credentials(receipt: dict) -> dict[str, str]:
    import base64
    pwd_map: dict[str, str] = {}
    if not isinstance(receipt, dict):
        return pwd_map
    blobs = [receipt]
    cr = receipt.get("credentialReceipt")
    if isinstance(cr, dict):
        blobs.append(cr)
        b64 = cr.get("contentBase64")
        if b64:
            raw = base64.b64decode(b64)
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h or "").strip() for h in rows[0]]
                for row in rows[1:]:
                    cells = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
                    ln = cells.get("工号/学号") or cells.get("loginName") or cells.get("账号")
                    pwd = cells.get("初始密码") or cells.get("password") or cells.get("initialPassword")
                    if ln and pwd:
                        pwd_map[str(ln).strip()] = str(pwd).strip()
    for blob in blobs:
        for key in ("accounts", "credentials", "items", "studentCredentials",
                    "teacherCredentials", "rows", "passwords"):
            items = blob.get(key)
            if isinstance(items, dict):
                for ln, pwd in items.items():
                    if isinstance(pwd, str) and pwd:
                        pwd_map[str(ln)] = pwd
                    elif isinstance(pwd, dict):
                        p = pwd.get("password") or pwd.get("initialPassword")
                        if p:
                            pwd_map[str(ln)] = p
                continue
            if not isinstance(items, list):
                continue
            for it in items:
                if not isinstance(it, dict):
                    continue
                ln = it.get("loginName") or it.get("accountNo") or it.get("工号/学号") or it.get("userNo")
                pwd = it.get("password") or it.get("initialPassword") or it.get("初始密码")
                if ln and pwd:
                    pwd_map[str(ln)] = str(pwd)
    return pwd_map


def activate_passwords(token: str, receipt: dict) -> dict:
    """Reset + change-password to STABLE for all E2E logins (rate-limit aware)."""
    pwd_map = _extract_credentials(receipt)
    if CRED_PATH.exists():
        try:
            prev = json.loads(CRED_PATH.read_text(encoding="utf-8"))
            if isinstance(prev.get("passwords"), dict):
                pwd_map.update({k: v for k, v in prev["passwords"].items() if v})
        except Exception:  # noqa: BLE001
            pass
    pwd_map["admin2"] = ADMIN[1]

    # list users
    users = {}
    page = 1
    while True:
        r = _req("GET", f"/system/users?page={page}&page_size=50", token=token)
        items = (r.get("data") or {}).get("list") or (r.get("data") or {}).get("items") or []
        for u in items:
            ln = str(u.get("loginName") or "")
            if ln in ALL_LOGINS or ln.startswith("e2e_ix_") or ln.startswith("E2EIX"):
                users[ln] = u
        total = int((r.get("data") or {}).get("total") or 0)
        if page * 50 >= total or not items:
            break
        page += 1

    results = []
    final = {"admin2": ADMIN[1]}
    for i, ln in enumerate(ALL_LOGINS):
        if ln == "admin2":
            results.append({"loginName": ln, "ok": True, "role": "SCHOOL_ADMIN", "skipped": True})
            continue
        # try stable first
        lg = _req("POST", "/auth/login", body={
            "loginName": ln, "password": STABLE, "tenantCode": TENANT,
        })
        if lg.get("code") == 0:
            final[ln] = STABLE
            role = (lg["data"].get("currentRole") or {}).get("roleCode")
            results.append({"loginName": ln, "ok": True, "role": role, "skipped": True})
            continue
        if lg.get("bizCode") == "RATE_LIMITED":
            print("sleep rate"); time.sleep(65)
        if i % 3 == 0:
            token = login_admin()
        uid = (users.get(ln) or {}).get("id")
        if not uid:
            results.append({"loginName": ln, "ok": False, "message": "user_not_found"})
            continue
        reset = _req("POST", f"/system/users/{uid}/reset-password", token=token, body={})
        temp = (reset.get("data") or {}).get("tempPassword")
        if not temp:
            results.append({"loginName": ln, "ok": False, "message": "no_temp", "reset": reset})
            time.sleep(2)
            continue
        time.sleep(2)
        lg = _req("POST", "/auth/login", body={
            "loginName": ln, "password": temp, "tenantCode": TENANT,
        })
        tries = 0
        while lg.get("bizCode") == "RATE_LIMITED" and tries < 3:
            time.sleep(65); tries += 1
            lg = _req("POST", "/auth/login", body={
                "loginName": ln, "password": temp, "tenantCode": TENANT,
            })
        if lg.get("code") != 0:
            results.append({"loginName": ln, "ok": False, "message": lg.get("message")})
            continue
        tok = lg["data"]["accessToken"]
        ch = _req("POST", "/auth/change-password", tok, body={
            "oldPassword": temp, "newPassword": STABLE, "confirmPassword": STABLE,
        })
        lg2 = _req("POST", "/auth/login", body={
            "loginName": ln, "password": STABLE, "tenantCode": TENANT,
        })
        tries = 0
        while lg2.get("bizCode") == "RATE_LIMITED" and tries < 3:
            time.sleep(65); tries += 1
            lg2 = _req("POST", "/auth/login", body={
                "loginName": ln, "password": STABLE, "tenantCode": TENANT,
            })
        ok = lg2.get("code") == 0
        if ok:
            final[ln] = STABLE
        results.append({
            "loginName": ln, "ok": ok,
            "role": ((lg2.get("data") or {}).get("currentRole") or {}).get("roleCode") if ok else None,
            "change": ch.get("code"),
            "message": None if ok else lg2.get("message"),
        })
        time.sleep(2)

    payload = {
        "tenantCode": TENANT,
        "stablePassword": STABLE,
        "passwords": final,
        "loginResults": results,
        "accounts": {
            "sysadmin": "admin2",
            "internshipAdmin": "e2e_ix_admin",
            "collegeA": "e2e_ix_college_a",
            "collegeB": "e2e_ix_college_b",
            "mentorA": "e2e_ix_mentor_a",
            "mentorB": "e2e_ix_mentor_b",
            "counselor": "e2e_ix_counselor",
            "scoreReviewer": "e2e_ix_score",
            "employment": "e2e_ix_employment",
            "studentA": "E2EIX20260001",
            "studentB": "E2EIX20260002",
            "studentC": "E2EIX20260003",
        },
        "orgNames": {
            "collegeA": COLLEGE_A, "collegeB": COLLEGE_B,
            "majorA": MAJOR_A, "majorB": MAJOR_B,
            "classA1": CLASS_A1, "classA2": CLASS_A2, "classB1": CLASS_B1,
        },
        "note": "企业导师无独立登录端；企业确认/企业评价由学校教师代录",
    }
    CRED_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {CRED_PATH}")
    ok_n = sum(1 for r in results if r.get("ok"))
    print(f"login_ok={ok_n}/{len(results)}")
    return payload


def main() -> int:
    token = login_admin()
    org = ensure_org(token)
    print("org:", json.dumps(org, ensure_ascii=False))
    imported = import_accounts(token)
    receipt = imported.get("detail") if imported.get("confirmed") else {}
    if CRED_PATH.exists() and not (isinstance(receipt, dict) and receipt):
        try:
            receipt = json.loads(CRED_PATH.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            receipt = {}
    token = login_admin()
    payload = activate_passwords(token, receipt if isinstance(receipt, dict) else {})
    ok_n = sum(1 for r in payload["loginResults"] if r.get("ok"))
    return 0 if ok_n >= 8 else 1


if __name__ == "__main__":
    sys.exit(main())
