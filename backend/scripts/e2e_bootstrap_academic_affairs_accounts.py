"""E2E: bootstrap academic-affairs multi-role accounts + org in sandbox-school.

Creates E2E教务测试-prefixed colleges/majors/classes and imports teachers/students
via official identity-import. Credentials only under backend/tmp/.
"""
from __future__ import annotations

import io
import json
import sys
import time
import uuid
from pathlib import Path

import urllib.error
import urllib.request
from openpyxl import load_workbook

BASE = "http://127.0.0.1:8000/api/v1"
TENANT = "sandbox-school"
ADMIN = ("admin2", "123456")
STABLE_PWD = "E2eTest@2026"
OUT_DIR = Path(__file__).resolve().parents[1] / "tmp"
OUT_DIR.mkdir(exist_ok=True)
CRED_PATH = OUT_DIR / "e2e_academic_affairs_credentials.local.json"
STATE_PATH = OUT_DIR / "e2e_academic_affairs_state.local.json"

COLLEGE_A = "E2E教务测试学院A"
COLLEGE_B = "E2E教务测试学院B"
MAJOR_A = "E2E教务测试软件技术"
MAJOR_B = "E2E教务测试机电一体化"
CLASS_A1 = "E2E教务测试软技2601班"
CLASS_A2 = "E2E教务测试软技2602班"
CLASS_B1 = "E2E教务测试机电2601班"
GRADE = "2026"

# loginName, displayName, roles CSV, scopeType, scopeRef, dept
# 注意：SCHOOL_ADMIN/SYS_ADMIN 不可经师生导入（teacherAssignable=false）；
# E2E系统管理员由 admin2 承担。
TEACHERS = [
    ("e2e_aa_admin", "E2E教务测试教务管理员", "ACADEMIC_ADMIN", "", "", "教务处"),
    ("e2e_aa_college_a", "E2E教务测试学院教务老师A", "COLLEGE_ADMIN", "COLLEGE", COLLEGE_A, COLLEGE_A),
    ("e2e_aa_college_b", "E2E教务测试学院教务老师B", "COLLEGE_ADMIN", "COLLEGE", COLLEGE_B, COLLEGE_B),
    ("e2e_aa_program", "E2E教务测试培养方案管理员", "ACADEMIC_ADMIN", "", "", "教务处"),
    ("e2e_aa_course", "E2E教务测试课程管理员", "ACADEMIC_ADMIN", "", "", "教务处"),
    ("e2e_aa_teacher_a", "E2E教务测试任课教师A", "ACADEMIC_TEACHER", "COLLEGE", COLLEGE_A, COLLEGE_A),
    ("e2e_aa_teacher_b", "E2E教务测试任课教师B", "ACADEMIC_TEACHER", "COLLEGE", COLLEGE_B, COLLEGE_B),
    ("e2e_aa_exam", "E2E教务测试考务管理员", "ACADEMIC_ADMIN", "", "", "教务处"),
    ("e2e_aa_grade", "E2E教务测试成绩审核发布", "ACADEMIC_ADMIN", "", "", "教务处"),
    ("e2e_aa_resource", "E2E教务测试教学资源管理员", "ACADEMIC_ADMIN", "", "", "教务处"),
    ("e2e_aa_quality", "E2E教务测试教学评价质量管理员", "ACADEMIC_ADMIN", "", "", "教务处"),
]

STUDENTS = [
    ("E2EAA20260001", "E2E教务测试学生A", COLLEGE_A, MAJOR_A, CLASS_A1, "男"),
    ("E2EAA20260002", "E2E教务测试学生B", COLLEGE_A, MAJOR_A, CLASS_A1, "女"),
    ("E2EAA20260003", "E2E教务测试学生C", COLLEGE_A, MAJOR_A, CLASS_A2, "男"),
]

ALL_LOGINS = ["admin2"] + [t[0] for t in TEACHERS] + [s[0] for s in STUDENTS]
# 文档角色：E2E系统管理员 = admin2（不单独导入）


def _req(method: str, path: str, token: str | None = None, body: dict | None = None,
         raw: bytes | None = None, headers: dict | None = None):
    data = raw
    hdrs = dict(headers or {})
    if token:
        hdrs["Authorization"] = f"Bearer {token}"
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        hdrs.setdefault("Content-Type", "application/json")
    req = urllib.request.Request(f"{BASE}{path}", data=data, headers=hdrs, method=method)
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            raw_body = resp.read()
            if not raw_body:
                return {"code": 0, "data": None}
            return json.loads(raw_body.decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        try:
            return json.loads(detail)
        except json.JSONDecodeError:
            return {"code": exc.code, "message": detail, "bizCode": str(exc.code)}


def login_admin() -> str:
    r = _req("POST", "/auth/login", body={
        "loginName": ADMIN[0], "password": ADMIN[1], "tenantCode": TENANT,
    })
    if r.get("code") != 0:
        raise SystemExit(f"admin2 login failed: {r}")
    return r["data"]["accessToken"]


def _find_node(nodes, name: str):
    for n in nodes or []:
        if n.get("name") == name:
            return n
        hit = _find_node(n.get("children") or [], name)
        if hit:
            return hit
    return None


def ensure_org(token: str) -> dict:
    tree = _req("GET", "/system/org-tree", token=token)
    nodes = tree.get("data") or []
    created = []
    ids = {}

    def ensure_college(name: str, code: str) -> dict:
        nonlocal nodes
        hit = _find_node(nodes, name)
        if hit:
            return hit
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "COLLEGE", "name": name, "code": code,
        })
        assert r.get("code") == 0, r
        created.append(name)
        tree2 = _req("GET", "/system/org-tree", token=token)
        nodes = tree2.get("data") or []
        return _find_node(nodes, name)

    def ensure_under(parent_id: int, name: str, code: str, typ: str, parent_node: dict) -> dict:
        for c in parent_node.get("children") or []:
            if c.get("name") == name:
                return c
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": typ, "name": name, "code": code, "parentId": parent_id,
        })
        assert r.get("code") == 0, r
        created.append(name)
        tree2 = _req("GET", "/system/org-tree", token=token)
        refreshed = _find_node(tree2.get("data") or [], parent_node["name"])
        for c in (refreshed or {}).get("children") or []:
            if c.get("name") == name:
                return c
        return {"id": r["data"]["id"], "name": name, "children": []}

    col_a = ensure_college(COLLEGE_A, "E2E-AA-COL-A")
    col_b = ensure_college(COLLEGE_B, "E2E-AA-COL-B")
    # refresh after colleges
    tree = _req("GET", "/system/org-tree", token=token)
    nodes = tree.get("data") or []
    col_a = _find_node(nodes, COLLEGE_A)
    col_b = _find_node(nodes, COLLEGE_B)

    maj_a = ensure_under(col_a["id"], MAJOR_A, "E2E-AA-MAJ-A", "MAJOR", col_a)
    tree = _req("GET", "/system/org-tree", token=token)
    col_a = _find_node(tree.get("data") or [], COLLEGE_A)
    maj_a = next(m for m in (col_a.get("children") or []) if m.get("name") == MAJOR_A)

    maj_b = ensure_under(col_b["id"], MAJOR_B, "E2E-AA-MAJ-B", "MAJOR", col_b)
    tree = _req("GET", "/system/org-tree", token=token)
    col_b = _find_node(tree.get("data") or [], COLLEGE_B)
    maj_b = next(m for m in (col_b.get("children") or []) if m.get("name") == MAJOR_B)

    cls_a1 = ensure_under(maj_a["id"], CLASS_A1, "E2E-AA-CLS-A1", "CLASS", maj_a)
    tree = _req("GET", "/system/org-tree", token=token)
    maj_a = _find_node(tree.get("data") or [], MAJOR_A)
    cls_a2 = ensure_under(maj_a["id"], CLASS_A2, "E2E-AA-CLS-A2", "CLASS", maj_a)
    tree = _req("GET", "/system/org-tree", token=token)
    maj_b = _find_node(tree.get("data") or [], MAJOR_B)
    cls_b1 = ensure_under(maj_b["id"], CLASS_B1, "E2E-AA-CLS-B1", "CLASS", maj_b)

    ids = {
        "collegeAId": col_a["id"], "collegeBId": col_b["id"],
        "majorAId": maj_a["id"], "majorBId": maj_b["id"],
        "classA1Id": cls_a1["id"], "classA2Id": cls_a2["id"], "classB1Id": cls_b1["id"],
        "created": created,
        "names": {
            "collegeA": COLLEGE_A, "collegeB": COLLEGE_B,
            "majorA": MAJOR_A, "majorB": MAJOR_B,
            "classA1": CLASS_A1, "classA2": CLASS_A2, "classB1": CLASS_B1,
            "grade": GRADE,
        },
    }
    return ids


def multipart(content: bytes, filename: str) -> tuple[bytes, str]:
    boundary = "----Bound" + uuid.uuid4().hex
    parts = [
        f"--{boundary}".encode(),
        f'Content-Disposition: form-data; name="file"; filename="{filename}"'.encode(),
        b"Content-Type: application/vnd.openxmlformats.spreadsheetml.sheet".replace(
            b"openxmlformats.spreadsheetml", b"openxmlformats-officedocument.spreadsheetml"
        ),
        b"",
        content,
        f"--{boundary}--".encode(),
        b"",
    ]
    # fix content-type properly
    parts[2] = b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return b"\r\n".join(parts), boundary


def build_xlsx(token: str) -> bytes:
    tpl = urllib.request.urlopen(urllib.request.Request(
        f"{BASE}/system/identity-import/template",
        headers={"Authorization": f"Bearer {token}"},
    )).read()
    wb = load_workbook(io.BytesIO(tpl))
    ws = wb["导入模板"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for login, name, roles, scope_t, scope_r, dept in TEACHERS:
        # SCHOOL_ADMIN may be blocked on teacher import — try anyway; bootstrap falls back to admin2
        ws.append([
            "TEACHER", login, name, "", "", "", "", "",
            dept, name, roles, scope_t, scope_r,
        ])
    for sno, name, college, major, cls, gender in STUDENTS:
        ws.append([
            "STUDENT", sno, name, college, major, cls, GRADE, gender,
            "", "", "", "", "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_accounts(token: str) -> dict:
    content = build_xlsx(token)
    body, boundary = multipart(content, "e2e_academic_affairs_identity.xlsx")
    validated = _req(
        "POST", "/system/identity-import/validate-file", token=token, raw=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    print("validate:", json.dumps({
        "code": validated.get("code"),
        "message": validated.get("message"),
        "summary": {k: (validated.get("data") or {}).get(k) for k in (
            "batchNo", "total", "valid", "invalid", "errorCount", "canConfirm", "status"
        )} if isinstance(validated.get("data"), dict) else validated.get("data"),
        "errors": ((validated.get("data") or {}).get("errors")
                   or (validated.get("data") or {}).get("errorRows") or [])[:20],
    }, ensure_ascii=False, indent=2))
    if validated.get("code") != 0:
        return {"confirmed": False, "detail": validated}
    data = validated.get("data") or {}
    batch_no = data.get("batchNo")
    if not batch_no:
        return {"confirmed": False, "detail": validated}
    confirmed = _req("POST", "/system/identity-import/confirm-batch", token=token, body={
        "batchNo": batch_no,
    })
    print("confirm:", json.dumps({
        "code": confirmed.get("code"),
        "message": confirmed.get("message"),
        "keys": list((confirmed.get("data") or {}).keys()) if isinstance(confirmed.get("data"), dict) else None,
    }, ensure_ascii=False, indent=2))
    return {"batchNo": batch_no, "confirmed": confirmed.get("code") == 0, "detail": confirmed}


def list_users(token: str) -> dict[str, dict]:
    out: dict[str, dict] = {}
    page = 1
    while True:
        r = _req("GET", f"/system/users?page={page}&page_size=50", token=token)
        items = (r.get("data") or {}).get("list") or []
        for u in items:
            ln = str(u.get("loginName") or u.get("userNo") or "")
            if ln:
                out[ln] = u
        total = int((r.get("data") or {}).get("total") or 0)
        if page * 50 >= total or not items:
            break
        page += 1
    return out


def activate_passwords(token: str) -> dict[str, str]:
    by = list_users(token)
    pwd_map = {"admin2": ADMIN[1]}
    results = []
    for i, ln in enumerate(ALL_LOGINS):
        if ln == "admin2":
            results.append({"loginName": ln, "ok": True, "skipped": True, "role": "SCHOOL_ADMIN"})
            continue
        time.sleep(2.5)
        if i % 4 == 0:
            token = login_admin()
        lg0 = _req("POST", "/auth/login", body={
            "loginName": ln, "password": STABLE_PWD, "tenantCode": TENANT,
        })
        if lg0.get("code") == 0:
            pwd_map[ln] = STABLE_PWD
            results.append({
                "loginName": ln, "ok": True, "skipped": True,
                "role": (lg0["data"].get("currentRole") or {}).get("roleCode"),
            })
            continue
        u = by.get(ln)
        if not u:
            # e2e_aa_sysadmin may be rejected by import — document fallback
            results.append({"loginName": ln, "ok": False, "message": "missing_in_list"})
            continue
        reset = _req("POST", f"/system/users/{u['id']}/reset-password", token=token, body={})
        temp = (reset.get("data") or {}).get("tempPassword")
        if not temp:
            results.append({"loginName": ln, "ok": False, "message": "no_temp", "reset": reset})
            continue
        time.sleep(1)
        lg = _req("POST", "/auth/login", body={
            "loginName": ln, "password": temp, "tenantCode": TENANT,
        })
        if lg.get("code") != 0:
            results.append({"loginName": ln, "ok": False, "message": lg.get("message")})
            time.sleep(3)
            continue
        ch = _req("POST", "/auth/change-password", token=lg["data"]["accessToken"], body={
            "oldPassword": temp, "newPassword": STABLE_PWD, "confirmPassword": STABLE_PWD,
        })
        if ch.get("code") != 0:
            ch = _req("POST", "/auth/password/change", token=lg["data"]["accessToken"], body={
                "oldPassword": temp, "newPassword": STABLE_PWD,
            })
        lg2 = _req("POST", "/auth/login", body={
            "loginName": ln, "password": STABLE_PWD, "tenantCode": TENANT,
        })
        ok = lg2.get("code") == 0
        if ok:
            pwd_map[ln] = STABLE_PWD
        results.append({
            "loginName": ln, "ok": ok,
            "role": ((lg2.get("data") or {}).get("currentRole") or {}).get("roleCode") if ok else None,
            "message": None if ok else (lg2.get("message") or ch.get("message")),
        })
    # system admin fallback note
    note = (
        "E2E系统管理员由 admin2/SCHOOL_ADMIN 承担；"
        "若 e2e_aa_sysadmin 导入被拒（teacherAssignable=false），属预期。"
    )
    CRED_PATH.write_text(json.dumps({
        "tenantCode": TENANT,
        "stablePassword": STABLE_PWD,
        "passwords": pwd_map,
        "loginResults": results,
        "note": note,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(results, ensure_ascii=False, indent=2))
    return pwd_map


def main() -> int:
    token = login_admin()
    org = ensure_org(token)
    print("org:", json.dumps(org, ensure_ascii=False, indent=2))
    imported = import_accounts(token)
    time.sleep(2)
    token = login_admin()
    pwds = activate_passwords(token)
    state = {
        "tenantCode": TENANT,
        "org": org,
        "import": {"batchNo": imported.get("batchNo"), "confirmed": imported.get("confirmed")},
        "accounts": ALL_LOGINS,
        "passwordFile": str(CRED_PATH),
        "loginOk": sorted(pwds.keys()),
    }
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"state -> {STATE_PATH}")
    print(f"creds -> {CRED_PATH}")
    # require core roles at minimum
    required = ["e2e_aa_admin", "e2e_aa_college_a", "e2e_aa_college_b",
                "e2e_aa_teacher_a", "E2EAA20260001", "E2EAA20260002", "E2EAA20260003"]
    ok_req = sum(1 for ln in required if ln in pwds)
    print(f"required_login_ok={ok_req}/{len(required)} total_ok={len(pwds)}")
    return 0 if ok_req >= 6 else 1


if __name__ == "__main__":
    sys.exit(main())
