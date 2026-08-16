"""Bootstrap dedicated graduation/internship E2E identities through canonical Data Exchange.

The retired mixed identity parser is intentionally not used. CI builds the dedicated
TEACHER and STUDENT workbooks from the same production template builders exposed to
schools, runs each file through FileObject -> explicit process/staging -> canonical
confirm, and keeps any credential receipt in memory only. Passwords are normalized by
the official reset/change endpoints in the following workflow step and are never written
to an artifact.
"""
from __future__ import annotations

import io
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.identity_import_file_service import (  # noqa: E402
    build_student_template,
    build_teacher_template,
)
from scripts.e2e_bootstrap_graduation_accounts import (  # noqa: E402
    CLASS,
    COLLEGE,
    MAJOR,
    _req,
    ensure_org,
    login,
    multipart,
)


def _workbook_with_rows(template: bytes, rows: list[list[str]]) -> bytes:
    """Populate the canonical Import sheet without changing its production headers."""
    wb = load_workbook(io.BytesIO(template))
    ws = wb["导入模板"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def _teacher_workbook() -> bytes:
    return _workbook_with_rows(
        build_teacher_template(),
        [
            ["e2e_academic_admin", "E2E教务管理员", "教务处", "教务管理员",
             "ACADEMIC_ADMIN,GRADUATION_ADMIN", "", ""],
            ["e2e_college_secretary", "E2E学院秘书", COLLEGE, "学院秘书",
             "GD_COLLEGE_ADMIN", "COLLEGE", COLLEGE],
            ["e2e_major_admin", "E2E专业负责人", COLLEGE, "专业负责人",
             "GD_MAJOR_ADMIN", "MAJOR", MAJOR],
            ["e2e_advisor_a", "E2E指导教师A", COLLEGE, "指导教师",
             "GD_MENTOR,INTERN_MENTOR", "", ""],
            ["e2e_advisor_b", "E2E指导教师B", COLLEGE, "指导教师",
             "GD_MENTOR", "", ""],
            ["e2e_reviewer", "E2E评阅教师", COLLEGE, "评阅教师",
             "GD_REVIEWER", "", ""],
            ["e2e_defense_a", "E2E答辩专家A", COLLEGE, "答辩专家",
             "GD_DEFENSE_EXPERT", "", ""],
            ["e2e_defense_b", "E2E答辩专家B", COLLEGE, "答辩专家",
             "GD_DEFENSE_EXPERT", "", ""],
        ],
    )


def _student_workbook() -> bytes:
    return _workbook_with_rows(
        build_student_template(),
        [
            ["E2E20260001", "E2E学生A", COLLEGE, MAJOR, CLASS, "2024", "男", ""],
            ["E2E20260002", "E2E学生B", COLLEGE, MAJOR, CLASS, "2024", "女", ""],
            ["E2E20260003", "E2E学生C", COLLEGE, MAJOR, CLASS, "2024", "男", ""],
        ],
    )


def _canonical_import(token: str, *, kind: str, content: bytes) -> dict:
    """Run upload -> explicit worker process -> canonical job confirmation."""
    if kind not in {"teachers", "students"}:
        raise ValueError(f"unsupported canonical identity kind: {kind}")
    filename = f"e2e_interaction_{kind}.xlsx"
    body, boundary = multipart(content, filename)
    upload_key = f"e2e-graduation-{kind}-canonical-v3"
    created = _req(
        "POST",
        f"/data-exchange/imports/identity/{kind}/validate-file",
        token=token,
        raw=body,
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Idempotency-Key": upload_key,
        },
    )
    if created.get("code") != 0:
        raise SystemExit(
            f"canonical {kind} identity upload failed: "
            + json.dumps(created, ensure_ascii=False)
        )
    item = dict(created.get("data") or {})
    job_id = str(item.get("id") or item.get("jobId") or "").strip()
    if not job_id:
        raise SystemExit(f"canonical {kind} identity upload returned no job id")

    status = str(item.get("status") or "").upper()
    if status == "SUCCEEDED":
        print(f"[e2e-bootstrap] canonical {kind} import replayed succeeded job {job_id}")
        return item
    if status != "VALIDATED":
        processed = _req(
            "POST",
            f"/data-exchange/imports/{job_id}/process",
            token=token,
        )
        if processed.get("code") != 0:
            raise SystemExit(
                f"canonical {kind} identity process failed: "
                + json.dumps(processed, ensure_ascii=False)
            )
        item = dict(processed.get("data") or {})
        status = str(item.get("status") or "").upper()
    if status != "VALIDATED":
        raise SystemExit(
            f"canonical {kind} identity job did not validate: "
            + json.dumps(
                {
                    "jobId": job_id,
                    "status": item.get("status"),
                    "message": item.get("errorMessage") or item.get("message"),
                    "invalidRows": item.get("invalidRows"),
                },
                ensure_ascii=False,
            )
        )

    expected_version = int(item.get("version") or 0)
    confirmed = _req(
        "POST",
        f"/data-exchange/imports/{job_id}/confirm",
        token=token,
        body={"expectedVersion": expected_version},
        headers={"Idempotency-Key": f"e2e-graduation-{kind}-confirm-v3"},
    )
    if confirmed.get("code") != 0:
        raise SystemExit(
            f"canonical {kind} identity confirm failed: "
            + json.dumps(confirmed, ensure_ascii=False)
        )
    receipt = dict(confirmed.get("data") or {})
    print(
        f"[e2e-bootstrap] canonical {kind} identity import confirmed",
        json.dumps(
            {
                "jobId": job_id,
                "receiptKeys": sorted(receipt.keys()),
            },
            ensure_ascii=False,
        ),
    )
    return receipt


def import_accounts_without_receipt_file(token: str) -> dict:
    teacher_receipt = _canonical_import(token, kind="teachers", content=_teacher_workbook())
    student_receipt = _canonical_import(token, kind="students", content=_student_workbook())
    return {
        "teachers": str(teacher_receipt.get("id") or teacher_receipt.get("jobId") or "confirmed"),
        "students": str(student_receipt.get("id") or student_receipt.get("jobId") or "confirmed"),
    }


def main() -> int:
    token = login()
    org = ensure_org(token)
    print("org:", json.dumps(org, ensure_ascii=False))
    imported = import_accounts_without_receipt_file(token)
    print("[e2e-bootstrap] ready:", json.dumps(imported, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
