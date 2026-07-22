"""E2E: bootstrap graduation multi-role accounts via official identity-import API.

Creates org (if needed) + imports teachers/students with E2E markers into sandbox-school.
Does NOT print passwords into git-tracked files; writes credentials to a local ignored path.
"""
from __future__ import annotations

import io
import json
import sys
import uuid
from pathlib import Path

import urllib.error
import urllib.request
from openpyxl import load_workbook

BASE = "http://127.0.0.1:8000/api/v1"
TENANT = "sandbox-school"
ADMIN = ("admin2", "123456")
OUT_DIR = Path(__file__).resolve().parents[1] / "tmp"
OUT_DIR.mkdir(exist_ok=True)
CRED_PATH = OUT_DIR / "e2e_graduation_credentials.local.json"

COLLEGE = "E2E智能制造学院"
MAJOR = "E2E工业机器人技术"
CLASS = "E2E机器人2401班"


def _req(method: str, path: str, token: str | None = None, body: dict | None = None,
         raw: bytes | None = None, headers: dict | None = None):
    data = raw
    hdrs = dict(headers or {})
    if token:
        hdrs["Authorization"] = f"Bearer {token}"
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        hdrs.setdefault("Content-Type", "application/json")
    req = urllib.request.Request(f"{BASE}{path}", data=data, headers=hdrs, method=method)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw_body = resp.read()
            if not raw_body:
                return {"code": 0, "data": None}
            return json.loads(raw_body.decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        try:
            return json.loads(detail)
        except json.JSONDecodeError:
            return {"code": exc.code, "message": detail}


def login() -> str:
    r = _req("POST", "/auth/login", body={
        "loginName": ADMIN[0], "password": ADMIN[1], "tenantCode": TENANT,
    })
    if r.get("code") != 0:
        raise SystemExit(f"admin2 login failed: {r}")
    return r["data"]["accessToken"]


def ensure_org(token: str) -> dict:
    tree = _req("GET", "/system/org-tree", token=token)
    nodes = tree.get("data") or []

    def find_college(name: str):
        for c in nodes:
            if c.get("name") == name:
                return c
        return None

    college = find_college(COLLEGE)
    if college is None:
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "COLLEGE", "name": COLLEGE, "code": "E2E-COL-GD",
        })
        assert r.get("code") == 0, r
        college_id = r["data"]["id"]
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "MAJOR", "name": MAJOR, "code": "E2E-MAJ-GD", "parentId": college_id,
        })
        assert r.get("code") == 0, r
        major_id = r["data"]["id"]
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "CLASS", "name": CLASS, "code": "E2E-CLS-GD", "parentId": major_id,
        })
        assert r.get("code") == 0, r
        return {"collegeId": college_id, "majorId": major_id, "classId": r["data"]["id"], "created": True}

    major = next((m for m in college.get("children") or [] if m.get("name") == MAJOR), None)
    if major is None:
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "MAJOR", "name": MAJOR, "code": "E2E-MAJ-GD", "parentId": college["id"],
        })
        assert r.get("code") == 0, r
        major_id = r["data"]["id"]
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "CLASS", "name": CLASS, "code": "E2E-CLS-GD", "parentId": major_id,
        })
        assert r.get("code") == 0, r
        return {"collegeId": college["id"], "majorId": major_id, "classId": r["data"]["id"], "created": True}

    clazz = next((c for c in major.get("children") or [] if c.get("name") == CLASS), None)
    if clazz is None:
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "CLASS", "name": CLASS, "code": "E2E-CLS-GD", "parentId": major["id"],
        })
        assert r.get("code") == 0, r
        return {"collegeId": college["id"], "majorId": major["id"], "classId": r["data"]["id"], "created": True}
    return {"collegeId": college["id"], "majorId": major["id"], "classId": clazz["id"], "created": False}


def build_xlsx(token: str) -> bytes:
    tpl = urllib.request.urlopen(urllib.request.Request(
        f"{BASE}/system/identity-import/template",
        headers={"Authorization": f"Bearer {token}"},
    )).read()
    wb = load_workbook(io.BytesIO(tpl))
    ws = wb["导入模板"]
    # Drop any leftover sample data rows
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    # SYS_ADMIN/SCHOOL_ADMIN 不可经师生导入分配（teacherAssignable=false）；
    # e2e_sysadmin 由平台侧 create_school_admin 或现有 admin2 承担。
    rows = [
        ["TEACHER", "e2e_academic_admin", "E2E教务管理员", "", "", "", "", "", "教务处", "教务管理员", "ACADEMIC_ADMIN,GRADUATION_ADMIN", "", ""],
        ["TEACHER", "e2e_college_secretary", "E2E学院秘书", "", "", "", "", "", COLLEGE, "学院秘书", "GD_COLLEGE_ADMIN", "COLLEGE", COLLEGE],
        ["TEACHER", "e2e_advisor_a", "E2E指导教师A", "", "", "", "", "", COLLEGE, "指导教师", "GD_MENTOR", "", ""],
        ["TEACHER", "e2e_advisor_b", "E2E指导教师B", "", "", "", "", "", COLLEGE, "指导教师", "GD_MENTOR", "", ""],
        ["TEACHER", "e2e_reviewer", "E2E评阅教师", "", "", "", "", "", COLLEGE, "评阅教师", "GD_REVIEWER", "", ""],
        ["TEACHER", "e2e_defense_a", "E2E答辩专家A", "", "", "", "", "", COLLEGE, "答辩专家", "GD_DEFENSE_EXPERT", "", ""],
        ["TEACHER", "e2e_defense_b", "E2E答辩专家B", "", "", "", "", "", COLLEGE, "答辩专家", "GD_DEFENSE_EXPERT", "", ""],
        ["STUDENT", "E2E20260001", "E2E学生A", COLLEGE, MAJOR, CLASS, "2024", "男", "", "", "", "", ""],
        ["STUDENT", "E2E20260002", "E2E学生B", COLLEGE, MAJOR, CLASS, "2024", "女", "", "", "", "", ""],
        ["STUDENT", "E2E20260003", "E2E学生C", COLLEGE, MAJOR, CLASS, "2024", "男", "", "", "", "", ""],
    ]
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def multipart(content: bytes, filename: str) -> tuple[bytes, str]:
    boundary = "----Bound" + uuid.uuid4().hex
    parts = []
    parts.append(f"--{boundary}".encode())
    parts.append(
        f'Content-Disposition: form-data; name="file"; filename="{filename}"'.encode()
    )
    parts.append(b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    parts.append(b"")
    parts.append(content)
    parts.append(f"--{boundary}--".encode())
    parts.append(b"")
    body = b"\r\n".join(parts)
    return body, boundary


def import_accounts(token: str) -> dict:
    content = build_xlsx(token)
    body, boundary = multipart(content, "e2e_graduation_identity.xlsx")
    validated = _req(
        "POST", "/system/identity-import/validate-file", token=token, raw=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    print("validate:", json.dumps({
        "code": validated.get("code"),
        "message": validated.get("message"),
        "keys": list((validated.get("data") or {}).keys()) if isinstance(validated.get("data"), dict) else None,
        "summary": {k: (validated.get("data") or {}).get(k) for k in (
            "batchNo", "total", "valid", "invalid", "errorCount", "canConfirm", "status"
        )} if isinstance(validated.get("data"), dict) else validated.get("data"),
    }, ensure_ascii=False, indent=2))
    if validated.get("code") != 0:
        raise SystemExit(validated)

    data = validated["data"] or {}
    batch_no = data.get("batchNo")
    if not batch_no:
        raise SystemExit(f"no batchNo: {validated}")

    # If already imported (idempotent re-run), skip confirm when invalid only due to duplicates
    errors = data.get("errors") or data.get("errorRows") or []
    if data.get("canConfirm") is False and errors:
        print("precheck errors sample:", json.dumps(errors[:8], ensure_ascii=False, indent=2))

    confirmed = _req("POST", "/system/identity-import/confirm-batch", token=token, body={
        "batchNo": batch_no,
    })
    print("confirm:", json.dumps({
        "code": confirmed.get("code"),
        "message": confirmed.get("message"),
        "keys": list((confirmed.get("data") or {}).keys()) if isinstance(confirmed.get("data"), dict) else None,
    }, ensure_ascii=False, indent=2))
    if confirmed.get("code") != 0:
        # Re-run may fail if accounts already exist — treat as soft OK when login works
        print("confirm failed; will verify logins anyway:", confirmed.get("message"))
        return {"batchNo": batch_no, "confirmed": False, "detail": confirmed}

    # Persist credentials locally (gitignored via tmp/)
    receipt = confirmed.get("data") or {}
    CRED_PATH.write_text(json.dumps(receipt, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"credentials written to {CRED_PATH}")
    return {"batchNo": batch_no, "confirmed": True, "detail": receipt}


def _extract_credentials(receipt: dict) -> dict[str, str]:
    """Normalize credential receipt shapes from identity-import / platform APIs."""
    import base64
    from openpyxl import load_workbook

    pwd_map: dict[str, str] = {}
    if not isinstance(receipt, dict):
        return pwd_map
    blobs = [receipt]
    cr = receipt.get("credentialReceipt")
    if isinstance(cr, dict):
        blobs.append(cr)
        b64 = cr.get("contentBase64")
        if b64:
            raw = base64.b64decode(b64)
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h or "").strip() for h in rows[0]]
                for row in rows[1:]:
                    cells = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
                    ln = cells.get("工号/学号") or cells.get("loginName") or cells.get("账号")
                    pwd = cells.get("初始密码") or cells.get("password") or cells.get("initialPassword")
                    if ln and pwd:
                        pwd_map[str(ln).strip()] = str(pwd).strip()
    for blob in blobs:
        for key in ("accounts", "credentials", "items", "studentCredentials", "teacherCredentials", "rows"):
            items = blob.get(key)
            if isinstance(items, dict):
                for ln, pwd in items.items():
                    if isinstance(pwd, str) and pwd:
                        pwd_map[str(ln)] = pwd
                    elif isinstance(pwd, dict):
                        p = pwd.get("password") or pwd.get("initialPassword")
                        if p:
                            pwd_map[str(ln)] = p
                continue
            if not isinstance(items, list):
                continue
            for it in items:
                if not isinstance(it, dict):
                    continue
                ln = it.get("loginName") or it.get("accountNo") or it.get("工号/学号") or it.get("userNo")
                pwd = it.get("password") or it.get("initialPassword") or it.get("初始密码")
                if ln and pwd:
                    pwd_map[str(ln)] = str(pwd)
    if receipt.get("loginName") and receipt.get("initialPassword"):
        pwd_map[str(receipt["loginName"])] = str(receipt["initialPassword"])
    return pwd_map


def _change_password_if_needed(login_name: str, old_pwd: str, token: str, login_data: dict) -> str:
    """If first-login forces password change, set a stable E2E test password via official API."""
    must = ((login_data.get("user") or {}).get("mustChangePassword")
            or login_data.get("mustChangePassword"))
    if not must:
        return old_pwd
    new_pwd = "E2eTest@2026"
    r = _req("POST", "/auth/change-password", token=token, body={
        "oldPassword": old_pwd, "newPassword": new_pwd, "confirmPassword": new_pwd,
    })
    if r.get("code") != 0:
        # try alternate payload keys used by some clients
        r2 = _req("POST", "/auth/password/change", token=token, body={
            "oldPassword": old_pwd, "newPassword": new_pwd,
        })
        if r2.get("code") != 0:
            raise RuntimeError(f"change-password failed for {login_name}: {r} / {r2}")
    return new_pwd


def verify_logins(receipt: dict) -> list[dict]:
    results = []
    pwd_map = _extract_credentials(receipt)
    known = [
        "admin2", "e2e_academic_admin", "e2e_college_secretary",
        "e2e_advisor_a", "e2e_advisor_b", "e2e_reviewer",
        "e2e_defense_a", "e2e_defense_b",
        "E2E20260001", "E2E20260002", "E2E20260003",
    ]
    if "admin2" not in pwd_map:
        pwd_map["admin2"] = ADMIN[1]
    final_pwds = dict(pwd_map)
    for ln in known:
        pwd = pwd_map.get(ln)
        if not pwd:
            results.append({"loginName": ln, "ok": False, "reason": "no_password_in_receipt"})
            continue
        r = _req("POST", "/auth/login", body={
            "loginName": ln, "password": pwd, "tenantCode": TENANT,
        })
        ok = r.get("code") == 0
        if not ok:
            results.append({"loginName": ln, "ok": False, "role": None, "message": r.get("message")})
            continue
        data = r["data"]
        role = (data.get("currentRole") or {}).get("roleCode")
        token = data["accessToken"]
        try:
            final_pwds[ln] = _change_password_if_needed(ln, pwd, token, data)
            # re-login with new password to confirm
            r2 = _req("POST", "/auth/login", body={
                "loginName": ln, "password": final_pwds[ln], "tenantCode": TENANT,
            })
            ok2 = r2.get("code") == 0
            results.append({
                "loginName": ln, "ok": ok2, "role": role,
                "message": None if ok2 else r2.get("message"),
                "passwordChanged": final_pwds[ln] != pwd,
            })
        except Exception as exc:  # noqa: BLE001
            results.append({"loginName": ln, "ok": False, "role": role, "message": str(exc)})
    # rewrite local credential file with usable passwords (still gitignored via tmp/)
    CRED_PATH.write_text(json.dumps({
        "tenantCode": TENANT,
        "passwords": final_pwds,
        "loginResults": results,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    return results


def main() -> int:
    token = login()
    org = ensure_org(token)
    print("org:", org)
    imported = import_accounts(token)
    receipt = imported.get("detail") if imported.get("confirmed") else {}
    if CRED_PATH.exists() and not receipt:
        receipt = json.loads(CRED_PATH.read_text(encoding="utf-8"))
    results = verify_logins(receipt if isinstance(receipt, dict) else {})
    print("login_results:", json.dumps(results, ensure_ascii=False, indent=2))
    ok_count = sum(1 for x in results if x["ok"])
    print(f"login_ok={ok_count}/{len(results)}")
    return 0 if ok_count >= 6 else 1


if __name__ == "__main__":
    sys.exit(main())
