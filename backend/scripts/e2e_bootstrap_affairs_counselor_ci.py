"""Bootstrap the single student-affairs counselor needed by Playwright CI.

This fixture intentionally uses only official identity and password APIs.  It does not
persist credentials, does not relax login throttling, and avoids the broad student-affairs
acceptance bootstrap which repeatedly authenticates the school admin.
"""
from __future__ import annotations

import io
import json
import sys
import time
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.e2e_bootstrap_graduation_accounts import (  # noqa: E402
    BASE,
    CLASS,
    COLLEGE,
    TENANT,
    _req,
    ensure_org,
    login,
    multipart,
)

LOGIN_NAME = "e2e_counselor_a"
DISPLAY_NAME = "E2E辅导员A"
STABLE_PASSWORD = "E2eTest@2026"


def build_counselor_xlsx(token: str) -> bytes:
    request = urllib.request.Request(
        f"{BASE}/system/identity-import/template",
        headers={"Authorization": f"Bearer {token}"},
    )
    content = urllib.request.urlopen(request, timeout=60).read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb["导入模板"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    ws.append([
        "TEACHER",
        LOGIN_NAME,
        DISPLAY_NAME,
        "",
        "",
        "",
        "",
        "",
        COLLEGE,
        "辅导员",
        "COUNSELOR",
        "CLASS",
        CLASS,
    ])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def import_counselor(token: str) -> None:
    body, boundary = multipart(build_counselor_xlsx(token), "e2e_affairs_counselor.xlsx")
    validated = _req(
        "POST",
        "/system/identity-import/validate-file",
        token=token,
        raw=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    if validated.get("code") != 0:
        raise SystemExit("counselor identity validation failed: " + json.dumps(validated, ensure_ascii=False))
    data = validated.get("data") or {}
    batch_no = data.get("batchNo")
    if not batch_no:
        raise SystemExit("counselor identity validation returned no batchNo")
    confirmed = _req(
        "POST",
        "/system/identity-import/confirm-batch",
        token=token,
        body={"batchNo": batch_no},
    )
    if confirmed.get("code") != 0:
        raise SystemExit("counselor identity import failed: " + json.dumps(confirmed, ensure_ascii=False))
    print("[e2e-affairs-counselor] identity import confirmed", batch_no)


def find_user(token: str) -> dict:
    page = 1
    while True:
        result = _req("GET", f"/system/users?page={page}&page_size=50", token=token)
        if result.get("code") != 0:
            raise SystemExit("list users failed: " + json.dumps(result, ensure_ascii=False))
        data = result.get("data") or {}
        rows = data.get("list") or []
        for user in rows:
            login_name = str(user.get("loginName") or user.get("userNo") or "")
            if login_name == LOGIN_NAME:
                return user
        total = int(data.get("total") or 0)
        if not rows or page * 50 >= total:
            break
        page += 1
    raise SystemExit(f"{LOGIN_NAME} missing after identity import")


def normalize_password(admin_token: str) -> None:
    user = find_user(admin_token)
    reset = _req("POST", f"/system/users/{user['id']}/reset-password", token=admin_token, body={})
    if reset.get("code") != 0:
        raise SystemExit("counselor reset-password failed: " + json.dumps(reset, ensure_ascii=False))
    temp_password = (reset.get("data") or {}).get("tempPassword")
    if not temp_password:
        raise SystemExit("counselor reset-password returned no temporary password")

    time.sleep(1)
    logged = _req("POST", "/auth/login", body={
        "loginName": LOGIN_NAME,
        "password": temp_password,
        "tenantCode": TENANT,
    })
    if logged.get("code") != 0:
        raise SystemExit("counselor temporary login failed: " + json.dumps(logged, ensure_ascii=False))

    access_token = (logged.get("data") or {}).get("accessToken")
    changed = _req("POST", "/auth/change-password", token=access_token, body={
        "oldPassword": temp_password,
        "newPassword": STABLE_PASSWORD,
        "confirmPassword": STABLE_PASSWORD,
    })
    if changed.get("code") != 0:
        changed = _req("POST", "/auth/password/change", token=access_token, body={
            "oldPassword": temp_password,
            "newPassword": STABLE_PASSWORD,
        })
    if changed.get("code") != 0:
        raise SystemExit("counselor password normalization failed: " + json.dumps(changed, ensure_ascii=False))
    print("[e2e-affairs-counselor] password normalized without credential output")


def main() -> int:
    admin_token = login()
    ensure_org(admin_token)
    import_counselor(admin_token)
    normalize_password(admin_token)
    print(f"[e2e-affairs-counselor] ready: {TENANT}:{LOGIN_NAME}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
