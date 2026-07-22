"""E2E: bootstrap student-affairs multi-role accounts + dorm + org in sandbox-school.

Reuses official identity-import + dorm APIs. Credentials written only under backend/tmp/.
Does not overwrite non-E2E org data.
"""
from __future__ import annotations

import io
import json
import sys
import time
import uuid
from pathlib import Path

import urllib.error
import urllib.request
from openpyxl import load_workbook

BASE = "http://127.0.0.1:8000/api/v1"
TENANT = "sandbox-school"
ADMIN = ("admin2", "123456")
STABLE_PWD = "E2eTest@2026"
OUT_DIR = Path(__file__).resolve().parents[1] / "tmp"
OUT_DIR.mkdir(exist_ok=True)
CRED_PATH = OUT_DIR / "e2e_student_affairs_credentials.local.json"
STATE_PATH = OUT_DIR / "e2e_student_affairs_state.local.json"

COLLEGE = "E2E智能制造学院"
MAJOR = "E2E工业机器人技术"
CLASS_A = "E2E机器人2401班"
CLASS_B = "E2E机器人2402班"
DORM_NAME = "E2E学生公寓1号楼"
DORM_CODE = "E2E-DORM-1"
STUDENT_ORG = "E2E青年志愿服务队"

# loginName -> (displayName, roles CSV, scopeType, scopeRef, dept)
TEACHERS = [
    ("e2e_sa_admin", "E2E学工处管理员", "STUDENT_AFFAIRS_ADMIN", "", "", "学工处"),
    ("e2e_college_admin", "E2E学院管理员", "COLLEGE_ADMIN", "COLLEGE", COLLEGE, COLLEGE),
    ("e2e_counselor_a", "E2E辅导员A", "COUNSELOR", "CLASS", CLASS_A, COLLEGE),
    ("e2e_counselor_b", "E2E辅导员B", "COUNSELOR", "CLASS", CLASS_B, COLLEGE),
    ("e2e_orientation_teacher", "E2E迎新老师", "STUDENT_AFFAIRS", "COLLEGE", COLLEGE, "学工处"),
    ("e2e_dorm_manager", "E2E宿管老师", "DORM_MANAGER", "DORM_BUILDING", DORM_NAME, "后勤处"),
    ("e2e_funding_teacher", "E2E资助老师", "FUNDING_TEACHER", "", "", "学工处"),
    ("e2e_mental_teacher", "E2E心理老师", "PSYCHOLOGY_TEACHER", "", "", "心理健康中心"),
    ("e2e_activity_manager", "E2E活动管理员", "YOUTH_LEAGUE", "", "", "团委"),
]

STUDENTS = [
    ("E2E20260001", "E2E学生A", CLASS_A, "男"),
    ("E2E20260002", "E2E学生B", CLASS_A, "女"),
    ("E2E20260003", "E2E学生C", CLASS_A, "男"),
    ("E2E20260004", "E2E新生D", CLASS_A, "女"),
]

ALL_LOGINS = ["admin2"] + [t[0] for t in TEACHERS] + [s[0] for s in STUDENTS]


def _req(method: str, path: str, token: str | None = None, body: dict | None = None,
         raw: bytes | None = None, headers: dict | None = None, params: dict | None = None):
    data = raw
    hdrs = dict(headers or {})
    if token:
        hdrs["Authorization"] = f"Bearer {token}"
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        hdrs.setdefault("Content-Type", "application/json")
    url = f"{BASE}{path}"
    if params:
        from urllib.parse import urlencode
        qs = urlencode({k: v for k, v in params.items() if v is not None and v != ""})
        if qs:
            url = f"{url}?{qs}"
    req = urllib.request.Request(url, data=data, headers=hdrs, method=method)
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            raw_body = resp.read()
            if not raw_body:
                return {"code": 0, "data": None}
            return json.loads(raw_body.decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        try:
            return json.loads(detail)
        except json.JSONDecodeError:
            return {"code": exc.code, "message": detail, "bizCode": str(exc.code)}


def login_admin() -> str:
    r = _req("POST", "/auth/login", body={
        "loginName": ADMIN[0], "password": ADMIN[1], "tenantCode": TENANT,
    })
    if r.get("code") != 0:
        raise SystemExit(f"admin2 login failed: {r}")
    return r["data"]["accessToken"]


def ensure_org(token: str) -> dict:
    tree = _req("GET", "/system/org-tree", token=token)
    nodes = tree.get("data") or []

    def find(name: str, items=None):
        for n in (items if items is not None else nodes):
            if n.get("name") == name:
                return n
            hit = find(name, n.get("children") or [])
            if hit:
                return hit
        return None

    college = find(COLLEGE)
    created = []
    if college is None:
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "COLLEGE", "name": COLLEGE, "code": "E2E-COL-SA",
        })
        assert r.get("code") == 0, r
        college_id = r["data"]["id"]
        created.append("college")
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "MAJOR", "name": MAJOR, "code": "E2E-MAJ-SA", "parentId": college_id,
        })
        assert r.get("code") == 0, r
        major_id = r["data"]["id"]
        created.append("major")
    else:
        college_id = college["id"]
        major = next((m for m in (college.get("children") or []) if m.get("name") == MAJOR), None)
        if major is None:
            r = _req("POST", "/system/org-nodes", token=token, body={
                "type": "MAJOR", "name": MAJOR, "code": "E2E-MAJ-SA", "parentId": college_id,
            })
            assert r.get("code") == 0, r
            major_id = r["data"]["id"]
            created.append("major")
            major = {"id": major_id, "children": []}
        else:
            major_id = major["id"]

    # refresh tree for class creation under major
    tree = _req("GET", "/system/org-tree", token=token)
    nodes = tree.get("data") or []
    college = find(COLLEGE)
    major = next((m for m in (college.get("children") or []) if m.get("name") == MAJOR), None)
    major_id = major["id"]
    class_ids = {}
    for cls_name, code in ((CLASS_A, "E2E-CLS-SA-A"), (CLASS_B, "E2E-CLS-SA-B")):
        existing = next((c for c in (major.get("children") or []) if c.get("name") == cls_name), None)
        if existing:
            class_ids[cls_name] = existing["id"]
            continue
        r = _req("POST", "/system/org-nodes", token=token, body={
            "type": "CLASS", "name": cls_name, "code": code, "parentId": major_id,
        })
        assert r.get("code") == 0, r
        class_ids[cls_name] = r["data"]["id"]
        created.append(cls_name)

    return {
        "collegeId": college_id, "majorId": major_id, "classIds": class_ids, "created": created,
    }


def ensure_dorm(token: str) -> dict:
    listed = _req("GET", "/student-affairs/dorm/buildings?pageSize=100", token=token)
    items = ((listed.get("data") or {}).get("list")
             or (listed.get("data") or {}).get("items") or [])
    for b in items:
        if b.get("buildingName") == DORM_NAME or b.get("buildingCode") == DORM_CODE:
            return {"buildingId": b.get("buildingId") or b.get("id"), "created": False, "raw": b}
    r = _req("POST", "/student-affairs/dorm/buildings", token=token, body={
        "buildingName": DORM_NAME,
        "buildingCode": DORM_CODE,
        "genderLimit": "MIXED",
        "floors": 1,
        "roomsPerFloor": 2,
        "bedsPerRoom": 4,
    })
    if r.get("code") != 0:
        raise SystemExit(f"create dorm failed: {r}")
    bid = (r.get("data") or {}).get("buildingId") or (r.get("data") or {}).get("id")
    # rooms will be 101/102 under E2E building — treat as E2E-101 / E2E-102
    rooms = _req("GET", f"/student-affairs/dorm/buildings/{bid}/rooms?pageSize=50", token=token)
    return {"buildingId": bid, "created": True, "rooms": rooms.get("data"), "raw": r.get("data")}


def ensure_student_org(token: str) -> dict:
    listed = _req("GET", "/student-affairs/organizations?pageSize=100", token=token)
    items = ((listed.get("data") or {}).get("list")
             or (listed.get("data") or {}).get("items") or [])
    for o in items:
        if o.get("orgName") == STUDENT_ORG or o.get("name") == STUDENT_ORG:
            return {"orgId": o.get("id") or o.get("orgId"), "created": False}
    r = _req("POST", "/student-affairs/organizations", token=token, body={
        "orgName": STUDENT_ORG, "orgType": "VOLUNTEER",
    })
    if r.get("code") != 0:
        # soft-fail: record gap
        return {"orgId": None, "created": False, "error": r}
    return {"orgId": (r.get("data") or {}).get("id") or (r.get("data") or {}).get("orgId"),
            "created": True, "raw": r.get("data")}


def multipart(content: bytes, filename: str) -> tuple[bytes, str]:
    boundary = "----Bound" + uuid.uuid4().hex
    parts = [
        f"--{boundary}".encode(),
        f'Content-Disposition: form-data; name="file"; filename="{filename}"'.encode(),
        b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        b"",
        content,
        f"--{boundary}--".encode(),
        b"",
    ]
    return b"\r\n".join(parts), boundary


def build_xlsx(token: str) -> bytes:
    tpl = urllib.request.urlopen(urllib.request.Request(
        f"{BASE}/system/identity-import/template",
        headers={"Authorization": f"Bearer {token}"},
    )).read()
    wb = load_workbook(io.BytesIO(tpl))
    ws = wb["导入模板"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for login, name, roles, scope_t, scope_r, dept in TEACHERS:
        ws.append([
            "TEACHER", login, name, "", "", "", "", "",
            dept, name, roles, scope_t, scope_r,
        ])
    for sno, name, cls, gender in STUDENTS:
        ws.append([
            "STUDENT", sno, name, COLLEGE, MAJOR, cls, "2024", gender,
            "", "", "", "", "",
        ])
    # relations sheet if present
    if "业务关系" in wb.sheetnames:
        rs = wb["业务关系"]
        if rs.max_row > 1:
            rs.delete_rows(2, rs.max_row - 1)
        rs.append(["COUNSELOR_CLASS", "e2e_counselor_a", CLASS_A, "", "E2E"])
        rs.append(["COUNSELOR_CLASS", "e2e_counselor_b", CLASS_B, "", "E2E"])
        rs.append(["DORM_MANAGER_BUILDING", "e2e_dorm_manager", DORM_CODE, "", "E2E"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_accounts(token: str) -> dict:
    content = build_xlsx(token)
    body, boundary = multipart(content, "e2e_student_affairs_identity.xlsx")
    validated = _req(
        "POST", "/system/identity-import/validate-file", token=token, raw=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    print("validate:", json.dumps({
        "code": validated.get("code"),
        "message": validated.get("message"),
        "summary": {k: (validated.get("data") or {}).get(k) for k in (
            "batchNo", "total", "valid", "invalid", "errorCount", "canConfirm", "status"
        )} if isinstance(validated.get("data"), dict) else validated.get("data"),
        "errors": ((validated.get("data") or {}).get("errors")
                   or (validated.get("data") or {}).get("errorRows") or [])[:12],
    }, ensure_ascii=False, indent=2))
    if validated.get("code") != 0:
        return {"confirmed": False, "detail": validated}
    data = validated.get("data") or {}
    batch_no = data.get("batchNo")
    if not batch_no:
        return {"confirmed": False, "detail": validated}
    if data.get("canConfirm") is False:
        # still try confirm for partial; duplicates may soft-fail later
        print("canConfirm=false; attempting confirm anyway for new rows")
    confirmed = _req("POST", "/system/identity-import/confirm-batch", token=token, body={
        "batchNo": batch_no,
    })
    print("confirm:", json.dumps({
        "code": confirmed.get("code"),
        "message": confirmed.get("message"),
        "keys": list((confirmed.get("data") or {}).keys()) if isinstance(confirmed.get("data"), dict) else None,
    }, ensure_ascii=False, indent=2))
    return {"batchNo": batch_no, "confirmed": confirmed.get("code") == 0, "detail": confirmed}


def list_users(token: str) -> dict[str, dict]:
    out: dict[str, dict] = {}
    page = 1
    while True:
        r = _req("GET", f"/system/users?page={page}&page_size=50", token=token)
        items = (r.get("data") or {}).get("list") or []
        for u in items:
            ln = str(u.get("loginName") or u.get("userNo") or "")
            if ln:
                out[ln] = u
        total = int((r.get("data") or {}).get("total") or 0)
        if page * 50 >= total or not items:
            break
        page += 1
    return out


def activate_passwords(token: str) -> dict[str, str]:
    by = list_users(token)
    pwd_map = {"admin2": ADMIN[1]}
    results = []
    for i, ln in enumerate(ALL_LOGINS):
        if ln == "admin2":
            results.append({"loginName": ln, "ok": True, "skipped": True, "role": "SCHOOL_ADMIN"})
            continue
        time.sleep(3)
        if i % 4 == 0:
            token = login_admin()
        # already stable?
        lg0 = _req("POST", "/auth/login", body={
            "loginName": ln, "password": STABLE_PWD, "tenantCode": TENANT,
        })
        if lg0.get("code") == 0:
            pwd_map[ln] = STABLE_PWD
            results.append({
                "loginName": ln, "ok": True, "skipped": True,
                "role": (lg0["data"].get("currentRole") or {}).get("roleCode"),
            })
            continue
        u = by.get(ln)
        if not u:
            results.append({"loginName": ln, "ok": False, "message": "missing_in_list"})
            continue
        reset = _req("POST", f"/system/users/{u['id']}/reset-password", token=token, body={})
        temp = (reset.get("data") or {}).get("tempPassword")
        if not temp:
            results.append({"loginName": ln, "ok": False, "message": "no_temp", "reset": reset})
            time.sleep(2)
            continue
        time.sleep(1)
        lg = _req("POST", "/auth/login", body={
            "loginName": ln, "password": temp, "tenantCode": TENANT,
        })
        if lg.get("code") != 0:
            results.append({"loginName": ln, "ok": False, "message": lg.get("message"), "login": lg})
            time.sleep(5)
            continue
        ch = _req("POST", "/auth/change-password", token=lg["data"]["accessToken"], body={
            "oldPassword": temp, "newPassword": STABLE_PWD, "confirmPassword": STABLE_PWD,
        })
        if ch.get("code") != 0:
            ch = _req("POST", "/auth/password/change", token=lg["data"]["accessToken"], body={
                "oldPassword": temp, "newPassword": STABLE_PWD,
            })
        # force clear must_change if still required by re-login
        lg2 = _req("POST", "/auth/login", body={
            "loginName": ln, "password": STABLE_PWD, "tenantCode": TENANT,
        })
        ok = lg2.get("code") == 0
        if ok and ((lg2["data"].get("user") or {}).get("mustChangePassword") or lg2["data"].get("mustChangePassword")):
            _req("POST", "/auth/change-password", token=lg2["data"]["accessToken"], body={
                "oldPassword": STABLE_PWD, "newPassword": STABLE_PWD, "confirmPassword": STABLE_PWD,
            })
            lg2 = _req("POST", "/auth/login", body={
                "loginName": ln, "password": STABLE_PWD, "tenantCode": TENANT,
            })
            ok = lg2.get("code") == 0
        if ok:
            pwd_map[ln] = STABLE_PWD
        results.append({
            "loginName": ln, "ok": ok,
            "role": ((lg2.get("data") or {}).get("currentRole") or {}).get("roleCode") if ok else None,
            "message": None if ok else (lg2.get("message") or ch.get("message")),
            "change": ch.get("code"),
        })
    CRED_PATH.write_text(json.dumps({
        "tenantCode": TENANT,
        "schoolName": "E2E测试职业学院(sandbox-school租户内E2E组织)",
        "passwords": pwd_map,
        "loginResults": results,
        "note": "E2E系统管理员由 admin2/SCHOOL_ADMIN 承担（SYS_ADMIN 不可经师生导入分配）",
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(results, ensure_ascii=False, indent=2))
    return pwd_map


def main() -> int:
    token = login_admin()
    org = ensure_org(token)
    print("org:", org)
    dorm = ensure_dorm(token)
    print("dorm:", {k: dorm[k] for k in dorm if k != "rooms"})
    org_stu = ensure_student_org(token)
    print("student_org:", org_stu)
    imported = import_accounts(token)
    time.sleep(2)
    token = login_admin()
    pwds = activate_passwords(token)
    state = {
        "tenantCode": TENANT,
        "org": org,
        "dorm": {k: dorm[k] for k in dorm if k != "rooms"},
        "studentOrg": org_stu,
        "import": {"batchNo": imported.get("batchNo"), "confirmed": imported.get("confirmed")},
        "accounts": ALL_LOGINS,
        "passwordFile": str(CRED_PATH),
        "loginOk": sorted(pwds.keys()),
    }
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"state -> {STATE_PATH}")
    print(f"creds -> {CRED_PATH}")
    ok = sum(1 for ln in ALL_LOGINS if ln in pwds)
    print(f"login_ok={ok}/{len(ALL_LOGINS)}")
    return 0 if ok >= 10 else 1


if __name__ == "__main__":
    sys.exit(main())
