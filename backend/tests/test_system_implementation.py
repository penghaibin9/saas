"""实施与预设中心：真实数据库、状态机、租户隔离和验收阻断。"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def test_implementation_snapshot_and_readiness_blocker(client, auth_headers, db_mode):
    catalog = client.get("/api/v1/system/implementation/preset-catalog", headers=auth_headers).json()
    assert catalog["code"] == 0
    assert len(catalog["data"]["profiles"]) == 5
    assert len(catalog["data"]["sections"]) == 12
    assert catalog["data"]["standardPolicy"]["professionalStandards"] is True

    created = client.post("/api/v1/system/implementation/projects", headers=auth_headers,
                          json={"projectName": "测试学校实施", "profileCode": "HIGHER_VOCATIONAL"}).json()
    assert created["code"] == 0
    project = created["data"]
    assert project["status"] == "CONFIGURING"
    assert len(project["sections"]) == 12

    preview = client.post(f"/api/v1/system/implementation/projects/{project['id']}/preview",
                          headers=auth_headers).json()
    assert preview["code"] == 0
    assert preview["data"]["preview"]["blocked"] is False
    assert "not-allowed" not in str(preview["data"]["preview"]).lower()

    bad = client.post(f"/api/v1/system/implementation/projects/{project['id']}/apply",
                      headers=auth_headers, json={"reason": "首次实施", "confirmText": "错误文本"}).json()
    assert bad["code"] != 0

    applied = client.post(f"/api/v1/system/implementation/projects/{project['id']}/apply",
                          headers=auth_headers, json={"reason": "首次实施", "confirmText": "确认应用"}).json()
    assert applied["code"] == 0
    assert applied["data"]["snapshotOnly"] is False
    assert applied["data"]["runtime"]["workflows"]["created"] > 0
    assert applied["data"]["runtime"]["workbenches"]["created"] == 12
    assert applied["data"]["runtime"]["notifications"]["created"] == 16

    runtime = client.get(
        f"/api/v1/system/implementation/projects/{project['id']}/runtime-presets",
        headers=auth_headers,
    ).json()
    assert runtime["code"] == 0
    assert len(runtime["data"]["workbenches"]) == 12
    assert len(runtime["data"]["notifications"]) == 16
    current_project = client.get(
        "/api/v1/system/implementation/projects/current", headers=auth_headers
    ).json()["data"]
    stale = client.put(
        f"/api/v1/system/implementation/projects/{project['id']}/runtime-presets/workbenches/SCHOOL_ADMIN",
        headers=auth_headers, json={"projectVersion": max(current_project["version"] - 1, 0), "title": "过期修改"},
    ).json()
    assert stale["code"] != 0
    wb = client.put(
        f"/api/v1/system/implementation/projects/{project['id']}/runtime-presets/workbenches/SCHOOL_ADMIN",
        headers=auth_headers, json={"projectVersion": current_project["version"], "title": "学校管理工作台（定制）", "status": "ENABLED"},
    ).json()
    assert wb["code"] == 0 and wb["data"]["title"] == "学校管理工作台（定制）"
    notice = runtime["data"]["notifications"][0]
    updated_notice = client.put(
        f"/api/v1/system/implementation/projects/{project['id']}/runtime-presets/notifications/{notice['templateCode']}/{notice['channel']}",
        headers=auth_headers,
        json={"projectVersion": current_project["version"] + 1, "title": "账号已创建（定制）", "content": notice["content"], "enabled": True},
    ).json()
    assert updated_notice["code"] == 0

    from sqlalchemy import select
    from app.core.exceptions import AppException
    from app.core.security import hash_password
    from app.db.session import get_sessionmaker
    from app.models import (NotificationTemplate, RoleWorkbenchConfig, User,
                            WorkflowDefinition, WorkflowNodeDefinition)
    from app.services.runtime_preset_install_service import ensure_workflow_enabled
    from app.services.saas_role_service import ensure_builtin_roles, ensure_user_roles
    db = get_sessionmaker()()
    assert len(db.scalars(select(WorkflowDefinition)).all()) == 24
    assert len(db.scalars(select(RoleWorkbenchConfig)).all()) == 12
    assert len(db.scalars(select(NotificationTemplate)).all()) == 16
    try:
        ensure_workflow_enabled(db, 1000000000000000001, "AFFAIRS_LEAVE")
        assert False, "未确认流程必须阻止发起"
    except AppException as exc:
        assert exc.code == "DATA_CONFLICT"
    ensure_builtin_roles(db, 1000000000000000001)
    responsible = User(tenant_id=1000000000000000001, login_name="workflow-owner",
                       real_name="流程责任人", password_hash=hash_password("Test@123456"),
                       user_type="TEACHER", status="ACTIVE")
    db.add(responsible); db.flush()
    role_codes = sorted(set(db.scalars(select(WorkflowNodeDefinition.approver_role_code)).all()))
    ensure_user_roles(db, 1000000000000000001, responsible.id, role_codes)
    db.commit(); db.close()

    policy = client.post(
        f"/api/v1/system/implementation/projects/{project['id']}/runtime-presets/workflows/confirm-policy",
        headers=auth_headers,
        json={"reason": "学校集中确认首批流程责任与时限", "confirmText": "确认启用学校流程政策"},
    ).json()
    assert policy["code"] == 0
    assert policy["data"]["enabled"] == 24

    db = get_sessionmaker()()
    ensure_workflow_enabled(db, 1000000000000000001, "AFFAIRS_LEAVE")
    db.close()

    changed = client.put(
        f"/api/v1/system/implementation/projects/{project['id']}/runtime-presets/workflows/AFFAIRS_LEAVE",
        headers=auth_headers, json={"timeoutHours": 36}).json()
    assert changed["code"] == 0
    assert changed["data"]["status"] == "PENDING_CONFIRMATION"
    reconfirmed = client.post(
        f"/api/v1/system/implementation/projects/{project['id']}/runtime-presets/workflows/confirm-policy",
        headers=auth_headers,
        json={"workflowCodes": ["AFFAIRS_LEAVE"], "reason": "确认调整后的36小时时限",
              "confirmText": "确认启用学校流程政策"},
    ).json()
    assert reconfirmed["code"] == 0 and reconfirmed["data"]["enabled"] == 1

    repeated = client.post(f"/api/v1/system/implementation/projects/{project['id']}/apply",
                           headers=auth_headers, json={"reason": "重复提交", "confirmText": "确认应用"}).json()
    assert repeated["code"] == 0 and repeated["data"]["idempotent"] is True

    checked = client.post(f"/api/v1/system/implementation/projects/{project['id']}/checks/run",
                          headers=auth_headers).json()
    assert checked["code"] == 0
    assert checked["data"]["ready"] is False
    assert len(checked["data"]["checks"]) == 12
    assert all(item["severity"] in {"BLOCKER", "WARNING"} and item["ownerRole"] == "SCHOOL_ADMIN"
               for item in checked["data"]["checks"])
    assert any(x["code"] == "organization" and x["result"] == "FAIL" for x in checked["data"]["checks"])
    current_after_check = client.get("/api/v1/system/implementation/projects/current",
                                     headers=auth_headers).json()["data"]
    confirmed = client.post(
        f"/api/v1/system/implementation/projects/{project['id']}/checks/security_audit/confirm",
        headers=auth_headers,
        json={"projectVersion": current_after_check["version"], "comment": "信息中心已核对安全策略", "confirmText": "确认责任"},
    ).json()
    assert confirmed["code"] == 0 and confirmed["data"]["comment"] == "信息中心已核对安全策略"

    rejected = client.post(f"/api/v1/system/implementation/projects/{project['id']}/accept",
                           headers=auth_headers, json={"comment": "同意上线", "confirmText": "确认验收"}).json()
    assert rejected["code"] != 0


def test_implementation_rejects_sensitive_snapshot_keys(client, auth_headers, db_mode):
    created = client.post("/api/v1/system/implementation/projects", headers=auth_headers,
                          json={"projectName": "敏感字段测试", "profileCode": "DEMO"}).json()["data"]
    response = client.put(
        f"/api/v1/system/implementation/projects/{created['id']}/sections/security_audit",
        headers=auth_headers,
        json={"projectVersion": created["version"], "config": {"password": "not-allowed"}},
    ).json()
    assert response["code"] != 0


def test_acceptance_freezes_summary_and_cannot_be_reaccepted(client, auth_headers, db_mode):
    created = client.post("/api/v1/system/implementation/projects", headers=auth_headers,
                          json={"projectName": "验收摘要测试", "profileCode": "DEMO"}).json()["data"]
    from app.db.session import get_sessionmaker
    from app.models import SystemImplementationCheck, SystemImplementationProject
    db = get_sessionmaker()()
    project = db.get(SystemImplementationProject, int(created["id"]))
    project.status = "READY_FOR_ACCEPTANCE"
    for code, severity in ((code, "WARNING" if code in {"school_opening", "dictionary_numbering", "security_audit", "module_business"} else "BLOCKER")
                           for code in ("school_opening", "role_permission", "organization", "identity_import", "business_relation", "workflow", "dictionary_numbering", "security_audit", "menu_workbench", "message_notification", "go_live_check", "module_business")):
        db.add(SystemImplementationCheck(tenant_id=project.tenant_id, project_id=project.id, check_code=code,
                                         category_code=code, check_name=code, severity=severity, result="PASS",
                                         evidence_json={"test": True}))
    db.commit(); db.close()
    accepted = client.post(f"/api/v1/system/implementation/projects/{created['id']}/accept",
                           headers=auth_headers,
                           json={"comment": "学校已完成验收", "confirmText": "确认验收"}).json()
    assert accepted["code"] == 0
    assert len(accepted["data"]["acceptanceDigest"]) == 64
    assert len(accepted["data"]["acceptanceSummary"]["checks"]) == 12
    repeat = client.post(f"/api/v1/system/implementation/projects/{created['id']}/accept",
                         headers=auth_headers,
                         json={"comment": "重复验收", "confirmText": "确认验收"}).json()
    assert repeat["code"] != 0
    frozen_check = client.post(f"/api/v1/system/implementation/projects/{created['id']}/checks/security_audit/confirm",
                               headers=auth_headers,
                               json={"comment": "封板后修改", "confirmText": "确认责任"}).json()
    assert frozen_check["code"] != 0
    frozen_runtime = client.put(f"/api/v1/system/implementation/projects/{created['id']}/runtime-presets/workflows/AFFAIRS_LEAVE",
                                headers=auth_headers,
                                json={"timeoutHours": 48}).json()
    assert frozen_runtime["code"] != 0
    frozen_workbench = client.put(f"/api/v1/system/implementation/projects/{created['id']}/runtime-presets/workbenches/SCHOOL_ADMIN",
                                  headers=auth_headers,
                                  json={"title": "封板后修改"}).json()
    assert frozen_workbench["code"] != 0
    frozen_notification = client.put(f"/api/v1/system/implementation/projects/{created['id']}/runtime-presets/notifications/IMPORT_COMPLETED/IN_APP",
                                     headers=auth_headers,
                                     json={"title": "封板后修改", "content": "{batchNo}"}).json()
    assert frozen_notification["code"] != 0


def test_implementation_permission_is_fail_closed(client):
    login = client.post("/api/v1/auth/mock-login", json={"loginName": "student01", "password": "any"}).json()
    token = login["data"]["accessToken"]
    response = client.get("/api/v1/system/implementation/preset-catalog",
                          headers={"Authorization": f"Bearer {token}"})
    assert response.status_code == 403
    assert response.json()["code"] != 0


def test_change_project_inherits_installed_snapshot(client, auth_headers, db_mode):
    created = client.post("/api/v1/system/implementation/projects", headers=auth_headers,
                          json={"projectName": "变更继承测试", "profileCode": "DEMO"}).json()["data"]
    preview = client.post(f"/api/v1/system/implementation/projects/{created['id']}/preview",
                          headers=auth_headers).json()
    assert preview["code"] == 0
    applied = client.post(f"/api/v1/system/implementation/projects/{created['id']}/apply",
                          headers=auth_headers,
                          json={"reason": "初始安装", "confirmText": "确认应用"}).json()
    assert applied["code"] == 0
    from sqlalchemy import select, update
    from app.db.session import get_sessionmaker
    from app.models import SystemImplementationProject, SystemPresetInstallation
    db = get_sessionmaker()()
    db.execute(update(SystemImplementationProject).where(SystemImplementationProject.id == int(created["id"])).values(status="ACCEPTED"))
    db.commit()
    installation = db.scalars(select(SystemPresetInstallation).where(
        SystemPresetInstallation.project_id == int(created["id"]))).first()
    db.close()
    change = client.post(f"/api/v1/system/implementation/installations/{installation.id}/changes",
                         headers=auth_headers, json={"projectName": "演示版变更项目"}).json()
    assert change["code"] == 0
    assert change["data"]["projectName"] == "演示版变更项目"
    assert all(section["source"] == "INHERITED" for section in change["data"]["sections"])
    assert change["data"]["changeSourceInstallationId"] == str(installation.id)
    analyzed = client.post(f"/api/v1/system/implementation/projects/{change['data']['id']}/changes/analyze",
                           headers=auth_headers).json()
    assert analyzed["code"] == 0
    assert analyzed["data"]["riskLevel"] == "NONE"
    saved = client.put(f"/api/v1/system/implementation/projects/{change['data']['id']}/sections/security_audit",
                       headers=auth_headers,
                       json={"projectVersion": analyzed["data"]["projectVersion"],
                             "config": {"firstLoginChangePassword": True, "exportWatermark": False,
                                        "sessionMinutes": 60}}).json()
    assert saved["code"] == 0
    changed = client.post(f"/api/v1/system/implementation/projects/{change['data']['id']}/changes/analyze",
                          headers=auth_headers).json()
    assert changed["code"] == 0
    assert changed["data"]["riskLevel"] == "HIGH"
    assert "security_audit" in changed["data"]["changedSections"]


def test_identity_file_discovers_and_installs_org_role_mapping(client, auth_headers, db_mode):
    created = client.post("/api/v1/system/implementation/projects", headers=auth_headers,
                          json={"projectName": "智能匹配测试", "profileCode": "HIGHER_VOCATIONAL"}).json()["data"]

    template = client.get("/api/v1/system/identity-import/template", headers=auth_headers)
    assert template.status_code == 200
    workbook = load_workbook(BytesIO(template.content))
    sheet = workbook["导入模板"]
    columns = {str(cell.value or "").rstrip(" *"): index for index, cell in enumerate(sheet[1], 1)}

    student = {"账号类型": "STUDENT", "工号/学号": "20260001", "姓名": "张同学",
               "所属学院（学生）": "信息工程学院", "所属专业（学生）": "软件技术",
               "班级名称（学生）": "软件2601", "年级（学生）": "2026", "性别（学生）": "男"}
    teacher = {"账号类型": "TEACHER", "工号/学号": "T2026001", "姓名": "李老师",
               "所属部门（教师）": "教务处", "岗位名称（教师）": "教务管理员"}
    for row_number, values in ((2, student), (3, teacher)):
        for header, value in values.items():
            sheet.cell(row=row_number, column=columns[header], value=value)
    output = BytesIO(); workbook.save(output); workbook.close()

    validation = client.post(
        "/api/v1/system/identity-import/validate-file", headers=auth_headers,
        files={"file": ("师生导入.xlsx", output.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    assert validation["code"] == 0
    batch = validation["data"]

    discovered = client.post(
        f"/api/v1/system/implementation/projects/{created['id']}/mapping/discover",
        headers=auth_headers, json={"batchNo": batch["batchNo"]},
    ).json()
    assert discovered["code"] == 0
    mapping = discovered["data"]
    assert mapping["summary"] == {"students": 1, "teachers": 1, "organizations": 4, "blockers": 0}
    assert mapping["roleSuggestions"][0]["suggestedRoleCodes"] == ["ACADEMIC_ADMIN"]
    assert {item["recommendation"]["action"] for item in mapping["candidates"]} == {"CREATE"}

    confirmed = client.put(
        f"/api/v1/system/implementation/projects/{created['id']}/mapping/decisions",
        headers=auth_headers,
        json={"projectVersion": mapping["projectVersion"], "acceptRecommendations": True,
              "roleDecisions": [{"loginName": "T2026001", "roleCodes": ["ACADEMIC_ADMIN"]}]},
    ).json()
    assert confirmed["code"] == 0
    assert confirmed["data"]["status"] == "CONFIRMED"

    applied = client.post(
        f"/api/v1/system/implementation/projects/{created['id']}/mapping/apply",
        headers=auth_headers,
        json={"projectVersion": confirmed["data"]["projectVersion"], "reason": "首次开户",
              "confirmText": "确认安装组织与角色"},
    ).json()
    assert applied["code"] == 0
    assert applied["data"]["mapping"]["created"] == {"colleges": 2, "majors": 1, "classes": 1}
    assert applied["data"]["refreshedPreview"]["invalid"] == 0

    retried = client.post(
        f"/api/v1/system/implementation/projects/{created['id']}/mapping/apply",
        headers=auth_headers,
        json={"reason": "网络重试", "confirmText": "确认安装组织与角色"},
    ).json()
    assert retried["code"] == 0 and retried["data"]["idempotent"] is True
    assert retried["data"]["refreshedPreview"]["invalid"] == 0

    project = client.get("/api/v1/system/implementation/projects/current", headers=auth_headers).json()["data"]
    identity_config = next(item for item in project["sections"] if item["code"] == "identity_import")["config"]
    assert identity_config["mapping"]["status"] == "APPLIED"


def test_business_relation_installs_real_class_relation_and_rolls_back(client, auth_headers, db_mode):
    """开户文件关系只在账号/组织就绪后写 t_class，并保留安全回滚账本。"""
    created = client.post("/api/v1/system/implementation/projects", headers=auth_headers,
                          json={"projectName": "业务关系安装测试", "profileCode": "HIGHER_VOCATIONAL"}).json()["data"]
    template = client.get("/api/v1/system/identity-import/template", headers=auth_headers)
    workbook = load_workbook(BytesIO(template.content))
    sheet = workbook["导入模板"]
    columns = {str(cell.value or "").rstrip(" *"): index for index, cell in enumerate(sheet[1], 1)}
    rows = (
        {"账号类型": "STUDENT", "工号/学号": "20260002", "姓名": "王同学",
         "所属学院（学生）": "智能工程学院", "所属专业（学生）": "智能制造技术",
         "班级名称（学生）": "智造2601", "年级（学生）": "2026"},
        {"账号类型": "TEACHER", "工号/学号": "T2026002", "姓名": "王辅导",
         "所属部门（教师）": "学生工作处", "岗位名称（教师）": "辅导员",
         "预设角色编码（教师）": "COUNSELOR", "数据范围类型（教师）": "CLASS",
         "数据范围引用（教师）": "智造2601"},
    )
    for row_number, values in enumerate(rows, 2):
        for header, value in values.items():
            sheet.cell(row=row_number, column=columns[header], value=value)
    relation_sheet = workbook["业务关系"]
    relation_columns = {str(cell.value or "").rstrip(" *"): index
                        for index, cell in enumerate(relation_sheet[1], 1)}
    relation = {"关系类型": "COUNSELOR_CLASS", "主体工号": "T2026002",
                "对象编号/学号": "智造2601", "备注": "开户首批关系"}
    for header, value in relation.items():
        relation_sheet.cell(row=2, column=relation_columns[header], value=value)
    output = BytesIO(); workbook.save(output); workbook.close()

    validation = client.post(
        "/api/v1/system/identity-import/validate-file", headers=auth_headers,
        files={"file": ("师生及关系.xlsx", output.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    assert validation["code"] == 0
    batch = validation["data"]
    assert batch["relations"] == {"total": 1, "suggested": 1, "invalid": 0, "errors": []}

    mapping = client.post(
        f"/api/v1/system/implementation/projects/{created['id']}/mapping/discover",
        headers=auth_headers, json={"batchNo": batch["batchNo"]}).json()["data"]
    confirmed = client.put(
        f"/api/v1/system/implementation/projects/{created['id']}/mapping/decisions",
        headers=auth_headers,
        json={"projectVersion": mapping["projectVersion"], "acceptRecommendations": True,
              "roleDecisions": [{"loginName": "T2026002", "roleCodes": ["COUNSELOR"]}]},
    ).json()["data"]
    refreshed = client.post(
        f"/api/v1/system/implementation/projects/{created['id']}/mapping/apply",
        headers=auth_headers,
        json={"projectVersion": confirmed["projectVersion"], "reason": "首次开户",
              "confirmText": "确认安装组织与角色"}).json()["data"]["refreshedPreview"]
    assert refreshed["invalid"] == 0
    account_result = client.post(
        "/api/v1/system/identity-import/confirm-batch", headers=auth_headers,
        json={"batchNo": batch["batchNo"]}).json()
    assert account_result["code"] == 0
    repeated = client.post(
        "/api/v1/system/identity-import/confirm-batch", headers=auth_headers,
        json={"batchNo": batch["batchNo"]}).json()
    assert repeated["code"] == 0 and repeated["data"]["alreadyConfirmed"] is True
    assert repeated["data"]["credentialReceipt"] is None

    relation_batch = client.post(
        f"/api/v1/system/implementation/projects/{created['id']}/relations/discover",
        headers=auth_headers, json={"batchNo": batch["batchNo"]}).json()
    assert relation_batch["code"] == 0
    relation_data = relation_batch["data"]
    assert relation_data["summary"] == {"total": 1, "ready": 1, "already": 0,
                                         "conflicts": 0, "blocked": 0}
    decision = client.put(
        f"/api/v1/system/implementation/projects/{created['id']}/relations/decisions",
        headers=auth_headers,
        json={"batchNo": relation_data["batchNo"], "projectVersion": relation_data["projectVersion"],
              "acceptRecommendations": True}).json()["data"]
    installed = client.post(
        f"/api/v1/system/implementation/projects/{created['id']}/relations/apply",
        headers=auth_headers,
        json={"batchNo": relation_data["batchNo"], "projectVersion": decision["projectVersion"],
              "reason": "学校确认辅导员分班", "confirmText": "确认安装业务关系"}).json()
    assert installed["code"] == 0
    assert installed["data"]["summary"]["installed"] == 1

    from sqlalchemy import select
    from app.db.session import get_sessionmaker
    from app.models import SchoolClass, SystemBusinessRelationInstallItem, User
    db = get_sessionmaker()()
    school_class = db.scalars(select(SchoolClass).where(SchoolClass.class_name == "智造2601")).one()
    teacher = db.scalars(select(User).where(User.login_name == "T2026002")).one()
    assert school_class.counselor_id == teacher.id
    assert db.scalars(select(SystemBusinessRelationInstallItem).where(
        SystemBusinessRelationInstallItem.status == "APPLIED")).one().target_table == "t_class"
    db.close()

    rolled_back = client.post(
        f"/api/v1/system/implementation/projects/{created['id']}/relations/{relation_data['batchNo']}/rollback",
        headers=auth_headers,
        json={"reason": "验收测试回滚", "confirmText": "确认回滚业务关系"}).json()
    assert rolled_back["code"] == 0
    db = get_sessionmaker()()
    assert db.scalars(select(SchoolClass).where(SchoolClass.class_name == "智造2601")).one().counselor_id is None
    assert db.scalars(select(SystemBusinessRelationInstallItem).where(
        SystemBusinessRelationInstallItem.status == "ROLLED_BACK")).one()
    db.close()
