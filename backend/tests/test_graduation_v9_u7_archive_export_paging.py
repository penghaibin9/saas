"""U7 archive XLSX export must cross the 200-row SQL page boundary on real MySQL."""
from __future__ import annotations

import base64
import io
import uuid

from openpyxl import load_workbook

from app.db.session import get_sessionmaker
from app.models import GraduationArchiveRecord, GraduationBatch, GraduationStudent

MAIN_TENANT_ID = 1000000000000000001
ARCHIVE_EXPORT = "/api/v1/graduation/gd-archives/export"


def _seed_archive_export_rows(count: int = 205) -> int:
    db = get_sessionmaker()()
    try:
        suffix = uuid.uuid4().hex[:10]
        batch = GraduationBatch(
            tenant_id=MAIN_TENANT_ID,
            batch_name=f"U7 archive export-{suffix}",
            batch_no=f"U7-ARCH-EXPORT-{suffix}",
            academic_year="2026-2027",
            grade_year="2027届",
            planned_count=count,
            status="RUNNING",
        )
        db.add(batch)
        db.flush()
        students = [
            GraduationStudent(
                tenant_id=MAIN_TENANT_ID,
                batch_id=batch.id,
                student_no=f"U7X{idx:04d}",
                name=f"U7导出学生{idx:04d}",
                class_name="软件01班",
                topic_title=f"U7导出课题{idx:04d}",
                stage="FINAL_CHECK",
                record_status="ACTIVE",
            )
            for idx in range(1, count + 1)
        ]
        db.add_all(students)
        db.flush()
        db.add_all([
            GraduationArchiveRecord(
                tenant_id=MAIN_TENANT_ID,
                gd_student_id=student.id,
                status="PENDING_SUBMIT",
                checklist_json=[],
                missing_items=[],
            )
            for student in students
        ])
        db.commit()
        return int(batch.id)
    finally:
        db.close()


def test_u7_mysql_archive_xlsx_export_streams_past_200_rows(
    db_mode, graduation_client, auth_headers
):
    batch_id = _seed_archive_export_rows(205)
    response = graduation_client.post(
        ARCHIVE_EXPORT,
        headers=auth_headers,
        params={"batchId": batch_id},
    )
    body = response.json()
    assert body["code"] == 0, body
    payload = body["data"]
    assert payload["rowCount"] == 205
    assert payload["filename"].endswith(".xlsx")

    workbook = load_workbook(
        io.BytesIO(base64.b64decode(payload["contentBase64"])),
        read_only=True,
        data_only=True,
    )
    sheet = workbook[workbook.sheetnames[0]]
    assert sheet.max_row == 207, "title + header + all 205 archive rows must be present"
    workbook.close()
