"""DB 鐪熷疄妯″紡绔埌绔細涓存椂 SQLite 寤鸿〃+绉嶅瓙 鈫?瀛︾敓/瀹℃壒/寰呭姙/娑堟伅/瀹¤ 鍏ㄨ蛋鏁版嵁搴撱€?""
from __future__ import annotations

import pytest

from app.core.config import settings
from app.db.session import reset_state


def test_students_from_db(client, auth_headers, db_mode):
    body = client.get("/api/v1/students", headers=auth_headers).json()
    assert body["code"] == 0
    items = body["data"]["items"]
    assert body["data"]["total"] == 1 and items[0]["studentNo"] == "2023115001"
    assert items[0]["phoneMasked"] == "138****0001"
    detail = client.get(f"/api/v1/students/{db_mode['student']}", headers=auth_headers).json()["data"]
    assert detail["contacts"] and detail["timeline"] is not None


def test_students_filter_and_paginate_in_database(client, auth_headers, db_mode):
    """缁勭粐绛涢€夈€佸垎椤靛拰鑱旂郴鏂瑰紡鎵归噺鍔犺浇蹇呴』淇濇寔绉熸埛闅旂涓庡噯纭€绘暟銆?""
    from app.db.session import get_sessionmaker
    from app.models import College, Major, SchoolClass, StudentContact, StudentProfile

    tenant_id = 1000000000000000001
    other_tenant = 1000000000000000002
    db = get_sessionmaker()()
    college = College(tenant_id=tenant_id, college_name="淇℃伅宸ョ▼瀛﹂櫌", status="ACTIVE")
    db.add(college); db.flush()
    major = Major(tenant_id=tenant_id, college_id=college.id, major_name="杞欢鎶€鏈?, status="ACTIVE")
    db.add(major); db.flush()
    school_class = SchoolClass(tenant_id=tenant_id, major_id=major.id,
                               class_name="杞欢涓€鐝?, status="ACTIVE")
    db.add(school_class); db.flush()
    for index in (1, 2):
        student = StudentProfile(
            tenant_id=tenant_id, student_no=f"2026DB{index:03d}", real_name=f"鍒嗛〉瀛︾敓{index}",
            college_id=college.id, major_id=major.id, class_id=school_class.id,
            current_stage="ENROLLED", student_status="NORMAL", status="ACTIVE")
        db.add(student); db.flush()
        db.add(StudentContact(tenant_id=tenant_id, student_id=student.id, contact_type="PHONE",
                              contact_value_encrypted=f"1390000000{index}", is_primary=True,
                              verified_status="VERIFIED"))
    db.add(StudentProfile(tenant_id=other_tenant, student_no="2026OTHER", real_name="澶栨牎瀛︾敓",
                          current_stage="ENROLLED", student_status="NORMAL", status="ACTIVE"))
    db.commit(); db.close()

    first = client.get("/api/v1/students?college=淇℃伅宸ョ▼&page=1&pageSize=1",
                       headers=auth_headers).json()["data"]
    second = client.get("/api/v1/students?college=淇℃伅宸ョ▼&page=2&pageSize=1",
                        headers=auth_headers).json()["data"]
    assert first["total"] == second["total"] == 2
    assert first["items"][0]["id"] != second["items"][0]["id"]
    assert first["items"][0]["collegeName"] == "淇℃伅宸ョ▼瀛﹂櫌"
    assert first["items"][0]["majorName"] == "杞欢鎶€鏈?
    assert first["items"][0]["className"] == "杞欢涓€鐝?
    assert first["items"][0]["phoneMasked"].startswith("139****")
    by_class = client.get("/api/v1/students?className=杞欢涓€鐝?pageSize=20",
                          headers=auth_headers).json()["data"]
    assert by_class["total"] == 2
    assert all(item["studentNo"] != "2026OTHER" for item in by_class["items"])


def test_student_create_void_db(client, auth_headers, db_mode):
    from app.db.session import get_sessionmaker
    from app.models import College, Major, SchoolClass
    db = get_sessionmaker()()
    college = College(tenant_id=1000000000000000001, college_name="寤烘。娴嬭瘯瀛﹂櫌", status="ACTIVE")
    db.add(college); db.flush()
    major = Major(tenant_id=college.tenant_id, college_id=college.id,
                  major_name="寤烘。娴嬭瘯涓撲笟", status="ACTIVE")
    db.add(major); db.flush()
    school_class = SchoolClass(tenant_id=college.tenant_id, major_id=major.id,
                               class_name="寤烘。娴嬭瘯鐝?, grade="2026", status="ACTIVE",
                               class_status="NORMAL")
    db.add(school_class); db.commit(); db.close()
    created = client.post("/api/v1/students", headers=auth_headers,
                          json={"studentNo": "2099115999", "realName": "搴撲腑鏂扮敓",
                                "phone": "13800001111",
                                "collegeId": str(college.id),
                                "majorId": str(major.id),
                                "classId": str(school_class.id)}).json()["data"]
    assert created["phoneMasked"] == "138****1111"
    void = client.post(f"/api/v1/students/{created['id']}/void", headers=auth_headers,
                       json={"reason": "閲嶅寤烘。闇€瑕佷綔搴?}).json()["data"]
    assert void["physicalDelete"] is False
    items = client.get("/api/v1/students", headers=auth_headers).json()["data"]["items"]
    assert all(r["id"] != created["id"] for r in items)  # 閫昏緫鍒犻櫎鍚庡垪琛ㄤ笉鍙


def test_approval_flow_db(client, auth_headers, db_mode):
    tasks = client.get("/api/v1/approvals/tasks", headers=auth_headers).json()["data"]["items"]
    assert len(tasks) == 1 and tasks[0]["status"] == "PENDING"
    tid = tasks[0]["taskId"]
    version = tasks[0]["version"]
    no_reason = client.post(f"/api/v1/approvals/tasks/{tid}/reject", headers=auth_headers,
                            json={"version": version})
    assert no_reason.json()["code"] in (400001, 422001)
    ok = client.post(f"/api/v1/approvals/tasks/{tid}/approve", headers=auth_headers,
                     json={"comment": "鍚屾剰", "version": version}).json()["data"]
    assert ok["status"] == "APPROVED"
    processed = client.get("/api/v1/approvals/processed", headers=auth_headers).json()["data"]["items"]
    assert any(p["taskId"] == tid for p in processed)


def test_todos_messages_db(client, auth_headers, db_mode):
    todos = client.get("/api/v1/todos", headers=auth_headers).json()["data"]
    assert todos["total"] == 1
    todo_id = todos["items"][0]["todoId"]
    assert client.post(f"/api/v1/todos/{todo_id}/done", headers=auth_headers).json()["data"]["status"] == "DONE"
    msgs = client.get("/api/v1/messages", headers=auth_headers).json()["data"]
    assert msgs["total"] == 1
    mid = msgs["items"][0]["messageId"]
    assert client.post(f"/api/v1/messages/{mid}/read", headers=auth_headers).json()["data"]["status"] == "READ"


def test_audit_persisted_db(client, auth_headers, db_mode):
    # 鐢ㄧ湡瀹炶瀹¤鍔ㄤ綔锛堢櫥鍑猴級浜х敓璁板綍锛?audit/mock-record 宸插垹闄ゃ€?
    client.post("/api/v1/authz/logout", headers=auth_headers)
    body = client.get("/api/v1/audit/logs", headers=auth_headers).json()["data"]
    assert body["total"] >= 1  # 浠?t_security_audit_log 璇诲嚭
    from app.db.session import get_sessionmaker
    from app.models import SecurityAuditLog
    from sqlalchemy import select
    db = get_sessionmaker()()
    rows = db.scalars(select(SecurityAuditLog)).all()
    db.close()
    assert len(rows) >= 1


def test_p4_upload_real(client, auth_headers, db_mode, tmp_path):
    import io as _io
    resp = client.post("/api/v1/files", headers=auth_headers,
                       files={"file": ("娴嬭瘯.txt", _io.BytesIO("浣犲ソ".encode()), "text/plain")}).json()
    assert resp["code"] == 0
    meta = resp["data"]
    assert meta["fileId"].isdigit() and meta["sha256"] and meta["sizeBytes"] > 0
    got = client.get(f"/api/v1/files/{meta['fileId']}", headers=auth_headers).json()["data"]
    assert got["fileName"] == "娴嬭瘯.txt"
    bad = client.post("/api/v1/files", headers=auth_headers,
                      files={"file": ("evil.exe", _io.BytesIO(b"x"), "application/octet-stream")}).json()
    assert bad["code"] != 0  # 榛戝悕鍗曟墿灞曡鎷?


# 鍘?test_p4_import_dry_run_and_confirm 宸查殢 /import/students/* 鍒犻櫎锛?
# 瀛︾敓鎵归噺瀵煎叆鐜扮敱銆岀郴缁熺鐞?鈥?瀛︾敓瀵煎叆涓庤处鍙峰紑閫氥€嶆壙鎷咃紝瑕嗙洊瑙?
# tests/test_student_import_entries.py锛涙湰鏂囦欢缁х画瑕嗙洊瀵煎嚭涓庡叾瀹?DB 妯″紡琛屼负銆?


def test_p4_export_and_download(client, auth_headers, db_mode):
    t = client.post("/api/v1/export/students", headers=auth_headers,
                    json={"purpose": "瀛﹂櫌渚嬩細鍚嶅崟鏍稿"}).json()["data"]
    assert t["status"] == "SUCCESS" and t["taskId"].isdigit()
    resp = client.get(f"/api/v1/export/tasks/{t['taskId']}/download", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.content[:2] == b"PK"  # xlsx = zip 瀹瑰櫒
    short = client.post("/api/v1/export/students", headers=auth_headers, json={"purpose": "鐭?}).json()
    assert short["code"] in (400001, 422001)  # 鐢ㄩ€?<5 瀛楄鎷?


def test_p4_export_uses_all_database_pages(client, auth_headers, db_mode, monkeypatch):
    """灏忓垎椤垫ā鎷熷ぇ瀵煎嚭锛氫笉寰楀彧鍐欑涓€椤碉紝涔熶笉寰椾负鍚庣画椤甸噸澶嶆墽琛屾€绘暟缁熻銆?""
    from io import BytesIO

    from openpyxl import load_workbook

    from app.db.session import get_sessionmaker
    from app.models import StudentProfile
    from app.services import import_export_service as ie

    db = get_sessionmaker()()
    for index in (2, 3):
        db.add(StudentProfile(tenant_id=1000000000000000001,
                              student_no=f"2026EXPORT{index}", real_name=f"瀵煎嚭鍒嗛〉{index}",
                              current_stage="ENROLLED", student_status="NORMAL", status="ACTIVE"))
    db.commit(); db.close()
    monkeypatch.setattr(ie, "EXPORT_PAGE_SIZE", 1)

    task = client.post("/api/v1/export/students", headers=auth_headers,
                       json={"purpose": "楠岃瘉鍒嗛〉瀵煎嚭瀹屾暣鎬?}).json()["data"]
    assert task["rowCount"] == 3
    response = client.get(f"/api/v1/export/tasks/{task['taskId']}/download", headers=auth_headers)
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert len(rows) == 5  # 姘村嵃 + 琛ㄥご + 3 鍚嶅鐢?
    assert {row[0] for row in rows[2:]} == {"2023115001", "2026EXPORT2", "2026EXPORT3"}


def test_p4_audit_filters(client, auth_headers, db_mode):
    client.post("/api/v1/authz/logout", headers=auth_headers)
    body = client.get("/api/v1/audit/logs?action=LOGOUT", headers=auth_headers).json()["data"]
    assert body["total"] >= 1
    assert all(i["action"] == "LOGOUT" for i in body["items"])
    empty = client.get("/api/v1/audit/logs?operator=涓嶅瓨鍦ㄧ殑浜?, headers=auth_headers).json()["data"]
    assert empty["total"] == 0

