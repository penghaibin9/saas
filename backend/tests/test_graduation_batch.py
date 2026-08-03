"""毕业设计中心 · 毕设批次测试（MySQL 真库 via db_mode）：CRUD + 状态机 + 作废 + 阶段/规则 + 统计 + Excel 导出。"""
from __future__ import annotations

import base64

BASE = "/api/v1/graduation/batches"


def _create(graduation_client, h, **over):
    body = {"batchName": "2026届毕业设计", "batchNo": "GD-2026-1", "academicYear": "2025-2026",
            "gradeYear": "2026届", "plannedCount": 120}
    body.update(over)
    return graduation_client.post(BASE, headers=h, json=body).json()


def test_create_and_list(graduation_client, auth_headers, db_mode):
    r = _create(graduation_client, auth_headers)
    assert r["code"] == 0 and r["data"]["status"] == "DRAFT"
    lst = graduation_client.get(BASE, headers=auth_headers).json()
    assert lst["code"] == 0 and lst["data"]["total"] >= 1
    # 详情含默认阶段(8)与规则
    d = graduation_client.get(f"{BASE}/{r['data']['id']}", headers=auth_headers).json()["data"]
    assert len(d["stages"]) == 8 and "plagiarism" in d["rules"]


def test_batch_no_dedup_and_required(graduation_client, auth_headers, db_mode):
    assert _create(graduation_client, auth_headers, batchNo="GD-DUP-1")["code"] == 0
    assert _create(graduation_client, auth_headers, batchName="另一批", batchNo="GD-DUP-1")["code"] != 0
    assert graduation_client.post(BASE, headers=auth_headers, json={"batchName": "", "batchNo": ""}).json()["code"] != 0


def test_state_machine(graduation_client, auth_headers, db_mode):
    bid = _create(graduation_client, auth_headers, batchNo="GD-SM-1")["data"]["id"]
    # 草稿直接结束 → 非法
    assert graduation_client.post(f"{BASE}/{bid}/close", headers=auth_headers).json()["code"] != 0
    assert graduation_client.post(f"{BASE}/{bid}/activate", headers=auth_headers).json()["data"]["status"] == "RUNNING"
    # 已进行中再启用 → 非法
    assert graduation_client.post(f"{BASE}/{bid}/activate", headers=auth_headers).json()["code"] != 0
    assert graduation_client.post(f"{BASE}/{bid}/close", headers=auth_headers).json()["data"]["status"] == "CLOSED"
    # 已结束不可编辑
    assert graduation_client.put(f"{BASE}/{bid}", headers=auth_headers, json={"remark": "x"}).json()["code"] != 0
    assert graduation_client.post(f"{BASE}/{bid}/archive", headers=auth_headers).json()["data"]["status"] == "ARCHIVED"
    # 已归档不可再归档
    assert graduation_client.post(f"{BASE}/{bid}/archive", headers=auth_headers).json()["code"] != 0


def test_void_only_draft(graduation_client, auth_headers, db_mode):
    bid = _create(graduation_client, auth_headers, batchNo="GD-VOID-1")["data"]["id"]
    # 原因过短拒
    assert graduation_client.post(f"{BASE}/{bid}/void", headers=auth_headers, json={"reason": "短"}).json()["code"] != 0
    assert graduation_client.post(f"{BASE}/{bid}/void", headers=auth_headers, json={"reason": "编号录入错误需作废"}).json()["data"]["status"] == "VOIDED"
    # 进行中不能作废
    b2 = _create(graduation_client, auth_headers, batchNo="GD-VOID-2")["data"]["id"]
    graduation_client.post(f"{BASE}/{b2}/activate", headers=auth_headers)
    assert graduation_client.post(f"{BASE}/{b2}/void", headers=auth_headers, json={"reason": "已进行中不能作废吧"}).json()["code"] != 0


def test_stages_and_rules(graduation_client, auth_headers, db_mode):
    bid = _create(graduation_client, auth_headers, batchNo="GD-CFG-1")["data"]["id"]
    st = graduation_client.post(f"{BASE}/{bid}/stages", headers=auth_headers, json={"stages": [
        {"code": "TOPIC", "name": "选题", "startDate": "2025-09-01", "endDate": "2025-09-30"}]}).json()
    assert st["code"] == 0 and len(st["data"]["stages"]) == 1
    rl = graduation_client.post(f"{BASE}/{bid}/rules", headers=auth_headers, json={"rules": {"plagiarism": {"thresholdPercent": 20}}}).json()
    assert rl["code"] == 0 and rl["data"]["rules"]["plagiarism"]["thresholdPercent"] == 20


def test_stats(graduation_client, auth_headers, db_mode):
    _create(graduation_client, auth_headers, batchNo="GD-STAT-1")
    s = graduation_client.get(f"{BASE}/stats", headers=auth_headers).json()
    assert s["code"] == 0 and s["data"]["total"] >= 1
    assert any(x["status"] == "DRAFT" for x in s["data"]["byStatus"])


def test_export_xlsx(graduation_client, auth_headers, db_mode):
    import io
    from openpyxl import load_workbook
    _create(graduation_client, auth_headers, batchNo="GD-EXP-1")
    # 按当前筛选导出（status=DRAFT）
    ex = graduation_client.post(f"{BASE}/export?status=DRAFT", headers=auth_headers).json()
    assert ex["code"] == 0 and ex["data"]["rowCount"] >= 1
    # 文件名含「毕设批次台账」+ 时间，且为 .xlsx（非 CSV）
    assert "毕设批次台账" in ex["data"]["filename"] and ex["data"]["filename"].endswith(".xlsx")
    assert ex["data"]["mediaType"].endswith("spreadsheetml.sheet")
    raw = base64.b64decode(ex["data"]["contentBase64"])
    assert raw[:2] == b"PK"  # xlsx = zip
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    assert "毕设批次台账" in str(ws["A1"].value)  # 抬头/来源行
    assert ws["A2"].value == "批次名称"            # 第 2 行为表头


def test_detail_not_found(graduation_client, auth_headers, db_mode):
    assert graduation_client.get(f"{BASE}/99999999", headers=auth_headers).json()["code"] != 0
