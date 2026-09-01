"""X1 production XLSX/import/audit closure (executed by final X2 gate)."""
from __future__ import annotations

import base64
import importlib.util
import io
from pathlib import Path

from openpyxl import load_workbook


def _resource_xlsx(rows: list[list]) -> bytes:
    from app.services.xlsx_util import build_template_xlsx
    return build_template_xlsx(
        ["楼栋编码", "楼栋名称", "性别属性", "楼层", "房号", "房型", "容量", "床号", "房间状态"],
        samples=rows,
        required=["楼栋编码", "楼栋名称", "性别属性", "楼层", "房号", "容量", "床号", "房间状态"],
    )


def test_x1_dorm_resource_template_dry_run_confirm_and_export(client, db_mode, auth_headers):
    template_response = client.get("/api/v1/import/domain/dorm/template", headers=auth_headers)
    assert template_response.status_code == 200
    template = template_response.json()["data"]
    assert template["filename"].endswith(".xlsx")
    workbook = load_workbook(io.BytesIO(base64.b64decode(template["contentBase64"])), read_only=True)
    assert workbook["导入模板"]["A1"].value == "楼栋编码 *"
    workbook.close()

    content = _resource_xlsx([
        ["X1-DORM", "X1学生公寓", "混合", 1, "101", "标准双人间", 2, "101-1", "启用"],
        ["X1-DORM", "X1学生公寓", "混合", 1, "101", "标准双人间", 2, "101-2", "启用"],
    ])
    preview_response = client.post(
        "/api/v1/import/domain/dorm/validate-file", headers=auth_headers,
        files={"file": ("dorm-x1.xlsx", content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()["data"]
    assert preview["status"] == "DRY_RUN_PASSED"
    assert preview["okRows"] == 2 and preview["errorRows"] == 0

    confirm_response = client.post(
        "/api/v1/import/domain/confirm", headers={**auth_headers, "Idempotency-Key": "x1-dorm-import"},
        json={"domain": "dorm", "batchNo": preview["batchNo"]},
    )
    assert confirm_response.status_code == 200
    receipt = confirm_response.json()["data"]
    assert receipt["createdBuildings"] == 1
    assert receipt["createdRooms"] == 1
    assert receipt["createdBeds"] == 2

    task_response = client.post(
        "/api/v1/export/domain/dorm", headers={**auth_headers, "Idempotency-Key": "x1-dorm-export"},
        json={"reportType": "resources", "purpose": "X1宿舍房源验收导出"},
    )
    assert task_response.status_code == 200
    task = task_response.json()["data"]
    assert task["fileName"] == "房源台账.xlsx"
    assert task["rowCount"] >= 2
    download = client.get(f"/api/v1/export/tasks/{task['taskId']}/download", headers=auth_headers)
    assert download.status_code == 200
    exported = load_workbook(io.BytesIO(download.content), read_only=True)
    sheet = exported.active
    assert "范围：" in sheet["A1"].value
    assert any(row[0].value == "X1-DORM" for row in sheet.iter_rows(min_row=3))
    exported.close()


def test_x1_dorm_dry_run_error_workbook_and_atomic_reject(client, db_mode, auth_headers):
    bad = _resource_xlsx([
        ["X1-BAD", "X1不完整公寓", "男", 1, "101", "标准四人间", 4, "101-1", "启用"],
    ])
    preview_response = client.post(
        "/api/v1/import/domain/dorm/validate-file", headers=auth_headers,
        files={"file": ("dorm-x1-bad.xlsx", bad,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    preview = preview_response.json()["data"]
    assert preview["status"] == "DRY_RUN_FAILED"
    assert preview["errorRows"] >= 1
    error_book = client.get(preview["errorWorkbookUrl"], headers=auth_headers)
    assert error_book.status_code == 200
    workbook = load_workbook(io.BytesIO(error_book.content), read_only=True)
    assert workbook.active["A2"].value == "Excel行号"
    workbook.close()
    confirm = client.post(
        "/api/v1/import/domain/confirm", headers=auth_headers,
        json={"domain": "dorm", "batchNo": preview["batchNo"]},
    )
    assert confirm.status_code >= 400 or confirm.json()["code"] != 0


def test_x1_orientation_production_report_variants(client, db_mode, auth_headers, monkeypatch):
    from test_orientation_import_export_a1 import _authority
    from app.api.v1 import import_export as import_export_api
    from app.db.session import get_sessionmaker
    from app.models import OrientationStudent

    monkeypatch.setattr(import_export_api, "_limit_operation", lambda *args, **kwargs: None)
    ids = _authority()
    db = get_sessionmaker()()
    try:
        student = OrientationStudent(
            tenant_id=1000000000000000001, batch_id=ids["batch"],
            name="X1迎新报表生", admission_no="X1-ORI-001", gender="男",
            college_id=ids["college"], major_id=ids["major"], class_id=ids["class"],
            college_name="A1信息学院", major_name="A1软件专业", class_name="A1软件2601班",
            grade="2026", stage="ADMITTED", report_status="NOT_REPORTED",
            payment_status="UNPAID", green_channel_status="NOT_APPLIED",
            material_status="NOT_UPLOADED", dorm_status="UNASSIGNED", risk_level="LOW",
            record_status="ACTIVE", source_type="MANUAL", source_record_id="X1-ORI-001",
            identity_status="UNLINKED",
        )
        db.add(student)
        db.commit()
    finally:
        db.close()

    for report_type, filename in (
        ("students", "迎新新生台账.xlsx"), ("progress", "报到进度.xlsx"),
        ("no-show", "未报到.xlsx"), ("exceptions", "迎新异常.xlsx"),
    ):
        response = client.post(
            "/api/v1/export/domain/orientation", headers=auth_headers,
            json={"batchId": ids["batch"], "reportType": report_type,
                  "purpose": f"X1{report_type}迎新台账验收"},
        )
        assert response.status_code == 200, response.text
        assert response.json()["data"]["fileName"] == filename


def test_x1_consistency_audit_is_read_only(client, db_mode):
    from app.db.session import get_sessionmaker
    from app.models import StudentProfile
    script_path = Path(__file__).resolve().parents[2] / "scripts" / "audit_school_authority_consistency.py"
    spec = importlib.util.spec_from_file_location("x1_consistency_audit", script_path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)

    db = get_sessionmaker()()
    try:
        before = db.query(StudentProfile).count()
    finally:
        db.close()
    report = module.audit(1000000000000000001)
    assert report["readOnly"] is True
    assert isinstance(report["issues"], list)
    db = get_sessionmaker()()
    try:
        assert db.query(StudentProfile).count() == before
    finally:
        db.close()
