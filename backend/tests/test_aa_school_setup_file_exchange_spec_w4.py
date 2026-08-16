"""A-W4 Course Catalog File Exchange spec helper contracts."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


def _spec():
    from app.modules.academic_affairs.services import academic_affairs_school_setup_file_exchange_spec as spec
    return spec


def test_course_template_is_real_xlsx_with_stable_key_columns_and_required_markers():
    spec = _spec()
    content = spec.build_course_catalog_import_template()
    assert content.startswith(b"PK")

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True, keep_links=False)
    try:
        sheet = workbook["导入模板"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        assert headers[0] == "课程代码 *"
        assert headers[1] == "版本 *"
        assert "课程名称 *" in headers
        assert "课程类别 *" in headers
        assert "课程性质 *" in headers
        assert "学分 *" in headers
        assert "考核方式 *" in headers
        assert "先修课代码" in headers
        sample = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert sample[0] == "CS101"
        assert sample[1] == 1
        assert sample[2] == "Python程序设计"
        assert "填写说明" in workbook.sheetnames
    finally:
        workbook.close()


def test_required_headers_derive_from_single_contract_map_not_second_schema():
    spec = _spec()
    from app.modules.academic_affairs.services import academic_affairs_school_setup_import_contract as contract

    expected = [
        title
        for title, field in contract.COURSE_HEADER_MAP.items()
        if field in contract.COURSE_REQUIRED_FIELDS
    ]
    assert spec.course_catalog_template_headers() == list(contract.COURSE_HEADER_MAP.keys())
    assert spec.course_catalog_required_headers() == expected


def test_parser_adapter_uses_caller_security_gated_reader_and_fixed_query_dry_run(monkeypatch):
    spec = _spec()
    calls = {}
    rows = [{"courseCode": "CS101", "version": "1", "courseName": "Python程序设计"}]

    def fake_reader(path, header_map):
        calls["path"] = path
        calls["headerMap"] = dict(header_map)
        return rows

    expected_preview = {
        "totalRows": 1,
        "validRows": 1,
        "invalidRows": 0,
        "createRows": 1,
        "reuseRows": 0,
        "conflictRows": 0,
        "rejectRows": 0,
        "items": [],
        "errors": [],
    }

    def fake_dry_run(received_rows, user):
        calls["rows"] = received_rows
        calls["user"] = user
        return expected_preview

    monkeypatch.setattr(spec, "course_catalog_dry_run", fake_dry_run)
    source = Path("/tmp/security-gated-course.xlsx")
    user = {"currentRoleCode": "ACADEMIC_ADMIN"}
    parsed_rows, preview = spec.parse_and_validate_course_catalog(
        source,
        user=user,
        reader=fake_reader,
    )

    assert calls["path"] == source
    assert calls["headerMap"] == spec.COURSE_HEADER_MAP
    assert calls["rows"] is rows
    assert calls["user"] is user
    assert parsed_rows is rows
    assert preview is expected_preview


def test_empty_course_workbook_is_rejected_before_domain_dry_run(monkeypatch):
    spec = _spec()
    calls = {"dryRun": 0}

    def fake_reader(_path, _header_map):
        return []

    def forbidden_dry_run(_rows, _user):
        calls["dryRun"] += 1
        raise AssertionError("empty workbook must not enter DB dry-run")

    monkeypatch.setattr(spec, "course_catalog_dry_run", forbidden_dry_run)
    rows, preview = spec.parse_and_validate_course_catalog(
        Path("/tmp/header-only-course.xlsx"),
        user={"currentRoleCode": "ACADEMIC_ADMIN"},
        reader=fake_reader,
    )

    assert rows == []
    assert calls["dryRun"] == 0
    assert preview["totalRows"] == 0
    assert preview["validRows"] == 0
    assert preview["invalidRows"] == 1
    assert preview["rejectRows"] == 1
    assert preview["items"][0]["action"] == "REJECT"
    assert preview["items"][0]["code"] == "COURSE_SOURCE_EMPTY"
    assert preview["errors"][0]["field"] == "file"
    assert preview["errors"][0]["code"] == "COURSE_SOURCE_EMPTY"


def test_spec_explicitly_stays_internal_until_int_shared_confirm_is_ready():
    contract = _spec().course_catalog_file_exchange_contract()
    assert contract["templateVersion"] == "course-catalog-v1"
    assert contract["publicImportEnabled"] is False
    assert contract["confirmOwner"] == "INT_SHARED_DATA_EXCHANGE"
