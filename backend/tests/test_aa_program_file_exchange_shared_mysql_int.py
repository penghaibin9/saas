"""INT MySQL lifecycle proof for Program through the shared File Exchange owner."""
from __future__ import annotations

import hashlib
import uuid
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.context import set_current_user, set_tenant
from app.db.session import get_sessionmaker

TID = 1000000000000000001
USER_ID = 81001


def _user() -> dict:
    return {
        "tenantId": str(TID),
        "userId": str(USER_ID),
        "realName": "Program shared lifecycle tester",
        "userType": "TEACHER",
        "currentRoleCode": "ACADEMIC_ADMIN",
        "permissions": ["*"],
        "dataScope": "ALL",
    }


def _patch_program_scope(monkeypatch) -> None:
    from app.modules.academic_affairs.services import academic_affairs_school_setup_program_binding_confirm_service as binding
    from app.modules.academic_affairs.services import academic_affairs_school_setup_program_definition_confirm_service as definition
    from app.modules.academic_affairs.services import academic_affairs_school_setup_program_preview_service as preview

    context = lambda _user, _db: SimpleNamespace(  # noqa: E731 - frozen TENANT_ALL fixture
        scope_type="TENANT_ALL",
        college_ids=set(),
        class_ids=set(),
    )
    monkeypatch.setattr(preview, "build_affairs_context", context)
    monkeypatch.setattr(definition, "build_affairs_context", context)
    monkeypatch.setattr(binding, "build_affairs_context", context)


def _seed_authorities_and_file(workbook: bytes) -> tuple[int, str, int]:
    from app.models import AaCourse, Major, Tenant
    from app.models.file import FileObject
    from app.services import data_exchange_job_service as jobs
    from app.services.file_scan_constants import SCAN_NOT_REQUIRED

    suffix = uuid.uuid4().hex[:10].upper()
    db = get_sessionmaker()()
    try:
        if db.get(Tenant, TID) is None:
            db.add(Tenant(
                id=TID,
                tenant_code=f"program-shared-{suffix.lower()}",
                school_name=f"Program shared lifecycle school {suffix}",
                status="ACTIVE",
            ))
            db.flush()
        major = Major(
            tenant_id=TID,
            college_id=889001,
            major_name=f"Program shared software {suffix}",
            code=f"PS{suffix}",
            status="ACTIVE",
            education_years=3,
            enroll_status="ENROLLING",
        )
        course = AaCourse(
            tenant_id=TID,
            course_code=f"PSC{suffix}",
            course_name=f"Program shared course {suffix}",
            category="MAJOR_CORE",
            nature="REQUIRED",
            credit=3,
            exam_mode="EXAM",
            is_core=True,
            prerequisite_codes_json="[]",
            applicable_majors_json="[]",
            is_all_major=False,
            version=1,
            status="ENABLED",
        )
        db.add_all([major, course])
        db.flush()
        now = jobs._now()
        file_row = FileObject(
            tenant_id=TID,
            file_key=f"program-shared/{suffix}/source.xlsx",
            file_name="academic_program_import.xlsx",
            ext="xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            size_bytes=len(workbook),
            sha256=hashlib.sha256(workbook).hexdigest(),
            biz_type="ACADEMIC_PROGRAM_IMPORT_SOURCE",
            owner_user_id=USER_ID,
            created_by=USER_ID,
            visibility="PRIVATE",
            security_level="SENSITIVE",
            status="AVAILABLE",
            storage_backend="local",
            storage_zone="IMPORT",
            upload_source="USER",
            scan_required=False,
            scan_status=SCAN_NOT_REQUIRED,
            scan_attempts=0,
            scanned_at=now,
            available_at=now,
        )
        db.add(file_row)
        db.commit()
        return int(major.id), str(course.course_code), int(file_row.id)
    finally:
        db.close()


def _workbook_bytes(*, major_id: int, course_code: str, series_key: str) -> bytes:
    from app.modules.academic_affairs.services import academic_affairs_school_setup_program_workbook_adapter as workbook

    content = workbook.build_program_import_template()
    wb = load_workbook(BytesIO(content))
    try:
        for sheet_name in ("培养方案", "方案课程", "学分要求", "实践环节", "毕业要求", "适用范围"):
            wb[sheet_name]["A2"] = series_key
        wb["培养方案"]["C2"] = "Program shared 2026 definition"
        wb["培养方案"]["D2"] = major_id
        wb["方案课程"]["C2"] = course_code
        wb["适用范围"]["C2"] = major_id
        # PRACTICE is optional; keep this contract focused on shared job lifecycle.
        wb["实践环节"].delete_rows(2, 1)
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    finally:
        wb.close()


def _replace_file_payload(file_id: int, workbook: bytes) -> None:
    from app.models.file import FileObject

    db = get_sessionmaker()()
    try:
        row = db.get(FileObject, file_id)
        assert row is not None
        row.size_bytes = len(workbook)
        row.sha256 = hashlib.sha256(workbook).hexdigest()
        db.commit()
    finally:
        db.close()


@pytest.mark.usefixtures("db_mode")
def test_program_shared_importjob_definition_then_binding_real_mysql_lifecycle(tmp_path, monkeypatch):
    from app.models import AaProgram, AaProgramBinding
    from app.models.data_exchange import ImportJob
    from app.modules.academic_affairs.services import academic_file_exchange_service as exchange
    from app.services import data_exchange_confirm_service as confirm
    from app.services import file_scan_service

    user = _user()
    set_tenant({"tenantId": str(TID)})
    set_current_user(user)
    try:
        _patch_program_scope(monkeypatch)

        # Seed authorities first, then replace the template's placeholder IDs with
        # the real Major/Course values.  The FileObject is real; only the physical
        # storage/scan edge is substituted because File Center owns that contract.
        placeholder = b"program-shared-placeholder"
        major_id, course_code, file_id = _seed_authorities_and_file(placeholder)
        series_key = f"INT-SHARED-{uuid.uuid4().hex[:12].upper()}"
        workbook = _workbook_bytes(
            major_id=major_id,
            course_code=course_code,
            series_key=series_key,
        )
        _replace_file_payload(file_id, workbook)
        source_path = Path(tmp_path) / "academic_program_import.xlsx"
        source_path.write_bytes(workbook)
        monkeypatch.setattr(exchange, "_source_file_path", lambda _row, _user: source_path)
        monkeypatch.setattr(file_scan_service, "assert_file_ready_for_business", lambda *_args, **_kwargs: None)

        definition_job = exchange.create_academic_import_job(
            filename=source_path.name,
            source_file_id=file_id,
            import_type=exchange.ACADEMIC_PROGRAM_IMPORT,
            context={"phase": "DEFINITION"},
            user=user,
        )
        assert definition_job["status"] == "VALIDATED"
        assert definition_job["validRows"] == 5
        assert definition_job["invalidRows"] == 0
        assert definition_job["preview"]["invalidRows"] == 0

        definition_done = confirm.confirm_import_job(
            definition_job["id"],
            expected_version=definition_job["version"],
            user=user,
        )
        assert definition_done["status"] == "SUCCEEDED"
        assert definition_done["confirmedRows"] == 5
        assert definition_done["result"]["phase"] == "DEFINITION"

        db = get_sessionmaker()()
        try:
            program = db.scalars(select(AaProgram).where(
                AaProgram.tenant_id == TID,
                AaProgram.series_key == series_key,
                AaProgram.version == 1,
                AaProgram.is_deleted.is_(False),
            )).one()
            program_id = int(program.id)
            assert program.status == "DRAFT"
            definition_row = db.get(ImportJob, int(definition_job["id"]))
            assert definition_row is not None
            assert definition_row.status == "SUCCEEDED"
            assert definition_row.lease_token is None
            assert definition_row.lease_started_at is None
            assert dict(definition_row.source_snapshot_json or {})["context"] == {"phase": "DEFINITION"}
            assert dict(definition_row.source_snapshot_json or {}).get("rowDigest")

            # Approval is intentionally outside ordinary import.  BINDING may only
            # activate a definition that is already PUBLISHED/ENABLED.
            program.status = "PUBLISHED"
            db.commit()
        finally:
            db.close()

        binding_job = exchange.create_academic_import_job(
            filename=source_path.name,
            source_file_id=file_id,
            import_type=exchange.ACADEMIC_PROGRAM_IMPORT,
            context={"phase": "BINDING"},
            user=user,
        )
        assert binding_job["status"] == "VALIDATED"
        assert binding_job["validRows"] == 5
        assert binding_job["invalidRows"] == 0

        binding_done = confirm.confirm_import_job(
            binding_job["id"],
            expected_version=binding_job["version"],
            user=user,
        )
        assert binding_done["status"] == "SUCCEEDED"
        assert binding_done["confirmedRows"] == 5
        assert binding_done["result"]["phase"] == "BINDING"

        db = get_sessionmaker()()
        try:
            program = db.get(AaProgram, program_id)
            assert program is not None and program.status == "ENABLED"
            bindings = db.scalars(select(AaProgramBinding).where(
                AaProgramBinding.tenant_id == TID,
                AaProgramBinding.major_id == major_id,
                AaProgramBinding.grade_year == "2026",
                AaProgramBinding.class_id.is_(None),
                AaProgramBinding.is_deleted.is_(False),
            )).all()
            assert len(bindings) == 1
            assert int(bindings[0].program_id) == program_id
            assert bindings[0].status == "ACTIVE"

            binding_row = db.get(ImportJob, int(binding_job["id"]))
            assert binding_row is not None
            assert binding_row.status == "SUCCEEDED"
            assert binding_row.lease_token is None
            assert binding_row.lease_started_at is None
            assert dict(binding_row.source_snapshot_json or {})["context"] == {"phase": "BINDING"}
            assert dict(binding_row.source_snapshot_json or {}).get("rowDigest")
        finally:
            db.close()
    finally:
        set_current_user(None)
        set_tenant(None)


def _representative_course_plan(*, major_code: str, major_name: str):
    """Build the same 37-course/140-credit school plan used by the pure Gold contract."""
    from decimal import Decimal

    from app.services import sandbox_school_academic_affairs_seed as seed
    from app.services import sandbox_school_curriculum_closure as closure

    courses = []
    for code, name, _category, _nature, credit, _hours, _exam in seed.PUBLIC_COURSES:
        courses.append({
            "code": code,
            "name": name,
            "credit": Decimal(str(credit)),
            "module": "PUBLIC_BASIC",
        })
    for code, name, credit in closure.PUBLIC_EXPANSION:
        courses.append({
            "code": code,
            "name": name,
            "credit": Decimal(str(credit)),
            "module": "PUBLIC_BASIC",
        })
    for suffix, label, _category, _nature, credit, _hours, _exam in seed.MAJOR_COURSE_TEMPLATES:
        courses.append({
            "code": f"{major_code}-{suffix}",
            "name": f"{major_name}{label}",
            "credit": Decimal(str(credit)),
            "module": "MAJOR_CORE",
        })
    for index, label in enumerate(closure.ADVANCED_MAJOR_COURSE_LABELS, start=7):
        courses.append({
            "code": f"{major_code}-{index:02d}",
            "name": f"{major_name}{label}",
            "credit": Decimal("4"),
            "module": "MAJOR_CORE",
        })
    for index, label in enumerate(closure.MAJOR_EXTENSION_LABELS, start=10):
        courses.append({
            "code": f"{major_code}-{index:02d}",
            "name": f"{major_name}{label}",
            "credit": Decimal("4"),
            "module": "MAJOR_CORE",
        })
    for index, (label, credit, _segment_type, _weeks) in enumerate(
        closure.PRACTICE_LABELS,
        start=18,
    ):
        courses.append({
            "code": f"{major_code}-{index:02d}",
            "name": f"{major_name}{label}",
            "credit": Decimal(str(credit)),
            "module": "PRACTICE",
        })

    assert len(courses) == 37
    assert len({row["code"] for row in courses}) == 37
    module_credit = {
        module: sum(
            (row["credit"] for row in courses if row["module"] == module),
            Decimal("0"),
        )
        for module in ("PUBLIC_BASIC", "MAJOR_CORE", "PRACTICE")
    }
    assert module_credit == {
        "PUBLIC_BASIC": Decimal("30"),
        "MAJOR_CORE": Decimal("64"),
        "PRACTICE": Decimal("46"),
    }
    assert sum((row["credit"] for row in courses), Decimal("0")) == Decimal("140")

    public_codes = [row["code"] for row in courses if row["module"] == "PUBLIC_BASIC"]
    major_codes = [row["code"] for row in courses if row["module"] != "PUBLIC_BASIC"]
    term_by_code = closure._term_assignments("2026", public_codes, major_codes)
    assert set(term_by_code) == {row["code"] for row in courses}
    return courses, term_by_code


def _seed_representative_authorities_and_file(
    workbook: bytes,
    *,
    major_code: str,
    major_name: str,
    courses: list[dict],
) -> tuple[int, int]:
    from app.models import AaCourse, Major, Tenant
    from app.models.file import FileObject
    from app.services import data_exchange_job_service as jobs
    from app.services.file_scan_constants import SCAN_NOT_REQUIRED

    suffix = uuid.uuid4().hex[:10].upper()
    db = get_sessionmaker()()
    try:
        if db.get(Tenant, TID) is None:
            db.add(Tenant(
                id=TID,
                tenant_code=f"program-school-{suffix.lower()}",
                school_name=f"Representative school {suffix}",
                status="ACTIVE",
            ))
            db.flush()
        major = Major(
            tenant_id=TID,
            college_id=889002,
            major_name=major_name,
            code=major_code,
            status="ACTIVE",
            education_years=3,
            enroll_status="ENROLLING",
        )
        db.add(major)
        db.flush()
        db.add_all([
            AaCourse(
                tenant_id=TID,
                course_code=str(row["code"]),
                course_name=str(row["name"]),
                category=str(row["module"]),
                nature="REQUIRED",
                credit=row["credit"],
                exam_mode="EXAM",
                is_core=str(row["module"]) != "PUBLIC_BASIC",
                prerequisite_codes_json="[]",
                applicable_majors_json="[]",
                is_all_major=False,
                version=1,
                status="ENABLED",
            )
            for row in courses
        ])
        db.flush()
        now = jobs._now()
        file_row = FileObject(
            tenant_id=TID,
            file_key=f"program-representative/{suffix}/source.xlsx",
            file_name="representative_school_program.xlsx",
            ext="xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            size_bytes=len(workbook),
            sha256=hashlib.sha256(workbook).hexdigest(),
            biz_type="ACADEMIC_PROGRAM_IMPORT_SOURCE",
            owner_user_id=USER_ID,
            created_by=USER_ID,
            visibility="PRIVATE",
            security_level="SENSITIVE",
            status="AVAILABLE",
            storage_backend="local",
            storage_zone="IMPORT",
            upload_source="USER",
            scan_required=False,
            scan_status=SCAN_NOT_REQUIRED,
            scan_attempts=0,
            scanned_at=now,
            available_at=now,
        )
        db.add(file_row)
        db.commit()
        return int(major.id), int(file_row.id)
    finally:
        db.close()


def _representative_grouped(
    *,
    major_id: int,
    major_code: str,
    major_name: str,
    series_key: str,
    courses: list[dict],
    term_by_code: dict[str, int],
) -> dict[str, list[dict]]:
    from app.services import sandbox_school_curriculum_closure as closure

    graduation_items = (
        ("KNOWLEDGE", "掌握{major}专业基础理论、技术标准和岗位知识体系。"),
        ("ABILITY", "能够完成{major}典型岗位任务、项目实施与质量改进。"),
        ("QUALITY", "具备职业道德、团队协作、安全意识、数字素养和持续学习能力。"),
        ("CERTIFICATE", "鼓励取得与{major}相关的职业技能等级证书或行业认证。"),
    )
    return {
        "MAIN": [{
            "programSeriesKey": series_key,
            "programVersion": 1,
            "programName": f"{major_name}2026级人才培养方案",
            "majorId": major_id,
            "gradeYear": "2026",
            "totalCredits": 140,
            "educationYears": 3,
        }],
        "COURSE": [{
            "programSeriesKey": series_key,
            "programVersion": 1,
            "courseCode": row["code"],
            "courseVersion": 1,
            "openTermNo": term_by_code[row["code"]],
            "module": row["module"],
            "formationMode": "ADMIN_FIXED",
            "creditSnapshot": "",
        } for row in courses],
        "CREDIT_REQUIREMENT": [{
            "programSeriesKey": series_key,
            "programVersion": 1,
            "module": module,
            "creditTarget": credit_target,
        } for module, credit_target in closure.CREDIT_STRUCTURE],
        "PRACTICE": [{
            "programSeriesKey": series_key,
            "programVersion": 1,
            "segmentName": f"{major_name}{label}",
            "segmentType": segment_type,
            "openTermNo": term_by_code[f"{major_code}-{18 + index:02d}"],
            "weeks": weeks,
            "credit": credit,
            "orgMode": "DISTRIBUTED" if segment_type == "POST_INTERNSHIP" else "CENTRALIZED",
            "assessmentMode": "CHECK",
            "location": "校内实训中心/合作企业",
            "sortOrder": index,
        } for index, (label, credit, segment_type, weeks) in enumerate(closure.PRACTICE_LABELS)],
        "GRADUATION": [{
            "programSeriesKey": series_key,
            "programVersion": 1,
            "category": category,
            "content": template.format(major=major_name),
            "sortOrder": index,
        } for index, (category, template) in enumerate(graduation_items, start=1)],
        "BINDING": [{
            "programSeriesKey": series_key,
            "programVersion": 1,
            "majorId": major_id,
            "gradeYear": "2026",
            "bindingScope": "MAJOR_GRADE",
            "classId": "",
        }],
    }


def _grouped_workbook_bytes(grouped: dict[str, list[dict]]) -> bytes:
    from app.modules.academic_affairs.services import academic_affairs_school_setup_program_file_exchange_spec as spec
    from app.modules.academic_affairs.services import academic_affairs_school_setup_program_workbook_adapter as workbook

    wb = load_workbook(BytesIO(workbook.build_program_import_template()))
    try:
        for sheet_name in spec.PROGRAM_SHEET_ORDER:
            ws = wb[sheet_name]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            group = spec.PROGRAM_GROUP_BY_SHEET[sheet_name]
            fields = list(spec.PROGRAM_HEADER_MAP_BY_GROUP[group].values())
            for row in grouped[group]:
                ws.append([row.get(field, "") for field in fields])
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    finally:
        wb.close()


@pytest.mark.usefixtures("db_mode")
def test_program_shared_representative_school_140_credit_create_bind_and_replay(tmp_path, monkeypatch):
    import json
    from decimal import Decimal

    from app.models import (
        AaProgram,
        AaProgramBinding,
        AaProgramCourse,
        AaProgramGraduationRequirement,
        AaProgramPracticeSegment,
    )
    from app.models.data_exchange import ImportJob
    from app.modules.academic_affairs.services import academic_file_exchange_service as exchange
    from app.services import data_exchange_confirm_service as confirm
    from app.services import file_scan_service
    from app.services import sandbox_school_curriculum_closure as closure

    user = _user()
    set_tenant({"tenantId": str(TID)})
    set_current_user(user)
    try:
        _patch_program_scope(monkeypatch)
        major_code = f"SBX{uuid.uuid4().hex[:8].upper()}"
        major_name = "软件技术"
        courses, term_by_code = _representative_course_plan(
            major_code=major_code,
            major_name=major_name,
        )
        placeholder = b"representative-school-program-placeholder"
        major_id, file_id = _seed_representative_authorities_and_file(
            placeholder,
            major_code=major_code,
            major_name=major_name,
            courses=courses,
        )
        series_key = f"SBX-SOFTWARE-2026-{uuid.uuid4().hex[:8].upper()}"
        grouped = _representative_grouped(
            major_id=major_id,
            major_code=major_code,
            major_name=major_name,
            series_key=series_key,
            courses=courses,
            term_by_code=term_by_code,
        )
        workbook = _grouped_workbook_bytes(grouped)
        _replace_file_payload(file_id, workbook)
        source_path = Path(tmp_path) / "representative_school_program.xlsx"
        source_path.write_bytes(workbook)
        monkeypatch.setattr(exchange, "_source_file_path", lambda _row, _user: source_path)
        monkeypatch.setattr(file_scan_service, "assert_file_ready_for_business", lambda *_args, **_kwargs: None)

        definition_job = exchange.create_academic_import_job(
            filename=source_path.name,
            source_file_id=file_id,
            import_type=exchange.ACADEMIC_PROGRAM_IMPORT,
            context={"phase": "DEFINITION"},
            user=user,
        )
        assert definition_job["status"] == "VALIDATED"
        assert definition_job["validRows"] == 52
        assert definition_job["invalidRows"] == 0

        definition_done = confirm.confirm_import_job(
            definition_job["id"],
            expected_version=definition_job["version"],
            user=user,
        )
        assert definition_done["status"] == "SUCCEEDED"
        assert definition_done["confirmedRows"] == 52
        definition_result = definition_done["result"]
        assert definition_result["phase"] == "DEFINITION"
        assert definition_result["domainMutationWriteCount"] == 48
        metrics = definition_result["preflight"]["quality"]["programMetrics"][0]
        assert Decimal(str(metrics["courseCreditSum"])) == Decimal("140")
        assert Decimal(str(metrics["practiceCreditSum"])) == Decimal("46")
        assert Decimal(str(metrics["actualCreditSum"])) == Decimal("140")

        db = get_sessionmaker()()
        try:
            program = db.scalars(select(AaProgram).where(
                AaProgram.tenant_id == TID,
                AaProgram.series_key == series_key,
                AaProgram.version == 1,
                AaProgram.is_deleted.is_(False),
            )).one()
            program_id = int(program.id)
            assert program.status == "DRAFT"
            assert Decimal(str(program.total_credits)) == Decimal("140")

            relations = db.scalars(select(AaProgramCourse).where(
                AaProgramCourse.tenant_id == TID,
                AaProgramCourse.program_id == program_id,
                AaProgramCourse.is_deleted.is_(False),
            ).order_by(AaProgramCourse.id)).all()
            assert len(relations) == 37
            assert {str(row.formation_mode) for row in relations} == {"ADMIN_FIXED"}
            assert sum(
                (Decimal(str(row.credit_snapshot or 0)) for row in relations),
                Decimal("0"),
            ) == Decimal("140")

            practices = db.scalars(select(AaProgramPracticeSegment).where(
                AaProgramPracticeSegment.tenant_id == TID,
                AaProgramPracticeSegment.program_id == program_id,
                AaProgramPracticeSegment.is_deleted.is_(False),
                AaProgramPracticeSegment.status == "ACTIVE",
            ).order_by(AaProgramPracticeSegment.sort_order, AaProgramPracticeSegment.id)).all()
            assert len(practices) == 6
            assert sum(
                (Decimal(str(row.credit or 0)) for row in practices),
                Decimal("0"),
            ) == Decimal("46")

            graduations = db.scalars(select(AaProgramGraduationRequirement).where(
                AaProgramGraduationRequirement.tenant_id == TID,
                AaProgramGraduationRequirement.program_id == program_id,
                AaProgramGraduationRequirement.is_deleted.is_(False),
                AaProgramGraduationRequirement.status == "ACTIVE",
            )).all()
            assert len(graduations) == 4

            requirement = json.loads(program.requirement_json or "{}")
            actual_credit_structure = {
                str(item["module"]): Decimal(str(item["creditTarget"]))
                for item in requirement["creditStructure"]
            }
            assert actual_credit_structure == {
                str(module): Decimal(str(target))
                for module, target in closure.CREDIT_STRUCTURE
            }

            program.status = "PUBLISHED"
            db.commit()
        finally:
            db.close()

        binding_job = exchange.create_academic_import_job(
            filename=source_path.name,
            source_file_id=file_id,
            import_type=exchange.ACADEMIC_PROGRAM_IMPORT,
            context={"phase": "BINDING"},
            user=user,
        )
        assert binding_job["status"] == "VALIDATED"
        assert binding_job["validRows"] == 52
        assert binding_job["invalidRows"] == 0
        binding_done = confirm.confirm_import_job(
            binding_job["id"],
            expected_version=binding_job["version"],
            user=user,
        )
        assert binding_done["status"] == "SUCCEEDED"
        assert binding_done["confirmedRows"] == 52
        assert binding_done["result"]["phase"] == "BINDING"
        assert binding_done["result"]["domainMutationWriteCount"] == 3
        assert binding_done["result"]["reconciliation"]["createdBindings"] == 1

        definition_replay_job = exchange.create_academic_import_job(
            filename=source_path.name,
            source_file_id=file_id,
            import_type=exchange.ACADEMIC_PROGRAM_IMPORT,
            context={"phase": "DEFINITION"},
            user=user,
        )
        assert definition_replay_job["status"] == "VALIDATED"
        assert definition_replay_job["validRows"] == 52
        definition_replay_done = confirm.confirm_import_job(
            definition_replay_job["id"],
            expected_version=definition_replay_job["version"],
            user=user,
        )
        assert definition_replay_done["status"] == "SUCCEEDED"
        assert definition_replay_done["result"]["domainMutationWriteCount"] == 0
        assert definition_replay_done["result"]["reconciliation"]["reusedPrograms"] == 1

        binding_replay_job = exchange.create_academic_import_job(
            filename=source_path.name,
            source_file_id=file_id,
            import_type=exchange.ACADEMIC_PROGRAM_IMPORT,
            context={"phase": "BINDING"},
            user=user,
        )
        assert binding_replay_job["status"] == "VALIDATED"
        assert binding_replay_job["validRows"] == 52
        binding_replay_done = confirm.confirm_import_job(
            binding_replay_job["id"],
            expected_version=binding_replay_job["version"],
            user=user,
        )
        assert binding_replay_done["status"] == "SUCCEEDED"
        assert binding_replay_done["result"]["domainMutationWriteCount"] == 0
        assert binding_replay_done["result"]["reconciliation"]["reusedBindings"] == 1

        db = get_sessionmaker()()
        try:
            program = db.get(AaProgram, program_id)
            assert program is not None and program.status == "ENABLED"
            relations = db.scalars(select(AaProgramCourse).where(
                AaProgramCourse.tenant_id == TID,
                AaProgramCourse.program_id == program_id,
                AaProgramCourse.is_deleted.is_(False),
            )).all()
            practices = db.scalars(select(AaProgramPracticeSegment).where(
                AaProgramPracticeSegment.tenant_id == TID,
                AaProgramPracticeSegment.program_id == program_id,
                AaProgramPracticeSegment.is_deleted.is_(False),
            )).all()
            graduations = db.scalars(select(AaProgramGraduationRequirement).where(
                AaProgramGraduationRequirement.tenant_id == TID,
                AaProgramGraduationRequirement.program_id == program_id,
                AaProgramGraduationRequirement.is_deleted.is_(False),
            )).all()
            bindings = db.scalars(select(AaProgramBinding).where(
                AaProgramBinding.tenant_id == TID,
                AaProgramBinding.major_id == major_id,
                AaProgramBinding.grade_year == "2026",
                AaProgramBinding.class_id.is_(None),
                AaProgramBinding.is_deleted.is_(False),
            )).all()
            assert (len(relations), len(practices), len(graduations), len(bindings)) == (37, 6, 4, 1)
            assert bindings[0].status == "ACTIVE"
            assert int(bindings[0].program_id) == program_id

            jobs = [
                db.get(ImportJob, int(item["id"]))
                for item in (
                    definition_job,
                    binding_job,
                    definition_replay_job,
                    binding_replay_job,
                )
            ]
            assert all(row is not None and row.status == "SUCCEEDED" for row in jobs)
            assert all(row.lease_token is None and row.lease_started_at is None for row in jobs)
            digests = {
                str(dict(row.source_snapshot_json or {}).get("rowDigest") or "")
                for row in jobs
            }
            assert len(digests) == 1
            assert next(iter(digests))
        finally:
            db.close()
    finally:
        set_current_user(None)
        set_tenant(None)
