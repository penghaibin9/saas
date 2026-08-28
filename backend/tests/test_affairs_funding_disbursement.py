"""13A-C 奖助发放台账端到端（真实 DB）。

覆盖：按批次为 GRANTED 申请生成发放台账(幂等)→标记已发放/失败→失败重试→发放概览；
重复发放冲突、失败原因过短、已发放不可置失败、金额两位小数、异步 XLSX 的筛选/用途/水印/下载。
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from affairs_contract_test_support import role_headers

TID = 1000000000000000001
BASE = "/api/v1/student-affairs"


def _seed_granted(sid, n=2):
    from app.db.session import get_sessionmaker
    from app.models import FundingApplication, FundingBatch, FundingProject, StudentProfile
    db = get_sessionmaker()()
    amount = Decimal("3300.75")
    p = FundingProject(tenant_id=TID, project_name="国家助学金", project_type="GRANT",
                       amount=amount, quota=10, status="ENABLED")
    db.add(p); db.flush()
    b = FundingBatch(tenant_id=TID, project_id=p.id, project_type="GRANT", year_code="2025-2026",
                     quota=10, status="OPEN")
    db.add(b); db.flush()
    base_student = db.get(StudentProfile, int(sid))
    assert base_student is not None
    student_ids = [int(sid)]
    for i in range(1, n):
        other = StudentProfile(
            tenant_id=TID, student_no=f"DISB{sid}-{i}", real_name=f"发放测试学生{i}",
            class_id=base_student.class_id, college_id=base_student.college_id,
            current_stage=base_student.current_stage or "ON_CAMPUS",
            student_status="NORMAL", status="ACTIVE", is_deleted=False, version=0,
        )
        db.add(other); db.flush(); student_ids.append(int(other.id))
    applications = [
        FundingApplication(
            tenant_id=TID, batch_id=b.id, student_id=student_id,
            apply_source="SELF", project_type="GRANT", amount=amount,
            status="PUBLICITY", is_deleted=False, version=0,
        )
        for student_id in student_ids
    ]
    db.add_all(applications)
    # 真实迁移禁止 INSERT 直接伪造 GRANTED。先落 PUBLICITY，再让 MySQL 的
    # package-10 trigger 执行正式的 PUBLICITY→GRANTED 金额/额度原子占用约束。
    db.flush()
    for application in applications:
        application.status = "GRANTED"
    db.flush()
    db.commit()
    bid = b.id
    db.close()
    return bid


def test_disbursement_full_flow(client, db_mode):
    # 发放属于学工/资助真实业务职责，禁止用 SCHOOL_ADMIN 通配权限冒充业务经办人；
    # 也禁止 mock-login 对未知登录名按 userType 回退成演示教师。这里直接签发已落库的
    # STUDENT_AFFAIRS_ADMIN 角色上下文，走与生产 SYSTEM RoleTemplate 相同的权限解析器。
    hdr = role_headers(
        "STUDENT_AFFAIRS_ADMIN", login_name="sa_disbursement_admin",
        real_name="资助发放测试学工管理员",
    )
    bid = _seed_granted(db_mode["student"], n=2)
    # 生成发放台账
    g = client.post(f"{BASE}/funding/batches/{bid}/disbursements/generate", headers=hdr).json()
    assert g["code"] == 0 and g["data"]["generated"] == 2
    # 幂等：再生成 0 条
    g2 = client.post(f"{BASE}/funding/batches/{bid}/disbursements/generate", headers=hdr).json()
    assert g2["data"]["generated"] == 0
    # 台账列表（PENDING 2 条；金额必须保留两位小数）
    lst = client.get(f"{BASE}/funding/disbursements", headers=hdr, params={"batchId": bid}).json()
    items = lst["data"]["items"]
    assert len(items) == 2 and all(x["bankStatus"] == "PENDING" for x in items)
    assert all(str(x["amount"]) == "3300.75" for x in items)
    row1, row2 = items[0], items[1]
    d1, d2 = row1["disbursementId"], row2["disbursementId"]
    # 标记已发放：显式提交列表中可见版本
    iss = client.post(f"{BASE}/funding/disbursements/{d1}/issue", headers=hdr,
                      json={"disburseNo": "FB2026-001", "bankLast4": "6411",
                            "version": row1["version"]}).json()
    assert iss["code"] == 0 and iss["data"]["bankStatus"] == "ISSUED"
    assert iss["data"]["bankLast4"] == "6411"  # 仅后4位
    # 重复发放 → 状态冲突（使用服务返回的新版本，避免把状态冲突误测成版本冲突）
    assert client.post(f"{BASE}/funding/disbursements/{d1}/issue", headers=hdr,
                       json={"version": iss["data"]["version"]}).json()["code"] != 0
    # 已发放不可置失败
    assert client.post(f"{BASE}/funding/disbursements/{d1}/fail", headers=hdr,
                       json={"reason": "银行退回卡号有误",
                             "version": iss["data"]["version"]}).json()["code"] != 0
    # 另一条置失败（原因≥5）
    assert client.post(f"{BASE}/funding/disbursements/{d2}/fail", headers=hdr,
                       json={"reason": "短", "version": row2["version"]}).json()["code"] != 0
    f2 = client.post(f"{BASE}/funding/disbursements/{d2}/fail", headers=hdr,
                     json={"reason": "银行账号信息缺失待补",
                           "version": row2["version"]}).json()
    assert f2["code"] == 0 and f2["data"]["bankStatus"] == "FAILED"
    # 失败重试：FAILED 必须允许再次真实发放，而不是新增第二条台账。
    retry = client.post(f"{BASE}/funding/disbursements/{d2}/issue", headers=hdr,
                        json={"disburseNo": "FB2026-002", "bankLast4": "7522",
                              "version": f2["data"]["version"]}).json()
    assert retry["code"] == 0 and retry["data"]["bankStatus"] == "ISSUED"
    after_retry = client.get(f"{BASE}/funding/disbursements", headers=hdr,
                             params={"batchId": bid}).json()["data"]["items"]
    assert len(after_retry) == 2
    assert {x["disbursementId"] for x in after_retry} == {d1, d2}

    # 发放概览：两笔真实金额精确到分，合计不能由 float 漂移。
    st = client.get(f"{BASE}/funding/disbursements/stats", headers=hdr).json()["data"]
    assert st["total"] == 2
    assert any(x["key"] == "ISSUED" and x["count"] == 2 for x in st["byStatus"])
    assert st["issuedAmountTotal"] == "6601.50"

    # XLSX 必须显式用途；请求只创建任务，真正生成由学工 scheduler worker 完成。
    bad_export = client.post(
        f"{BASE}/funding/disbursements/export", headers=hdr,
        json={"batchId": bid, "purpose": "短"},
    ).json()
    assert bad_export["code"] != 0

    created = client.post(
        f"{BASE}/funding/disbursements/export", headers=hdr,
        json={"batchId": bid, "bankStatus": "ISSUED", "purpose": "财务发放结果复核归档"},
    ).json()
    assert created["code"] == 0 and created["data"]["queued"] is True
    job_id = created["data"]["jobId"]

    from app.core.context import set_tenant
    from app.services import affairs_funding_export_service as export_svc
    set_tenant({"tenantId": str(TID)})
    try:
        run = export_svc.run_pending(limit=2, worker_id="pytest-funding-export")
    finally:
        set_tenant(None)
    assert run == {"claimed": 1, "succeeded": 1, "failed": 0}

    job = client.get(
        f"{BASE}/funding/disbursements/export-jobs/{job_id}", headers=hdr,
    ).json()["data"]
    assert job["status"] == "SUCCEEDED" and job["rowCount"] == 2
    ticket = client.post(
        f"{BASE}/funding/disbursements/export-jobs/{job_id}/download-ticket",
        headers=hdr, json={"expectedVersion": job["version"]},
    ).json()["data"]["ticket"]
    response = client.get(
        f"{BASE}/funding/disbursements/export-jobs/{job_id}/download",
        headers=hdr, params={"ticket": ticket},
    )
    assert response.status_code == 200
    assert "spreadsheetml.sheet" in response.headers.get("content-type", "")

    wb = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    assert wb.sheetnames == ["资助发放台账"]
    ws = wb["资助发放台账"]
    watermark = [str(x or "") for x in next(ws.iter_rows(values_only=True))]
    assert any("用途：财务发放结果复核归档" in x for x in watermark)
    headers = list(next(ws.iter_rows(values_only=True)))
    assert headers == ["学号", "姓名", "项目类型", "发放金额", "银行卡后4位", "发放状态",
                       "发放批次号", "发放时间", "失败原因"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 2
    assert all(Decimal(str(row[3])).quantize(Decimal("0.01")) == Decimal("3300.75") for row in rows)
    assert {str(row[4]) for row in rows} == {"6411", "7522"}
    assert all(str(row[5]) == "已发放" for row in rows)
