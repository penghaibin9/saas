"""公共 Excel 底座 V1.1 单元/集成测试。

覆盖：模板生成、文件读取、必填/类型/格式校验、文件内重复、错误行收集、错误行 Excel、
统一预校验结构、确认导入模式、导入记录写入 MySQL、Excel 导出、脱敏导出、
拒绝 CSV 冒充、拒绝 SQLite 回落（TEST_DATABASE_URL 必须 MySQL）。

不改四个已完成业务模块；仅验证底座本身可复用。
"""
from __future__ import annotations

import io
import os

from openpyxl import Workbook, load_workbook

from app.services import xlsx_util
from app.services.excel import (ColumnSpec, ExportSpec, ImportSpec, build_error_rows, build_export,
                                build_template, confirm_import, pre_validate, read_upload)

# ── 一个演示用导入契约（模拟业务模块「只写配置」）──
_STORE: list = []


def _persist(rows):
    _STORE.extend(rows)
    return {"created": len(rows)}


def _dup_check(rows):
    """库内查重扩展点：编号已在 _STORE 视为重复。"""
    exist = {r.get("code") for r in _STORE}
    return {i + 1: f"编号 {r.get('code')} 已存在" for i, r in enumerate(rows)
            if r.get("code") in exist}


def _mask_phone(v):
    v = str(v)
    return v[:3] + "****" + v[-4:] if len(v) >= 7 else "***"


def _import_spec() -> ImportSpec:
    return ImportSpec(
        module_key="demo", biz_type="person", template_name="人员导入",
        columns=[
            ColumnSpec("name", "姓名", required=True, max_length=20, example="张三"),
            ColumnSpec("code", "编号", required=True, unique_in_file=True, example="P001"),
            ColumnSpec("phone", "手机号", type="phone", example="13800000000"),
            ColumnSpec("age", "年龄", type="int", example="18"),
            ColumnSpec("joinDate", "入职日期", type="date", example="2026-07-07"),
            ColumnSpec("level", "级别", type="enum", options=["A", "B"], example="A"),
        ],
        duplicate_check=_dup_check, persist_rows=_persist,
        audit_action="导入人员",
    )


def _export_spec() -> ExportSpec:
    return ExportSpec(
        module_key="demo", biz_type="person", sheet_title="人员台账",
        columns=[
            ColumnSpec("name", "姓名"),
            ColumnSpec("phone", "手机号", mask=_mask_phone),
        ],
    )


def setup_function(_):
    _STORE.clear()


# ═══════════ 1. 模板生成 ═══════════

def test_build_template_has_headers_and_required_mark():
    data = build_template(_import_spec())
    wb = load_workbook(io.BytesIO(data))
    assert "导入模板" in wb.sheetnames
    ws = wb["导入模板"]
    headers = [c.value for c in ws[1]]
    assert "姓名 *" in headers and "编号 *" in headers   # 必填带 *
    assert "手机号" in headers                            # 非必填不带 *
    assert "填写说明" in wb.sheetnames


# ═══════════ 2. 文件读取（表头映射，去必填 *）═══════════

def test_read_upload_maps_headers():
    wb = Workbook(); ws = wb.active; ws.title = "导入模板"
    ws.append(["姓名 *", "编号 *", "手机号"])
    ws.append(["张三", "P001", "13800000000"])
    buf = io.BytesIO(); wb.save(buf)
    rows = read_upload(_import_spec(), buf.getvalue())
    assert rows == [{"name": "张三", "code": "P001", "phone": "13800000000"}]


# ═══════════ 3~6. 必填/类型/格式/文件内重复 + 错误行收集 ═══════════

def test_required_type_format_and_dup_in_file():
    rows = [
        {"name": "", "code": "P001", "phone": "138", "age": "x", "joinDate": "2026/13/40", "level": "C"},
        {"name": "李四", "code": "P002", "phone": "13800000000", "age": "20", "joinDate": "2026-07-07", "level": "A"},
        {"name": "王五", "code": "P002", "phone": "", "age": "", "joinDate": "", "level": ""},  # 编号文件内重复
    ]
    res = pre_validate(_import_spec(), rows)
    assert res["total"] == 3
    msgs = " ".join(e["message"] for e in res["errors"])
    assert "姓名 必填" in msgs
    assert "手机号 手机号格式不正确" in msgs
    assert "年龄 必须为整数" in msgs
    assert "日期格式不正确" in msgs
    assert "级别 取值须为" in msgs
    assert "文件内重复" in msgs
    assert res["invalidRows"] == 2 and res["validRows"] == 1


# ═══════════ 7. 库内重复扩展点 ═══════════

def test_duplicate_check_against_store():
    _STORE.append({"code": "P001"})
    res = pre_validate(_import_spec(), [{"name": "张三", "code": "P001", "level": "A"}])
    assert res["invalidRows"] == 1
    assert any("已存在" in e["message"] for e in res["errors"])


# ═══════════ 8. 错误行 Excel 生成 ═══════════

def test_build_error_rows_xlsx_only_bad_rows():
    rows = [{"name": "", "code": "P001"}, {"name": "李四", "code": "P002", "level": "A"}]
    res = pre_validate(_import_spec(), rows)
    packed = build_error_rows(_import_spec(), rows, res["errors"])
    assert packed["filename"].endswith(".xlsx")
    import base64
    wb = load_workbook(io.BytesIO(base64.b64decode(packed["contentBase64"])))
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[2]]  # 第一行水印，第二行表头
    assert "错误字段" in header and "错误原因" in header


# ═══════════ 9. 统一预校验结构 ═══════════

def test_pre_validate_unified_shape():
    res = pre_validate(_import_spec(), [{"name": "张三", "code": "P001", "level": "A"}])
    for k in ("moduleKey", "bizType", "templateVersion", "total", "validRows",
              "invalidRows", "passed", "rows", "errors"):
        assert k in res
    assert res["moduleKey"] == "demo" and res["passed"] is True


# ═══════════ 10. 确认导入模式（脏数据被拦截）═══════════

def test_confirm_import_rejects_invalid():
    import pytest
    with pytest.raises(Exception):
        confirm_import(_import_spec(), [{"name": "", "code": "P001"}])


def test_confirm_import_persists_valid():
    res = confirm_import(_import_spec(), [{"name": "张三", "code": "P001", "level": "A"}])
    assert res["created"] == 1 and len(_STORE) == 1


# ═══════════ 11. 导入记录写入 MySQL ═══════════

def test_import_job_record_written_to_db(client, auth_headers, db_mode):
    # record_import 直接调用（不经过 HTTP 中间件），租户上下文不会自动注入；
    # 缺租户上下文时 _tid() 抛异常，被 record_import 自己的 except Exception 静默吞掉
    # 变成返回 None，看着像"没接库"，其实是真异常被吃了——这里手动设租户避免误判。
    from app.core.context import set_tenant
    from app.services.excel import job_service
    from tests.conftest import MAIN_TENANT_ID
    pre = {"total": 3, "validRows": 3, "invalidRows": 0}
    set_tenant({"tenantId": str(MAIN_TENANT_ID)})
    try:
        row = job_service.record_import("demo", "person", file_name="人员.xlsx",
                                        pre=pre, result={"created": 3}, status="IMPORTED")
    finally:
        set_tenant(None)
    assert row is not None and row["status"] == "IMPORTED" and row["successRows"] == 3
    # 通过通用端点回读
    r = client.get("/api/v1/excel/import-jobs?moduleKey=demo", headers=auth_headers).json()
    assert r["code"] == 0
    assert any(it["fileName"] == "人员.xlsx" for it in r["data"]["items"])


# ═══════════ 12~13. Excel 导出 + 脱敏导出 ═══════════

def test_build_export_ledger_with_mask():
    import base64
    items = [{"name": "张三", "phone": "13812345678"}]
    packed = build_export(_export_spec(), items, operator_name="测试员")
    assert packed["filename"].endswith(".xlsx") and packed["rowCount"] == 1
    wb = load_workbook(io.BytesIO(base64.b64decode(packed["contentBase64"])))
    ws = wb[wb.sheetnames[0]]
    all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "138****5678" in all_text        # 手机号脱敏生效
    assert "13812345678" not in all_text     # 明文不出现
    assert "导出人：测试员" in all_text        # 水印


# ═══════════ 14. 不允许 CSV 冒充（底座只产出 xlsx）═══════════

def test_no_csv_output():
    tpl = build_template(_import_spec())
    assert tpl[:2] == b"PK"  # xlsx 是 zip(PK) 开头，绝非 CSV 文本
    packed = build_export(_export_spec(), [{"name": "张三", "phone": "13812345678"}])
    assert packed["filename"].endswith(".xlsx")
    assert "csv" not in packed["mediaType"].lower()


# ═══════════ 15. 不允许 SQLite 回落 ═══════════

def test_test_db_is_mysql_not_sqlite():
    url = os.environ.get("TEST_DATABASE_URL", "")
    assert url and not url.startswith("sqlite"), "TEST_DATABASE_URL 必须为 MySQL，禁止回落 SQLite"
    assert url.startswith("mysql")
