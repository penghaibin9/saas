"""SaaS 角色模板与开局预检纯单元测试；明确不连接数据库。"""
from __future__ import annotations

import pytest
from types import SimpleNamespace
import io
import zipfile
import base64

from openpyxl import Workbook, load_workbook

from app.core.exceptions import AppException
from app.core.permissions import ROLE_PERMISSIONS, has_permission
from app.services import school_onboarding_service as onboarding
from app.services import identity_import_service
from app.services import identity_import_file_service as identity_files
from app.services import mobile_teacher_service
from app.services.saas_role_templates import (BUILTIN_ROLE_TEMPLATES,
                                               role_catalog,
                                               role_codes_from_row)
from app.services import saas_role_service
from pathlib import Path
import ast


class _Rows:
    def __init__(self, values):
        self.values = values

    def all(self):
        return self.values


class _FakeDb:
    def __init__(self, scalar_results):
        self.scalar_results = list(scalar_results)
        self.added = []
        self.flush_count = 0

    def scalars(self, statement):
        return _Rows(self.scalar_results.pop(0))

    def add(self, value):
        self.added.append(value)

    def flush(self):
        self.flush_count += 1


def test_every_school_permission_role_has_a_versioned_template():
    template_codes = {item["roleCode"] for item in BUILTIN_ROLE_TEMPLATES}
    assert set(ROLE_PERMISSIONS) - {"PLATFORM_SUPER_ADMIN"} == template_codes


def test_teacher_row_accepts_multiple_codes_and_chinese_position_aliases():
    assert role_codes_from_row({"roleCodes": ["辅导员", "毕设导师", "INTERN_MENTOR"]}) == [
        "COUNSELOR", "GD_MENTOR", "INTERN_MENTOR",
    ]


def test_legacy_academic_alias_resolves_to_least_privilege_teacher_role():
    assert role_codes_from_row({"roleCode": "ACADEMIC"}) == ["ACADEMIC_TEACHER"]


def test_teacher_import_cannot_assign_school_admin():
    with pytest.raises(AppException) as exc:
        role_codes_from_row({"roleCode": "SCHOOL_ADMIN"})
    assert exc.value.code == "VALIDATION_ERROR"


def test_teacher_catalog_excludes_student_and_platform_roles():
    codes = {item["roleCode"] for item in role_catalog(teacher_only=True)["items"]}
    assert "STUDENT" not in codes
    assert "SCHOOL_ADMIN" not in codes
    assert "SYS_ADMIN" not in codes
    assert "SECURITY_AUDITOR" not in codes
    assert "PLATFORM_SUPER_ADMIN" not in codes


def test_identity_import_rejects_non_identity_entities(monkeypatch):
    monkeypatch.setattr(onboarding, "db_enabled", lambda: False)
    with pytest.raises(AppException) as exc:
        identity_import_service.run_identity_import(
            {"userType": "PLATFORM_SUPER_ADMIN"},
            {"tenantId": "1001", "colleges": [{"name": "信息学院"}]},
            dry_run=True,
        )
    assert exc.value.code == "VALIDATION_ERROR"
    assert "colleges" in exc.value.message


def test_identity_import_requires_at_least_one_account_row():
    with pytest.raises(AppException) as exc:
        identity_import_service.run_identity_import(
            {"userType": "PLATFORM_SUPER_ADMIN"}, {"tenantId": "1001"}, dry_run=True)
    assert exc.value.code == "VALIDATION_ERROR"


def test_identity_import_cannot_disable_atomic_transaction():
    with pytest.raises(AppException) as exc:
        identity_import_service.run_identity_import(
            {"userType": "PLATFORM_SUPER_ADMIN"},
            {"tenantId": "1001", "atomic": False,
             "students": [{"studentNo": "S001", "name": "学生"}]},
            dry_run=True,
        )
    assert exc.value.code == "VALIDATION_ERROR"
    assert "整批" in exc.value.message


def test_identity_import_confirm_fails_closed_when_database_is_disabled(monkeypatch):
    monkeypatch.setattr(onboarding, "db_enabled", lambda: False)
    with pytest.raises(AppException) as exc:
        identity_import_service.run_identity_import(
            {"userType": "PLATFORM_SUPER_ADMIN"},
            {"tenantId": "1001", "students": [{"studentNo": "S001", "name": "学生"}]},
            dry_run=False,
        )
    assert exc.value.code == "SERVER_ERROR"


def test_onboarding_dry_run_reports_accounts_roles_and_scopes_without_db(monkeypatch):
    monkeypatch.setattr(onboarding, "db_enabled", lambda: False)
    body = {
        "tenantId": "1001",
        "students": [{"studentNo": "S001", "name": "张同学"}],
        "teachers": [{
            "loginName": "T001", "name": "李老师",
            "roleCodes": ["辅导员", "毕设导师"],
            "scopeType": "CLASS", "scopeRef": "软件2401",
        }],
    }
    report = identity_import_service.run_identity_import(
        {"userType": "PLATFORM_SUPER_ADMIN"}, body, dry_run=True)
    assert report["errors"] == []
    assert report["entities"]["studentAccounts"]["created"] == 1
    assert report["entities"]["teachers"]["created"] == 1
    assert report["entities"]["roleBindings"]["created"] == 3
    assert report["entities"]["scopes"]["created"] == 1


def test_onboarding_rejects_unknown_role_instead_of_creating_it(monkeypatch):
    monkeypatch.setattr(onboarding, "db_enabled", lambda: False)
    report = identity_import_service.run_identity_import(
        {"userType": "PLATFORM_SUPER_ADMIN"},
        {"tenantId": "1001", "teachers": [{
            "loginName": "T001", "name": "李老师", "roleCode": "SUPER_TEACHER",
        }]},
        dry_run=True,
    )
    assert any(e["field"] == "roleCodes" for e in report["errors"])


def test_onboarding_public_channel_cannot_create_student_or_teacher_accounts():
    with pytest.raises(AppException) as exc:
        onboarding.run_onboarding(
            {"userType": "PLATFORM_SUPER_ADMIN"},
            {"tenantId": "1001", "students": [{"studentNo": "S001", "name": "学生"}]},
            dry_run=True,
        )
    assert exc.value.code == "VALIDATION_ERROR"
    assert "系统管理" in exc.value.message


def test_student_number_and_teacher_login_collision_is_rejected():
    errors = onboarding._validate_rows({
        "students": [{"studentNo": "10001", "name": "学生"}],
        "teachers": [{"loginName": "10001", "name": "老师", "roleCode": "ACADEMIC_TEACHER"}],
    })
    assert any("学号与教师登录名冲突" in e["error"] for e in errors)


def test_builtin_role_initialization_is_idempotent_without_committing():
    first_db = _FakeDb([[]])
    first = saas_role_service.ensure_builtin_roles(first_db, 1001)
    assert first["created"] == len(BUILTIN_ROLE_TEMPLATES)
    assert len(first_db.added) == len(BUILTIN_ROLE_TEMPLATES)
    existing = [SimpleNamespace(
        role_code=item["roleCode"], role_type="SYSTEM", is_deleted=False,
        status="ACTIVE", version=0, remark="") for item in BUILTIN_ROLE_TEMPLATES]
    second_db = _FakeDb([existing])
    second = saas_role_service.ensure_builtin_roles(second_db, 1001)
    assert second["unchanged"] == len(BUILTIN_ROLE_TEMPLATES)
    assert second_db.added == []


def test_disabled_builtin_role_is_restored_by_initialization():
    roles = [SimpleNamespace(
        role_code=item["roleCode"], role_type="SYSTEM", is_deleted=False,
        status="DISABLED" if item["roleCode"] == "STUDENT" else "ACTIVE",
        version=0, remark="", role_name="旧名称") for item in BUILTIN_ROLE_TEMPLATES]
    db = _FakeDb([roles])
    report = saas_role_service.ensure_builtin_roles(db, 1001)
    assert report["restored"] == 1
    student = next(role for role in roles if role.role_code == "STUDENT")
    assert student.status == "ACTIVE"
    assert student.role_name == "学生"


def test_only_system_administrators_have_identity_import_permission():
    assert has_permission({"currentRoleCode": "SCHOOL_ADMIN"}, "systemAdmin.user.import")
    assert has_permission({"currentRoleCode": "SYS_ADMIN"}, "systemAdmin.user.import")
    assert not has_permission({"currentRoleCode": "ACADEMIC_TEACHER"}, "systemAdmin.user.import")


def test_existing_user_gets_only_missing_role_binding():
    roles = [
        SimpleNamespace(id=1, role_code="COUNSELOR"),
        SimpleNamespace(id=2, role_code="GD_MENTOR"),
    ]
    existing_link = SimpleNamespace(
        role_id=1, is_deleted=False, status="ACTIVE", version=0)
    db = _FakeDb([roles, [existing_link]])
    report = saas_role_service.ensure_user_roles(
        db, 1001, 7, ["COUNSELOR", "GD_MENTOR"])
    assert report == {"created": 1, "restored": 0, "unchanged": 1}
    assert len(db.added) == 1


def _user_constructor_sites() -> set[tuple[str, str]]:
    """返回生产代码中 User(...) 的文件和最外层业务函数，兼容导入别名。"""
    app_dir = Path(__file__).resolve().parents[1] / "app"
    sites: set[tuple[str, str]] = set()
    for path in app_dir.rglob("*.py"):
        tree = ast.parse(path.read_text(encoding="utf-8"))
        user_aliases: set[str] = set()
        model_aliases: set[str] = set()
        for node in ast.walk(tree):
            if isinstance(node, ast.ImportFrom) and node.module in ("app.models", "app.models.rbac"):
                user_aliases.update(
                    item.asname or item.name for item in node.names if item.name == "User")
            elif isinstance(node, ast.Import):
                model_aliases.update(
                    item.asname or item.name for item in node.names if item.name == "app.models")

        function_stack: list[str] = []

        class Visitor(ast.NodeVisitor):
            def _visit_function(self, node):
                function_stack.append(node.name)
                self.generic_visit(node)
                function_stack.pop()

            visit_FunctionDef = _visit_function
            visit_AsyncFunctionDef = _visit_function

            def visit_Call(self, node):
                direct = isinstance(node.func, ast.Name) and node.func.id in user_aliases
                qualified = (
                    isinstance(node.func, ast.Attribute)
                    and node.func.attr == "User"
                    and isinstance(node.func.value, ast.Name)
                    and node.func.value.id in model_aliases
                )
                if direct or qualified:
                    rel = path.relative_to(app_dir).as_posix()
                    sites.add((rel, function_stack[0] if function_stack else "<module>"))
                self.generic_visit(node)

        Visitor().visit(tree)
    return sites


def test_all_production_login_account_constructors_match_the_frozen_allowlist():
    # 师生批量开户、平台创建首位学校管理员、隔离演示沙箱种子是仅有的三类例外。
    assert _user_constructor_sites() == {
        ("services/school_onboarding_service.py", "run_onboarding"),
        ("services/platform_service.py", "create_school_admin"),
        ("services/sandbox_service.py", "seed_sandbox"),
    }


def test_identity_channel_can_only_be_opened_by_the_fixed_import_service():
    app_dir = Path(__file__).resolve().parents[1] / "app"
    callers = set()
    for path in app_dir.rglob("*.py"):
        tree = ast.parse(path.read_text(encoding="utf-8"))
        for node in ast.walk(tree):
            if not isinstance(node, ast.Call):
                continue
            keyword = next((kw for kw in node.keywords if kw.arg == "identity_channel"), None)
            if keyword is not None:
                callers.add((path.relative_to(app_dir).as_posix(), ast.dump(keyword.value)))
    assert callers == {
        ("services/identity_import_service.py", "Constant(value=True)"),
    }


def test_identity_import_api_has_no_raw_json_account_creation_bypass():
    root = Path(__file__).resolve().parents[2]
    system_api = root / "backend/app/api/v1/system.py"
    tree = ast.parse(system_api.read_text(encoding="utf-8"))
    routes = set()
    for node in ast.walk(tree):
        if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            continue
        for decorator in node.decorator_list:
            if not isinstance(decorator, ast.Call) or not isinstance(decorator.func, ast.Attribute):
                continue
            if not decorator.args or not isinstance(decorator.args[0], ast.Constant):
                continue
            route = decorator.args[0].value
            if isinstance(route, str) and route.startswith("/system/identity-import"):
                routes.add((decorator.func.attr.upper(), route))
    assert routes == {
        ("GET", "/system/identity-import/role-templates"),
        ("GET", "/system/identity-import/template"),
        ("POST", "/system/identity-import/validate-file"),
        ("POST", "/system/identity-import/confirm-batch"),
        ("GET", "/system/identity-import/batches/{batch_no}/errors"),
    }


def test_internship_mentor_scope_uses_user_id_and_blocks_same_name_teacher():
    # resolve_teacher_scope 现会把 _tid() 写进返回的 scope，纯单元测试须自备租户上下文。
    from app.core.context import set_tenant
    set_tenant({"tenantId": "1000000000000000001"})
    try:
        scope = mobile_teacher_service.resolve_teacher_scope({
            "userId": "db-17", "userType": "TEACHER",
            "currentRoleCode": "INTERN_MENTOR", "realName": "同名老师",
        })
    finally:
        set_tenant(None)
    assert scope["mode"] == "SCOPED"
    assert scope["advisorUserIds"] == {"17"}
    scope["classNames"].add("软件2401")
    assert mobile_teacher_service.scope_match_row(
        scope, advisor_user_id=17, advisor_name="同名老师")
    assert not mobile_teacher_service.scope_match_row(
        scope, advisor_user_id=18, advisor_name="同名老师", class_name="软件2401")
    # 仅对尚未回填 advisor_user_id 的历史数据保留姓名兼容。
    assert mobile_teacher_service.scope_match_row(scope, advisor_name="同名老师")


def test_non_advisor_role_cannot_inherit_advisor_relation_from_same_account():
    scope = mobile_teacher_service.resolve_teacher_scope({
        "userId": "db-17", "userType": "TEACHER",
        "currentRoleCode": "COUNSELOR", "realName": "兼任老师",
    })
    scope["classNames"].add("软件2401")
    assert not mobile_teacher_service.scope_match_row(
        scope, advisor_user_id=17, advisor_name="兼任老师", class_name="机电2401")
    assert mobile_teacher_service.scope_match_row(
        scope, advisor_user_id=18, advisor_name="其他老师", class_name="软件2401")


def test_academic_teacher_is_not_a_tenant_wide_administrator():
    scope = mobile_teacher_service.resolve_teacher_scope({
        "userId": "db-21", "userType": "TEACHER",
        "currentRoleCode": "ACADEMIC_TEACHER", "realName": "任课教师",
    })
    assert scope["mode"] == "SCOPED"
    assert scope["by"] == "DEFAULT_DENY"
    assert not mobile_teacher_service.scope_match_row(scope, student_no="S001")


def test_academic_admin_is_the_explicit_tenant_wide_academic_role():
    scope = mobile_teacher_service.resolve_teacher_scope({
        "userId": "db-22", "userType": "TEACHER",
        "currentRoleCode": "ACADEMIC_ADMIN", "realName": "教务处管理员",
    })
    assert scope["mode"] == "ADMIN_TENANT"


def test_internship_import_persists_advisor_account_id_not_only_display_name():
    root = Path(__file__).resolve().parents[2]
    source = (root / "backend/app/modules/internship/services/internship_student_service.py").read_text(
        encoding="utf-8")
    assert "advisor_user_id=advisor.id if advisor else None" in source
    assert "所选教师尚未分配“岗位实习指导教师”角色" in source


def test_frontend_exposes_batch_account_creation_only_at_fixed_system_route():
    root = Path(__file__).resolve().parents[2]
    dashboard = (root / "frontend/src/modules/system/views/SystemDashboardView.vue").read_text(
        encoding="utf-8")
    users = (root / "frontend/src/modules/system/views/SystemUserListView.vue").read_text(
        encoding="utf-8")
    routes = (root / "frontend/src/modules/system/system.routes.js").read_text(encoding="utf-8")
    assert "this.$router.push('/admin/system/identity-import')" in dashboard
    assert "this.$router.replace('/admin/system/identity-import')" in users
    assert "path: 'identity-import'" in routes
    assert "if (a.key === 'importUsers') return this.isIdentityImport" in users


def test_business_import_screens_show_the_non_account_creation_boundary():
    root = Path(__file__).resolve().parents[2]
    notice = (root / "frontend/src/components/common/AccountImportBoundaryNotice.vue").read_text(
        encoding="utf-8")
    assert "此处不会创建登录账号" in notice
    assert 'to="/admin/system/identity-import"' in notice

    direct_notice_files = [
        "frontend/src/views/admin/student/StudentImportExportView.vue",
        "frontend/src/modules/academicAffairs/components/ImportDrawer.vue",
        "frontend/src/modules/campusService/components/ImportDrawer.vue",
        "frontend/src/modules/orientation/components/ImportDialog.vue",
        "frontend/src/modules/employment/components/ImportDialog.vue",
    ]
    for relative in direct_notice_files:
        assert "AccountImportBoundaryNotice" in (root / relative).read_text(encoding="utf-8")

    guarded_excel_imports = [
        "frontend/src/modules/graduation/views/GraduationStudentListView.vue",
        "frontend/src/modules/graduation/views/GraduationMentorListView.vue",
        "frontend/src/modules/internship/views/InternshipStudentListView.vue",
        "frontend/src/modules/internship/views/InternshipMatchListView.vue",
    ]
    for relative in guarded_excel_imports:
        assert "show-account-boundary" in (root / relative).read_text(encoding="utf-8")


def _identity_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append(list(identity_files.HEADERS))
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_identity_template_is_empty_xlsx_with_roles_and_no_csv_path():
    content = identity_files.build_template()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    assert wb.sheetnames == ["导入模板", "填写说明", "填写示例（不要导入）", "教师预设角色", "业务关系", "业务关系说明"]
    assert wb["导入模板"].max_row == 1
    role_codes = {row[0] for row in wb["教师预设角色"].iter_rows(min_row=2, values_only=True)}
    assert "ACADEMIC_TEACHER" in role_codes
    assert "SYS_ADMIN" not in role_codes
    assert "SCHOOL_ADMIN" not in role_codes
    with pytest.raises(AppException) as exc:
        identity_files.parse_xlsx(b"accountType,accountNo", "teachers.csv")
    assert exc.value.code == "FILE_TYPE_NOT_ALLOWED"


def test_identity_xlsx_parses_student_and_teacher_with_original_excel_rows():
    content = _identity_xlsx([
        ["STUDENT", "001234", "张同学", "", "软件2601", "2026", "男", "", ""],
        ["TEACHER", "T001", "李老师", "辅导员，毕设导师", "", "", "", "CLASS", "软件2601"],
    ])
    parsed = identity_files.parse_xlsx(content, "师生账号.xlsx")
    assert parsed["totalRows"] == 2
    assert parsed["students"][0] == {
        "_rowNo": 2, "studentNo": "001234", "name": "张同学",
        "collegeName": "", "majorName": "",
        "className": "软件2601", "grade": "2026", "gender": "男",
    }
    assert parsed["teachers"][0]["_rowNo"] == 3
    assert parsed["teachers"][0]["roleCodes"] == "辅导员，毕设导师"
    assert parsed["errors"] == []


def test_identity_xlsx_reports_invalid_account_type_and_student_role():
    content = _identity_xlsx([
        ["PARENT", "P001", "家长", "", "", "", "", "", ""],
        ["STUDENT", "S001", "学生", "SYS_ADMIN", "", "", "", "", ""],
    ])
    parsed = identity_files.parse_xlsx(content, "师生账号.xlsx")
    assert [item["row"] for item in parsed["errors"]] == [2, 3]
    assert parsed["students"][0]["_rowNo"] == 3


def test_identity_xlsx_rejects_nonstandard_headers():
    wb = Workbook()
    ws = wb.active
    ws.append(["账号类型", "账号", "姓名"])
    ws.append(["STUDENT", "S001", "学生"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(AppException) as exc:
        identity_files.parse_xlsx(buf.getvalue(), "bad.xlsx")
    assert exc.value.code == "VALIDATION_ERROR"
    assert "最新版模板" in exc.value.message


def test_identity_xlsx_rejects_macro_payload_even_when_renamed_xlsx():
    content = _identity_xlsx([
        ["STUDENT", "S001", "学生", "", "", "", "", "", ""],
    ])
    buf = io.BytesIO(content)
    with zipfile.ZipFile(buf, "a") as archive:
        archive.writestr("xl/vbaProject.bin", b"macro")
    with pytest.raises(AppException) as exc:
        identity_files.parse_xlsx(buf.getvalue(), "renamed.xlsx")
    assert exc.value.code == "FILE_TYPE_NOT_ALLOWED"
    assert "宏" in exc.value.message


def test_identity_xlsx_preview_batch_is_bound_to_user_and_tenant(monkeypatch, db_mode):
    monkeypatch.setattr(onboarding, "db_enabled", lambda: False)
    parsed = identity_files.parse_xlsx(_identity_xlsx([
        ["STUDENT", "S001", "学生", "", "", "2026", "", "", ""],
    ]), "students.xlsx")
    user = {"userId": "u-1", "userType": "PLATFORM_SUPER_ADMIN"}
    report = identity_import_service.preview_identity_import(
        user, {"tenantId": "1001", "students": parsed["students"],
               "teachers": parsed["teachers"], "atomic": True},
        pre_errors=parsed["errors"])
    preview = identity_files.create_batch(user, parsed, report)
    assert preview["total"] == 1 and preview["invalid"] == 0
    assert identity_files.get_batch(user, "1001", preview["batchNo"], require_valid=True)
    with pytest.raises(AppException):
        identity_files.get_batch({**user, "userId": "u-2"}, "1001", preview["batchNo"])
    with pytest.raises(AppException):
        identity_files.get_batch(user, "2002", preview["batchNo"])

    # 模拟请求被负载均衡到另一实例（模块内状态全部重新加载），批次仍可读取和抢占。
    import importlib
    reloaded = importlib.reload(identity_files)
    restored = reloaded.get_batch(user, "1001", preview["batchNo"], require_valid=True)
    assert restored["fileSha256"] == parsed["fileSha256"]
    claimed, token, already_confirmed = reloaded.claim_batch(user, "1001", preview["batchNo"])
    assert claimed["status"] == "CONFIRMING" and token and not already_confirmed
    with pytest.raises(AppException) as exc:
        reloaded.claim_batch(user, "1001", preview["batchNo"])
    assert exc.value.code == "DATA_CONFLICT"
    reloaded.release_claim(user, "1001", preview["batchNo"], token, "模拟实例退出")
    assert reloaded.get_batch(user, "1001", preview["batchNo"])["status"] == "VALIDATED"


def test_global_preview_error_still_blocks_batch_confirmation(db_mode):
    parsed = {
        "students": [], "teachers": [], "rawRows": [{
            "row": 2, "accountType": "STUDENT", "accountNo": "S001", "name": "学生"}],
        "errors": [], "totalRows": 1, "fileName": "students.xlsx", "fileSha256": "abc",
    }
    user = {"userId": "u-global"}
    report = {
        "tenantId": "1001", "errors": [{"row": 0, "entity": "account", "field": "loginName",
                                           "error": "跨类型账号冲突"}],
        "entities": {}, "roleTemplateVersion": "test",
    }
    preview = identity_files.create_batch(user, parsed, report)
    assert preview["invalid"] == 1
    with pytest.raises(AppException):
        identity_files.get_batch(user, "1001", preview["batchNo"], require_valid=True)


def test_initial_passwords_are_packaged_as_one_time_xlsx_receipt():
    entry = {
        "batchNo": "IDIMP-TEST",
        "payload": {
            "students": [{"studentNo": "S001", "name": "张同学"}],
            "teachers": [{"loginName": "T001", "name": "李老师"}],
        },
    }
    receipt = identity_files.build_credential_receipt(entry, {
        "studentCredentials": [{"studentNo": "S001", "initialPassword": "Stu@secret"}],
        "teacherCredentials": [{"loginName": "T001", "initialPassword": "Tmp@secret"}],
    })
    assert receipt and receipt["rowCount"] == 2
    wb = load_workbook(io.BytesIO(base64.b64decode(receipt["contentBase64"])), read_only=True)
    rows = list(wb.active.iter_rows(min_row=3, values_only=True))
    assert rows[0][:4] == ("STUDENT", "S001", "张同学", "Stu@secret")
    assert rows[1][:4] == ("TEACHER", "T001", "李老师", "Tmp@secret")


def test_fixed_identity_import_frontend_accepts_xlsx_only():
    root = Path(__file__).resolve().parents[2]
    dialog = (root / "frontend/src/modules/system/components/ImportDialog.vue").read_text(
        encoding="utf-8")
    mock = (root / "frontend/src/mocks/system/system.mock.js").read_text(encoding="utf-8")
    assert 'accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"' in dialog
    assert "只支持标准 .xlsx" in dialog
    assert "不提供 CSV 导入" in mock
