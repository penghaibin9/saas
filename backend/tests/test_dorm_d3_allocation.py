"""D3 dorm allocation authority: dry-run, frozen pools, reservation and atomic self-select."""
from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

TID = 1000000000000000001
BASE = "/api/v1/student-affairs/dorm/allocation-batches"
MOBILE = "/api/v1/mobile/affairs/dorm"


def _admin(client):
    data = client.post("/api/v1/auth/mock-login", json={
        "loginName": "school_admin01", "password": "any",
    }).json()["data"]
    return {"Authorization": f"Bearer {data['accessToken']}"}


def _student_token(name: str, no: str):
    from app.core.security import create_access_token
    token = create_access_token({
        "userId": f"d3-{no}", "realName": name, "studentNo": no,
        "userType": "STUDENT", "tenantId": str(TID), "tid": str(TID),
        "activeContextId": f"d3-{no}", "currentRoleCode": "STUDENT", "clientType": "MP",
    })
    return {"Authorization": f"Bearer {token}"}


def _seed_authorities(*, students=2, beds=2, unlinked=0):
    from app.db.session import get_sessionmaker
    from app.models import (
        DormBed, DormBuilding, DormRoom, OrientationBatch, OrientationFlowStep,
        OrientationFlowVersion, OrientationStudent, OrientationStudentStep,
        SchoolClass, StudentProfile,
    )

    now = datetime.utcnow()
    db = get_sessionmaker()()
    school_class = SchoolClass(
        tenant_id=TID, major_id=1, class_name="D3软件2601", grade="2026",
        status="ACTIVE", class_status="NORMAL",
    )
    db.add(school_class); db.flush()
    profiles = []
    for index in range(students):
        profile = StudentProfile(
            tenant_id=TID, student_no=f"D3-2026-{index + 1:03d}",
            real_name=f"D3学生{index + 1}", class_id=school_class.id, gender="M",
            current_stage="ORIENTATION", student_status="NORMAL", status="ACTIVE",
        )
        db.add(profile); profiles.append(profile)
    db.flush()
    version = OrientationFlowVersion(
        tenant_id=TID, version_no=3003, version_name="D3住宿流程",
        status="PUBLISHED", source_type="MANUAL", published_at=now,
    )
    db.add(version); db.flush()
    step = OrientationFlowStep(
        tenant_id=TID, flow_version_id=version.id, step_key="DORM", step_name="宿舍安排",
        enabled=True, required=True, sort_order=1,
    )
    db.add(step); db.flush()
    ori_batch = OrientationBatch(
        tenant_id=TID, batch_name="2026级 D3 迎新", batch_no="ORI-D3-2026",
        year="2026", status="ACTIVE", planned_count=students + unlinked,
        flow_version_id=version.id,
    )
    db.add(ori_batch); db.flush()
    orientation_students = []
    for index, profile in enumerate(profiles):
        ori = OrientationStudent(
            tenant_id=TID, batch_id=ori_batch.id, student_id=profile.id,
            name=profile.real_name, admission_no=f"D3-ADMIT-{index + 1:03d}",
            source_type="IMPORT", source_record_id=f"D3-ADMIT-{index + 1:03d}",
            identity_status="LINKED", record_status="ACTIVE", dorm_status="UNASSIGNED",
            report_status="NOT_REPORTED", steps_json={"DORM": "TODO"},
        )
        db.add(ori); orientation_students.append(ori)
    db.flush()
    missing_orientation_students = []
    for index in range(unlinked):
        ori = OrientationStudent(
            tenant_id=TID, batch_id=ori_batch.id, student_id=None,
            name=f"D3待绑定学生{index + 1}", admission_no=f"D3-MISSING-{index + 1:03d}",
            source_type="IMPORT", source_record_id=f"D3-MISSING-{index + 1:03d}",
            identity_status="UNLINKED", record_status="ACTIVE", dorm_status="UNASSIGNED",
            report_status="NOT_REPORTED", steps_json={"DORM": "TODO"},
        )
        db.add(ori); missing_orientation_students.append(ori)
    db.flush()
    for ori in orientation_students + missing_orientation_students:
        db.add(OrientationStudentStep(
            tenant_id=TID, orientation_student_id=ori.id, flow_version_id=version.id,
            flow_step_id=step.id, step_key="DORM", status="NOT_STARTED",
            status_source="PROCESS_FACT", status_changed_at=now,
        ))
    building = DormBuilding(
        tenant_id=TID, building_name="D3紫荆1号楼", building_code="D3-ZJ-1",
        gender_limit="MALE", floor_count=1, status="ENABLED",
    )
    db.add(building); db.flush()
    room = DormRoom(
        tenant_id=TID, building_id=building.id, floor_no=1, room_no="101",
        capacity=beds, room_type="STANDARD", status="ENABLED",
    )
    db.add(room); db.flush()
    bed_rows = []
    for index in range(beds):
        bed = DormBed(
            tenant_id=TID, building_id=building.id, room_id=room.id,
            bed_no=str(index + 1), status="VACANT",
        )
        db.add(bed); bed_rows.append(bed)
    db.flush()
    result = {
        "students": [(row.id, row.real_name, row.student_no) for row in profiles],
        "orientationStudents": [row.id for row in orientation_students],
        "orientationBatchId": ori_batch.id, "buildingId": building.id,
        "roomId": room.id, "bedIds": [row.id for row in bed_rows],
    }
    db.commit(); db.close()
    return result


def _create(client, headers, seeded, mode, suffix):
    now = datetime.utcnow()
    response = client.post(BASE, headers=headers, json={
        "batchNo": f"D3-{suffix}", "name": f"D3 {suffix} 分配",
        "academicYear": "2026-2027", "sourceType": "ORIENTATION",
        "orientationBatchId": str(seeded["orientationBatchId"]), "mode": mode,
        "openAt": (now - timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S"),
        "closeAt": (now + timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S"),
        "rules": {"sameCollege": True, "minimizeVacancy": True},
        "resourceScope": {"buildingIds": [seeded["buildingId"]]}, "studentScope": {},
    })
    assert response.status_code == 200, response.text
    return response.json()["data"]["batchId"]


def test_d3_auto_requires_dry_run_and_publishes_reservations(client, db_mode):
    seeded = _seed_authorities(students=2, beds=3, unlinked=1)
    headers = _admin(client)
    batch_id = _create(client, headers, seeded, "ADMIN_AUTO", "AUTO")
    before = client.post(f"{BASE}/{batch_id}/publish", headers=headers)
    assert before.status_code == 400 and "Dry Run" in before.text

    dry = client.post(f"{BASE}/{batch_id}/dry-run", headers=headers)
    assert dry.status_code == 200
    summary = dry.json()["data"]["summary"]
    assert summary["proposed"] == 2 and summary["unassigned"] == 1
    assert summary["reasonCounts"]["DATA_MISSING"] == 1
    workbook = client.get(f"{BASE}/{batch_id}/conflicts.xlsx", headers=headers)
    assert workbook.status_code == 200
    sheet = load_workbook(BytesIO(workbook.content)).active
    assert sheet.title == "住宿分配异常"
    assert any(row[4].value == "DATA_MISSING" for row in sheet.iter_rows(min_row=2))

    published = client.post(f"{BASE}/{batch_id}/publish", headers=headers)
    assert published.status_code == 200
    assert published.json()["data"]["status"] == "PUBLISHED"

    from app.db.session import get_sessionmaker
    from app.models import DormAllocationItem, DormBed, DormStay, OrientationStudent, OrientationStudentStep
    db = get_sessionmaker()()
    items = db.query(DormAllocationItem).filter_by(allocation_batch_id=int(batch_id)).all()
    assert len(items) == 2 and {row.status for row in items} == {"RESERVED"}
    bed_rows = db.query(DormBed).filter(DormBed.id.in_(seeded["bedIds"])).all()
    assert sum(row.status == "LOCKED" for row in bed_rows) == 2
    assert sum(row.status == "VACANT" for row in bed_rows) == 1
    assert all(row.student_id is None for row in bed_rows)  # D3 不伪造 D4 正式入住
    stays = db.query(DormStay).filter_by(source_type="ALLOCATION").all()
    assert len(stays) == 2 and {row.status for row in stays} == {"RESERVED"}
    ori_rows = db.query(OrientationStudent).filter(OrientationStudent.id.in_(seeded["orientationStudents"])).all()
    assert {row.dorm_status for row in ori_rows} == {"ASSIGNED"}
    canonical = db.query(OrientationStudentStep).filter(
        OrientationStudentStep.orientation_student_id.in_(seeded["orientationStudents"]),
    ).all()
    assert {row.status for row in canonical} == {"DONE"}
    assert all(row.source_biz_id.startswith("dorm-allocation:") for row in canonical)
    db.close()

    # A later manual plan must recheck the canonical stay ledger, not rely on an earlier Dry Run.
    now = datetime.utcnow()
    later = client.post(BASE, headers=headers, json={
        "batchNo": "D3-MANUAL-RECHECK", "name": "D3 人工重复分配拦截",
        "academicYear": "2026-2027", "sourceType": "ORIENTATION",
        "orientationBatchId": str(seeded["orientationBatchId"]), "mode": "ADMIN_MANUAL",
        "openAt": (now + timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S"),
        "closeAt": (now + timedelta(days=4)).strftime("%Y-%m-%d %H:%M:%S"),
        "rules": {}, "resourceScope": {"buildingIds": [seeded["buildingId"]]},
        "studentScope": {},
    })
    assert later.status_code == 200, later.text
    later_id = later.json()["data"]["batchId"]
    proposal = client.post(f"{BASE}/{later_id}/manual-assign", headers=headers, json={
        "studentId": str(seeded["students"][0][0]), "bedId": str(seeded["bedIds"][2]),
    })
    assert proposal.status_code == 200, proposal.text
    duplicate = client.post(f"{BASE}/{later_id}/publish", headers=headers)
    assert duplicate.status_code == 409
    assert "已有生效或预留床位" in duplicate.text


def test_d3_student_select_same_bed_is_atomic_and_returns_one_conflict(client, db_mode):
    seeded = _seed_authorities(students=2, beds=1)
    headers = _admin(client)
    batch_id = _create(client, headers, seeded, "STUDENT_SELECT", "SELF")
    assert client.post(f"{BASE}/{batch_id}/publish", headers=headers).status_code == 200
    from app.db.session import get_sessionmaker
    from app.models import DormAllocationItem, DormBed, DormStay
    db = get_sessionmaker()()
    late_bed = DormBed(
        tenant_id=TID, building_id=seeded["buildingId"], room_id=seeded["roomId"],
        bed_no="LATE", status="VACANT",
    )
    db.add(late_bed); db.commit(); db.close()
    student_headers = [_student_token(name, no) for _sid, name, no in seeded["students"]]
    assert client.get(BASE, headers=student_headers[0]).status_code == 403
    for own_headers in student_headers:
        options = client.get(f"{MOBILE}/select-options", headers=own_headers)
        assert options.status_code == 200 and options.json()["data"]["canSelfSelect"] is True
        beds = client.get(f"{MOBILE}/rooms/{seeded['roomId']}/beds", headers=own_headers)
        assert {row["bedId"] for row in beds.json()["data"]["items"]} == {str(seeded["bedIds"][0])}

    from app.main import app
    def choose(own_headers):
        with TestClient(app) as own_client:
            result = own_client.post(
                f"{MOBILE}/beds/{seeded['bedIds'][0]}/self-select", headers=own_headers,
            )
            return result.status_code, result.json()

    with ThreadPoolExecutor(max_workers=2) as pool:
        results = list(pool.map(choose, student_headers))
    assert sorted(status for status, _body in results) == [200, 409]
    success = next(body["data"] for status, body in results if status == 200)
    assert success["status"] == "CONFIRMED"

    db = get_sessionmaker()()
    items = db.query(DormAllocationItem).filter_by(allocation_batch_id=int(batch_id)).all()
    assert sum(row.status == "CONFIRMED" for row in items) == 1
    assert sum(row.status == "PENDING" for row in items) == 1
    bed = db.get(DormBed, seeded["bedIds"][0])
    assert bed.status == "LOCKED" and bed.student_id is None
    assert db.query(DormStay).filter_by(status="RESERVED").count() == 1
    db.close()


def test_d3_post_checkin_publish_hides_location_until_report(client, db_mode):
    seeded = _seed_authorities(students=1, beds=1)
    headers = _admin(client)
    batch_id = _create(client, headers, seeded, "POST_CHECKIN_PUBLISH", "HIDDEN")
    assert client.post(f"{BASE}/{batch_id}/dry-run", headers=headers).status_code == 200
    assert client.post(f"{BASE}/{batch_id}/publish", headers=headers).status_code == 200
    _sid, name, no = seeded["students"][0]
    student_headers = _student_token(name, no)
    hidden = client.get(f"{MOBILE}/my", headers=student_headers).json()["data"]
    assert hidden["hasAllocation"] is True
    assert hidden["allocation"]["hiddenUntilCheckin"] is True
    assert "bedId" not in hidden["allocation"]

    from app.db.session import get_sessionmaker
    from app.models import OrientationStudent
    db = get_sessionmaker()()
    ori = db.get(OrientationStudent, seeded["orientationStudents"][0])
    ori.report_status = "CHECKED_IN"
    db.commit(); db.close()
    visible = client.get(f"{MOBILE}/my", headers=student_headers).json()["data"]
    assert visible["allocation"]["hiddenUntilCheckin"] is False
    assert visible["allocation"]["bedId"] == str(seeded["bedIds"][0])


def test_d3_migration_is_serial_and_preflights_runtime_conflicts():
    migration = (Path(__file__).parents[1] / "alembic" / "versions" /
                 "20260901_dorm_allocation_d3.py").read_text(encoding="utf-8")
    assert 'down_revision = "20260901_orientation_flow_o2"' in migration
    assert "duplicate bed proposals" in migration
    assert "published allocation batch lacks published_at" in migration
    assert "uk_dorm_alloc_item_bed" in migration
    assert "ck_dorm_alloc_batch_publish_time" in migration
    assert "D3 downgrade blocked" in migration
