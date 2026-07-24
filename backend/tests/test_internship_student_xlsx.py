"""P0-E · 实习学生 Excel(.xlsx) 导入导出（正式交付格式，替代 CSV）。

覆盖：模板下载 / 正常导入 / 学号缺失 / 学号不存在 / 企业不存在 / 日期格式错 / 错误行反馈 / 导出脱敏台账。
本轮：导入/导出必须带 batchId。
"""
from __future__ import annotations

import io
from uuid import uuid4

TID = 1000000000000000001
BASE = "/api/v1/internship/intern-students"


def _admin(client):
    d = client.post("/api/v1/auth/mock-login",
                    json={"loginName": "school_admin01", "password": "any"}).json()["data"]
    return {"Authorization": f"Bearer {d['accessToken']}"}


def _seed(db_mode):
    """一名学生（可建档）+ 一家企业（存在性校验用）。"""
    from app.db.session import get_sessionmaker
    from app.models import EmpCompany, StudentProfile
    db = get_sessionmaker()()
    try:
        s = StudentProfile(tenant_id=TID, student_no="XLSX-001", real_name="导入生",
                           current_stage="INTERNSHIP", student_status="NORMAL", status="ACTIVE")
        db.add(s)
        db.add(EmpCompany(tenant_id=TID, name="真实科技有限公司", coop_status="ACTIVE"))
        db.commit()
    finally:
        db.close()


def _mk_running_batch(client):
    h = _admin(client)
    b = client.post("/api/v1/internship/batches", headers=h, json={
        "batchName": f"xlsx批次-{uuid4().hex[:6]}", "batchNo": f"XLS-{uuid4().hex[:8]}",
        "startDate": "2026-03-01", "endDate": "2026-08-31", "plannedCount": 5,
    }).json()
    assert b["code"] == 0
    bid = b["data"]["id"]
    assert client.post(f"/api/v1/internship/batches/{bid}/activate", headers=h).json()["code"] == 0
    return bid


def _make_xlsx(rows: list[dict]) -> bytes:
    """按 15 列模板表头构造上传用 .xlsx（sheet 名 = 导入模板，与 read_xlsx 对齐）。"""
    from openpyxl import Workbook
    from app.modules.internship.services.internship_student_service import IMPORT_HEADERS, IMPORT_HEADER_MAP
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append(IMPORT_HEADERS)
    for r in rows:
        ws.append([r.get(IMPORT_HEADER_MAP[h], "") for h in IMPORT_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, rows, batch_id=None):
    content = _make_xlsx(rows)
    bid = batch_id or _mk_running_batch(client)
    return client.post(f"{BASE}/import/xlsx", headers=_admin(client),
                       params={"batchId": bid},
                       files={"file": ("import.xlsx", content,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})


def test_template_download(client, db_mode):
    r = client.get(f"{BASE}/import/template", headers=_admin(client))
    assert r.status_code == 200
    assert "spreadsheetml.sheet" in r.headers.get("content-type", "")
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert "导入模板" in wb.sheetnames
    headers = [c.value for c in wb["导入模板"][1]]
    # P0-6：模板仅保留真实可写入字段（学号必填；不含企业/岗位等未写入列）
    assert any(str(h).startswith("学号") for h in headers)
    assert "企业名称" not in headers and "岗位名称" not in headers
    assert len(headers) == 5


def test_import_valid(client, db_mode):
    _seed(db_mode)
    # 不填顾问姓名：顾问为选填，填了则必须匹配唯一在职指导教师账号
    body = _upload(client, [{"studentNo": "XLSX-001",
                             "startDate": "2026-03-02",
                             "endDate": "2026-08-28", "remark": "OK"}]).json()
    assert body["code"] == 0, body
    assert body["data"]["validRows"] == 1 and body["data"]["invalidRows"] == 0
    assert body["data"].get("batchId")


def test_import_missing_studentno(client, db_mode):
    _seed(db_mode)
    body = _upload(client, [{"advisorName": "刘强"}]).json()
    assert body["data"]["invalidRows"] == 1
    assert body["data"]["errors"][0]["field"] == "studentNo"


def test_import_student_not_found(client, db_mode):
    _seed(db_mode)
    body = _upload(client, [{"studentNo": "NOPE-999"}]).json()
    assert body["data"]["invalidRows"] == 1
    assert "未匹配到学生" in body["data"]["errors"][0]["message"]


def test_import_ignores_enterprise_column_not_in_template(client, db_mode):
    """P0-6：缩模板后企业字段不再校验；多余字段被忽略，学号有效即通过。"""
    _seed(db_mode)
    body = _upload(client, [{"studentNo": "XLSX-001", "enterpriseName": "不存在企业"}]).json()
    assert body["code"] == 0
    assert body["data"]["validRows"] == 1


def test_import_bad_date(client, db_mode):
    _seed(db_mode)
    body = _upload(client, [{"studentNo": "XLSX-001", "startDate": "2026年3月"}]).json()
    assert body["data"]["invalidRows"] == 1
    assert body["data"]["errors"][0]["field"] == "startDate"


def test_import_errors_xlsx_download(client, db_mode):
    _seed(db_mode)
    dry = _upload(client, [{"advisorName": "无学号"}]).json()["data"]
    r = client.post(f"{BASE}/import/errors-xlsx", headers=_admin(client),
                    json={"rows": [{"advisorName": "无学号"}], "errors": dry["errors"]})
    assert r.status_code == 200 and "spreadsheetml.sheet" in r.headers.get("content-type", "")


def test_export_xlsx(client, db_mode):
    _seed(db_mode)
    bid = _mk_running_batch(client)
    body = client.post(f"{BASE}/export", headers=_admin(client), params={"batchId": bid}).json()
    assert body["code"] == 0
    d = body["data"]
    assert d["filename"].endswith(".xlsx")
    assert "spreadsheetml.sheet" in d["mediaType"]
    assert "contentBase64" in d and d["contentBase64"]
    assert d.get("batchId") == str(bid)
