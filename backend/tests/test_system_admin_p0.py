"""系统管理中心 P0/P1 真实后端回归：登录/操作日志分流、账号停用启用/重置密码、
角色改名/停用、组织节点停用。直接调 API 端点函数（传 user dict 绕过 HTTP/Depends），
在 db_mode 真库上种最小数据后断言真实持久化与守卫。权限门禁另由 test_permissions* 覆盖。"""
from __future__ import annotations

from datetime import datetime

import pytest

from app.core.context import set_current_user, set_tenant
from app.core.exceptions import AppException

TID = 1000000000000000001
ADMIN = {"userId": "db-1", "realName": "测试管理员", "currentRoleCode": "SCHOOL_ADMIN",
         "tenantId": str(TID), "userType": "ADMIN"}


@pytest.fixture()
def seeded(db_mode):
    """在干净真库里种：1 管理员账号(id=1) + 1 目标教师账号(id=2) + 1 自定义角色 + 学院/班级/学生 + 审计若干。"""
    from app.services.db_service import session
    from app.core.security import hash_password
    from app.models import (College, Role, SchoolClass, SecurityAuditLog, StudentProfile, User, UserRole)
    set_tenant({"tenantId": TID})
    set_current_user(ADMIN)
    with session() as db:
        db.add(User(id=1, tenant_id=TID, login_name="admin01", real_name="测试管理员",
                    password_hash=hash_password("Init@123"), user_type="ADMIN", status="ACTIVE"))
        db.add(User(id=2, tenant_id=TID, login_name="t2020019", real_name="李敏",
                    password_hash=hash_password("Init@123"), user_type="TEACHER", status="ACTIVE"))
        role = Role(id=90, tenant_id=TID, role_code="CUSTOM_AAA", role_name="自定义辅导员",
                    role_type="CUSTOM", status="ACTIVE", remark="SAAS_CUSTOM;scope=COLLEGE;permMode=DB")
        role_sys = Role(id=91, tenant_id=TID, role_code="SCHOOL_ADMIN", role_name="学校管理员",
                        role_type="SYSTEM", status="ACTIVE", remark="")
        db.add(role); db.add(role_sys)
        college = College(id=10, tenant_id=TID, college_name="信息工程学院", status="ACTIVE")
        db.add(college)
        cls_empty = SchoolClass(id=20, tenant_id=TID, major_id=1, class_name="软件2301", status="ACTIVE")
        cls_full = SchoolClass(id=21, tenant_id=TID, major_id=1, class_name="软件2302", status="ACTIVE")
        db.add(cls_empty); db.add(cls_full)
        db.add(StudentProfile(id=500, tenant_id=TID, student_no="S001", real_name="学生甲",
                              class_id=21, college_id=10, current_stage="ENROLLED"))
        # 审计：2 条登录（含中文动作）+ 1 条登录失败 + 2 条业务操作
        for a, res, name in [("登录", "SUCCESS", "李敏"), ("登录", "SUCCESS", "测试管理员"),
                             ("LOGIN_FAIL", "FAIL", "未知")]:
            db.add(SecurityAuditLog(tenant_id=TID, action=a, operator_name=name, result=res,
                                    resource="auth", ip="10.1.2.3", created_at=datetime.utcnow()))
        for a in ("ROLE_PERMISSION_CONFIG", "EXPORT"):
            db.add(SecurityAuditLog(tenant_id=TID, action=a, operator_name="测试管理员", result="SUCCESS",
                                    resource=f"res:{a}", ip="10.1.2.3", created_at=datetime.utcnow()))
        db.commit()
    set_tenant({"tenantId": TID})
    set_current_user(ADMIN)
    yield


def test_login_operation_log_split(seeded):
    from app.api.v1 import system
    login = system.list_login_logs(user=ADMIN)["data"]
    op = system.list_operation_logs(user=ADMIN)["data"]
    login_actions_present = login["total"]
    # 登录日志含 3 条（2 登录 + 1 LOGIN_FAIL），不含业务操作
    assert login_actions_present == 3, login
    assert op["total"] == 2, op
    # 登录日志字段契约
    row = login["list"][0]
    for key in ("userName", "userNo", "roleName", "time", "result", "resultLabel", "ip", "device"):
        assert key in row
    # 登录失败被正确标记
    results = {r["result"] for r in login["list"]}
    assert "FAILED" in results and "SUCCESS" in results
    # 操作日志不含任何登录动作
    assert all(r["action"] not in ("登录", "LOGIN_FAIL") for r in op["list"])


def test_disable_enable_account(seeded):
    from app.api.v1 import system
    from app.services.db_service import session
    from app.models import User
    # 停用原因不足 5 字 → 拒绝
    with pytest.raises(AppException):
        system.set_system_user_status(2, {"action": "DISABLE", "reason": "离职"}, user=ADMIN)
    # 不能停用本人（admin db-1 → user_id 1）
    with pytest.raises(AppException):
        system.set_system_user_status(1, {"action": "DISABLE", "reason": "测试自我停用"}, user=ADMIN)
    # 正常停用
    res = system.set_system_user_status(2, {"action": "DISABLE", "reason": "离职交接完成"}, user=ADMIN)["data"]
    assert res["status"] == "DISABLED"
    with session() as db:
        assert db.get(User, 2).status == "DISABLED"
    # 启用
    res2 = system.set_system_user_status(2, {"action": "ENABLE"}, user=ADMIN)["data"]
    assert res2["status"] == "ACTIVE"
    with session() as db:
        assert db.get(User, 2).status == "ACTIVE"


def test_reset_password(seeded):
    from app.api.v1 import system
    from app.services.db_service import session
    from app.core.security import verify_password
    from app.models import User
    with session() as db:
        old_hash = db.get(User, 2).password_hash
    res = system.reset_system_user_password(2, user=ADMIN)["data"]
    assert res["tempPassword"] and res["mustChangePassword"] is True
    with session() as db:
        u = db.get(User, 2)
        assert u.password_hash != old_hash
        assert u.must_change_password is True
        assert verify_password(res["tempPassword"], u.password_hash)  # 临时密码可登录


def test_role_rename_and_disable(seeded):
    from app.api.v1 import system
    from app.services.db_service import session
    from app.models import Role, UserRole
    # 预设角色不可改名
    with pytest.raises(AppException):
        system.update_system_role(91, {"name": "改预设"}, user=ADMIN)
    # 自定义角色改名
    system.update_system_role(90, {"name": "自定义辅导员改"}, user=ADMIN)
    with session() as db:
        assert db.get(Role, 90).role_name == "自定义辅导员改"
    # 有成员时不可停用
    with session() as db:
        db.add(UserRole(tenant_id=TID, user_id=2, role_id=90, status="ACTIVE")); db.commit()
    with pytest.raises(AppException):
        system.set_system_role_status(90, {"action": "DISABLE", "reason": "不再使用该角色"}, user=ADMIN)
    # 改派掉成员后可停用
    with session() as db:
        link = db.query(UserRole).filter_by(role_id=90).first()
        link.status = "DISABLED"; link.is_deleted = True; db.commit()
    res = system.set_system_role_status(90, {"action": "DISABLE", "reason": "不再使用该角色"}, user=ADMIN)["data"]
    assert res["status"] == "DISABLED"


def test_org_node_disable_guards(seeded):
    from app.api.v1 import system
    from app.services.db_service import session
    from app.models import SchoolClass
    # 有在籍学生的班级不可停用
    with pytest.raises(AppException):
        system.set_system_org_node_status(21, {"type": "CLASS", "action": "DISABLE",
                                               "reason": "撤销该班级建制"}, user=ADMIN)
    # 空班级可停用
    res = system.set_system_org_node_status(20, {"type": "CLASS", "action": "DISABLE",
                                                 "reason": "撤销该班级建制"}, user=ADMIN)["data"]
    assert res["status"] == "DISABLED"
    with session() as db:
        assert db.get(SchoolClass, 20).status == "DISABLED"


def test_system_config_effective(seeded):
    """系统配置真实生效：保存后强制层 get_int 读到新值；越界拒绝。"""
    from app.api.v1 import system
    from app.services import system_config_service as cfg
    assert cfg.get_int("SEC_LOCK_MAX_FAIL", 5) == 5  # 默认
    system.save_system_config("SEC_LOCK_MAX_FAIL", {"valueText": "7", "reason": "收紧登录策略"}, user=ADMIN)
    assert cfg.get_int("SEC_LOCK_MAX_FAIL", 5) == 7  # 登录失败锁定逻辑真实读取此值
    with pytest.raises(AppException):  # 越界拒绝
        system.save_system_config("SEC_LOCK_MAX_FAIL", {"valueText": "99", "reason": "非法"}, user=ADMIN)
    with pytest.raises(AppException):  # 未知键拒绝
        system.save_system_config("UNKNOWN_KEY", {"valueText": "1", "reason": "x"}, user=ADMIN)
    items = system.list_system_configs(user=ADMIN)["data"]["list"]
    assert any(i["key"] == "SEC_LOCK_MAX_FAIL" and i["valueText"] == "7" for i in items)


def test_brand_edit_takes_effect(seeded):
    """品牌真实生效：保存后 get_brand（顶栏/登录页读取）反映新主色。"""
    from app.api.v1 import system
    from app.services import mock_tenant_service
    system.save_system_brand({"brandColor": "#FF6600", "watermarkText": "测试水印",
                              "loginSlogan": "知行合一", "reason": "品牌统一"}, user=ADMIN)
    form = system.get_system_brand(user=ADMIN)["data"]
    assert form["brandColor"] == "#FF6600" and form["watermarkText"] == "测试水印"
    brand = mock_tenant_service.get_brand()  # 顶栏/登录页真实数据源
    assert brand.get("primaryColor") == "#FF6600"
    # 非法主色拒绝
    with pytest.raises(AppException):
        system.save_system_brand({"brandColor": "红色"}, user=ADMIN)
    reset = system.reset_system_brand({"reason": "恢复平台默认品牌"}, user=ADMIN)["data"]
    assert reset["brandColor"] == "#2563EB"
    assert reset["watermarkText"] == ""


def test_export_users_real_xlsx(seeded):
    """导出账号台账：真实 xlsx（可被 openpyxl 打开），含种子里的2个账号数据行。"""
    import asyncio, io
    from openpyxl import load_workbook
    from app.api.v1 import system

    async def _body(resp):
        chunks = []
        async for c in resp.body_iterator:
            chunks.append(c if isinstance(c, bytes) else c.encode())
        return b"".join(chunks)

    resp = system.export_system_users(user=ADMIN)
    data = asyncio.run(_body(resp))
    ws = load_workbook(io.BytesIO(data)).active
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert "工号/学号" in flat  # 表头
    assert any("admin01" in s or "t2020019" in s for s in flat)  # 真实账号数据行


def test_export_role_config_and_configs(seeded):
    """导出角色配置(JSON,真实权限点)+系统配置快照(JSON)+范围规则清单(xlsx)。"""
    import asyncio, io, json
    from openpyxl import load_workbook
    from app.api.v1 import system

    async def _body(resp):
        chunks = []
        async for c in resp.body_iterator:
            chunks.append(c if isinstance(c, bytes) else c.encode())
        return b"".join(chunks)

    # 角色配置 JSON（种子自定义角色 id=90）
    rc = json.loads(asyncio.run(_body(system.export_role_config(90, user=ADMIN))))
    assert rc["roleCode"] == "CUSTOM_AAA" and "permissions" in rc
    # 系统配置快照 JSON
    cfg = json.loads(asyncio.run(_body(system.export_configs(user=ADMIN))))
    assert "systemConfigs" in cfg and "brand" in cfg
    # 范围规则清单 xlsx（先建一条规则确保有数据行）
    system.save_scope_rule({"name": "本学院范围", "scopeCode": "COLLEGE", "remark": "x"}, user=ADMIN)
    ws = load_workbook(io.BytesIO(asyncio.run(_body(system.export_scope_rules(user=ADMIN))))).active
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert "规则名称" in flat and any("本学院范围" in s for s in flat)


def test_scope_rule_catalog(seeded):
    """数据范围规则真实可编辑目录：引用角色/影响用户由后端按角色 scopeCode 真实计算；被引用不可作废。"""
    from app.api.v1 import system
    # 种子里 role id=90 remark 含 ;scope=COLLEGE → COLLEGE 规则应显示被该角色引用
    rule = system.save_scope_rule({"name": "本学院范围", "scopeCode": "COLLEGE", "remark": "跨院不可见"}, user=ADMIN)["data"]
    assert rule["scopeLabel"] == "本学院"
    assert "自定义辅导员" in rule["appliedRoles"]  # 真实引用角色
    # 被角色引用时不可作废
    with pytest.raises(AppException):
        system.set_scope_rule_status(int(rule["id"]), {"action": "DISABLE", "reason": "不再使用该范围"}, user=ADMIN)


def test_user_detail_update_and_batch_disable(seeded):
    """账号详情/编辑/批量停用走真库，禁止假成功。"""
    from app.api.v1 import system
    from app.services.db_service import session
    from app.models import User
    detail = system.get_system_user(2, user=ADMIN)["data"]
    assert detail["userNo"] == "t2020019" and detail["id"] == "2"
    updated = system.update_system_user(2, {"name": "李敏改", "phone": "13800138000"}, user=ADMIN)["data"]
    assert updated["name"] == "李敏改"
    with session() as db:
        assert db.get(User, 2).real_name == "李敏改"
        from app.core.field_crypto import decrypt_field
        stored_phone = db.get(User, 2).phone_encrypted
        assert stored_phone != "13800138000"
        assert decrypt_field(stored_phone) == "13800138000"
    batch = system.batch_set_system_user_status(
        {"action": "DISABLE", "ids": [2], "reason": "批量停用验证用例"}, user=ADMIN)["data"]
    assert batch["count"] == 1
    with session() as db:
        assert db.get(User, 2).status == "DISABLED"


def test_staff_student_account_boundaries(seeded):
    """正式账号页按类型分流；学生角色、主档姓名和混合批量操作均由后端拒绝。"""
    from app.api.v1 import system
    from app.core.security import hash_password
    from app.models import Role, StudentAccountLink, User, UserRole
    from app.services.db_service import session

    with session() as db:
        student = User(id=3, tenant_id=TID, login_name="S001", real_name="学生甲",
                       password_hash=hash_password("Init@123"), user_type="STUDENT", status="ACTIVE")
        legacy_student = User(
            id=4, tenant_id=TID, login_name="LEGACY_S002", real_name="历史学生账号",
            password_hash=hash_password("Init@123"), user_type="TEACHER", status="ACTIVE",
        )
        student_role = Role(id=92, tenant_id=TID, role_code="STUDENT", role_name="学生",
                            role_type="SYSTEM", status="ACTIVE", remark="")
        db.add(student); db.add(legacy_student); db.add(student_role); db.flush()
        db.add(UserRole(tenant_id=TID, user_id=3, role_id=92, status="ACTIVE"))
        db.add(UserRole(tenant_id=TID, user_id=4, role_id=92, status="ACTIVE"))
        db.add(StudentAccountLink(tenant_id=TID, student_id=500, user_id=3,
                                  link_status="ACTIVE", source="BACKFILL",
                                  bound_login_name="S001", bound_student_no="S001"))
        db.commit()

    staff = system.list_system_users(account_type="STAFF", user=ADMIN)["data"]
    students = system.list_system_users(account_type="STUDENT", user=ADMIN)["data"]
    assert {row["id"] for row in staff["list"]} == {"1", "2"}
    assert {row["id"] for row in students["list"]} == {"3", "4"}
    bound = next(row for row in students["list"] if row["id"] == "3")
    assert bound["profileBound"] is True
    assert bound["className"] == "软件2302"

    with pytest.raises(AppException):
        system.assign_system_user_roles(3, {"roleCodes": ["CUSTOM_AAA"]}, user=ADMIN)
    with pytest.raises(AppException):
        system.assign_system_user_roles(4, {"roleCodes": ["CUSTOM_AAA"]}, user=ADMIN)
    with pytest.raises(AppException):
        system.update_system_user(3, {"name": "改错姓名", "phone": ""}, user=ADMIN)
    with pytest.raises(AppException):
        system.batch_set_system_user_status(
            {"action": "DISABLE", "ids": [2, 3], "reason": "混合类型批量停用",
             "accountType": "STAFF"},
            user=ADMIN,
        )

    scoped = system.batch_set_system_user_status(
        {"action": "DISABLE", "scope": "CLASS", "filters": {"classId": "21"},
         "reason": "软件2302班毕业停用", "accountType": "STUDENT"},
        user=ADMIN,
    )["data"]
    assert scoped["count"] == 1 and scoped["scope"] == "CLASS"
    with session() as db:
        assert db.get(User, 3).status == "DISABLED"
        assert db.get(User, 4).status == "ACTIVE"  # 无稳定主档绑定，不属于班级范围

    with pytest.raises(AppException):
        system.batch_set_system_user_status(
            {"action": "DISABLE", "scope": "SCHOOL", "reason": "全校学生毕业批量停用",
             "accountType": "STUDENT"},
            user=ADMIN,
        )
    school = system.batch_set_system_user_status(
        {"action": "DISABLE", "scope": "SCHOOL", "reason": "全校学生毕业批量停用",
         "accountType": "STUDENT", "confirmSchoolScope": True},
        user=ADMIN,
    )["data"]
    assert school["count"] == 1 and school["scope"] == "SCHOOL"
    with session() as db:
        assert db.get(User, 4).status == "DISABLED"


def test_permission_tree_and_role_save_merge(seeded):
    """权限树为真实 permissionCode；保存只替换可见码，页外码保留。"""
    from app.api.v1 import system
    from app.services.db_service import session
    from app.models import Permission, RolePermission
    from sqlalchemy import select
    tree_res = system.get_permission_tree(user=ADMIN)["data"]
    tree = tree_res["tree"]
    visible = set(tree_res["visibleCodes"])
    assert tree and any(m.get("key", "").startswith("systemAdmin.") for g in tree for m in g.get("children") or [])
    outside = "employment.definitely.outside.tree.code"
    with session() as db:
        p = db.scalars(select(Permission).where(Permission.permission_code == outside)).first()
        if p is None:
            p = Permission(permission_code=outside, permission_name=outside, module_code="employment", action="view")
            db.add(p); db.flush()
        db.add(RolePermission(tenant_id=TID, role_id=90, permission_id=p.id, status="ACTIVE", is_deleted=False))
        db.commit()
    assert outside not in visible
    submit = sorted(c for c in visible if c.startswith("systemAdmin.user."))[:3] or sorted(visible)[:3]
    saved = system.save_system_role_permissions(
        90, {"permissionCodes": submit, "visiblePermissionCodes": sorted(visible), "scopeCode": "COLLEGE"},
        user=ADMIN)["data"]
    codes = set(saved["permissionCodes"])
    assert outside in codes
    assert set(submit).issubset(codes)


def test_copy_school_admin_expands_star(seeded):
    """复制 SCHOOL_ADMIN 不得落库字面量 *。"""
    from app.api.v1 import system
    from app.services.db_service import session
    from app.models import Permission, RolePermission
    from sqlalchemy import select
    copied = system.copy_system_role(91, user=ADMIN)["data"]
    role_id = int(copied["id"])
    with session() as db:
        codes = set(db.scalars(select(Permission.permission_code).join(
            RolePermission, RolePermission.permission_id == Permission.id).where(
            RolePermission.role_id == role_id, RolePermission.status == "ACTIVE")).all())
    assert "*" not in codes
    assert codes


def test_system_info_requires_auth_and_real_caps(seeded):
    from app.api.v1 import system
    info = system.system_info(user=ADMIN)["data"]
    assert info["capabilities"]["auth"] == "real"
    assert info["capabilities"]["rbac"] == "real"


def test_governance_delegation_and_integration(seeded):
    from app.api.v1 import system
    from datetime import datetime, timedelta
    expires = (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S")
    row = system.api_create_delegation({
        "granteeUserNo": "t2020019", "roleCode": "COUNSELOR",
        "expiresAt": expires, "reason": "临时顶岗一周"
    }, user=ADMIN)["data"]
    assert row["status"] == "ACTIVE"
    listed = system.api_list_delegations(user=ADMIN)["data"]["list"]
    assert any(x["id"] == row["id"] for x in listed)
    system.api_revoke_delegation(row["id"], {"reason": "提前结束顶岗"}, user=ADMIN)
    integ = system.api_save_integration({
        "name": "教务同步", "endpoint": "https://example.edu/api", "credential": "secret-token-01"
    }, user=ADMIN)["data"]
    assert integ["hasCredential"] is True
    job = system.api_enqueue_sync_job({"name": "全量同步", "integrationId": integ["id"], "forceFail": True},
                                     user=ADMIN)["data"]
    assert job["status"] == "FAILED"
    retried = system.api_retry_sync_job(job["id"], user=ADMIN)["data"]
    assert retried["status"] == "PENDING"
    assert retried["status"] != "SUCCESS"  # 没有真实 adapter/worker 时禁止伪造成功


def test_account_exceptions_and_scope_users(seeded):
    from app.api.v1 import system
    system.set_system_user_status(2, {"action": "DISABLE", "reason": "异常中心验证用例"}, user=ADMIN)
    exc = system.list_account_exceptions(user=ADMIN)["data"]
    assert exc["total"] >= 1
    assert any(r["userNo"] == "t2020019" for r in exc["list"])
    rule = system.save_scope_rule({"name": "影响用户规则", "scopeCode": "COLLEGE", "remark": "x"}, user=ADMIN)["data"]
    users = system.list_scope_affected_users(int(rule["id"]), user=ADMIN)["data"]
    assert isinstance(users, list)
