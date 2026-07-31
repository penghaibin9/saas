from __future__ import annotations

import hashlib
import io
import json
import os
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select

TENANT_ID = 1000000000000000001
OTHER_TENANT_ID = 1000000000000000002


def _configure() -> None:
    os.environ.setdefault("DB_ENABLED", "true")
    os.environ.setdefault("DB_DRIVER", "mysql")
    os.environ.setdefault(
        "DATABASE_URL",
        "mysql+pymysql://root:root@127.0.0.1:3306/student_lifecycle_test?charset=utf8mb4",
    )
    temp_root = Path(tempfile.gettempdir()) / "phase6-graduation-material-center"
    temp_root.mkdir(parents=True, exist_ok=True)
    os.environ["UPLOAD_DIR"] = str(temp_root)
    os.environ["FILE_STORAGE_BACKEND"] = "local"


_configure()

from app.core.context import set_current_user, set_tenant
from app.core.exceptions import AppException
from app.db.session import get_sessionmaker
from app.models import (
    ArchiveManifest,
    ArchiveManifestItem,
    ExportJob,
    FileAsset,
    FileBinding,
    FileObject,
    FileVersion,
    GraduationArchiveRecord,
    GraduationBatch,
    GraduationDefenseGroup,
    GraduationDefenseScore,
    GraduationFinal,
    GraduationGuidance,
    GraduationMaterialBackfillCheckpoint,
    GraduationMaterialRule,
    GraduationMidterm,
    GraduationPlagiarismCheck,
    GraduationProposal,
    GraduationReview,
    GraduationStudent,
    GraduationStudentMaterial,
    GraduationTaskBook,
    GraduationTemplate,
    GraduationTemplateAssetPolicy,
    GraduationTopic,
    StudentProfile,
)
from app.modules.graduation.services import (
    graduation_material_catalog_service as catalog,
    graduation_material_center_service as center,
    graduation_material_export_service as export_service,
    graduation_material_ticket_service as tickets,
    graduation_structured_snapshot_service as structured_snapshots,
)
from app.services.storage import get_backend


STUDENT_NO = "GD-PHASE6-001"
LEGACY_STUDENT_NO = "GD-PHASE6-LEGACY"


def _admin_user() -> dict:
    return {
        "tenantId": str(TENANT_ID),
        "userId": "6201",
        "realName": "阶段六管理员",
        "userType": "TEACHER",
        "currentRoleCode": "GRADUATION_ADMIN",
        "permissions": ["*"],
        "dataScope": "ALL",
        "graduationDataScope": "ALL",
    }


def _teacher_user() -> dict:
    return {
        "tenantId": str(TENANT_ID),
        "userId": "6202",
        "realName": "阶段六指导教师",
        "userType": "TEACHER",
        "currentRoleCode": "GD_MENTOR",
        "permissions": [
            "graduationDesign.view",
            "graduationDesign.proposal.view",
            "graduationDesign.proposal.review",
            "graduationDesign.final.view",
            "graduationDesign.final.review",
            "graduationDesign.archive.view",
            "graduationDesign.archive.preview",
            "graduationDesign.archive.file",
            "graduationDesign.archive.export",
            "graduationDesign.template.view",
        ],
        "dataScope": "ALL",
        "graduationDataScope": "ALL",
    }


def _student_user(student_id: int, batch_id: int) -> dict:
    return {
        "tenantId": str(TENANT_ID),
        "userId": "6301",
        "studentId": str(student_id),
        "studentNo": STUDENT_NO,
        "realName": "阶段六学生",
        "userType": "STUDENT",
        "currentRoleCode": "STUDENT",
        "graduationBatchId": str(batch_id),
    }


def _write_storage(key: str, body: bytes) -> tuple[str, int]:
    backend = get_backend()
    target = backend.staging_path(key)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(body)
    backend.promote(key)
    return hashlib.sha256(body).hexdigest(), len(body)


def _real_file(
    db,
    *,
    name: str,
    body: bytes,
    tenant_id: int = TENANT_ID,
    owner: int = 6301,
    status: str = "AVAILABLE",
    scan_status: str = "CLEAN",
    storage_zone: str = "ACTIVE",
) -> FileObject:
    key = f"phase6/{tenant_id}/{hashlib.sha256(name.encode('utf-8')).hexdigest()[:12]}-{name}"
    digest, size = _write_storage(key, body)
    row = FileObject(
        tenant_id=tenant_id,
        file_key=key,
        file_name=name,
        ext=name.rsplit(".", 1)[-1].lower() if "." in name else "",
        mime_type="application/pdf" if name.endswith(".pdf") else "application/zip",
        size_bytes=size,
        sha256=digest,
        biz_type="GRADUATION_MATERIAL",
        biz_id=STUDENT_NO,
        owner_user_id=owner,
        created_by=owner,
        visibility="BIZ_SCOPED",
        security_level="SENSITIVE",
        status=status,
        storage_backend="local",
        storage_zone=storage_zone,
        scan_required=scan_status != "NOT_REQUIRED",
        scan_status=scan_status,
        available_at=datetime.utcnow() if status == "AVAILABLE" else None,
    )
    db.add(row)
    db.flush()
    return row


def _read_file(file_id: int) -> tuple[FileObject, bytes]:
    db = get_sessionmaker()()
    try:
        row = db.get(FileObject, int(file_id))
        assert row is not None
        data = get_backend().open(row.file_key).read()
        db.expunge(row)
        return row, data
    finally:
        db.close()


def _expect_blocked(fn, *, codes: set[str] | None = None) -> None:
    try:
        fn()
    except AppException as exc:
        if codes:
            assert exc.code in codes, (exc.code, str(exc))
    else:
        raise AssertionError("operation unexpectedly passed the Stage 6 safety gate")


def _expect_not_found(fn) -> None:
    try:
        fn()
    except AppException as exc:
        assert exc.code in {"DATA_NOT_FOUND", "NO_PERMISSION"}, (exc.code, str(exc))
        assert getattr(exc, "http_status", 404) in {403, 404}
    else:
        raise AssertionError("enumeration attempt unexpectedly succeeded")


def _zip_payload(file_id: int):
    row, data = _read_file(file_id)
    with zipfile.ZipFile(io.BytesIO(data), "r") as archive:
        names = archive.namelist()
        entries = {name: archive.read(name) for name in names}
        manifest = json.loads(entries["manifest.json"].decode("utf-8"))
    return row, names, entries, manifest


def _seed() -> dict:
    db = get_sessionmaker()()
    try:
        suffix = datetime.utcnow().strftime("%H%M%S%f")
        batch = GraduationBatch(
            tenant_id=TENANT_ID,
            batch_name=f"阶段六材料中心验收-{suffix}",
            batch_no=f"GD6-{suffix}",
            academic_year="2026",
            status="ACTIVE",
            stage_config=[],
        )
        db.add(batch)
        db.flush()
        topic = GraduationTopic(
            tenant_id=TENANT_ID,
            batch_id=batch.id,
            title="阶段六公共文件中心毕业设计",
            source_type="TEACHER",
            advisor_name="阶段六指导教师",
            capacity=2,
            status="APPROVED",
        )
        db.add(topic)
        db.flush()
        profile = StudentProfile(
            tenant_id=TENANT_ID,
            student_no=STUDENT_NO,
            real_name="阶段六学生",
            current_stage="GRADUATION",
            student_status="NORMAL",
            status="ACTIVE",
        )
        legacy_profile = StudentProfile(
            tenant_id=TENANT_ID,
            student_no=LEGACY_STUDENT_NO,
            real_name="阶段六旧数据学生",
            current_stage="GRADUATION",
            student_status="NORMAL",
            status="ACTIVE",
        )
        db.add_all([profile, legacy_profile])
        db.flush()
        student = GraduationStudent(
            tenant_id=TENANT_ID,
            batch_id=batch.id,
            student_id=profile.id,
            student_no=STUDENT_NO,
            name="阶段六学生",
            class_name="软件2401",
            college_id="10",
            major_id="100",
            class_id="1001",
            advisor_name="阶段六指导教师",
            topic_id=topic.id,
            topic_title=topic.title,
            stage="TASKBOOK_CONFIRM",
            eligibility_status="QUALIFIED",
            risk_level="NONE",
            record_status="ACTIVE",
        )
        legacy_student = GraduationStudent(
            tenant_id=TENANT_ID,
            batch_id=batch.id,
            student_id=legacy_profile.id,
            student_no=LEGACY_STUDENT_NO,
            name="阶段六旧数据学生",
            class_name="软件2401",
            college_id="10",
            major_id="100",
            class_id="1001",
            advisor_name="阶段六指导教师",
            topic_id=topic.id,
            topic_title=topic.title,
            stage="FINAL_CHECK",
            eligibility_status="QUALIFIED",
            risk_level="NONE",
            record_status="ACTIVE",
        )
        db.add_all([student, legacy_student])
        db.flush()
        db.add_all([
            GraduationTaskBook(
                tenant_id=TENANT_ID,
                gd_student_id=student.id,
                taskbook_version=1,
                status="CONFIRMED",
                objective="完成公共文件中心版本验收",
                content="完成开题、初稿、定稿、作品、源代码和归档闭环",
                history_json=[],
                confirmed_at=datetime.utcnow(),
            ),
            GraduationGuidance(
                tenant_id=TENANT_ID,
                gd_student_id=student.id,
                guidance_date=datetime.utcnow(),
                content="阶段六指导记录",
                issues="无",
                next_plan="完成归档",
                hours=2,
                attachments_json=[],
                advisor_name="阶段六指导教师",
            ),
            GraduationMidterm(
                tenant_id=TENANT_ID,
                gd_student_id=student.id,
                status="CHECKED_PASS",
                progress="80%",
                issues="无",
                next_plan="完成定稿",
                check_comment="通过",
                conclusion="PASS",
                checked_at=datetime.utcnow(),
            ),
            GraduationArchiveRecord(
                tenant_id=TENANT_ID,
                gd_student_id=student.id,
                status="SUBMITTED",
                checklist_json=[],
                history_json=[],
            ),
        ])
        template = GraduationTemplate(
            tenant_id=TENANT_ID,
            template_type="PROPOSAL",
            template_name="阶段六开题模板",
            version="v1",
            status="ACTIVE",
            is_default=True,
            file_url="",
            variables_json=["studentName", "topicTitle"],
            description="公共模板版本验收",
        )
        db.add(template)
        db.commit()
        return {
            "batchId": int(batch.id),
            "studentId": int(student.id),
            "legacyStudentId": int(legacy_student.id),
            "topicId": int(topic.id),
            "templateId": int(template.id),
        }
    finally:
        db.close()


def main() -> None:
    set_tenant({"tenantId": str(TENANT_ID)})
    seeded = _seed()
    batch_id = seeded["batchId"]
    student_id = seeded["studentId"]
    legacy_student_id = seeded["legacyStudentId"]
    template_id = seeded["templateId"]
    admin_user = _admin_user()
    teacher_user = _teacher_user()
    student_user = _student_user(student_id, batch_id)

    set_current_user(admin_user)
    rule = catalog.ensure_rules(batch_id, admin_user)
    assert rule["itemCount"] == 18
    rule_codes = {item["materialCode"] for item in rule["items"]}
    assert len(rule_codes) == 18
    assert rule_codes == set(catalog.SPEC_BY_CODE)

    set_current_user(student_user)
    proposal_v1 = center.submit_proposal(student_user, {
        "background": "阶段六开题背景 v1",
        "plan": "阶段六计划 v1",
        "outcome": "阶段六预期成果 v1",
        "attachments": [],
    })
    p1_id = int(proposal_v1["id"])
    p1_versions = proposal_v1["currentSafeVersions"]
    assert len(p1_versions) == 1
    assert p1_versions[0]["materialCode"] == "PROPOSAL_SNAPSHOT"
    assert p1_versions[0]["scanStatus"] == "NOT_REQUIRED"

    set_current_user(teacher_user)
    detail_v1 = center.proposal_detail(p1_id)
    assert detail_v1["reviewReady"] is True
    rejected = center.review_proposal(
        p1_id,
        "REJECT",
        "研究方案缺少版本迁移和归档证据，请补充后重交",
        teacher_user,
    )
    assert rejected["status"] == "REJECTED"

    set_current_user(student_user)
    proposal_v2 = center.submit_proposal(student_user, {
        "background": "阶段六开题背景 v2",
        "plan": "阶段六计划 v2，补充 FileVersion 与 Manifest",
        "outcome": "完成可追溯归档包",
        "attachments": [],
    })
    p2_id = int(proposal_v2["id"])
    assert proposal_v2["version"] == "v2"

    set_current_user(teacher_user)
    detail_v2 = center.proposal_detail(p2_id)
    assert detail_v2["reviewReady"] is True
    approved = center.review_proposal(p2_id, "APPROVE", "开题材料完整，同意通过", teacher_user)
    assert approved["status"] == "APPROVED"

    db = get_sessionmaker()()
    try:
        old_versions = db.scalars(select(FileVersion).join(
            FileBinding, FileBinding.version_id == FileVersion.id,
        ).where(
            FileBinding.tenant_id == TENANT_ID,
            FileBinding.biz_type == "GRADUATION_MATERIAL",
            FileBinding.biz_id == str(p1_id),
            FileBinding.is_deleted.is_(False),
        )).all()
        assert old_versions
        assert all(not row.is_current and row.status == "INVALIDATED" for row in old_versions)
        current_versions = db.scalars(select(FileVersion).join(
            FileBinding, FileBinding.version_id == FileVersion.id,
        ).where(
            FileBinding.tenant_id == TENANT_ID,
            FileBinding.biz_type == "GRADUATION_MATERIAL",
            FileBinding.biz_id == str(p2_id),
            FileBinding.is_current.is_(True),
            FileBinding.is_deleted.is_(False),
        )).all()
        assert current_versions and all(row.is_current and row.status == "APPROVED" for row in current_versions)

        draft_file = _real_file(
            db,
            name="阶段六毕业设计初稿.pdf",
            body=b"%PDF-1.4\nphase6 graduation draft\n%%EOF\n",
        )
        final_file = _real_file(
            db,
            name="阶段六毕业设计定稿.pdf",
            body=b"%PDF-1.4\nphase6 graduation final approved\n%%EOF\n",
        )
        legacy_file = _real_file(
            db,
            name="旧开题附件.pdf",
            body=b"%PDF-1.4\nlegacy graduation material\n%%EOF\n",
            owner=6301,
        )
        pending_file = _real_file(
            db,
            name="扫描中设计作品.pdf",
            body=b"%PDF-1.4\npending design work\n%%EOF\n",
            status="QUARANTINED",
            scan_status="PENDING",
            storage_zone="QUARANTINE",
        )
        infected_file = _real_file(
            db,
            name="感染源代码.zip",
            body=b"PK\x03\x04phase6 infected fixture metadata only",
            status="REJECTED",
            scan_status="INFECTED",
            storage_zone="QUARANTINE",
        )
        clean_design = _real_file(
            db,
            name="安全设计作品.pdf",
            body=b"%PDF-1.4\nclean design work\n%%EOF\n",
        )
        clean_source = _real_file(
            db,
            name="安全源代码.zip",
            body=b"PK\x03\x04phase6 clean source archive",
        )
        cross_tenant_file = _real_file(
            db,
            name="跨租户材料.pdf",
            body=b"%PDF-1.4\nother tenant\n%%EOF\n",
            tenant_id=OTHER_TENANT_ID,
            owner=9999,
        )
        legacy_student = db.get(GraduationStudent, legacy_student_id)
        legacy_proposal = GraduationProposal(
            tenant_id=TENANT_ID,
            gd_student_id=int(legacy_student.id),
            version="v1",
            is_resubmit=False,
            submit_at=datetime.utcnow(),
            background="旧背景",
            plan="旧计划",
            outcome="旧成果",
            attachments_json=[int(legacy_file.id)],
            status="APPROVED",
            active_key=None,
        )
        db.add(legacy_proposal)
        proposal_v2_row = db.get(GraduationProposal, p2_id)
        proposal_v2_row.defense_result = "PASS"
        proposal_v2_row.defense_comment = "开题答辩通过"
        proposal_v2_row.defense_at = datetime.utcnow()
        db.get(GraduationStudent, student_id).stage = "FINAL_CHECK"
        db.commit()

        draft_file_id = int(draft_file.id)
        final_file_id = int(final_file.id)
        pending_file_id = int(pending_file.id)
        infected_file_id = int(infected_file.id)
        clean_design_id = int(clean_design.id)
        clean_source_id = int(clean_source.id)
        cross_tenant_file_id = int(cross_tenant_file.id)
        legacy_proposal_id = int(legacy_proposal.id)
    finally:
        db.close()

    set_current_user(student_user)
    draft = center.submit_final(student_user, {"finalType": "初稿", "attachments": [draft_file_id]})
    draft_id = int(draft["id"])
    set_current_user(teacher_user)
    draft_detail = center.final_detail(draft_id)
    assert draft_detail["reviewReady"] is True
    center.review_final(draft_id, "APPROVE", "初稿结构完整，同意通过当前安全版本", teacher_user)

    set_current_user(student_user)
    final = center.submit_final(student_user, {"finalType": "定稿", "attachments": [final_file_id]})
    final_id = int(final["id"])
    db = get_sessionmaker()()
    try:
        db.add(GraduationPlagiarismCheck(
            tenant_id=TENANT_ID,
            gd_student_id=student_id,
            gd_final_id=final_id,
            submit_at=datetime.utcnow(),
            status="DONE",
            active_key=None,
            rate="8.5%",
            threshold=30,
            over_threshold=False,
            dispute_status="NONE",
        ))
        db.add(GraduationReview(
            tenant_id=TENANT_ID,
            gd_student_id=student_id,
            gd_final_id=final_id,
            reviewer_name="阶段六评阅教师",
            status="COMPLETED",
            score=88,
            opinion="评阅通过",
            reviewed_at=datetime.utcnow(),
        ))
        db.add(GraduationDefenseScore(
            tenant_id=TENANT_ID,
            gd_student_id=student_id,
            defense_group_id=77001,
            judge_name="阶段六答辩评委",
            status="SCORED",
            score=90,
            comment="答辩通过",
            round_no=1,
        ))
        db.commit()
    finally:
        db.close()
    set_current_user(teacher_user)
    final_detail = center.final_detail(final_id)
    assert final_detail["reviewReady"] is True
    center.review_final(final_id, "APPROVE", "定稿文件、查重和评阅材料均通过", teacher_user)

    set_current_user(student_user)
    _expect_blocked(
        lambda: catalog.submit_material(student_id, "DESIGN_WORK", pending_file_id, 0, student_user),
        codes={"DATA_CONFLICT"},
    )
    _expect_blocked(
        lambda: catalog.submit_material(student_id, "SOURCE_CODE", infected_file_id, 0, student_user),
        codes={"DATA_CONFLICT"},
    )
    _expect_not_found(
        lambda: catalog.submit_material(student_id, "DESIGN_WORK", cross_tenant_file_id, 0, student_user)
    )
    design = catalog.submit_material(student_id, "DESIGN_WORK", clean_design_id, 0, student_user)
    source = catalog.submit_material(student_id, "SOURCE_CODE", clean_source_id, 0, student_user)
    assert design["version"] == 1 and source["version"] == 1
    set_current_user(teacher_user)
    catalog.review_material(
        design["materialId"], "APPROVE", "设计作品完整", design["fileVersionId"], teacher_user,
    )
    catalog.review_material(
        source["materialId"], "APPROVE", "源代码归档完整", source["fileVersionId"], teacher_user,
    )

    set_current_user(admin_user)
    backfill_1 = catalog.backfill_legacy_attachments(
        admin_user,
        batch_id=batch_id,
        page_size=1,
        dry_run=False,
        checkpoint_key="phase6-backfill",
    )
    backfill_2 = catalog.backfill_legacy_attachments(
        admin_user,
        batch_id=batch_id,
        page_size=20,
        dry_run=False,
        checkpoint_key="phase6-backfill",
    )
    assert backfill_1["processed"] <= 1
    assert backfill_2["checkpoint"]["status"] in {"COMPLETED", "PARTIAL_FAILED"}
    repeat = catalog.backfill_legacy_attachments(
        admin_user,
        batch_id=batch_id,
        page_size=20,
        dry_run=False,
        checkpoint_key="phase6-backfill-repeat",
    )
    assert repeat["createdBindings"] >= 0
    db = get_sessionmaker()()
    try:
        legacy_bindings = db.scalars(select(FileBinding).where(
            FileBinding.tenant_id == TENANT_ID,
            FileBinding.biz_type == "GRADUATION_MATERIAL",
            FileBinding.biz_id == str(legacy_proposal_id),
            FileBinding.is_deleted.is_(False),
        )).all()
        assert len(legacy_bindings) == len({(row.asset_id, row.version_id, row.file_id) for row in legacy_bindings})
        checkpoints = db.scalars(select(GraduationMaterialBackfillCheckpoint).where(
            GraduationMaterialBackfillCheckpoint.tenant_id == TENANT_ID,
            GraduationMaterialBackfillCheckpoint.batch_id == batch_id,
        )).all()
        assert checkpoints
    finally:
        db.close()

    set_current_user(admin_user)
    template_file_v1 = None
    template_file_v2 = None
    db = get_sessionmaker()()
    try:
        template_file_v1 = _real_file(
            db,
            name="阶段六开题报告模板.docx",
            body=b"PK\x03\x04docx-template-v1",
            owner=6201,
        )
        template_file_v1.ext = "docx"
        template_file_v1.mime_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        template_file_v1.biz_type = "GRADUATION_TEMPLATE"
        template_file_v1.biz_id = str(template_id)
        template_file_v2 = _real_file(
            db,
            name="阶段六开题报告模板-v2.docx",
            body=b"PK\x03\x04docx-template-v2",
            owner=6201,
        )
        template_file_v2.ext = "docx"
        template_file_v2.mime_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        template_file_v2.biz_type = "GRADUATION_TEMPLATE"
        template_file_v2.biz_id = str(template_id)
        db.commit()
        template_file_v1_id = int(template_file_v1.id)
        template_file_v2_id = int(template_file_v2.id)
    finally:
        db.close()
    policy_v1 = catalog.publish_template_version(
        template_id,
        template_file_v1_id,
        admin_user,
        batch_id=batch_id,
        template_code="GD_PROPOSAL_REPORT",
        variable_schema={"variables": [{"name": "studentName", "type": "string"}]},
    )
    policy_v2 = catalog.publish_template_version(
        template_id,
        template_file_v2_id,
        admin_user,
        batch_id=batch_id,
        template_code="GD_PROPOSAL_REPORT",
        variable_schema={"variables": [{"name": "studentName", "type": "string"}]},
    )
    assert policy_v1["versionNo"] == 1
    assert policy_v2["versionNo"] == 2
    db = get_sessionmaker()()
    try:
        policy = db.scalars(select(GraduationTemplateAssetPolicy).where(
            GraduationTemplateAssetPolicy.tenant_id == TENANT_ID,
            GraduationTemplateAssetPolicy.template_code == "GD_PROPOSAL_REPORT",
            GraduationTemplateAssetPolicy.batch_id == batch_id,
            GraduationTemplateAssetPolicy.is_deleted.is_(False),
        )).one()
        assert int(policy.current_version_id) == int(policy_v2["fileVersionId"])
        versions = db.scalars(select(FileVersion).where(
            FileVersion.tenant_id == TENANT_ID,
            FileVersion.asset_id == int(policy.asset_id),
            FileVersion.is_deleted.is_(False),
        ).order_by(FileVersion.version_no)).all()
        assert len(versions) == 2
        assert versions[0].is_current is False and versions[1].is_current is True
    finally:
        db.close()

    set_current_user(admin_user)
    structured_snapshots.prepare_all(student_id, admin_user)
    set_current_user(teacher_user)
    manifest_v1 = export_service.freeze_manifest(student_id, f"GD6-ARCHIVE-{batch_id}", teacher_user)
    assert manifest_v1["revision"] == 1
    assert manifest_v1["itemCount"] > 0
    manifest_item_version_ids = {item["fileVersionId"] for item in manifest_v1["items"]}
    assert all(item["scanResult"] in {"CLEAN", "NOT_REQUIRED"} for item in manifest_v1["items"])
    assert all(item["reviewStatus"] in {"APPROVED", "NOT_REQUIRED"} for item in manifest_v1["items"])

    export_job = export_service.create_export_job(
        teacher_user,
        scope_type="STUDENT",
        scope_id=student_id,
        export_format="ZIP_XLSX",
        batch_id=batch_id,
    )
    completed = export_service.run_export_job(int(export_job["jobId"]), teacher_user)
    assert completed["status"] == "SUCCEEDED"
    assert completed["rowCount"] == manifest_v1["itemCount"]
    assert completed["result"]["fileCount"] == manifest_v1["itemCount"]
    zip_file_id = int(completed["result"]["zipFileId"])
    xlsx_file_id = int(completed["result"]["xlsxFileId"])
    zip_row, zip_names, zip_entries, zip_manifest = _zip_payload(zip_file_id)
    assert "manifest.json" in zip_names and "档案清单.xlsx" in zip_names
    material_names = [name for name in zip_names if name not in {"manifest.json", "档案清单.xlsx"}]
    assert len(material_names) == manifest_v1["itemCount"]
    assert zip_manifest["fileCount"] == manifest_v1["itemCount"]
    for item in manifest_v1["items"]:
        matching = [name for name in material_names if name.endswith(item["fileName"])]
        assert len(matching) == 1
        assert hashlib.sha256(zip_entries[matching[0]]).hexdigest() == item["sha256"]
    assert hashlib.sha256(get_backend().open(zip_row.file_key).read()).hexdigest() == completed["result"]["zipSha256"]

    _, xlsx_bytes = _read_file(xlsx_file_id)
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) - 1 == manifest_v1["itemCount"]
    headers = list(rows[0])
    assert headers == export_service.XLSX_HEADERS
    assert any("'=" in str(value) or "阶段六" in str(value) for row in rows[1:] for value in row)

    db = get_sessionmaker()()
    try:
        frozen = db.get(ArchiveManifest, int(manifest_v1["manifestId"]))
        persisted_items = db.scalars(select(ArchiveManifestItem).where(
            ArchiveManifestItem.tenant_id == TENANT_ID,
            ArchiveManifestItem.manifest_id == frozen.id,
            ArchiveManifestItem.is_deleted.is_(False),
        )).all()
        assert len(persisted_items) == manifest_v1["itemCount"]
        assert {str(item.version_id) for item in persisted_items} == manifest_item_version_ids
        assert all(item.sha256_snapshot and len(item.sha256_snapshot) == 64 for item in persisted_items)
        job_row = db.get(ExportJob, int(export_job["jobId"]))
        assert job_row.status == "SUCCEEDED"
    finally:
        db.close()

    ticket = tickets.issue_export_ticket(int(export_job["jobId"]), teacher_user)
    resolved = tickets.consume_export_ticket(ticket["ticket"], teacher_user)
    assert int(resolved["fileId"]) == zip_file_id
    revoked = export_service.revoke_manifest(int(manifest_v1["manifestId"]), "阶段六撤销重归档验收", teacher_user)
    assert revoked["status"] == "REVOKED"
    _expect_not_found(lambda: tickets.consume_export_ticket(ticket["ticket"], teacher_user))
    db = get_sessionmaker()()
    try:
        zip_row_db = db.get(FileObject, zip_file_id)
        xlsx_row_db = db.get(FileObject, xlsx_file_id)
        assert zip_row_db.status == "INVALIDATED" and xlsx_row_db.status == "INVALIDATED"
        assert db.get(ExportJob, int(export_job["jobId"])).status == "REVOKED"
    finally:
        db.close()

    set_current_user(admin_user)
    structured_snapshots.prepare_all(student_id, admin_user)
    set_current_user(teacher_user)
    manifest_v2 = export_service.freeze_manifest(student_id, f"GD6-ARCHIVE-{batch_id}-R2", teacher_user)
    assert manifest_v2["revision"] == 2
    assert manifest_v2["manifestId"] != manifest_v1["manifestId"]

    db = get_sessionmaker()()
    try:
        assets = db.scalars(select(FileAsset).where(
            FileAsset.tenant_id == TENANT_ID,
            FileAsset.module_code == "GRADUATION",
            FileAsset.is_deleted.is_(False),
        )).all()
        assert assets
        assert all(asset.current_version_id for asset in assets)
        current_counts = db.execute(select(
            FileVersion.asset_id,
            func.sum(FileVersion.is_current == True),  # noqa: E712
        ).where(
            FileVersion.tenant_id == TENANT_ID,
            FileVersion.is_deleted.is_(False),
        ).group_by(FileVersion.asset_id)).all()
        assert all(int(count or 0) <= 1 for _asset_id, count in current_counts)
        manifests = db.scalars(select(ArchiveManifest).where(
            ArchiveManifest.tenant_id == TENANT_ID,
            ArchiveManifest.module_code == "GRADUATION",
            ArchiveManifest.target_id == str(student_id),
            ArchiveManifest.is_deleted.is_(False),
        ).order_by(ArchiveManifest.revision)).all()
        assert [row.revision for row in manifests][-2:] == [1, 2]
    finally:
        db.close()

    print(json.dumps({
        "phase": 6,
        "rules": 18,
        "proposalV1": p1_id,
        "proposalV2": p2_id,
        "finalVersionId": final_detail["currentSafeVersions"][0]["versionId"],
        "manifestRevision1": manifest_v1["manifestId"],
        "manifestRevision2": manifest_v2["manifestId"],
        "exportJob": export_job["jobId"],
        "zipFileId": zip_file_id,
        "xlsxFileId": xlsx_file_id,
        "status": "PASS",
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
