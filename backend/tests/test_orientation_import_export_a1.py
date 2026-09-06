"""A1 precise contract: real orientation xlsx import/export with canonical auth and scope reuse."""
from __future__ import annotations

import base64
import io

from openpyxl import load_workbook

TID = 1000000000000000001


def _headers(role: str) -> dict[str, str]:
    from app.core.security import create_access_token

    token = create_access_token({
        "userId": f"a1-{role}",
        "realName": f"A1-{role}",
        "userType": "TEACHER",
        "tid": "a1",
        "tenantId": str(TID),
        "activeContextId": "a1-context",
        "currentRoleCode": role,
        "clientType": "PC",
    })
    return {"Authorization": f"Bearer {token}"}


def _authority() -> dict[str, int]:
    from app.db.session import get_sessionmaker
    from app.models import College, Major, OrientationBatch, SchoolClass
    from app.services.orientation_flow_service import ensure_published_flow_version

    db = get_sessionmaker()()
    try:
        college = College(tenant_id=TID, college_name="A1信息学院", code="A1-COL", status="ACTIVE")
        db.add(college)
        db.flush()
        major = Major(tenant_id=TID, college_id=college.id, major_name="A1软件专业",
                      code="A1-MAJ", status="ACTIVE")
        db.add(major)
        db.flush()
        school_class = SchoolClass(
            tenant_id=TID, major_id=major.id, class_name="A1软件2601班",
            class_code="A1-CLS", grade="2026", status="ACTIVE",
        )
        flow_version = ensure_published_flow_version(db, TID)
        batch = OrientationBatch(
            tenant_id=TID, batch_name="A1 2026迎新", batch_no="A1-ORI-2026",
            year="2026", status="ACTIVE", planned_count=1,
            flow_version_id=flow_version.id,
        )
        db.add_all([school_class, batch])
        db.commit()
        return {"batch": batch.id, "college": college.id, "major": major.id, "class": school_class.id}
    finally:
        db.close()


def _xlsx(rows: list[list[str]]) -> bytes:
    from app.services.xlsx_util import build_template_xlsx

    return build_template_xlsx(
        ["迎新批次编号", "录取编号", "候选人编号", "姓名", "学院代码", "专业代码", "班级代码", "录取类型"],
        samples=rows,
        required=["迎新批次编号", "录取编号", "姓名", "学院代码", "专业代码", "班级代码"],
    )


def test_orientation_real_file_import_export_roundtrip(client, db_mode):
    headers = _headers("SCHOOL_ADMIN")
    ids = _authority()

    template_response = client.get(
        "/api/v1/import/domain/orientation/template",
        headers=headers,
    )
    assert template_response.status_code == 200
    template = template_response.json()["data"]
    assert template["filename"].endswith(".xlsx")
    workbook = load_workbook(io.BytesIO(base64.b64decode(template["contentBase64"])), read_only=True)
    assert workbook["导入模板"]["A1"].value == "迎新批次编号 *"
    assert workbook["导入模板"]["B1"].value == "录取编号 *"
    workbook.close()

    validate_response = client.post(
        "/api/v1/import/domain/orientation/validate-file",
        headers=headers,
        files={
            "file": (
                "orientation-a1.xlsx",
                _xlsx([["A1-ORI-2026", "A1-LQ-0001", "A1-CAND-0001", "A1测试新生",
                        "A1-COL", "A1-MAJ", "A1-CLS", "统招"]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validate_response.status_code == 200
    preview = validate_response.json()["data"]
    assert preview["status"] == "DRY_RUN_PASSED"
    assert preview["totalRows"] == preview["okRows"] == 1
    assert preview["errorRows"] == 0

    confirm_response = client.post(
        "/api/v1/import/domain/confirm",
        headers={**headers, "Idempotency-Key": "a1-orientation-import-confirm"},
        json={"domain": "orientation", "batchNo": preview["batchNo"]},
    )
    assert confirm_response.status_code == 200
    receipt = confirm_response.json()["data"]
    assert receipt["status"] == "SUCCESS"
    assert receipt["insertedRows"] == 1

    list_response = client.get(
        f"/api/v1/orientation/students?batchId={ids['batch']}", headers=headers,
    )
    assert list_response.status_code == 200
    assert [row["admissionNo"] for row in list_response.json()["data"]["items"]] == ["A1-LQ-0001"]

    export_response = client.post(
        "/api/v1/export/domain/orientation",
        headers={**headers, "Idempotency-Key": "a1-orientation-export"},
        json={"purpose": "A1迎新导出验收", "batchId": ids["batch"]},
    )
    assert export_response.status_code == 200
    task = export_response.json()["data"]
    assert task["status"] == "SUCCESS"
    assert task["rowCount"] == 1

    download_response = client.get(
        f"/api/v1/export/tasks/{task['taskId']}/download",
        headers=headers,
    )
    assert download_response.status_code == 200
    assert download_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    exported = load_workbook(io.BytesIO(download_response.content), read_only=True)
    sheet = exported.active
    assert "A1迎新导出验收" in sheet["A1"].value
    assert sheet["A3"].value == "A1-ORI-2026"
    assert sheet["B3"].value == "A1测试新生"
    exported.close()


def test_orientation_import_export_rejects_unrelated_dorm_role(client, db_mode):
    headers = _headers("DORM_MANAGER")
    assert client.get(
        "/api/v1/import/domain/orientation/template", headers=headers,
    ).status_code == 403
    assert client.post(
        "/api/v1/export/domain/orientation",
        headers=headers,
        json={"purpose": "无权导出验收"},
    ).status_code == 403
