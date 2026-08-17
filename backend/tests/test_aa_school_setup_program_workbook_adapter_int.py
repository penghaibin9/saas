"""INT six-sheet Program workbook adapter contract."""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
import pytest

from app.core.exceptions import AppException


def _adapter():
    from app.modules.academic_affairs.services import academic_affairs_school_setup_program_workbook_adapter as adapter
    return adapter


def test_program_template_is_real_xlsx_with_exact_six_business_sheets_and_notes():
    adapter = _adapter()
    content = adapter.build_program_import_template()
    assert content.startswith(b"PK")

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True, keep_links=False)
    try:
        assert wb.sheetnames == [
            "培养方案",
            "方案课程",
            "学分要求",
            "实践环节",
            "毕业要求",
            "适用范围",
            "填写说明",
        ]
        assert wb["培养方案"]["A1"].value == "培养方案系列键 *"
        assert wb["方案课程"]["G1"].value == "编班方式 *"
        assert wb["适用范围"]["E1"].value == "绑定范围 *"
        assert "programSeriesKey" in str(wb["填写说明"]["A2"].value)
    finally:
        wb.close()


def test_template_round_trip_returns_all_six_groups_and_frozen_normalized_rows():
    adapter = _adapter()
    grouped, normalized = adapter.parse_and_normalize_program_workbook(
        adapter.build_program_import_template()
    )

    assert list(grouped) == [
        "MAIN",
        "COURSE",
        "CREDIT_REQUIREMENT",
        "PRACTICE",
        "GRADUATION",
        "BINDING",
    ]
    assert {group: len(rows) for group, rows in grouped.items()} == {
        "MAIN": 1,
        "COURSE": 1,
        "CREDIT_REQUIREMENT": 1,
        "PRACTICE": 1,
        "GRADUATION": 1,
        "BINDING": 1,
    }
    assert len(normalized) == 6
    assert [row["logicalGroup"] for row in normalized] == list(grouped)
    assert all(row["programKey"] == "SERIES:SOFTWARE-2026:v1" for row in normalized)
    course = next(row for row in normalized if row["logicalGroup"] == "COURSE")
    assert course["payload"]["courseKey"] == "CS101@v1"
    assert course["payload"]["formationMode"] == "ADMIN_FIXED"


def test_unknown_required_header_is_rejected_instead_of_silently_dropping_column():
    adapter = _adapter()
    wb = load_workbook(BytesIO(adapter.build_program_import_template()))
    wb["方案课程"]["C1"] = "课程代号 *"
    output = BytesIO()
    wb.save(output)
    wb.close()

    with pytest.raises(AppException) as exc:
        adapter.read_program_workbook(output.getvalue())
    assert exc.value.code == "VALIDATION_ERROR"
    assert "缺少必填表头：课程代码" in exc.value.message
    assert "存在未知表头：课程代号" in exc.value.message


def test_missing_or_unexpected_business_sheet_fails_closed():
    adapter = _adapter()
    wb = load_workbook(BytesIO(adapter.build_program_import_template()))
    wb["毕业要求"].title = "毕业要求旧版"
    output = BytesIO()
    wb.save(output)
    wb.close()

    with pytest.raises(AppException) as exc:
        adapter.read_program_workbook(output.getvalue())
    assert exc.value.code == "VALIDATION_ERROR"
    assert "缺少工作表：毕业要求" in exc.value.message
    assert "存在未知工作表：毕业要求旧版" in exc.value.message


def test_workbook_adapter_reuses_shared_xlsx_package_security_owner():
    import inspect

    source = inspect.getsource(_adapter())
    assert "xlsx_util.validate_xlsx_package(file_bytes)" in source
    assert "zipfile.ZipFile" not in source
    assert "read_safe_upload" not in source
    assert "FileObject" in source  # docstring explicitly declares non-ownership
    assert "ImportJob" in source
