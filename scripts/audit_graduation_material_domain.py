#!/usr/bin/env python3
"""Read-only MySQL audit for the graduation material domain.

The command intentionally owns no repair path.  It starts a read-only
transaction, calculates full-scope counters, and paginates the evidence rows.
JSON is written to stdout by default; XLSX output requires ``--output``.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.engine import Connection


ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))

REQUIRED_TABLES = {
    "t_gd_student",
    "t_gd_material_rule",
    "t_gd_material_item",
    "t_gd_student_material",
    "t_file_object",
    "t_file_asset",
    "t_file_version",
    "t_file_binding",
    "t_archive_manifest",
    "t_archive_manifest_item",
    "t_gd_proposal",
    "t_gd_final",
    "t_gd_guidance",
    "t_gd_topic",
}

ACTIVE_MANIFEST_STATUSES = "('PREPARED','FROZEN','PACKAGED')"
SAFE_SCAN_STATUSES = "('CLEAN','PASSED','NOT_REQUIRED')"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="只读审计毕业设计材料业务域")
    parser.add_argument("--tenant-id", type=int, required=True, help="租户 ID")
    parser.add_argument("--batch-id", type=int, help="可选批次 ID")
    parser.add_argument("--page", type=int, default=1, help="问题证据页码")
    parser.add_argument("--page-size", type=int, default=200, help="每页问题证据，最大 2000")
    parser.add_argument("--format", choices=("json", "xlsx"), default="json")
    parser.add_argument("--output", type=Path, help="输出路径；JSON 未指定时写 stdout")
    parser.add_argument(
        "--database-url",
        help="MySQL SQLAlchemy URL；也可使用 AUDIT_DATABASE_URL/DATABASE_URL/TEST_DATABASE_URL",
    )
    return parser.parse_args()


def database_url(args: argparse.Namespace) -> str:
    value = (
        args.database_url
        or os.getenv("AUDIT_DATABASE_URL")
        or os.getenv("DATABASE_URL")
        or os.getenv("TEST_DATABASE_URL")
        or ""
    ).strip()
    if not value.startswith("mysql+"):
        raise SystemExit("审计仅允许真实 MySQL：请提供 mysql+... DATABASE_URL")
    return value


def scalar(conn: Connection, sql: str, params: dict[str, Any]) -> int:
    return int(conn.execute(text(sql), params).scalar() or 0)


def scope(args: argparse.Namespace, alias: str = "s") -> tuple[str, dict[str, Any]]:
    params: dict[str, Any] = {"tenant_id": args.tenant_id}
    clause = (
        f"{alias}.tenant_id=:tenant_id AND {alias}.record_status='ACTIVE' "
        f"AND {alias}.is_deleted=0"
    )
    if args.batch_id:
        clause += f" AND {alias}.batch_id=:batch_id"
        params["batch_id"] = args.batch_id
    return clause, params


def current_rule_join(student_alias: str = "s", rule_alias: str = "r") -> str:
    return f"""
      JOIN t_gd_material_rule {rule_alias}
        ON {rule_alias}.tenant_id={student_alias}.tenant_id
       AND {rule_alias}.batch_id={student_alias}.batch_id
       AND {rule_alias}.status='ENABLED' AND {rule_alias}.enabled=1
       AND {rule_alias}.is_deleted=0
       AND {rule_alias}.rule_version=(
         SELECT MAX(r2.rule_version) FROM t_gd_material_rule r2
          WHERE r2.tenant_id={student_alias}.tenant_id
            AND r2.batch_id={student_alias}.batch_id
            AND r2.status='ENABLED' AND r2.enabled=1 AND r2.is_deleted=0
       )
    """


def metric_queries(student_scope: str) -> dict[str, str]:
    batch_rule = " AND batch_id=:batch_id" if ":batch_id" in student_scope else ""
    return {
        "activeGraduationStudents": f"SELECT COUNT(*) FROM t_gd_student s WHERE {student_scope}",
        "currentRules": f"""
          SELECT COUNT(*) FROM t_gd_material_rule
           WHERE tenant_id=:tenant_id{batch_rule}
             AND status='ENABLED' AND enabled=1 AND is_deleted=0
        """,
        "expectedMaterialRows": f"""
          SELECT COUNT(*) FROM t_gd_student s
          {current_rule_join()}
          JOIN t_gd_material_item i ON i.tenant_id=s.tenant_id AND i.rule_id=r.id
            AND i.enabled=1 AND i.is_deleted=0
          WHERE {student_scope}
        """,
        "actualMaterialRows": f"""
          SELECT COUNT(*) FROM t_gd_student_material m
          JOIN t_gd_student s ON s.id=m.gd_student_id AND s.tenant_id=m.tenant_id
          WHERE {student_scope} AND m.is_deleted=0
        """,
        "missingMaterialCatalog": f"""
          SELECT COUNT(*) FROM t_gd_student s
          {current_rule_join()}
          JOIN t_gd_material_item i ON i.tenant_id=s.tenant_id AND i.rule_id=r.id
            AND i.enabled=1 AND i.is_deleted=0
          LEFT JOIN t_gd_student_material m ON m.tenant_id=s.tenant_id
            AND m.batch_id=s.batch_id AND m.gd_student_id=s.id
            AND m.material_code=i.material_code AND m.is_deleted=0
          WHERE {student_scope} AND m.id IS NULL
        """,
        "duplicateMaterials": f"""
          SELECT COUNT(*) FROM (
            SELECT m.gd_student_id,m.material_code
              FROM t_gd_student_material m
              JOIN t_gd_student s ON s.id=m.gd_student_id AND s.tenant_id=m.tenant_id
             WHERE {student_scope} AND m.is_deleted=0
             GROUP BY m.tenant_id,m.batch_id,m.gd_student_id,m.material_code HAVING COUNT(*)>1
          ) duplicate_groups
        """,
        "orphanFileAssets": f"""
          SELECT COUNT(*) FROM t_file_asset a
           WHERE a.tenant_id=:tenant_id AND a.owner_type='GRADUATION_STUDENT_MATERIAL'
             AND a.is_deleted=0
             AND NOT EXISTS (SELECT 1 FROM t_gd_student_material m
                              WHERE m.tenant_id=a.tenant_id AND m.asset_id=a.id AND m.is_deleted=0)
             AND NOT EXISTS (SELECT 1 FROM t_file_binding b
                              WHERE b.tenant_id=a.tenant_id AND b.asset_id=a.id AND b.is_deleted=0)
        """,
        "orphanFileVersions": f"""
          SELECT COUNT(*) FROM t_file_version v
          JOIN t_file_asset a ON a.id=v.asset_id AND a.tenant_id=v.tenant_id
           WHERE v.tenant_id=:tenant_id AND a.owner_type='GRADUATION_STUDENT_MATERIAL'
             AND v.is_deleted=0
             AND NOT EXISTS (SELECT 1 FROM t_file_binding b
                              WHERE b.tenant_id=v.tenant_id AND b.version_id=v.id AND b.is_deleted=0)
             AND NOT EXISTS (SELECT 1 FROM t_gd_student_material m
                              WHERE m.tenant_id=v.tenant_id AND m.current_version_id=v.id AND m.is_deleted=0)
             AND NOT EXISTS (SELECT 1 FROM t_archive_manifest_item mi
                              WHERE mi.tenant_id=v.tenant_id AND mi.version_id=v.id AND mi.is_deleted=0)
        """,
        "orphanFileBindings": """
          SELECT COUNT(*) FROM t_file_binding b
          LEFT JOIN t_file_object f ON f.id=b.file_id AND f.tenant_id=b.tenant_id AND f.is_deleted=0
          LEFT JOIN t_file_asset a ON a.id=b.asset_id AND a.tenant_id=b.tenant_id AND a.is_deleted=0
          LEFT JOIN t_file_version v ON v.id=b.version_id AND v.tenant_id=b.tenant_id AND v.is_deleted=0
           WHERE b.tenant_id=:tenant_id AND UPPER(COALESCE(b.module_code,''))='GRADUATION'
             AND b.is_deleted=0 AND (f.id IS NULL OR a.id IS NULL OR v.id IS NULL)
        """,
        "currentVersionMismatches": f"""
          SELECT COUNT(*) FROM t_gd_student_material m
          JOIN t_gd_student s ON s.id=m.gd_student_id AND s.tenant_id=m.tenant_id
          LEFT JOIN t_file_asset a ON a.id=m.asset_id AND a.tenant_id=m.tenant_id AND a.is_deleted=0
          LEFT JOIN t_file_version v ON v.id=m.current_version_id AND v.tenant_id=m.tenant_id AND v.is_deleted=0
           WHERE {student_scope} AND m.is_deleted=0 AND m.current_version_id IS NOT NULL
             AND (a.id IS NULL OR v.id IS NULL OR a.current_version_id<>m.current_version_id
                  OR v.asset_id<>m.asset_id OR v.is_current<>1)
        """,
        "unmigratedAttachmentsJson": """
          SELECT SUM(problem_count) FROM (
            SELECT COUNT(*) problem_count FROM t_gd_proposal p
             WHERE p.tenant_id=:tenant_id AND p.is_deleted=0 AND JSON_LENGTH(p.attachments_json)>0
               AND NOT EXISTS (SELECT 1 FROM t_file_binding b WHERE b.tenant_id=p.tenant_id
                 AND UPPER(COALESCE(b.module_code,''))='GRADUATION' AND b.biz_id=CAST(p.id AS CHAR)
                 AND b.relation_type='GRADUATION_PROPOSAL_MATERIAL' AND b.is_deleted=0)
            UNION ALL
            SELECT COUNT(*) FROM t_gd_final f
             WHERE f.tenant_id=:tenant_id AND f.is_deleted=0 AND JSON_LENGTH(f.attachments_json)>0
               AND NOT EXISTS (SELECT 1 FROM t_file_binding b WHERE b.tenant_id=f.tenant_id
                 AND UPPER(COALESCE(b.module_code,''))='GRADUATION' AND b.biz_id=CAST(f.id AS CHAR)
                 AND b.relation_type='GRADUATION_FINAL_MATERIAL' AND b.is_deleted=0)
          ) legacy_attachment_counts
        """,
        "legacyDynamicMaterialCodes": f"""
          SELECT COUNT(*) FROM t_gd_student_material m
          JOIN t_gd_student s ON s.id=m.gd_student_id AND s.tenant_id=m.tenant_id
          LEFT JOIN t_gd_material_item i ON i.tenant_id=m.tenant_id AND i.rule_id=m.rule_id
            AND i.material_code=m.material_code AND i.enabled=1 AND i.is_deleted=0
           WHERE {student_scope} AND m.is_deleted=0 AND
             (i.id IS NULL OR m.material_code REGEXP '(_[0-9][0-9]|ATTACHMENT_[0-9]+)$')
        """,
        "v1Manifests": """
          SELECT COUNT(*) FROM t_archive_manifest
           WHERE tenant_id=:tenant_id AND module_code='graduation'
             AND archive_type='GRADUATION_STUDENT_ARCHIVE' AND is_deleted=0
        """,
        "v2Manifests": """
          SELECT COUNT(*) FROM t_archive_manifest
           WHERE tenant_id=:tenant_id AND module_code='GRADUATION'
             AND archive_type='GRADUATION_FILE_VERSION' AND is_deleted=0
        """,
        "duplicateActiveManifests": """
          SELECT COUNT(*) FROM (
            SELECT target_id FROM t_archive_manifest
             WHERE tenant_id=:tenant_id AND module_code='GRADUATION'
               AND archive_type='GRADUATION_FILE_VERSION'
               AND status IN ('PREPARED','FROZEN','PACKAGED') AND is_deleted=0
             GROUP BY target_id HAVING COUNT(*)>1
          ) duplicate_manifests
        """,
        "archivedVersionsInvalidated": """
          SELECT COUNT(DISTINCT v.id) FROM t_archive_manifest_item mi
          JOIN t_archive_manifest am ON am.id=mi.manifest_id AND am.tenant_id=mi.tenant_id
          JOIN t_file_version v ON v.id=mi.version_id AND v.tenant_id=mi.tenant_id
           WHERE mi.tenant_id=:tenant_id AND mi.is_deleted=0 AND am.is_deleted=0
             AND am.module_code='GRADUATION' AND am.archive_type='GRADUATION_FILE_VERSION'
             AND am.status IN ('FROZEN','PACKAGED','REVOKED','SUPERSEDED')
             AND v.status='INVALIDATED'
        """,
        "ruleMaterialMismatches": f"""
          SELECT COUNT(*) FROM t_gd_student_material m
          JOIN t_gd_student s ON s.id=m.gd_student_id AND s.tenant_id=m.tenant_id
          LEFT JOIN t_gd_material_rule r ON r.id=m.rule_id AND r.tenant_id=m.tenant_id
          LEFT JOIN t_gd_material_item i ON i.rule_id=m.rule_id AND i.tenant_id=m.tenant_id
            AND i.material_code=m.material_code AND i.is_deleted=0
           WHERE {student_scope} AND m.is_deleted=0
             AND (r.id IS NULL OR i.id IS NULL OR r.batch_id<>m.batch_id
                  OR r.rule_version<>m.rule_version OR r.status<>'ENABLED' OR r.enabled<>1)
        """,
        "fileHashOrSecurityAnomalies": f"""
          SELECT COUNT(*) FROM t_gd_student_material m
          JOIN t_gd_student s ON s.id=m.gd_student_id AND s.tenant_id=m.tenant_id
          JOIN t_file_version v ON v.id=m.current_version_id AND v.tenant_id=m.tenant_id AND v.is_deleted=0
          LEFT JOIN t_file_object f ON f.id=v.file_object_id AND f.tenant_id=v.tenant_id AND f.is_deleted=0
           WHERE {student_scope} AND m.is_deleted=0
             AND (f.id IS NULL OR COALESCE(f.sha256,'')='' OR f.status<>'AVAILABLE'
                  OR UPPER(COALESCE(f.scan_status,'')) NOT IN {SAFE_SCAN_STATUSES})
        """,
    }


def issue_queries(student_scope: str) -> list[str]:
    return [
        f"""SELECT 'MISSING_MATERIAL_CATALOG' issue_type,'GRADUATION_STUDENT' entity_type,
          CAST(s.id AS CHAR) entity_id,CAST(s.id AS CHAR) gd_student_id,i.material_code,
          CONCAT('rule=',r.id,',version=',r.rule_version) detail
          FROM t_gd_student s {current_rule_join()}
          JOIN t_gd_material_item i ON i.tenant_id=s.tenant_id AND i.rule_id=r.id
            AND i.enabled=1 AND i.is_deleted=0
          LEFT JOIN t_gd_student_material m ON m.tenant_id=s.tenant_id AND m.batch_id=s.batch_id
            AND m.gd_student_id=s.id AND m.material_code=i.material_code AND m.is_deleted=0
          WHERE {student_scope} AND m.id IS NULL""",
        f"""SELECT 'CURRENT_VERSION_MISMATCH' issue_type,'GRADUATION_MATERIAL' entity_type,
          CAST(m.id AS CHAR) entity_id,CAST(m.gd_student_id AS CHAR) gd_student_id,
          m.material_code material_code,
          CONCAT('material=',COALESCE(m.current_version_id,''),',asset=',COALESCE(a.current_version_id,''),
                 ',versionCurrent=',COALESCE(v.is_current,'')) detail
          FROM t_gd_student_material m
          JOIN t_gd_student s ON s.id=m.gd_student_id AND s.tenant_id=m.tenant_id
          LEFT JOIN t_file_asset a ON a.id=m.asset_id AND a.tenant_id=m.tenant_id AND a.is_deleted=0
          LEFT JOIN t_file_version v ON v.id=m.current_version_id AND v.tenant_id=m.tenant_id AND v.is_deleted=0
          WHERE {student_scope} AND m.is_deleted=0 AND m.current_version_id IS NOT NULL
            AND (a.id IS NULL OR v.id IS NULL OR a.current_version_id<>m.current_version_id
                 OR v.asset_id<>m.asset_id OR v.is_current<>1)""",
        """SELECT 'ORPHAN_FILE_ASSET' issue_type,'FILE_ASSET' entity_type,
          CAST(a.id AS CHAR) entity_id,'' gd_student_id,'' material_code,a.asset_code detail
          FROM t_file_asset a WHERE a.tenant_id=:tenant_id
            AND a.owner_type='GRADUATION_STUDENT_MATERIAL' AND a.is_deleted=0
            AND NOT EXISTS (SELECT 1 FROM t_gd_student_material m WHERE m.tenant_id=a.tenant_id
                             AND m.asset_id=a.id AND m.is_deleted=0)
            AND NOT EXISTS (SELECT 1 FROM t_file_binding b WHERE b.tenant_id=a.tenant_id
                             AND b.asset_id=a.id AND b.is_deleted=0)""",
        """SELECT 'ORPHAN_FILE_VERSION' issue_type,'FILE_VERSION' entity_type,
          CAST(v.id AS CHAR) entity_id,'' gd_student_id,'' material_code,
          CONCAT('asset=',v.asset_id,',file=',v.file_object_id) detail
          FROM t_file_version v JOIN t_file_asset a ON a.id=v.asset_id AND a.tenant_id=v.tenant_id
          WHERE v.tenant_id=:tenant_id AND a.owner_type='GRADUATION_STUDENT_MATERIAL'
            AND v.is_deleted=0
            AND NOT EXISTS (SELECT 1 FROM t_file_binding b WHERE b.tenant_id=v.tenant_id
                             AND b.version_id=v.id AND b.is_deleted=0)
            AND NOT EXISTS (SELECT 1 FROM t_gd_student_material m WHERE m.tenant_id=v.tenant_id
                             AND m.current_version_id=v.id AND m.is_deleted=0)
            AND NOT EXISTS (SELECT 1 FROM t_archive_manifest_item mi WHERE mi.tenant_id=v.tenant_id
                             AND mi.version_id=v.id AND mi.is_deleted=0)""",
        """SELECT 'ARCHIVED_VERSION_INVALIDATED' issue_type,'FILE_VERSION' entity_type,
          CAST(v.id AS CHAR) entity_id,'' gd_student_id,mi.material_code material_code,
          CONCAT('manifest=',mi.manifest_id) detail
          FROM t_archive_manifest_item mi
          JOIN t_archive_manifest am ON am.id=mi.manifest_id AND am.tenant_id=mi.tenant_id
          JOIN t_file_version v ON v.id=mi.version_id AND v.tenant_id=mi.tenant_id
          WHERE mi.tenant_id=:tenant_id AND mi.is_deleted=0 AND am.is_deleted=0
            AND am.module_code='GRADUATION' AND am.archive_type='GRADUATION_FILE_VERSION'
            AND am.status IN ('FROZEN','PACKAGED','REVOKED','SUPERSEDED') AND v.status='INVALIDATED'""",
        f"""SELECT 'FILE_HASH_OR_SECURITY' issue_type,'FILE_OBJECT' entity_type,
          CAST(COALESCE(f.id,v.file_object_id) AS CHAR) entity_id,
          CAST(m.gd_student_id AS CHAR) gd_student_id,m.material_code material_code,
          CONCAT('status=',COALESCE(f.status,'MISSING'),',scan=',COALESCE(f.scan_status,'MISSING'),
                 ',sha=',IF(COALESCE(f.sha256,'')='','MISSING','PRESENT')) detail
          FROM t_gd_student_material m
          JOIN t_gd_student s ON s.id=m.gd_student_id AND s.tenant_id=m.tenant_id
          JOIN t_file_version v ON v.id=m.current_version_id AND v.tenant_id=m.tenant_id AND v.is_deleted=0
          LEFT JOIN t_file_object f ON f.id=v.file_object_id AND f.tenant_id=v.tenant_id AND f.is_deleted=0
          WHERE {student_scope} AND m.is_deleted=0
            AND (f.id IS NULL OR COALESCE(f.sha256,'')='' OR f.status<>'AVAILABLE'
                 OR UPPER(COALESCE(f.scan_status,'')) NOT IN {SAFE_SCAN_STATUSES})""",
    ]


def serialize(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def audit(conn: Connection, args: argparse.Namespace) -> dict[str, Any]:
    available = set(inspect(conn).get_table_names())
    missing_tables = sorted(REQUIRED_TABLES - available)
    if missing_tables:
        raise SystemExit("数据库未升级到材料中心基线，缺少表：" + ", ".join(missing_tables))

    student_scope, params = scope(args)
    metrics = {
        name: scalar(conn, sql, params)
        for name, sql in metric_queries(student_scope).items()
    }
    queries = issue_queries(student_scope)
    issue_counts = [
        scalar(conn, f"SELECT COUNT(*) FROM ({query}) issue_rows", params)
        for query in queries
    ]
    issue_total = sum(issue_counts)
    evidence_params = {
        **params,
        "limit": min(2000, max(1, args.page_size)),
        "offset": (max(1, args.page) - 1) * min(2000, max(1, args.page_size)),
    }
    rows: list[Any] = []
    remaining_offset = evidence_params["offset"]
    remaining_limit = evidence_params["limit"]
    for query, query_count in zip(queries, issue_counts):
        if remaining_limit <= 0:
            break
        if remaining_offset >= query_count:
            remaining_offset -= query_count
            continue
        query_params = {
            **params,
            "limit": remaining_limit,
            "offset": remaining_offset,
        }
        page_rows = conn.execute(text(
            f"SELECT * FROM ({query}) issue_rows "
            "ORDER BY entity_type,entity_id LIMIT :limit OFFSET :offset"
        ), query_params).mappings().all()
        rows.extend(page_rows)
        remaining_limit -= len(page_rows)
        remaining_offset = 0
    return {
        "meta": {
            "schemaVersion": "graduation-material-domain-audit-v1",
            "generatedAt": datetime.now().astimezone().isoformat(),
            "readOnly": True,
            "tenantId": str(args.tenant_id),
            "batchId": str(args.batch_id or ""),
            "page": max(1, args.page),
            "pageSize": evidence_params["limit"],
            "issueTotal": issue_total,
        },
        "statistics": metrics,
        "issues": [{key: serialize(value) for key, value in row.items()} for row in rows],
    }


def write_json(report: dict[str, Any], output: Path | None) -> None:
    payload = json.dumps(report, ensure_ascii=False, indent=2) + "\n"
    if output:
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(payload, encoding="utf-8")
    else:
        sys.stdout.write(payload)


def write_xlsx(report: dict[str, Any], output: Path | None) -> None:
    if not output:
        raise SystemExit("--format xlsx 必须同时指定 --output")
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["metric", "value"])
    for key, value in report["statistics"].items():
        summary.append([key, value])
    metadata = workbook.create_sheet("Metadata")
    metadata.append(["key", "value"])
    for key, value in report["meta"].items():
        metadata.append([key, value])
    evidence = workbook.create_sheet("Issues")
    headers = ["issue_type", "entity_type", "entity_id", "gd_student_id", "material_code", "detail"]
    evidence.append(headers)
    for row in report["issues"]:
        evidence.append([row.get(key, "") for key in headers])
    evidence.freeze_panes = "A2"
    evidence.auto_filter.ref = evidence.dimensions
    workbook.save(output)


def main() -> int:
    args = parse_args()
    engine = create_engine(database_url(args), pool_pre_ping=True)
    with engine.connect() as conn:
        if conn.dialect.name != "mysql":
            raise SystemExit("审计仅允许真实 MySQL")
        conn.exec_driver_sql("SET TRANSACTION READ ONLY")
        try:
            report = audit(conn, args)
        finally:
            conn.rollback()
    if args.format == "xlsx":
        write_xlsx(report, args.output)
    else:
        write_json(report, args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
